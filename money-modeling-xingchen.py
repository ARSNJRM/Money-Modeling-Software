"""
Economic System Simulation Module 

This module provides classes and functions for simulating an economic system with various agents,
assets, liabilities, and financial instruments.
"""

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union
from enum import Enum
from datetime import datetime, timedelta
from copy import deepcopy

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl package not found. Excel export functionality will be disabled.")
    print("To enable Excel export, please install openpyxl using: pip install openpyxl")
    EXCEL_AVAILABLE = False

# Core Enums and Classes
class AgentType(Enum):
    BANK = "bank"
    COMPANY = "company"
    HOUSEHOLD = "household"
    TREASURY = "treasury"
    CENTRAL_BANK = "central_bank"
    OTHER = "other"

class EntryType(Enum):
    LOAN = "loan"
    DEPOSIT = "deposit"
    PAYABLE = "payable"
    BOND_ZERO_COUPON = "bond_zero_coupon"
    BOND_COUPON = "bond_coupon"
    BOND_AMORTIZING = "bond_amortizing"
    DELIVERY_CLAIM = "delivery_claim"
    NON_FINANCIAL = "non_financial"
    DEFAULT = "default"

class MaturityType(Enum):
    ON_DEMAND = "on_demand"
    FIXED_DATE = "fixed_date"
    PERPETUAL = "perpetual"

class SettlementType(Enum):
    """Types of settlement for financial instruments"""
    MEANS_OF_PAYMENT = "means_of_payment"
    SECURITIES = "securities"
    NON_FINANCIAL_ASSET = "non_financial_asset"
    SERVICES = "services"
    CRYPTO = "crypto"
    NONE = "none"

class BondType(Enum):
    """Types of bonds"""
    ZERO_COUPON = 0  # Zero-coupon bond
    COUPON = 1      # Regular coupon bond
    AMORTIZING = 2  # Amortizing bond

@dataclass
class SettlementDetails:
    type: SettlementType
    denomination: str  # Currency or unit of settlement

@dataclass
class BalanceSheetEntry:
    type: EntryType
    is_asset: bool  # True for assets, False for liabilities
    counterparty: Optional[str]  # Optional for non-financial entries
    amount: float
    denomination: str
    maturity_type: MaturityType
    maturity_date: Optional[datetime]  # Required for FIXED_DATE, None for others
    settlement_details: SettlementDetails
    name: Optional[str] = None
    issuance_time: str = 't0'
    book_value: Optional[float] = None
    expected_cash_flow: Optional[float] = None
    parent_bond: Optional[str] = None  # Reference to the main bond

    def matches(self, other: 'BalanceSheetEntry') -> bool:
        """Check if two entries match (used for removing entries)"""
        return (
            self.type == other.type and
            self.is_asset == other.is_asset and
            self.counterparty == other.counterparty and
            self.amount == other.amount and
            self.denomination == other.denomination and
            self.maturity_type == other.maturity_type and
            self.maturity_date == other.maturity_date and
            self.settlement_details.type == other.settlement_details.type and
            self.settlement_details.denomination == other.settlement_details.denomination and
            self.name == other.name and
            self.issuance_time == other.issuance_time
        )

    def __post_init__(self):
        if self.amount <= 0:
            raise ValueError("Amount must be positive")

        # Validate issuance time
        if self.issuance_time not in ['t0', 't1', 't2']:
            raise ValueError("Issuance time must be 't0', 't1', or 't2'")

        # Validate counterparty rules
        if self.type != EntryType.NON_FINANCIAL and not self.counterparty:
            raise ValueError("Counterparty is required for financial entries")
        if self.type == EntryType.NON_FINANCIAL and self.counterparty:
            raise ValueError("Non-financial entries cannot have a counterparty")

        # Validate name rules
        if self.type == EntryType.NON_FINANCIAL and not self.name:
            raise ValueError("Non-financial entries must have a name")

        # Validate payable rules
        if self.type == EntryType.PAYABLE and self.settlement_details.type != SettlementType.MEANS_OF_PAYMENT:
            raise ValueError("Payable entries must have means_of_payment settlement type")

class SettlementFailure(Exception):
    def __init__(self, agent_name: str, entry: BalanceSheetEntry, reason: str):
        self.agent_name = agent_name
        self.entry = entry
        self.reason = reason
        super().__init__(f"Settlement failure for {agent_name}: {reason}")

class Agent:
    def __init__(self, name: str, agent_type: AgentType):
        self.name = name
        self.type = agent_type
        self.assets: List[BalanceSheetEntry] = []
        self.liabilities: List[BalanceSheetEntry] = []
        self.status: str = "operating"  # operating or bankrupt
        self.creation_time: datetime = datetime.now()
        # Add settlement history
        self.settlement_history = {
            'as_asset_holder': [],  # Settlements where this agent was the creditor
            'as_liability_holder': []  # Settlements where this agent was the debtor
        }

    def add_asset(self, entry: BalanceSheetEntry):
        self.assets.append(entry)

    def add_liability(self, entry: BalanceSheetEntry):
        self.liabilities.append(entry)

    def remove_asset(self, entry: BalanceSheetEntry):
        self.assets = [e for e in self.assets if not e.matches(entry)]

    def remove_liability(self, entry: BalanceSheetEntry):
        self.liabilities = [e for e in self.liabilities if not e.matches(entry)]

    def get_balance_sheet(self) -> Dict:
        return {
            "assets": self.assets,
            "liabilities": self.liabilities
        }

    def get_total_assets(self) -> float:
        return sum(entry.amount for entry in self.assets)

    def get_total_liabilities(self) -> float:
        return sum(entry.amount for entry in self.liabilities)

    def get_net_worth(self) -> float:
        return self.get_total_assets() - self.get_total_liabilities()

    def get_type_specific_metrics(self) -> Dict:
        metrics = {
            "name": self.name,
            "type": self.type.value,
            "creation_time": self.creation_time,
            "status": self.status,
            "total_assets": self.get_total_assets(),
            "total_liabilities": self.get_total_liabilities(),
            "net_worth": self.get_net_worth()
        }

        if self.type == AgentType.BANK:
            metrics["capital_ratio"] = self.get_total_assets() / self.get_total_liabilities() if self.get_total_liabilities() > 0 else float('inf')
        elif self.type == AgentType.COMPANY:
            metrics["leverage_ratio"] = self.get_total_liabilities() / self.get_total_assets() if self.get_total_assets() > 0 else float('inf')
        elif self.type == AgentType.HOUSEHOLD:
            metrics["savings_rate"] = (self.get_total_assets() - self.get_total_liabilities()) / self.get_total_assets() if self.get_total_assets() > 0 else 0

        return metrics

    def record_settlement(self,
                         time_point: str,
                         original_entry: BalanceSheetEntry,
                         settlement_result: BalanceSheetEntry,
                         counterparty: str,
                         as_asset_holder: bool):
        """Record a settlement in the agent's history"""
        settlement_record = {
            'time_point': time_point,
            'original_entry': deepcopy(original_entry),
            'settlement_result': deepcopy(settlement_result),
            'counterparty': counterparty,
            'timestamp': datetime.now()
        }
        if as_asset_holder:
            self.settlement_history['as_asset_holder'].append(settlement_record)
        else:
            self.settlement_history['as_liability_holder'].append(settlement_record)

    def update_inventory(self, bond_delta, cash_delta):
        """Adjust bond and cash inventory."""
        self.inventory_bond += bond_delta
        self.inventory_cash += cash_delta
        
        # Update trading statistics
        if bond_delta != 0:  # Only count actual trades
            self.trade_count += 1
            self.total_volume += abs(bond_delta)
            self.trade_volumes.append(abs(bond_delta))
            self.trade_prices.append(-cash_delta / bond_delta)  # Price per bond

class AssetLiabilityPair:
    """Represents a pair of corresponding asset and liability entries"""
    def __init__(self, time: datetime, type: str, amount: float,                 denomination: str, maturity_type: MaturityType,
                 maturity_date: Optional[datetime], settlement_type: SettlementType,
                 settlement_denomination: str, asset_holder: Agent,
                 liability_holder: Optional[Agent] = None,
                 asset_name: Optional[str] = None,
                 bond_type: Optional[BondType] = None,
                 coupon_rate: Optional[float] = None):
        self.time = time
        self.type = type
        self.amount = amount
        self.denomination = denomination
        self.maturity_type = maturity_type
        self.maturity_date = maturity_date
        self.settlement_details = SettlementDetails(
            type=settlement_type,
            denomination=settlement_denomination
        )
        self.coupon_rate = coupon_rate
        self.bond_type = bond_type
        self.asset_holder = asset_holder
        self.liability_holder = liability_holder
        self.asset_name = asset_name
        self.initial_book_value = amount  # BV₀
        self.connected_claims = []  # Store related claims

        if type == EntryType.NON_FINANCIAL.value:
            if liability_holder is not None:
                raise ValueError("Non-financial entries cannot have a liability holder")
            if not asset_name:
                raise ValueError("Non-financial entries must have an asset name")
        else:
            if liability_holder is None:
                raise ValueError("Financial entries must have a liability holder")

    def create_entries(self) -> Tuple[BalanceSheetEntry, Optional[BalanceSheetEntry]]:
        """Create corresponding asset and liability entries"""
        # Verify that only banks can hold loans
        if self.type == EntryType.LOAN.value:
            if self.asset_holder.type != AgentType.BANK:
                raise ValueError("Only banks can hold loans as assets")
        
        # Create basic balance sheet entries
        asset_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=True,
            counterparty=self.liability_holder.name if self.liability_holder else None,
            amount=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=self.asset_name,
            book_value=self.initial_book_value,
            expected_cash_flow=self._calculate_expected_cash_flow()
        )

        if self.type == EntryType.NON_FINANCIAL.value:
            return asset_entry, None

        liability_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=False,
            counterparty=self.asset_holder.name,
            amount=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=None,
            book_value=self.initial_book_value,
            expected_cash_flow=self._calculate_expected_cash_flow()
        )

        # Create connected claims for coupon and amortizing bonds
        if self.type in [EntryType.BOND_COUPON.value, EntryType.BOND_AMORTIZING.value]:
            claims = self.create_bond_claims()
            self.connected_claims = claims
            
            # Add claims to asset holder's assets
            for claim in claims:
                self.asset_holder.add_asset(claim)
                
                # Create corresponding liability
                liability = BalanceSheetEntry(
                    type=claim.type,
                    is_asset=False,
                    counterparty=self.asset_holder.name,
                    amount=claim.amount,
                    denomination=claim.denomination,
                    maturity_type=claim.maturity_type,
                    maturity_date=claim.maturity_date,
                    settlement_details=claim.settlement_details,
                    book_value=claim.book_value,
                    expected_cash_flow=claim.expected_cash_flow
                )
                self.liability_holder.add_liability(liability)

        return asset_entry, liability_entry

    def _calculate_expected_cash_flow(self) -> float:
        """Calculate expected cash flow based on bond type"""
        if self.type not in [EntryType.BOND_ZERO_COUPON.value, 
                           EntryType.BOND_COUPON.value, 
                           EntryType.BOND_AMORTIZING.value]:
            return 1.0  # Default cash flow for non-bond entries

        if self.type == EntryType.BOND_ZERO_COUPON.value:
            # Zero-coupon bond: cash flow is the face value at maturity
            return self.amount / self.initial_book_value

        elif self.type == EntryType.BOND_COUPON.value:
            if not self.coupon_rate:
                raise ValueError("Coupon rate is required for coupon bonds")
            
            # For coupon bonds, cash flow includes periodic coupon payments and principal
            t1 = datetime(2050, 1, 1)
            t2 = datetime(2100, 1, 1)
            
            if self.maturity_date == t1:
                # Matures at t1: one coupon payment + principal
                return (self.amount * self.coupon_rate + self.amount) / self.initial_book_value
            else:
                # Matures at t2: two coupon payments + principal
                return (2 * self.amount * self.coupon_rate + self.amount) / self.initial_book_value

        elif self.type == EntryType.BOND_AMORTIZING.value:
            if not self.coupon_rate:
                raise ValueError("Interest rate is required for amortizing bonds")
            
            # For amortizing bonds, cash flow includes principal and interest payments
            t1 = datetime(2050, 1, 1)
            t2 = datetime(2100, 1, 1)
            
            if self.maturity_date == t1:
                # Matures at t1: one payment of principal + interest
                return (self.amount * (1 + self.coupon_rate)) / self.initial_book_value
            else:
                # Matures at t2: two payments of principal/2 + interest
                principal_t1 = self.amount / 2
                principal_t2 = self.amount / 2
                interest_t1 = self.amount * self.coupon_rate
                interest_t2 = principal_t2 * self.coupon_rate
                return (principal_t1 + interest_t1 + principal_t2 + interest_t2) / self.initial_book_value

    def create_bond_claims(self) -> List[BalanceSheetEntry]:
        """Create connected claims for bonds"""
        claims = []
        if self.type in [EntryType.BOND_COUPON.value, EntryType.BOND_AMORTIZING.value]:
            schedule = self._create_bond_payment_schedule()
            
            # Generate a unique identifier for this bond
            bond_id = f"bond_{self.type}_{id(self)}"  # 使用对象的id作为唯一标识符
            
            for date, amount, payment_type in schedule:
                # Calculate book value and expected cash flow
                bv = amount  # This should be calculated based on the formula
                cf = amount / self.initial_book_value  # As portion of BV₀
                
                # Create claim with correct maturity date
                claim = BalanceSheetEntry(
                    type=EntryType.PAYABLE,
                    is_asset=True,
                    counterparty=self.liability_holder.name,
                    amount=amount,
                    denomination=self.denomination,
                    maturity_type=MaturityType.FIXED_DATE,
                    maturity_date=date,
                    settlement_details=self.settlement_details,
                    name=f"{payment_type}_claim_{date.year}",  # 不再依赖bond的name属性
                    book_value=bv,
                    expected_cash_flow=cf,
                    parent_bond=bond_id  # 使用生成的bond_id
                )
                claims.append(claim)
                
                print(f"Creating claim for {payment_type} payment at {date.year}")
                
        return claims

    def _create_bond_payment_schedule(self) -> List[Tuple[datetime, float, str]]:
        """Create payment schedule for different types of bonds"""
        if not self.bond_type:
            raise ValueError("Bond type must be specified for bonds")

        schedule = []
        t1 = datetime(2050, 1, 1)
        t2 = datetime(2100, 1, 1)
        
        if self.type == EntryType.BOND_ZERO_COUPON.value:
            # Zero-coupon bond: only principal payment at maturity
            schedule.append((self.maturity_date, self.amount, "principal"))

        elif self.type == EntryType.BOND_COUPON.value:
            if not self.coupon_rate:
                raise ValueError("Coupon rate is required for coupon bonds")
            
            is_t2_maturity = self.maturity_date == t2
            
            if not is_t2_maturity:  # Matures at t1
                schedule.append((t1, self.amount * self.coupon_rate, 'coupon'))
                schedule.append((t1, self.amount, 'principal'))
            else:  # Matures at t2
                schedule.append((t1, self.amount * self.coupon_rate, 'coupon'))
                schedule.append((t2, self.amount * self.coupon_rate, 'coupon'))
                schedule.append((t2, self.amount, 'principal'))

        elif self.type == EntryType.BOND_AMORTIZING.value:
            if not self.coupon_rate:
                raise ValueError("Interest rate is required for amortizing bonds")
            
            is_t2_maturity = self.maturity_date == t2
            
            if is_t2_maturity:  # Matures at t2
                # Split payments between t1 and t2
                principal_payment_t1 = self.amount / 2
                interest_payment_t1 = self.amount * self.coupon_rate
                schedule.append((t1, principal_payment_t1, 'principal'))
                schedule.append((t1, interest_payment_t1, 'coupon'))
                
                principal_payment_t2 = self.amount / 2
                interest_payment_t2 = (self.amount / 2) * self.coupon_rate
                schedule.append((t2, principal_payment_t2, 'principal'))
                schedule.append((t2, interest_payment_t2, 'coupon'))
            else:  # Matures at t1
                schedule.append((t1, self.amount, 'principal'))
                schedule.append((t1, self.amount * self.coupon_rate, 'coupon'))

        # Debug information
        print(f"\nBond Payment Schedule Details:")
        print(f"Bond Type: {self.type}")
        print(f"Maturity Date: {'t2' if self.maturity_date == t2 else 't1'}")
        print("\nPayment Schedule:")
        for date, amount, payment_type in sorted(schedule):
            time_point = 't2' if date == t2 else 't1'
            print(f"- {time_point}: {amount:.2f} ({payment_type})")

        return schedule

    def transfer_to(self, new_holder: Agent):
        """Transfer bond and all connected claims to new holder"""
        # Transfer main bond entry
        self.asset_holder.remove_asset(self.asset_entry)
        new_holder.add_asset(self.asset_entry)
        
        # Transfer all connected claims
        for claim in self.connected_claims:
            self.asset_holder.remove_asset(claim)
            new_holder.add_asset(claim)
        
        self.asset_holder = new_holder

class EconomicSystem:
    def __init__(self, num_cycles=1):
        self.agents: Dict[str, Agent] = {}  # Current state
        self.asset_liability_pairs: List[AssetLiabilityPair] = []
        self.time_states: Dict[str, Dict[str, Agent]] = {}  # States at different time points
        self.current_time_state = "t0"  # Track current time state
        self.simulation_finalized = False  # Track if simulation is finalized
        self.num_cycles = num_cycles       # Number of simulation cycles
        # Initialize t0 state
        self.save_state('t0')

    def validate_time_point(self, time_point: str, allow_t0: bool = True) -> None:
        """Validate a time point string"""
        valid_points = ['t0', 't1', 't2'] if allow_t0 else ['t1', 't2']
        if time_point not in valid_points:
            raise ValueError(f"Time point must be {', '.join(valid_points)}")

    def add_agent(self, agent: Agent):
        self.agents[agent.name] = agent
        # Auto-save t0 state when agent is added
        if self.current_time_state == 't0':
            self.save_state('t0')

    def create_asset_liability_pair(self, pair: AssetLiabilityPair):
        self.asset_liability_pairs.append(pair)
        asset_entry, liability_entry = pair.create_entries()
        pair.asset_holder.add_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.add_liability(liability_entry)

        # When creating a loan, automatically create corresponding deposit
        if pair.type == EntryType.LOAN.value:
            # At this point, we're guaranteed that asset_holder is a bank
            # because create_entries() would have thrown an error otherwise
            deposit_pair = AssetLiabilityPair(
                time=datetime.now(),
                type=EntryType.DEPOSIT.value,
                amount=pair.amount,  # Deposit amount equals loan amount
                denomination=pair.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_type=SettlementType.MEANS_OF_PAYMENT,
                settlement_denomination=pair.denomination,
                asset_holder=pair.liability_holder,  # Borrower gets deposit asset
                liability_holder=pair.asset_holder,  # Bank holds deposit liability
            )
            
            deposit_asset, deposit_liability = deposit_pair.create_entries()
            deposit_pair.asset_holder.add_asset(deposit_asset)
            deposit_pair.liability_holder.add_liability(deposit_liability)
            self.asset_liability_pairs.append(deposit_pair)

        self.save_state(self.current_time_state)

    def save_state(self, time_point: str):
        if time_point not in ['t0', 't1', 't2']:
            raise ValueError("Invalid time point")
        self.time_states[time_point] = {}
        for name, agent in self.agents.items():
            agent_copy = Agent(agent.name, agent.type)
            agent_copy.assets = deepcopy(agent.assets)
            agent_copy.liabilities = deepcopy(agent.liabilities)
            self.time_states[time_point][name] = agent_copy
        self.current_time_state = time_point

    def display_balance_sheets(self, time_point: str):
        if not self.agents:
            print("\nNo agents in the system!")
            return

        current_agents = self.get_agents_at_time(time_point).values()
        print(f"\nBalance sheet at {time_point}:")
        print("Assets:")
        for agent in current_agents:
            for asset in agent.assets:
                maturity_info = ""
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    if asset.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif asset.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                entry_type = asset.type.value
                if asset.type == EntryType.PAYABLE:
                    entry_type = "Receivable"
                elif asset.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"Delivery claim {asset.name}" if asset.name else "Delivery claim"
                elif asset.type == EntryType.DEFAULT:
                    entry_type = f"Default claim ({asset.name})"

                print(f"  - {entry_type}: {asset.amount} {asset.denomination} "
                      f"(from {asset.counterparty if asset.counterparty else 'N/A'})"
                      f"{maturity_info} [issued at {asset.issuance_time}]")

            print("Liabilities:")
            for liability in agent.liabilities:
                maturity_info = ""
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    if liability.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif liability.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                entry_type = liability.type.value
                if liability.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"Delivery promise {liability.name}" if liability.name else "Delivery promise"
                elif liability.type == EntryType.DEFAULT:
                    entry_type = f"Default liability ({liability.name})"

                print(f"  - {entry_type}: {liability.amount} {liability.denomination} "
                      f"(to {liability.counterparty}){maturity_info} "
                      f"[issued at {liability.issuance_time}]")

            print(f"Total assets: {agent.get_total_assets()}")
            print(f"Total liabilities: {agent.get_total_liabilities()}")
            print(f"Net worth: {agent.get_net_worth()}")

    def get_agents_at_time(self, time_point: str) -> Dict[str, Agent]:
        if time_point not in ['t0', 't1', 't2']:
            raise ValueError("Invalid time point")
        if time_point == 't0':
            return {name: agent for name, agent in self.agents.items()}
        if time_point in self.time_states:
            return self.time_states[time_point]
        return {name: agent for name, agent in self.agents.items()}

    def run_simulation(self) -> bool:
        for cycle in range(self.num_cycles):
            # Initialize orders dictionary at the beginning of each cycle
            orders = {}
            
            # Create initial bond inventory for dealers if this is the first cycle and t0
            if cycle == 0 and self.current_time_state == 't0':
                print("\n=== Creating Initial Bond Inventory for Dealers ===")
                for agent in self.agents.values():
                    if isinstance(agent, SecurityDealer) and agent.inventory_bond == 0:
                        # Find a suitable liability issuer (preferably a Treasury or Bank)
                        issuers = [a for a in self.agents.values() if a.type in [AgentType.TREASURY, AgentType.BANK]]
                        if not issuers:  # If no Treasury/Bank, use any other agent
                            issuers = [a for a in self.agents.values() if a != agent and a.type != AgentType.OTHER]
                        
                        if issuers:
                            issuer = issuers[0]  # Take the first available issuer
                            initial_bond_qty = 20  # Give dealer 20 bonds to start with
                            
                            # Create a zero-coupon bond with dealer as asset holder
                            now = datetime.now()
                            maturity_date = datetime(2100, 1, 1)  # t2
                            pair = AssetLiabilityPair(
                                time=now,
                                type=EntryType.BOND_ZERO_COUPON.value,
                                amount=initial_bond_qty,
                                denomination="USD",
                                maturity_type=MaturityType.FIXED_DATE,
                                maturity_date=maturity_date,
                                settlement_type=SettlementType.MEANS_OF_PAYMENT,
                                settlement_denomination="USD",
                                asset_holder=agent,  # Dealer holds the bond
                                liability_holder=issuer  # Issuer (Treasury/Bank) issued it
                            )
                            self.create_asset_liability_pair(pair)
                            
                            # Calculate the cost of bonds using mid-price
                            bond_price = (agent.P_a + agent.P_b) / 2  # Using mid-price for initial inventory
                            total_cost = initial_bond_qty * bond_price
                            
                            # Use update_inventory method to properly track trading statistics
                            # Note: This is a "buy" transaction for the dealer (positive bond_delta, negative cash_delta)
                            # First, reset inventory to 0 to avoid double counting
                            agent.inventory_bond = 0  
                            agent.update_inventory(initial_bond_qty, -total_cost)
                            
                            print(f"Created initial inventory: {agent.name} now holds {initial_bond_qty} bonds from {issuer.name}")
                            print(f"Paid {total_cost:.2f} for initial inventory, remaining cash: {agent.inventory_cash:.2f}")
                            print(f"Transaction recorded in trading statistics as a buy of {initial_bond_qty} bonds @ {bond_price:.2f}")
            
            # Print initial state of all dealers
            print("\n=== Initial Dealer State ===")
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    print(f"{agent.name}: Bond Inventory = {agent.inventory_bond}, Cash = {agent.inventory_cash:.2f}")
                    print(f"  Inventory Limit = {agent.inventory_limit}, Base Price = {(agent.P_a + agent.P_b)/2:.2f}")
                    
            # === Stage tn2: Dealer quoting phase ===
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    bid, ask = agent.get_prices()
                    print(f"  {agent.name} bid = {bid:.2f}, ask = {ask:.2f}")
            # === Stage tn3: Dealer repricing phase ===
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    # 1. Read fundamental quotes from VBT stub
                    P_b = agent.P_b               # VBT bid price from tn2
                    P_a = agent.P_a               # VBT ask price from tn2

                    # 2. Read current inventory and parameters
                    X = agent.inventory_bond      # Current bond inventory
                    X_star = agent.inventory_limit  # Inventory limit (X*)
                    S = agent.S                   # Quoted size capacity for this round

                    # 3. Compute spread width
                    spread_width = (S / (2 * X_star)) * (P_a - P_b)

                    # 4. Compute price slope (delta)
                    delta = (P_a - P_b) / (2 * X_star + S)

                    # 5. Compute mid‐price based on current inventory
                    mid_price = (P_a + P_b) / 2 + delta * X

                    # 6. Derive bid and ask quotes
                    bid = mid_price - spread_width / 2
                    ask = mid_price + spread_width / 2

                    # 7. Store this round's quotes for use in tn5 and tn10
                    agent.current_bid = bid
                    agent.current_ask = ask

                    # 8. Log the repriced quotes
                    print(f"  {agent.name} repriced bid = {bid:.2f}, ask = {ask:.2f}")
            # === Stage tn4: Collect quote requests ===
            quote_requests = []
            for agent in self.agents.values():
                if not isinstance(agent, SecurityDealer):  # Only non-Dealer agents
                    for dealer in self.agents.values():
                        if isinstance(dealer, SecurityDealer):
                            # If agent has request_quote method, use it
                            if hasattr(agent, "request_quote"):
                                req = agent.request_quote(dealer.name)
                            else:
                                # Simulate quote request: buy if net worth > 0, else sell
                                side = 'buy' if agent.get_net_worth() > 0 else 'sell'
                                quantity = min(abs(agent.get_net_worth()) // 2, 10)
                                req = {
                                    'trader': agent.name,
                                    'dealer': dealer.name,
                                    'side': side,
                                    'quantity': quantity
                                }
                            quote_requests.append(req)
                            print(f"  {agent.name} requests {req['side']} {req['quantity']} from {dealer.name}")
            # === Stage tn5: Store dealer quotes ===
            dealer_quotes = {}
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    bid, ask = agent.get_prices()
                    dealer_quotes[agent.name] = (bid, ask)
                    print(f"  Stored for {agent.name}: bid={bid:.2f}, ask={ask:.2f}")
            # === Stage tn10: Simulating market order submissions ===
            # --- tn10 ---
            # Interactive order entry (only execute once per simulation cycle)
            print("\n--- Interactive Order Entry (tn10) ---")
            agent_list = list(self.agents.values())
            for idx, agent in enumerate(agent_list, 1):
                print(f"{idx}. {agent.name} ({agent.type.value})")
            # Select agent
            while True:
                try:
                    agent_idx = int(input(f"Select an agent as order initiator (default 1): ") or 1) - 1
                    if 0 <= agent_idx < len(agent_list):
                        order_agent = agent_list[agent_idx]
                        break
                    else:
                        print("Invalid selection. Please enter a valid number.")
                except ValueError:
                    print("Invalid input. Please enter a number.")
            # Select side
            while True:
                side = input("Order side? (buy/sell, default buy): ").strip().lower() or 'buy'
                if side in ['buy', 'sell']:
                    break
                print("Invalid side. Please enter 'buy' or 'sell'.")
            # Input quantity
            while True:
                try:
                    quantity = int(input("Order quantity (default 5): ") or 5)
                    if quantity > 0:
                        break
                    else:
                        print("Quantity must be positive.")
                except ValueError:
                    print("Invalid input. Please enter a positive integer.")
            # Input price
            while True:
                try:
                    price = float(input("Order price (default 101.0): ") or 101.0)
                    if price > 0:
                        break
                    else:
                        print("Price must be positive.")
                except ValueError:
                    print("Invalid input. Please enter a positive number.")
                    
            # Find security dealers in the system
            dealer_list = [agent for agent in self.agents.values() if isinstance(agent, SecurityDealer)]
            
            if not dealer_list:
                print("Error: No SecurityDealer found in the system. Order not executed.")
            else:
                # If there's only one dealer, use it directly
                if len(dealer_list) == 1:
                    dealer = dealer_list[0]
                    print(f"Using only available dealer: {dealer.name}")
                else:
                    # If multiple dealers, let user select
                    print("\nSelect a dealer to execute the order:")
                    for idx, d in enumerate(dealer_list, 1):
                        print(f"{idx}. {d.name}")
                    
                    # Get dealer selection
                    try:
                        dealer_idx = int(input(f"Select dealer (default 1): ") or 1) - 1
                        if 0 <= dealer_idx < len(dealer_list):
                            dealer = dealer_list[dealer_idx]
                        else:
                            print("Invalid selection. Using first dealer.")
                            dealer = dealer_list[0]
                    except ValueError:
                        print("Invalid input. Using first dealer.")
                        dealer = dealer_list[0]
                
                # Execute order directly, no need to store in orders dictionary
                # since we're directly calling handle_order here
                print(f"\nExecuting order: {side} {quantity} @ {price} from {order_agent.name} to {dealer.name}")
                
                # Pass system reference for creating asset-liability pairs
                dealer.handle_order(
                    side=side, 
                    quantity=quantity, 
                    price=price, 
                    system=self,  # Pass the economic system instance
                    counterparty=order_agent  # Pass the agent initiating the order
                )
                
                # Clear orders dictionary for this dealer to avoid duplicate processing in tn11
                orders = {}
            
            # --- tn11 ---
            # Skip, we already executed orders in tn10
            
            # --- tn12 ---
            # Only check inventory limits, no need to handle orders again
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    inv = agent.inventory_bond
                    limit = agent.inventory_limit
                    if abs(inv) > limit:  # Check both positive and negative excess
                        excess = inv - limit if inv > limit else inv + limit
                        side = 'sell' if inv > limit else 'buy'
                        print(f"\nLayoff check: {agent.name} inventory {inv} exceeds limit {limit}")
                        print(f"Inventory would need adjustment via {side} of {abs(excess)}")
                        # We don't actually process layoff here to avoid duplication
            # === Stage tn24: Dealer diagnostics ===
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    stats = agent.get_trading_statistics()
                    print(f"\n{agent.name} Trading Statistics:")
                    print(f"1. Inventory & Cash:")
                    print(f"   - Final Bond Inventory: {stats['final_inventory']}")
                    print(f"   - Final Cash Position: {agent.inventory_cash:.2f}")
                    print(f"   - P&L: {stats['pnl']:.2f}")
                    
                    print(f"\n2. Trading Volume:")
                    print(f"   - Total Volume: {stats['total_volume']}")
                    print(f"   - Number of Trades: {stats['trade_count']}")
                    print(f"   - Average Trade Volume: {stats['avg_trade_volume']:.2f}")
                    print(f"   - Layoff Volume: {stats['layoff_volume']}")
                    
                    print(f"\n3. Spread Statistics:")
                    print(f"   - Average Spread: {stats['avg_spread']:.4f}")
                    print(f"   - Minimum Spread: {stats['min_spread']:.4f}")
                    print(f"   - Maximum Spread: {stats['max_spread']:.4f}")
                    
                    print(f"\n4. Price Statistics:")
                    print(f"   - Average Trade Price: {stats['avg_trade_price']:.2f}")
                    print(f"   - Current Bid: {agent.P_b:.2f}")
                    print(f"   - Current Ask: {agent.P_a:.2f}")
            # === Stage tn15: VBT quote update ===
            for agent in self.agents.values():
                if isinstance(agent, SecurityDealer):
                    agent.update_vbt_quotes()
        return True

    def settle_entries(self, time_point: str):
        """Settle entries at a specific time point"""
        self.validate_time_point(time_point, allow_t0=False)

        # First save the current state at the previous time point
        prev_time = 't0' if time_point == 't1' else 't1'
        if prev_time not in self.time_states:
            self.save_state(prev_time)

        # Process all entries that are due at this time point
        for pair in self.asset_liability_pairs[:]:  # Create a copy to iterate over
            if (pair.maturity_type == MaturityType.FIXED_DATE and
                pair.maturity_date):
                # Check if the entry's maturity date matches our time state
                entry_time = 't1'
                if pair.maturity_date.year == 2100:  # t2
                    entry_time = 't2'

                if entry_time == time_point:
                    # Handle bond payments
                    if pair.type.startswith("bond_"):
                        self._settle_bond(pair, time_point)
                    else:
                        # Handle other types of settlement
                        self._settle_non_bond(pair, time_point)

        # Automatically save state after settlement
        self.save_state(time_point)
        self.current_time_state = time_point

    def _settle_bond(self, pair: AssetLiabilityPair, time_point: str):
        """Handle bond settlement"""
        # Get payment schedule
        payment_schedule = pair._create_bond_payment_schedule()
        
        # Find amounts to be paid at current time point
        current_payments = [
            (amount, payment_type) 
            for date, amount, payment_type in payment_schedule 
            if (time_point == 't1' and date.year == 2050) or 
               (time_point == 't2' and date.year == 2100)
        ]

        # Process each payment
        for amount, payment_type in current_payments:
            # Check if debtor has sufficient deposit
            debtor_deposit = next(
                (asset for asset in pair.liability_holder.assets
                 if asset.type == EntryType.DEPOSIT
                 and asset.amount >= amount),
                None
            )

            if not debtor_deposit:
                raise ValueError(f"Debtor does not have enough deposit to pay {payment_type}")

            # Get bank holding the deposit liability
            bank = next(a for a in self.agents.values() if a.name == debtor_deposit.counterparty)

            # Deduct payment amount from debtor's deposit
            pair.liability_holder.remove_asset(debtor_deposit)
            bank.remove_liability(next(
                l for l in bank.liabilities
                if l.counterparty == pair.liability_holder.name
                and l.amount == debtor_deposit.amount
            ))

            # Create new deposit for debtor (if there's remaining amount)
            if debtor_deposit.amount > amount:
                remainder_deposit = AssetLiabilityPair(
                    time=datetime.now(),
                    type=EntryType.DEPOSIT.value,
                    amount=debtor_deposit.amount - amount,
                    denomination=pair.denomination,
                    maturity_type=MaturityType.ON_DEMAND,
                    maturity_date=None,
                    settlement_type=SettlementType.NONE,
                    settlement_denomination=pair.denomination,
                    asset_holder=pair.liability_holder,
                    liability_holder=bank
                )
                remainder_asset, remainder_liability = remainder_deposit.create_entries()
                remainder_asset.issuance_time = time_point
                remainder_liability.issuance_time = time_point
                pair.liability_holder.add_asset(remainder_asset)
                bank.add_liability(remainder_liability)

            # Create new deposit for creditor
            creditor_deposit = AssetLiabilityPair(
                time=datetime.now(),
                type=EntryType.DEPOSIT.value,
                amount=amount,
                denomination=pair.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_type=SettlementType.NONE,
                settlement_denomination=pair.denomination,
                asset_holder=pair.asset_holder,
                liability_holder=bank
            )
            creditor_asset, creditor_liability = creditor_deposit.create_entries()
            creditor_asset.issuance_time = time_point
            creditor_liability.issuance_time = time_point
            pair.asset_holder.add_asset(creditor_asset)
            bank.add_liability(creditor_liability)

            # Record settlement history
            pair.asset_holder.record_settlement(
                time_point=time_point,
                original_entry=pair.create_entries()[0],  # Original bond asset
                settlement_result=creditor_asset,  # New deposit
                counterparty=pair.liability_holder.name,
                as_asset_holder=True
            )
            pair.liability_holder.record_settlement(
                time_point=time_point,
                original_entry=pair.create_entries()[1],  # Original bond liability
                settlement_result=remainder_asset if debtor_deposit.amount > amount else None,
                counterparty=pair.asset_holder.name,
                as_asset_holder=False
            )

        # If this is the last payment (maturity), remove original bond
        if time_point == 't2' or (time_point == 't1' and pair.maturity_date.year == 2050):
            # Remove original bond entries
            asset_entry, liability_entry = pair.create_entries()
            pair.asset_holder.remove_asset(asset_entry)
            if liability_entry:
                pair.liability_holder.remove_liability(liability_entry)
            
            # Remove bond pair from system
            if pair in self.asset_liability_pairs:
                self.asset_liability_pairs.remove(pair)

    def _settle_non_bond(self, pair: AssetLiabilityPair, time_point: str):
        """Handle non-bond type settlement"""
        # Remove original entries
        asset_entry, liability_entry = pair.create_entries()
        pair.asset_holder.remove_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.remove_liability(liability_entry)

        # Handle settlement based on type
        if pair.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
            # Find the bank deposit that will be used for settlement
            debtor_deposit = next(
                (asset for asset in pair.liability_holder.assets
                 if asset.type == EntryType.DEPOSIT
                 and asset.amount >= pair.amount
                 and asset.denomination == pair.denomination),
                None
            )

            if not debtor_deposit:
                raise ValueError(f"No suitable deposit found for settlement")

            # Get bank holding the deposit liability
            bank = next(a for a in self.agents.values() if a.name == debtor_deposit.counterparty)

            # Remove the original deposit from the debtor
            pair.liability_holder.remove_asset(debtor_deposit)

            # Remove the corresponding liability from the bank
            bank_liability = next(
                (l for l in bank.liabilities
                 if l.type == EntryType.DEPOSIT
                 and l.counterparty == pair.liability_holder.name
                 and l.amount == debtor_deposit.amount),
                None
            )
            if bank_liability:
                bank.remove_liability(bank_liability)

            # Create new deposit entry for the creditor
            settlement_pair = AssetLiabilityPair(
                time=datetime.now(),
                type=EntryType.DEPOSIT.value,
                amount=pair.amount,
                denomination=pair.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_type=SettlementType.NONE,
                settlement_denomination=pair.denomination,
                asset_holder=pair.asset_holder,  # Creditor gets the deposit
                liability_holder=bank,  # Bank keeps liability
                asset_name=None
            )

            # Create entries with current time point as issuance time
            new_asset_entry, new_liability_entry = settlement_pair.create_entries()
            new_asset_entry.issuance_time = time_point
            if new_liability_entry:
                new_liability_entry.issuance_time = time_point

            # Record settlement history
            pair.asset_holder.record_settlement(
                time_point=time_point,
                original_entry=asset_entry,
                settlement_result=new_asset_entry,
                counterparty=pair.liability_holder.name,
                as_asset_holder=True
            )
            pair.liability_holder.record_settlement(
                time_point=time_point,
                original_entry=liability_entry,
                settlement_result=debtor_deposit,  # The deposit that was used
                counterparty=pair.asset_holder.name,
                as_asset_holder=False
            )

            # Add entries
            settlement_pair.asset_holder.add_asset(new_asset_entry)
            if new_liability_entry:
                settlement_pair.liability_holder.add_liability(new_liability_entry)
            self.asset_liability_pairs.append(settlement_pair)

            # If there was remaining deposit amount, create a new deposit for the remainder
            if debtor_deposit.amount > pair.amount:
                remainder_pair = AssetLiabilityPair(
                    time=datetime.now(),
                    type=EntryType.DEPOSIT.value,
                    amount=debtor_deposit.amount - pair.amount,
                    denomination=pair.denomination,
                    maturity_type=MaturityType.ON_DEMAND,
                    maturity_date=None,
                    settlement_type=SettlementType.NONE,
                    settlement_denomination=pair.denomination,
                    asset_holder=pair.liability_holder,  # Debtor keeps remainder
                    liability_holder=bank,  # Bank keeps liability
                    asset_name=None
                )

                # Create entries with current time point as issuance time
                remainder_asset, remainder_liability = remainder_pair.create_entries()
                remainder_asset.issuance_time = time_point
                if remainder_liability:
                    remainder_liability.issuance_time = time_point

                # Add entries
                remainder_pair.asset_holder.add_asset(remainder_asset)
                if remainder_liability:
                    remainder_pair.liability_holder.add_liability(remainder_liability)
                self.asset_liability_pairs.append(remainder_pair)

        elif pair.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            # Find and remove the non-financial asset from the liability holder
            non_financial_asset = next(
                (asset for asset in pair.liability_holder.assets
                 if asset.type == EntryType.NON_FINANCIAL
                 and asset.name == pair.asset_name
                 and asset.amount >= pair.amount),
                None
            )

            if not non_financial_asset:
                raise ValueError(f"Non-financial asset {pair.asset_name} not found for settlement")

            # Remove the asset from the liability holder
            pair.liability_holder.remove_asset(non_financial_asset)

            # Create non-financial asset entry for the asset holder
            settlement_pair = AssetLiabilityPair(
                time=datetime.now(),
                type=EntryType.NON_FINANCIAL.value,
                amount=pair.amount,
                denomination=pair.settlement_details.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_type=SettlementType.NONE,
                settlement_denomination=pair.settlement_details.denomination,
                asset_holder=pair.asset_holder,  # Original creditor gets the goods
                liability_holder=None,  # Non-financial assets have no liability holder
                asset_name=pair.asset_name  # Use the asset_name directly
            )
            # Create entry with current time point as issuance time
            new_asset_entry, _ = settlement_pair.create_entries()
            new_asset_entry.issuance_time = time_point

            # Record settlement history
            pair.asset_holder.record_settlement(
                time_point=time_point,
                original_entry=asset_entry,
                settlement_result=new_asset_entry,
                counterparty=pair.liability_holder.name,
                as_asset_holder=True
            )
            pair.liability_holder.record_settlement(
                time_point=time_point,
                original_entry=liability_entry,
                settlement_result=non_financial_asset,  # The non-financial asset that was delivered
                counterparty=pair.asset_holder.name,
                as_asset_holder=False
            )

            # Add entry directly to avoid default t0 issuance time
            settlement_pair.asset_holder.add_asset(new_asset_entry)
            self.asset_liability_pairs.append(settlement_pair)

            # If there was remaining amount in the non-financial asset, create a new entry for it
            if non_financial_asset.amount > pair.amount:
                remainder_pair = AssetLiabilityPair(
                    time=datetime.now(),
                    type=EntryType.NON_FINANCIAL.value,
                    amount=non_financial_asset.amount - pair.amount,
                    denomination=non_financial_asset.denomination,
                    maturity_type=MaturityType.ON_DEMAND,
                    maturity_date=None,
                    settlement_type=SettlementType.NONE,
                    settlement_denomination=non_financial_asset.denomination,
                    asset_holder=pair.liability_holder,  # Original holder keeps remainder
                    liability_holder=None,
                    asset_name=non_financial_asset.name
                )

                # Create entry with current time point as issuance time
                remainder_asset, _ = remainder_pair.create_entries()
                remainder_asset.issuance_time = time_point

                # Add entry
                remainder_pair.asset_holder.add_asset(remainder_asset)
                self.asset_liability_pairs.append(remainder_pair)

        # Remove original pair from system
        if pair in self.asset_liability_pairs:
            self.asset_liability_pairs.remove(pair)

    def validate_time_point(self, time_point: str, allow_t0: bool = True) -> None:
        """Validate time point string"""
        valid_points = ['t0', 't1', 't2'] if allow_t0 else ['t1', 't2']
        if time_point not in valid_points:
            raise ValueError(f"Time point must be {', '.join(valid_points)}")

    def create_asset_liability_pair_interactive(self, entry_type_choice=None):
        """
        Interactive function to create an asset-liability pair.
        Handles user input for all necessary parameters and validates bank requirement for loans.
        """
        # Get asset holder first
        print("\nAvailable asset holders:")
        for i, agent in enumerate(self.agents.values(), 1):
            print(f"{i}. {agent.name} ({agent.type.value})")
        asset_holder_idx = int(input("Select asset holder (enter number): ")) - 1
        asset_holder = list(self.agents.values())[asset_holder_idx]

        # Get liability holder
        print("\nAvailable liability holders:")
        for i, agent in enumerate(self.agents.values(), 1):
            if agent != asset_holder:  # Cannot select self as liability holder
                print(f"{i}. {agent.name} ({agent.type.value})")
        liability_holder_idx = int(input("Select liability holder (enter number): ")) - 1
        liability_holder = list(self.agents.values())[liability_holder_idx]

        # Get entry type if not provided
        if entry_type_choice is None:
            print("\nIs this a financial or non-financial entry?")
            print("1. Financial (requires liability holder)")
            print("2. Non-financial (no liability holder)")
            entry_type_choice = int(input("Enter choice (1 or 2): "))

        entry_type = None
        bond_type = None
        coupon_rate = None

        # Handle different entry types
        if entry_type_choice == 1:  # financial
            print("\nSelect financial entry type:")
            print("1. Loan")
            print("2. Bond")
            print("3. Deposit")
            print("4. Payable")
            print("5. Delivery Claim")
            type_idx = int(input("Select entry type (enter number): ")) - 1
            
            if type_idx == 0:  # loan
                if asset_holder.type != AgentType.BANK:
                    print("Error: Only banks can hold loans as assets!")
                    return
                entry_type = EntryType.LOAN
            elif type_idx == 1:  # bond
                print("\nSelect bond type:")
                print("1. Zero-coupon Bond")
                print("2. Coupon Bond")
                print("3. Amortizing Bond")
                bond_type_choice = int(input("Enter choice (1-3): ")) - 1
                
                if bond_type_choice == 0:
                    entry_type = EntryType.BOND_ZERO_COUPON
                    bond_type = BondType.ZERO_COUPON
                elif bond_type_choice == 1:
                    entry_type = EntryType.BOND_COUPON
                    bond_type = BondType.COUPON
                elif bond_type_choice == 2:
                    entry_type = EntryType.BOND_AMORTIZING
                    bond_type = BondType.AMORTIZING
                
                # Get coupon rate for coupon and amortizing bonds
                if bond_type_choice in [1, 2]:
                    while True:
                        try:
                            coupon_rate = float(input("\nEnter coupon/interest rate (as decimal, e.g. 0.05 for 5%): "))
                            if coupon_rate <= 0:
                                print("Error: Rate must be positive")
                                continue
                            break
                        except ValueError:
                            print("Error: Please enter a valid number")
            elif type_idx == 2:  # deposit
                entry_type = EntryType.DEPOSIT
            elif type_idx == 3:  # payable
                entry_type = EntryType.PAYABLE
            elif type_idx == 4:  # delivery claim
                entry_type = EntryType.DELIVERY_CLAIM
            else:
                print("Invalid choice!")
                return
            
        elif entry_type_choice == 2:  # non-financial
            entry_type = EntryType.NON_FINANCIAL
        else:
            print("Invalid choice!")
            return

        if entry_type is None:
            print("Invalid entry type!")
            return

        # Get amount and denomination
        while True:
            try:
                amount = float(input("\nEnter amount (face value): "))
                if amount <= 0:
                    print("Error: Amount must be positive")
                    continue
                break
            except ValueError:
                print("Error: Please enter a valid number")
                
        denomination = input("Enter denomination (e.g., USD, EUR): ")

        # For bonds, we require fixed maturity date
        if entry_type in [EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
            maturity_type = MaturityType.FIXED_DATE
            print("\nSelect bond maturity:")
            print("1. t1 (2050)")
            print("2. t2 (2100)")
            time_point = int(input("Enter choice (1-2): "))
            maturity_date = datetime(2050, 1, 1) if time_point == 1 else datetime(2100, 1, 1)
        else:
            # Get maturity type and date for non-bond entries
            print("\nSelect maturity type:")
            print("1. On demand")
            print("2. Fixed date")
            print("3. Perpetual")
            maturity_idx = int(input("Enter choice (1-3): ")) - 1
            
            if maturity_idx == 0:
                maturity_type = MaturityType.ON_DEMAND
                maturity_date = None
            elif maturity_idx == 1:
                maturity_type = MaturityType.FIXED_DATE
                print("\nSelect maturity time point:")
                print("1. t1 (2050)")
                print("2. t2 (2100)")
                time_point = int(input("Enter choice (1-2): "))
                maturity_date = datetime(2050, 1, 1) if time_point == 1 else datetime(2100, 1, 1)
            else:
                maturity_type = MaturityType.PERPETUAL
                maturity_date = None

        # For bonds, default to means of payment settlement
        if entry_type in [EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
            settlement_type = SettlementType.MEANS_OF_PAYMENT
            settlement_denomination = denomination
        else:
            # Get settlement type and denomination for non-bond entries
            print("\nSelect settlement type:")
            print("1. Means of payment")
            print("2. Securities")
            print("3. Non-financial asset")
            print("4. Services")
            print("5. Crypto")
            print("6. None")
            settlement_idx = int(input("Enter choice (1-6): ")) - 1
            settlement_type = list(SettlementType)[settlement_idx]
            settlement_denomination = input("Enter settlement denomination (e.g., USD, EUR): ")

        try:
            # Create asset-liability pair
            pair = AssetLiabilityPair(
                time=datetime.now(),
                type=entry_type.value,
                amount=amount,
                denomination=denomination,
                maturity_type=maturity_type,
                maturity_date=maturity_date,
                settlement_type=settlement_type,
                settlement_denomination=settlement_denomination,
                asset_holder=asset_holder,
                liability_holder=liability_holder,
                asset_name=None,
                bond_type=bond_type,
                coupon_rate=coupon_rate
            )

            self.create_asset_liability_pair(pair)
            if entry_type in [EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
                print("\nBond created successfully!")
                # Display payment schedule
                schedule = pair._create_bond_payment_schedule()
                print("\nPayment schedule:")
                for date, amount, payment_type in schedule:
                    print(f"- {date.strftime('%Y')}: {amount} {denomination} ({payment_type})")
            else:
                print("\nAsset-liability pair created successfully!")
            
        except ValueError as e:
            print(f"\nError creating asset-liability pair: {str(e)}")
            return

    def get_time_points(self) -> List[str]:
        """Get all time points in order: t0, t1, t2"""
        return ['t0', 't1', 't2']

    def compute_changes(self, from_time: str, to_time: str) -> Dict[str, Dict[str, List]]:
        """Compute changes between two time points"""
        if from_time not in self.time_states or to_time not in self.time_states:
            raise ValueError(f"Missing state for time point {from_time} or {to_time}")

        changes = {}
        for name, to_agent in self.time_states[to_time].items():
            from_agent = self.time_states[from_time][name]

            # Find new and removed assets
            new_assets = [a for a in to_agent.assets if not any(a.matches(from_a) for from_a in from_agent.assets)]
            removed_assets = [a for a in from_agent.assets if not any(a.matches(to_a) for to_a in to_agent.assets)]

            # Find new and removed liabilities
            new_liabilities = [l for l in to_agent.liabilities if not any(l.matches(from_l) for from_l in from_agent.liabilities)]
            removed_liabilities = [l for l in from_agent.liabilities if not any(l.matches(to_l) for to_l in to_agent.liabilities)]

            changes[name] = {
                'new_assets': new_assets,
                'removed_assets': removed_assets,
                'new_liabilities': new_liabilities,
                'removed_liabilities': removed_liabilities
            }

        return changes

    def can_settle_entry(self, agent: Agent, entry: BalanceSheetEntry) -> Tuple[bool, str]:
        """Check if an agent can settle a liability"""
        if entry.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
            # Check for sufficient deposits
            deposits = sum(asset.amount for asset in agent.assets
                          if asset.type == EntryType.DEPOSIT
                          and asset.denomination == entry.denomination)
            if deposits < entry.amount:
                return False, f"Insufficient deposits: has {deposits} {entry.denomination}, needs {entry.amount}"

        elif entry.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            # Check for required non-financial asset
            has_asset = any(asset.type == EntryType.NON_FINANCIAL
                           and asset.name == entry.name
                           and asset.amount >= entry.amount
                           for asset in agent.assets)
            if not has_asset:
                return False, f"Missing required non-financial asset: {entry.name}"

        return True, "OK"

    def create_default_entries(self, failed_entry: BalanceSheetEntry) -> Tuple[BalanceSheetEntry, BalanceSheetEntry]:
        """Create default claim and liability entries when settlement fails"""
        # Create default claim for the creditor
        default_claim = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=failed_entry.counterparty,
            amount=failed_entry.amount,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        # Create default liability for the debtor
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=failed_entry.counterparty,
            amount=failed_entry.amount,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        return default_claim, default_liability

    def display_settlement_history(self):
        """Display settlement history for all agents"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        print("\nSettlement History:")
        for agent_name, agent in self.agents.items():
            print(f"\n{agent_name}'s Settlement History:")

            # Display settlements where agent was asset holder
            if agent.settlement_history.get('as_asset_holder', []):
                print("\n  As Asset Holder:")
                for settlement in agent.settlement_history['as_asset_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Asset: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].amount} {settlement['original_entry'].denomination}")
                    print(f"    Settled For: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].amount} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            # Display settlements where agent was liability holder
            if agent.settlement_history.get('as_liability_holder', []):
                print("\n  As Liability Holder:")
                for settlement in agent.settlement_history['as_liability_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Liability: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].amount} {settlement['original_entry'].denomination}")
                    print(f"    Settled With: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].amount} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            if (not agent.settlement_history.get('as_asset_holder') and
                not agent.settlement_history.get('as_liability_holder')):
                print("  No settlements recorded")

    def display_balance_sheets(self, time_point: str):
        """Display balance sheets for all agents at a specific time point"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        current_agents = self.get_agents_at_time(time_point).values()
        print(f"\nBalance sheets at {time_point}:")

        for agent in current_agents:
            print(f"\n{agent.name} ({agent.type.value}):")
            print("Assets:")
            for asset in agent.assets:
                maturity_info = ""
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    if asset.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif asset.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                # Show appropriate entry type
                if asset.type == EntryType.PAYABLE:
                    entry_type = "receivable"
                elif asset.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery claim for {asset.name}" if asset.name else "delivery claim"
                elif asset.type == EntryType.DEFAULT:
                    entry_type = f"default claim ({asset.name})"
                else:
                    entry_type = asset.type.value

                # Skip if entry has matured and been settled
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    entry_maturity = 't1' if asset.maturity_date.year == 2050 else 't2'
                    if time_point > entry_maturity:
                        continue

                print(f"  - {entry_type}: {asset.amount} {asset.denomination} "
                      f"(from {asset.counterparty if asset.counterparty else 'N/A'})"
                      f"{maturity_info} [issued at {asset.issuance_time}]")

            print("Liabilities:")
            for liability in agent.liabilities:
                maturity_info = ""
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    if liability.maturity_date.year == 2100:
                        maturity_info = " (matures at t2)"
                    elif liability.maturity_date.year == 2050:
                        maturity_info = " (matures at t1)"

                # Show appropriate entry type
                if liability.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery promise for {liability.name}" if liability.name else "delivery promise"
                elif liability.type == EntryType.DEFAULT:
                    entry_type = f"default liability ({liability.name})"
                else:
                    entry_type = liability.type.value

                # Skip if entry has matured and been settled
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    entry_maturity = 't1' if liability.maturity_date.year == 2050 else 't2'
                    if time_point > entry_maturity:
                        continue

                print(f"  - {entry_type}: {liability.amount} {liability.denomination} "
                      f"(to {liability.counterparty}){maturity_info} "
                      f"[issued at {liability.issuance_time}]")

class ExcelExporter:
    def __init__(self, system: EconomicSystem):
        self.system = system

    def create_t_table(self, sheet, row_start: int, col_start: int, agent: Agent, time_point: str):
        thick = Side(style='thick', color='000000')

        # Add time point header
        time_header = sheet.cell(row=row_start, column=1)
        time_header.value = f"Time: {time_point}"
        time_header.alignment = Alignment(horizontal="center")
        time_header.font = openpyxl.styles.Font(bold=True)

        name_cell = sheet.cell(row=row_start, column=col_start)
        name_cell.value = f"{agent.name} ({agent.type.value})"
        name_cell.alignment = Alignment(horizontal="center")

        # Set up headers and borders
        for i in range(10):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.border = Border(top=thick)

        for row in range(row_start + 1, row_start + 20):
            cell = sheet.cell(row=row, column=col_start + 4)
            cell.border = Border(right=thick)
            if row == row_start + 1:
                cell.border = Border(right=thick, top=thick)

        headers = ['Type', 'CP', 'Amount', 'Maturity', 'Settlement']
        for i, header in enumerate(headers):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")
            cell = sheet.cell(row=row_start + 1, column=col_start + i + 5)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")

        # Display balance sheet entries
        current_row = row_start + 2
        for entry in agent.assets:
            # Skip entries that were issued after the current time point
            time_points = ['t0', 't1', 't2']
            if time_points.index(entry.issuance_time) > time_points.index(time_point):
                continue

            # Skip matured entries if not at t0
            if time_point != 't0' and entry.maturity_type == MaturityType.FIXED_DATE:
                entry_time = 't1' if entry.maturity_date.year == 2050 else 't2'
                if time_point > entry_time:
                    continue

            # Show entry details
            entry_type = "receivable" if entry.type == EntryType.PAYABLE else entry.type.value
            if entry.type == EntryType.NON_FINANCIAL and entry.name:
                entry_type = f"{entry.type.value} ({entry.name})"
            sheet.cell(row=current_row, column=col_start).value = entry_type
            sheet.cell(row=current_row, column=col_start + 1).value = entry.counterparty if entry.counterparty else "N/A"
            sheet.cell(row=current_row, column=col_start + 2).value = f"+{entry.amount} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE:
                maturity = 't1' if entry.maturity_date.year == 2050 else 't2'
            sheet.cell(row=current_row, column=col_start + 3).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 4).value = settlement

            current_row += 1

        current_row = row_start + 2
        for entry in agent.liabilities:
            # Skip entries that were issued after the current time point
            time_points = ['t0', 't1', 't2']
            if time_points.index(entry.issuance_time) > time_points.index(time_point):
                continue

            # Skip matured entries if not at t0
            if time_point != 't0' and entry.maturity_type == MaturityType.FIXED_DATE:
                entry_time = 't1' if entry.maturity_date.year == 2050 else 't2'
                if time_point > entry_time:
                    continue

            # Show entry details
            entry_type = entry.type.value
            if entry.type == EntryType.DELIVERY_CLAIM:
                entry_type = f"delivery promise for {entry.name}" if entry.name else "delivery promise"
            sheet.cell(row=current_row, column=col_start + 5).value = entry_type
            sheet.cell(row=current_row, column=col_start + 6).value = entry.counterparty
            sheet.cell(row=current_row, column=col_start + 7).value = f"+{entry.amount} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE:
                maturity = 't1' if entry.maturity_date.year == 2050 else 't2'
            sheet.cell(row=current_row, column=col_start + 8).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 9).value = settlement

            current_row += 1

        # Add totals
        total_row = current_row + 2
        sheet.cell(row=total_row, column=col_start).value = "Total Assets:"
        sheet.cell(row=total_row, column=col_start + 2).value = agent.get_total_assets()
        sheet.cell(row=total_row + 1, column=col_start).value = "Total Liabilities:"
        sheet.cell(row=total_row + 1, column=col_start + 2).value = agent.get_total_liabilities()
        sheet.cell(row=total_row + 2, column=col_start).value = "Net Worth:"
        sheet.cell(row=total_row + 2, column=col_start + 2).value = agent.get_net_worth()

        return total_row + 4

    def export_balance_sheets(self, filename: str):
        """Export balance sheets for all time points vertically stacked"""
        # First, ensure all settlements are processed
        if self.system.simulation_finalized:
            # Save current state
            current_state = deepcopy(self.system.agents)
            current_time = self.system.current_time_state

            # Process settlements if not already done
            try:
                if 't1' not in self.system.time_states:
                    self.system.settle_entries('t1')
                if 't2' not in self.system.time_states:
                    self.system.settle_entries('t2')
            except Exception as e:
                print(f"\nWarning: Settlement processing failed ({str(e)})")

            # Restore original state
            self.system.agents = current_state
            self.system.current_time_state = current_time

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Balance Sheets Over Time"

        current_row = 1
        time_points = self.system.get_time_points()

        for time_point in time_points:
            agents = (self.system.time_states[time_point].values()
                     if time_point in self.system.time_states
                     else self.system.agents.values())

            # Add time point separator
            sheet.cell(row=current_row, column=1).value = "=" * 50
            current_row += 1

            col_start = 2
            max_row_in_timepoint = current_row

            for agent in agents:
                max_row_in_timepoint = max(
                    max_row_in_timepoint,
                    self.create_t_table(sheet, current_row, col_start, agent, time_point)
                )
                col_start += 10

            # Add system totals for this time point
            system_total_row = max_row_in_timepoint
            sheet.cell(row=system_total_row, column=1).value = f"System Totals at {time_point}:"
            sheet.cell(row=system_total_row + 1, column=1).value = "Total Assets:"
            sheet.cell(row=system_total_row + 1, column=2).value = sum(agent.get_total_assets() for agent in agents)
            sheet.cell(row=system_total_row + 2, column=1).value = "Total Liabilities:"
            sheet.cell(row=system_total_row + 2, column=2).value = sum(agent.get_total_liabilities() for agent in agents)
            sheet.cell(row=system_total_row + 3, column=1).value = "Total Net Worth:"
            sheet.cell(row=system_total_row + 3, column=2).value = sum(agent.get_net_worth() for agent in agents)

            current_row = system_total_row + 5  # Leave space between time points

        # Adjust column widths
        for i in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 15

        wb.save(filename)

        # NEED TO BE UNCOMMENTED IF RUNNING IN COLAB, OR MODIFIED IF RUNNING ELSEWHERE
        # # Add Colab download capability
        # try:
        #     from google.colab import files
        #     files.download(filename)
        #     print(f"\nExcel file has been created and downloaded: {filename}")
        # except ImportError:
        #     print(f"\nExcel file has been created: {filename}")

def get_user_date_input(prompt: str) -> Optional[datetime]:
    """Helper function to get time state input"""
    print(prompt)
    print("Enter a time state (t0, t1, t2)")
    time_state = input("Time state: ").strip().lower()

    if time_state not in ['t0', 't1', 't2']:
        print("Please enter a valid time state (t0/t1/t2)")
        return None

    # Map time states to representative dates
    if time_state == 't2':
        return datetime(2100, 1, 1)  # Far future for t2
    elif time_state == 't1':
        return datetime(2050, 1, 1)  # Middle future for t1
    else:  # t0
        return datetime(2000, 1, 1)  # Past for t0

def create_agent_interactive(system: EconomicSystem):
    """Interactive function to create an agent"""
    print("\nCreating new agent:")
    name = input("Enter agent name: ")

    if name in system.agents:
        print(f"Error: Agent '{name}' already exists!")
        return

    # Check if this should be a ValueBasedTrader (VBT)
    if "vbt" in name.lower() or name.lower().startswith("value"):
        # Create a ValueBasedTrader instead of a regular Agent
        print("\nDetected ValueBasedTrader naming pattern.")
        print("Creating a ValueBasedTrader with get_prices() capability.")
        
        # Get base price and spread if needed
        try:
            base_price = float(input("\nEnter base price (default 100.0): ") or 100.0)
            spread = float(input("Enter spread (default 0.02): ") or 0.02)
            agent = ValueBasedTrader(name, base_price=base_price, spread=spread)
            system.add_agent(agent)
            print(f"\nValueBasedTrader '{name}' created successfully!")
            print(f"- Base Price: {base_price}")
            print(f"- Spread: {spread}")
            print(f"- Bid: {base_price * (1 - spread/2):.2f}")
            print(f"- Ask: {base_price * (1 + spread/2):.2f}")
            return agent
        except ValueError:
            print("Error: Invalid input for price or spread. Using default values.")
            agent = ValueBasedTrader(name)
            system.add_agent(agent)
            print(f"\nValueBasedTrader '{name}' created with default values.")
            return agent
    
    # Regular Agent creation (original code)
    print("\nAvailable agent types:")
    for i, agent_type in enumerate(AgentType, 1):
        print(f"{i}. {agent_type.value}")

    try:
        type_idx = int(input("\nSelect agent type (enter number): ")) - 1
        agent_type = list(AgentType)[type_idx]

        agent = Agent(name, agent_type)
        system.add_agent(agent)
        print(f"\nAgent '{name}' of type '{agent_type.value}' created successfully!")
        return agent
    except (ValueError, IndexError):
        print("Error: Invalid agent type selection")
        return None

def create_non_financial_asset_interactive(system: EconomicSystem):
    """Interactive function to create a non-financial asset"""
    if len(system.agents) < 1:
        print("\nError: Need at least 1 agent to create a non-financial asset!")
        return

    print("\nCreating new non-financial asset:")

    # List available agents
    print("\nAvailable agents:")
    agents = list(system.agents.values())
    for i, agent in enumerate(agents, 1):
        print(f"{i}. {agent.name} ({agent.type.value})")

    try:
        # Get asset holder
        asset_idx = int(input("\nSelect asset holder (enter number): ")) - 1
        asset_holder = agents[asset_idx]

        # Get asset name
        while True:
            asset_name = input("\nEnter the name of the non-financial asset (e.g., 'machine', 'building'): ").strip()
            if asset_name:
                break
            print("Error: Non-financial asset requires a name!")

        # Get amount and denomination
        amount = float(input("\nEnter amount: "))
        denomination = input("Enter denomination (e.g., USD): ")

        # Create the asset-liability pair
        pair = AssetLiabilityPair(
            time=datetime.now(),
            type=EntryType.NON_FINANCIAL.value,
            amount=amount,
            denomination=denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_type=SettlementType.NONE,
            settlement_denomination=denomination,
            asset_holder=asset_holder,
            liability_holder=None,
            asset_name=asset_name,
            bond_type=None,
            coupon_rate=None
        )

        # Create entries and add them to the system
        system.create_asset_liability_pair(pair)
        print("\nNon-financial asset created successfully!")

    except (ValueError, IndexError) as e:
        print(f"\nError: {str(e)}")
        return

def create_security_dealer_interactive(system: EconomicSystem):
    """Interactive function to create a SecurityDealer with all necessary parameters"""
    print("\nCreating new Security Dealer:")
    
    try:
        # Get dealer name
        name = input("Enter dealer name: ")
        if name in system.agents:
            print(f"Error: Agent '{name}' already exists!")
            return
        
        # Get bond type
        print("\nSelect bond type to trade:")
        print("1. Zero-coupon Bond")
        print("2. Coupon Bond")
        print("3. Amortizing Bond")
        bond_type_choice = int(input("Enter choice (1-3): ")) - 1
        if bond_type_choice == 0:
            bond_type = BondType.ZERO_COUPON
        elif bond_type_choice == 1:
            bond_type = BondType.COUPON
        elif bond_type_choice == 2:
            bond_type = BondType.AMORTIZING
        else:
            print("Invalid bond type selection!")
            return
        
        # Dealer parameters
        capital_base = float(input("\nEnter capital base C (e.g., 10000): "))
        if capital_base <= 0:
            print("Error: Capital base must be positive!")
            return
        quoted_size = int(input("Enter quoted size S_max (e.g., 10): "))
        if quoted_size <= 0:
            print("Error: Quoted size must be positive!")
            return
        base_spread = float(input("Enter base spread (e.g., 0.001): "))
        if base_spread < 0:
            print("Error: Base spread cannot be negative!")
            return
        # Initial inventory and cash
        initial_bond_inventory = int(input("Enter initial bond inventory (e.g., 0): "))
        initial_cash = float(input("Enter initial cash (press Enter to use capital base): ") or capital_base)
        # Inventory limit calculation method
        print("\nSelect inventory limit calculation method:")
        print("1. Formula: C / p_ref")
        print("2. Manual input")
        inv_method_choice = int(input("Enter choice (1-2): "))
        if inv_method_choice == 2:
            inventory_limit_method = 'manual'
            manual_inventory_limit = int(input("Enter manual inventory limit X*: "))
        else:
            inventory_limit_method = 'formula'
            manual_inventory_limit = None
        # Value trader selection
        value_traders = [a for a in system.agents.values() if a.type == AgentType.OTHER and a.name != name]
        value_trader = None
        if not value_traders:
            print("\nNo agents of type AgentType.OTHER found in the system.")
            create_value_trader = input("Would you like to create a value-based trader now? (y/n): ").lower()
            if create_value_trader == 'y':
                while True:
                    value_trader_name = input("Enter value trader name: ")
                    if value_trader_name in system.agents:
                        print(f"Error: Agent '{value_trader_name}' already exists!")
                        continue
                    if value_trader_name == name:
                        print("Error: Cannot use the same name as the dealer!")
                        continue
                    break
                # Create a ValueBasedTrader by default
                try:
                    base_price = float(input("\nEnter base price for VBT (default 100.0): ") or 100.0)
                    spread = float(input("Enter spread for VBT (default 0.02): ") or 0.02)
                    value_trader = ValueBasedTrader(value_trader_name, base_price=base_price, spread=spread)
                    system.add_agent(value_trader)
                    print(f"\nValue trader '{value_trader_name}' (VBT) created successfully!")
                    print(f"- Base Price: {base_price}")
                    print(f"- Spread: {spread}")
                    print(f"- Bid: {base_price * (1 - spread/2):.2f}")
                    print(f"- Ask: {base_price * (1 + spread/2):.2f}")
                except ValueError:
                    print("Error: Invalid input for price or spread. Using default values.")
                    value_trader = ValueBasedTrader(value_trader_name)
                    system.add_agent(value_trader)
                    print(f"\nValue trader '{value_trader_name}' (VBT) created with default values.")
            else:
                print("Cannot create SecurityDealer without a value trader.")
                return
        else:
            print("\nSelect a value trader to handle excess orders:")
            for i, trader in enumerate(value_traders, 1):
                print(f"{i}. {trader.name}")
            try:
                trader_idx = int(input("Select value trader (enter number): ")) - 1
                if trader_idx < 0 or trader_idx >= len(value_traders):
                    print("Invalid selection. Aborting dealer creation.")
                    return
                value_trader = value_traders[trader_idx]
                
                # Check if selected trader is a ValueBasedTrader or at least has get_prices method
                if not isinstance(value_trader, ValueBasedTrader) and not hasattr(value_trader, 'get_prices'):
                    print(f"\nWarning: Selected agent '{value_trader.name}' is not a ValueBasedTrader and lacks get_prices() method.")
                    convert = input("Convert to ValueBasedTrader? (y/n): ").strip().lower()
                    if convert == 'y':
                        # Create a new ValueBasedTrader with the same name and replace the original
                        try:
                            base_price = float(input("\nEnter base price for VBT (default 100.0): ") or 100.0)
                            spread = float(input("Enter spread for VBT (default 0.02): ") or 0.02)
                            
                            # Remove the original agent
                            old_name = value_trader.name
                            system.agents.pop(old_name)
                            
                            # Create a new ValueBasedTrader with the same name
                            value_trader = ValueBasedTrader(old_name, base_price=base_price, spread=spread)
                            system.add_agent(value_trader)
                            print(f"\nAgent '{old_name}' converted to ValueBasedTrader successfully!")
                            print(f"- Base Price: {base_price}")
                            print(f"- Spread: {spread}")
                        except ValueError:
                            print("Error: Invalid input. Using default values.")
                            old_name = value_trader.name
                            system.agents.pop(old_name)
                            value_trader = ValueBasedTrader(old_name)
                            system.add_agent(value_trader)
                    else:
                        print("\nWarning: SecurityDealer may not function correctly without a proper ValueBasedTrader.")
                        continue_anyway = input("Continue anyway? (y/n): ").strip().lower()
                        if continue_anyway != 'y':
                            print("Aborting dealer creation.")
                            return
            except (ValueError, IndexError):
                print("Invalid selection. Aborting dealer creation.")
                return
        # Use VBT's quote for initial mid price and spread if possible
        if hasattr(value_trader, 'get_prices'):
            vbt_bid, vbt_ask = value_trader.get_prices()
            mid_price = (vbt_bid + vbt_ask) / 2
            spread = (vbt_ask - vbt_bid) / mid_price
            print(f"Auto-set dealer mid price from VBT: {mid_price:.2f}")
            print(f"Auto-set dealer spread from VBT: {spread:.4f}")
        else:
            mid_price = float(input("Enter initial mid price (e.g., 100): "))
            spread = float(input("Enter initial spread (e.g., 0.02 for 2%): "))
        P_b = mid_price * (1 - spread/2)
        P_a = mid_price * (1 + spread/2)
        # Create the SecurityDealer
        dealer = SecurityDealer(
            name=name,
            value_trader=value_trader,
            bond_type=bond_type,
            capital_base=capital_base,
            quoted_size=quoted_size,
            P_b=P_b,
            P_a=P_a,
            base_spread=base_spread,
            initial_bond_inventory=initial_bond_inventory,
            initial_cash=initial_cash,
            inventory_limit_method=inventory_limit_method,
            manual_inventory_limit=manual_inventory_limit
        )
        system.add_agent(dealer)
        print(f"\nSecurity Dealer '{name}' created successfully!")
        print(f"1. Bond Type: {bond_type.name}")
        print(f"2. Initial Pricing:")
        print(f"   - Mid Price: {mid_price}")
        print(f"   - Spread: {spread}")
        print(f"   - Bid: {P_b:.2f}")
        print(f"   - Ask: {P_a:.2f}")
        print(f"3. Dealer Parameters:")
        print(f"   - Capital Base: {capital_base}")
        print(f"   - Quoted Size: {quoted_size}")
        print(f"   - Base Spread: {base_spread}")
        print(f"   - Initial Bond Inventory: {initial_bond_inventory}")
        print(f"   - Initial Cash: {initial_cash}")
        print(f"   - Inventory Limit Method: {inventory_limit_method}")
        if inventory_limit_method == 'manual':
            print(f"   - Manual Inventory Limit: {manual_inventory_limit}")
        print(f"4. Value Trader: {value_trader.name}")
        print(f"5. Initial Inventory Limit: {dealer.inventory_limit}")
    except (ValueError, IndexError) as e:
        print(f"\nError: {str(e)}")
        return

def main():
    print("Welcome to the Economic Balance Sheet Simulation!")
    print("==============================================")
    print("\nThis simulation automatically saves states at each time point.")
    print("- t0: Initial state (auto-saved when creating agents and entries)")
    print("- t1: Intermediate state (auto-saved when settling entries)")
    print("- t2: Final state (auto-saved when settling entries)")

    # Get default denomination
    print("\nWould you like to set a default denomination (e.g., USD) for all entries?")
    print("You can still override this for individual entries if needed.")
    default_denomination = input("Enter default denomination (press Enter to skip): ").strip() or None

    system = EconomicSystem()

    while True:
        print("\nEconomic Balance Sheet Simulation")
        print("1. Create agent")
        print("2. Create asset-liability pair")
        print("3. View balance sheets")
        print("4. View settlement history")
        print("5. Simulate (finalize agents and pairs)")
        print("6. Export to Excel")
        print("7. Exit")
        print("8. Create a SecurityDealer")  # new function

        choice = input("\nEnter your choice (1-8): ")  

        if choice == '1':
            create_agent_interactive(system)
        elif choice == '2':
            print("\nIs this a financial or non-financial entry?")
            print("1. Financial (requires liability holder)")
            print("2. Non-financial (no liability holder)")
            entry_type_choice = int(input("Enter choice (1 or 2): "))
            
            if entry_type_choice == 1:
                system.create_asset_liability_pair_interactive(entry_type_choice)
            else:
                create_non_financial_asset_interactive(system)
        elif choice == '3':
            if system.simulation_finalized:
                time_point = input("\nEnter time point to view (t0/t1/t2): ").strip().lower()
                if time_point in ['t0', 't1', 't2']:
                    system.display_balance_sheets(time_point)
                else:
                    print("\nInvalid time point. Please enter t0, t1, or t2.")
            else:
                system.display_balance_sheets('t0')
        elif choice == '4':
            system.display_settlement_history()
        elif choice == '5':
            print("\nFinalizing simulation setup...")
            print("Processing settlements at t1 and t2...")
            system.simulation_finalized = True

            # Run the simulation
            success = system.run_simulation()
            if success:
                print("\nAll settlements processed successfully!")
                print("You can now view balance sheets at different time points (t0, t1, t2).")
            else:
                print("\nSimulation stopped due to settlement failure.")
                print("Check settlement history for details.")
            print("\nNote: No more agents or asset-liability pairs can be added.")

        elif choice == '6':
            if EXCEL_AVAILABLE:
                filename = "balance_sheets.xlsx"
                exporter = ExcelExporter(system)
                exporter.export_balance_sheets(filename)
            else:
                print("\nError: Excel export is not available. Please install openpyxl package.")
                print("Run: pip install openpyxl==3.1.2")
        elif choice == '7':
            print("\nExiting simulation. Goodbye!")
            break
        elif choice == '8':  # new function
            create_security_dealer_interactive(system)
        else:
            print("\nInvalid choice. Please try again.")

class SecurityDealer(Agent):
    """
    Passive market maker using Treynor inventory-based pricing.
    Only trades a single bond type and a single payment asset.
    """
    def __init__(self, name, value_trader, bond_type, capital_base=10000.0, quoted_size=10, base_spread=0.001, P_b=100.0, P_a=100.0, initial_bond_inventory=0, initial_cash=None, inventory_limit_method='formula', manual_inventory_limit=None):
        super().__init__(name, AgentType.OTHER)
        self.name = name
        self.value_trader = value_trader
        self.bond_type = bond_type  # Type of bond being traded
        self.capital_base = capital_base  # Capital base for inventory limit calculation
        self.quoted_size = quoted_size    # S_max, max quote size
        self.S = quoted_size
        self.base_spread = base_spread
        self.P_b = P_b
        self.P_a = P_a
        self.inventory_bond = initial_bond_inventory  # Initial bond inventory
        self.inventory_cash = initial_cash if initial_cash is not None else capital_base  # Initial cash, default to capital base
        self.inventory_limit_method = inventory_limit_method  # 'formula' or 'manual'
        self.manual_inventory_limit = manual_inventory_limit  # Used if method is manual
        # Trading statistics
        self.initial_cash = self.inventory_cash
        self.trade_count = 0
        self.total_volume = 0
        self.layoff_volume = 0
        self.spreads = []
        self.trade_prices = []
        self.trade_volumes = []
        self._update_inventory_limit()  # Initialize inventory limit

    def _update_inventory_limit(self):
        """Update inventory limit based on selected method."""
        if self.inventory_limit_method == 'manual' and self.manual_inventory_limit is not None:
            self.inventory_limit = int(self.manual_inventory_limit)
        else:
            p_ref = (self.P_a + self.P_b) / 2
            self.inventory_limit = int(self.capital_base / p_ref)  # Formula: C / p_ref
        self.X_star = self.inventory_limit

    def update_inventory(self, bond_delta, cash_delta):
        """Adjust bond and cash inventory."""
        self.inventory_bond += bond_delta
        self.inventory_cash += cash_delta
        
        # Update trading statistics
        if bond_delta != 0:  # Only count actual trades
            self.trade_count += 1
            self.total_volume += abs(bond_delta)
            self.trade_volumes.append(abs(bond_delta))
            self.trade_prices.append(-cash_delta / bond_delta)  # Price per bond

    def get_prices(self):
        """
        Dynamic inventory-based pricing according to professor's formula
        """
        try:
            # Try to get prices from value trader
            if hasattr(self.value_trader, 'get_prices'):
                P_b, P_a = self.value_trader.get_prices()
            else:
                # Fallback if value_trader has no get_prices method
                print(f"Warning: {self.value_trader.name} has no get_prices method. Using current P_b, P_a.")
                P_b, P_a = self.P_b, self.P_a
        except Exception as e:
            # Fallback in case of any error
            print(f"Error getting prices from value trader: {str(e)}. Using current values.")
            P_b, P_a = self.P_b, self.P_a
            
        self.P_b = P_b  # Update reference prices
        self.P_a = P_a
        self._update_inventory_limit()  # Update inventory limit
        
        X = self.inventory_bond
        X_star = self.inventory_limit
        S = self.quoted_size

        # Compute dynamic spread and slope
        spread_dyn = (S / (2 * X_star)) * (P_a - P_b)
        delta = (P_a - P_b) / (2 * X_star + S)

        # Compute dynamic mid-price
        mid_dyn = (P_a + P_b) / 2 + delta * X

        # Return bid and ask
        bid = mid_dyn - spread_dyn / 2
        ask = mid_dyn + spread_dyn / 2
        return bid, ask

    def handle_order(self, side, quantity, price=None, system=None, counterparty=None):
        """
        Process a 'buy' or 'sell' order within inventory limits.
        Forward any excess quantity to the value-based trader.
        
        Args:
            side: 'buy' or 'sell'
            quantity: Number of units to trade
            price: Optional explicit price, if None use standard bid/ask
            system: EconomicSystem instance to create asset-liability pairs
            counterparty: The agent initiating the trade
        """
        bid, ask = self.get_prices()
        self.spreads.append(ask - bid)  # Record current spreadni
        
        # If price is None, use standard bid or ask price
        trade_price = price if price is not None else (ask if side == 'buy' else bid)
        
        print(f"DEBUG: {self.name} handling {side} order for {quantity} @ {trade_price}")
        print(f"DEBUG: Current inventory: {self.inventory_bond}, limit: {self.inventory_limit}")
        
        if side == 'buy':
            # Trader buys bonds from dealer (dealer sells)
            available = self.inventory_bond
            if available <= 0:
                print(f"DEBUG: {self.name} has no inventory to sell, forwarding full order to VBT")
                self.layoff_volume += quantity
                try:
                    # Forward entire order to value trader
                    self.value_trader.handle_order(side, quantity, price=trade_price, system=system)
                except Exception as e:
                    print(f"Warning: Could not forward order to value trader: {str(e)}")
                return  # Exit early, nothing to process locally
                
            # Process what we can with our inventory
            exec_qty = min(quantity, available)
            print(f"DEBUG: {self.name} executing {exec_qty} of {quantity} {side} order")
            
            # Dealer cash increases by exec_qty * price, bond decreases
            self.update_inventory(-exec_qty, exec_qty * trade_price)
            print(f"DEBUG: After execution, inventory: {self.inventory_bond}, cash: {self.inventory_cash}")
            
            # Create asset-liability entries if system and counterparty are provided
            if system and counterparty and hasattr(system, 'create_asset_liability_pair'):
                # Create a zero-coupon bond (trader buys, dealer sells)
                now = datetime.now()
                maturity_date = datetime(2100, 1, 1)  # t2
                pair = AssetLiabilityPair(
                    time=now,
                    type=EntryType.BOND_ZERO_COUPON.value,
                    amount=exec_qty,
                    denomination="USD",
                    maturity_type=MaturityType.FIXED_DATE,
                    maturity_date=maturity_date,
                    settlement_type=SettlementType.MEANS_OF_PAYMENT,
                    settlement_denomination="USD",
                    asset_holder=counterparty,  # Buyer gets the bond
                    liability_holder=self,      # Dealer issues the bond
                )
                system.create_asset_liability_pair(pair)
                print(f"Created asset-liability pair: {counterparty.name} buys {exec_qty} bonds from {self.name}")
            
            # Only forward remainder if needed
            if quantity > exec_qty:
                remainder = quantity - exec_qty
                self.layoff_volume += remainder
                print(f"DEBUG: {self.name} forwarding remaining {remainder} to VBT")
                try:
                    # Forward excess order to value trader with same price
                    self.value_trader.handle_order(side, remainder, price=trade_price, system=system, counterparty=counterparty)
                except Exception as e:
                    print(f"Warning: Could not forward excess order to value trader: {str(e)}")
                    
        elif side == 'sell':
            # Trader sells bonds to dealer (dealer buys)
            space = self.inventory_limit - self.inventory_bond
            if space <= 0:
                print(f"DEBUG: {self.name} at inventory limit, forwarding full order to VBT")
                self.layoff_volume += quantity
                try:
                    # Forward entire order to value trader
                    self.value_trader.handle_order(side, quantity, price=trade_price, system=system, counterparty=counterparty)
                except Exception as e:
                    print(f"Warning: Could not forward order to value trader: {str(e)}")
                return  # Exit early, nothing to process locally
                
            # Process what we can within our limits
            exec_qty = min(quantity, space)
            print(f"DEBUG: {self.name} executing {exec_qty} of {quantity} {side} order")
            
            # Dealer bond increases, cash decreases by exec_qty * price
            self.update_inventory(exec_qty, -exec_qty * trade_price)
            print(f"DEBUG: After execution, inventory: {self.inventory_bond}, cash: {self.inventory_cash}")
            
            # Create asset-liability entries if system and counterparty are provided
            if system and counterparty and hasattr(system, 'create_asset_liability_pair'):
                # Create a zero-coupon bond (dealer buys, trader sells)
                now = datetime.now()
                maturity_date = datetime(2100, 1, 1)  # t2
                pair = AssetLiabilityPair(
                    time=now,
                    type=EntryType.BOND_ZERO_COUPON.value,
                    amount=exec_qty,
                    denomination="USD",
                    maturity_type=MaturityType.FIXED_DATE,
                    maturity_date=maturity_date,
                    settlement_type=SettlementType.MEANS_OF_PAYMENT,
                    settlement_denomination="USD",
                    asset_holder=self,          # Dealer gets the bond
                    liability_holder=counterparty,  # Seller issues the bond
                )
                system.create_asset_liability_pair(pair)
                print(f"Created asset-liability pair: {self.name} buys {exec_qty} bonds from {counterparty.name}")
            
            # Only forward remainder if needed
            if quantity > exec_qty:
                remainder = quantity - exec_qty
                self.layoff_volume += remainder
                print(f"DEBUG: {self.name} forwarding remaining {remainder} to VBT")
                try:
                    # Forward excess order to value trader with same price  
                    self.value_trader.handle_order(side, remainder, price=trade_price, system=system, counterparty=counterparty)
                except Exception as e:
                    print(f"Warning: Could not forward excess order to value trader: {str(e)}")
        else:
            raise ValueError("Order side must be 'buy' or 'sell'")

    def get_trading_statistics(self):
        """Calculate and return trading statistics"""
        if not self.trade_prices:
            return {
                'final_inventory': self.inventory_bond,
                'layoff_volume': self.layoff_volume,
                'total_volume': self.total_volume,
                'trade_count': self.trade_count,
                'pnl': self.inventory_cash - self.initial_cash,
                'avg_spread': 0,
                'min_spread': 0,
                'max_spread': 0,
                'avg_trade_price': 0,
                'avg_trade_volume': 0
            }
            
        return {
            'final_inventory': self.inventory_bond,
            'layoff_volume': self.layoff_volume,
            'total_volume': self.total_volume,
            'trade_count': self.trade_count,
            'pnl': self.inventory_cash - self.initial_cash,
            'avg_spread': sum(self.spreads) / len(self.spreads),
            'min_spread': min(self.spreads),
            'max_spread': max(self.spreads),
            'avg_trade_price': sum(self.trade_prices) / len(self.trade_prices),
            'avg_trade_volume': sum(self.trade_volumes) / len(self.trade_volumes)
        }

    def update_vbt_quotes(self):
        """Stub: update self.P_b, self.P_a for next cycle."""
        # currently: keep self.P_b, self.P_a unchanged
        pass

class ValueBasedTrader(Agent):
    """
    Value-Based Trader (VBT) stub.
    Returns simulated bid/ask prices for use by SecurityDealer.
    """
    def __init__(self, name, base_price=100.0, spread=0.02):
        super().__init__(name, AgentType.OTHER)
        self.base_price = base_price
        self.spread = spread
        self.inventory_bond = 0  # Track bonds inventory
        self.inventory_cash = 0.0  # Track cash

    def get_prices(self):
        """Return a simulated bid/ask around base_price."""
        bid = self.base_price * (1 - self.spread / 2)
        ask = self.base_price * (1 + self.spread / 2)
        return bid, ask
        
    def handle_order(self, side, quantity, price=None, system=None):
        """
        Process orders from dealers that exceed their inventory limits.
        
        Args:
            side: 'buy' or 'sell'
            quantity: Number of units to trade
            price: Optional price, if None use VBT prices
            system: EconomicSystem instance for creating asset-liability pairs
        """
        bid, ask = self.get_prices()
        
        # Determine execution price
        trade_price = price if price is not None else (ask if side == 'buy' else bid)
        
        if side == 'buy':
            # VBT sells bonds to the client
            # Increase cash, decrease bonds
            self.inventory_cash += quantity * trade_price
            self.inventory_bond -= quantity
            print(f"  {self.name} executed layoff order: sold {quantity} @ {trade_price:.2f}")
            
            # Create bond asset for the buyer if system is provided
            if system and hasattr(system, 'create_asset_liability_pair'):
                # Find counterparty (the dealer's counterparty)
                counterparties = [a for a in system.agents.values() if a.type == AgentType.COMPANY]
                if counterparties:
                    # Create Bond Zero Coupon (buyer gets bond, VBT issues liability)
                    now = datetime.now()
                    maturity_date = datetime(2100, 1, 1)  # t2
                    pair = AssetLiabilityPair(
                        time=now,
                        type=EntryType.BOND_ZERO_COUPON.value,
                        amount=quantity,
                        denomination="USD",
                        maturity_type=MaturityType.FIXED_DATE,
                        maturity_date=maturity_date,
                        settlement_type=SettlementType.MEANS_OF_PAYMENT,
                        settlement_denomination="USD",
                        asset_holder=counterparties[0],  # Buyer gets the bond
                        liability_holder=self,  # VBT issues the bond
                    )
                    system.create_asset_liability_pair(pair)
                    print(f"  Created bond asset-liability pair: {counterparties[0].name} <-> {self.name}")
        
        elif side == 'sell':
            # VBT buys bonds from the client
            # Decrease cash, increase bonds
            self.inventory_cash -= quantity * trade_price
            self.inventory_bond += quantity
            print(f"  {self.name} executed layoff order: bought {quantity} @ {trade_price:.2f}")
            
            # Create bond liability for the seller if system is provided
            if system and hasattr(system, 'create_asset_liability_pair'):
                # Find counterparty (the dealer's counterparty)
                counterparties = [a for a in system.agents.values() if a.type == AgentType.COMPANY]
                if counterparties:
                    # Create Bond Zero Coupon (VBT gets bond, seller issues liability)
                    now = datetime.now()
                    maturity_date = datetime(2100, 1, 1)  # t2
                    pair = AssetLiabilityPair(
                        time=now,
                        type=EntryType.BOND_ZERO_COUPON.value,
                        amount=quantity,
                        denomination="USD",
                        maturity_type=MaturityType.FIXED_DATE,
                        maturity_date=maturity_date,
                        settlement_type=SettlementType.MEANS_OF_PAYMENT,
                        settlement_denomination="USD",
                        asset_holder=self,  # VBT gets the bond
                        liability_holder=counterparties[0],  # Seller issues the bond
                    )
                    system.create_asset_liability_pair(pair)
                    print(f"  Created bond asset-liability pair: {self.name} <-> {counterparties[0].name}")
        else:
            raise ValueError("Order side must be 'buy' or 'sell'")

if __name__ == "__main__":
    main()
