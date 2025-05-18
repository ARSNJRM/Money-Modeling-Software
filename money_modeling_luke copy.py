# pip install openpyxl==3.1.2

# Economic Balance Sheet Simulation
# ============================
# To use this code in Google Colab:
# 1. First run: !pip install openpyxl==3.1.2
# 2. Then paste this entire code into a new cell
# 3. Finally run: main()
#
# The simulation allows you to:
# - Create agents (banks, companies, households, etc.)
# - Create asset-liability pairs between agents
# - Track balance sheets over time (t0, t1, t2)
# - Handle settlements and maturity
# - Export to Excel (requires openpyxl)

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl package not found. Excel export functionality will be disabled.")
    print("To enable Excel export, please install openpyxl using: pip install openpyxl")
    EXCEL_AVAILABLE = False

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union
from enum import Enum
from datetime import datetime, timedelta
from copy import deepcopy

# Core Enums and Classes
class AgentType(Enum): 
    BANK = "bank"
    COMPANY = "company"
    HOUSEHOLD = "household"
    TREASURY = "treasury"
    CENTRAL_BANK = "central_bank"
    OTHER = "other"

class EntryType(Enum):
    LOAN = "loan"
    DEPOSIT = "deposit"  # No maturity or settlement type
    PAYABLE = "payable"  # Always means of payment settlement
    BOND = "bond"
    DELIVERY_CLAIM = "delivery_claim"  # Always non-financial asset settlement
    NON_FINANCIAL = "non_financial"  # No maturity or settlement type
    DEFAULT = "default"  # Used when settlement fails

class MaturityType(Enum):
    ON_DEMAND = "on_demand"
    FIXED_DATE = "fixed_date"
    PERPETUAL = "perpetual"

class SettlementType(Enum):
    MEANS_OF_PAYMENT = "means_of_payment"  # Standard payment means (bank transfer, check, etc.)
    SECURITIES = "securities"  # Financial instruments
    NON_FINANCIAL_ASSET = "non_financial_asset"  # Physical assets
    SERVICES = "services"  # Services rendered
    CRYPTO = "crypto"  # Cryptocurrency
    NONE = "none"  # No settlement type

@dataclass
class SettlementDetails:
    type: SettlementType
    denomination: str  # Currency or unit of settlement

@dataclass
class BalanceSheetEntry:
    type: EntryType
    is_asset: bool  # True for assets, False for liabilities
    counterparty: Optional[str]  # Optional for non-financial entries
    initial_book_value: float
    # amount: float  ## we are trying to get rid of amount in favor of more detailed indicators
    denomination: str
    maturity_type: MaturityType
    maturity_date: Optional[int]  # Required for FIXED_DATE, None for others
    settlement_details: SettlementDetails
    cash_flow_at_maturity: float = 0
    name: Optional[str] = None  # For non-financial assets or special naming
    issuance_time: int = 0  # Time the entry was created
    current_book_value: float = 0 

    def matches(self, other: 'BalanceSheetEntry') -> bool:
        """Check if two entries match (used for removing entries)"""
        return (
            self.type == other.type and
            self.is_asset == other.is_asset and
            self.counterparty == other.counterparty and
            self.initial_book_value == other.initial_book_value and
            self.denomination == other.denomination and
            self.maturity_type == other.maturity_type and
            self.maturity_date == other.maturity_date and
            self.settlement_details.type == other.settlement_details.type and
            self.settlement_details.denomination == other.settlement_details.denomination and
            self.name == other.name and
            self.issuance_time == other.issuance_time
        )

    def __post_init__(self):
        if self.initial_book_value <= 0:
            raise ValueError("Amount must be positive")

        # Validate issuance time
        if self.issuance_time < 0: # might be an unnecessary check, left it in case
            raise ValueError("Issuance time must be non-negative")

        # Validate counterparty rules
        if self.type != EntryType.NON_FINANCIAL and not self.counterparty:
            raise ValueError("Counterparty is required for financial entries")
        if self.type == EntryType.NON_FINANCIAL and self.counterparty:
            raise ValueError("Non-financial entries cannot have a counterparty")

        # Validate name rules
        if self.type == EntryType.NON_FINANCIAL and not self.name:
            raise ValueError("Non-financial entries must have a name")

        # Validate payable rules
        if self.type == EntryType.PAYABLE and self.settlement_details.type != SettlementType.MEANS_OF_PAYMENT:
            raise ValueError("Payable entries must have means_of_payment settlement type")
        
        self.current_book_value = self.initial_book_value

class SettlementFailure(Exception):
    def __init__(self, agent_name: str, entry: BalanceSheetEntry, reason: str):
        self.agent_name = agent_name
        self.entry = entry
        self.reason = reason
        super().__init__(f"Settlement failure for {agent_name}: {reason}")

class Agent:
    def __init__(self, name: str, agent_type: AgentType):
        self.name = name
        self.type = agent_type
        self.assets: List[BalanceSheetEntry] = []
        self.liabilities: List[BalanceSheetEntry] = []
        self.status: str = "operating"  # operating or bankrupt
        self.creation_time: datetime = datetime.now()
        # Add settlement history
        self.settlement_history = {
            'as_asset_holder': [],  # Settlements where this agent was the creditor
            'as_liability_holder': []  # Settlements where this agent was the debtor
        }

    def add_asset(self, entry: BalanceSheetEntry):
        self.assets.append(entry)

    def add_liability(self, entry: BalanceSheetEntry):
        self.liabilities.append(entry)

    def remove_asset(self, entry: BalanceSheetEntry):
        self.assets = [e for e in self.assets if not e.matches(entry)]

    def remove_liability(self, entry: BalanceSheetEntry):
        self.liabilities = [e for e in self.liabilities if not e.matches(entry)]

    def get_balance_sheet(self) -> Dict:
        return {
            "assets": self.assets,
            "liabilities": self.liabilities
        }

    def get_total_assets(self) -> float:
        return sum(entry.current_book_value for entry in self.assets)

    def get_total_liabilities(self) -> float:
        return sum(entry.current_book_value for entry in self.liabilities)

    def get_net_worth(self) -> float:
        return self.get_total_assets() - self.get_total_liabilities()

    def get_type_specific_metrics(self) -> Dict:
        metrics = {
            "name": self.name,
            "type": self.type.value,
            "creation_time": self.creation_time,
            "status": self.status,
            "total_assets": self.get_total_assets(),
            "total_liabilities": self.get_total_liabilities(),
            "net_worth": self.get_net_worth()
        }

        if self.type == AgentType.BANK:
            metrics["capital_ratio"] = self.get_total_assets() / self.get_total_liabilities() if self.get_total_liabilities() > 0 else float('inf')
        elif self.type == AgentType.COMPANY:
            metrics["leverage_ratio"] = self.get_total_liabilities() / self.get_total_assets() if self.get_total_assets() > 0 else float('inf')
        elif self.type == AgentType.HOUSEHOLD:
            metrics["savings_rate"] = (self.get_total_assets() - self.get_total_liabilities()) / self.get_total_assets() if self.get_total_assets() > 0 else 0

        return metrics

    def record_settlement(self,
                         time_point: int,
                         original_entry: BalanceSheetEntry,
                         settlement_result: BalanceSheetEntry,
                         counterparty: str,
                         as_asset_holder: bool):
        """Record a settlement in the agent's history"""
        settlement_record = {
            'time_point': time_point,
            'original_entry': deepcopy(original_entry),
            'settlement_result': deepcopy(settlement_result),
            'counterparty': counterparty,
            'timestamp': datetime.now()
        }
        if as_asset_holder:
            self.settlement_history['as_asset_holder'].append(settlement_record)
        else:
            self.settlement_history['as_liability_holder'].append(settlement_record)

class AssetLiabilityPair:
    def __init__(self,
                 time: int,
                 type: str,
                 amount: float,
                 denomination: str,
                 maturity_type: MaturityType,
                 maturity_date: Optional[int],
                 settlement_type: SettlementType,
                 settlement_denomination: str,
                 asset_holder: Agent,
                 liability_holder: Optional[Agent] = None,
                 cash_flow_at_maturity: Optional[float] = 0,
                 asset_name: Optional[str] = None):  # New parameter for non-financial asset names
        self.time = time
        self.type = type
        self.amount = amount
        self.denomination = denomination
        self.maturity_type = maturity_type
        self.maturity_date = maturity_date
        self.cash_flow_at_maturity = cash_flow_at_maturity
        self.settlement_details = SettlementDetails(
            type=settlement_type,
            denomination=settlement_denomination
        )
        self.asset_holder = asset_holder
        self.liability_holder = liability_holder
        self.asset_name = asset_name

        if type == EntryType.NON_FINANCIAL.value:
            if liability_holder is not None:
                raise ValueError("Non-financial entries cannot have a liability holder")
            if not asset_name:
                raise ValueError("Non-financial entries must have an asset name")
        else:
            if liability_holder is None:
                raise ValueError("Financial entries must have a liability holder")

    def create_entries(self) -> Tuple[BalanceSheetEntry, Optional[BalanceSheetEntry]]:
        # For delivery claim entries
        if self.type == EntryType.DELIVERY_CLAIM.value:
            if not self.asset_name:
                raise ValueError("Delivery claim must specify the asset to be delivered")

            settlement_details = SettlementDetails(
                type=SettlementType.NON_FINANCIAL_ASSET,
                denomination=self.settlement_details.denomination
            )

            # Create delivery claim (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,  # Name of the asset to be delivered
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # Create delivery promise (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,  # Name of the asset to be delivered
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            return asset_entry, liability_entry

        # For payable entries (receivable-payable pairs)
        elif self.type == EntryType.PAYABLE.value:
            settlement_details = SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=self.settlement_details.denomination
            )

            # Create receivable (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE,  # Both sides are PAYABLE type
                is_asset=True,
                counterparty=self.liability_holder.name if self.liability_holder else None,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # Create payable (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=None,
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            return asset_entry, liability_entry

        # For non-financial entries
        if self.type == EntryType.NON_FINANCIAL.value:
            asset_entry = BalanceSheetEntry(
                type=EntryType.NON_FINANCIAL,
                is_asset=True,
                counterparty=None,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )
            return asset_entry, None

        # For all other types (LOAN, DEPOSIT, BOND)
        asset_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=True,
            counterparty=self.liability_holder.name if self.liability_holder else None,
            initial_book_value=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=self.asset_name,
            issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
            cash_flow_at_maturity=self.cash_flow_at_maturity
        )

        liability_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=False,
            counterparty=self.asset_holder.name,
            initial_book_value=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=None,  # Liabilities don't need names
            issuance_time=self.current_time_state if hasattr(self, 'current_time_state') else 0,
            cash_flow_at_maturity=self.cash_flow_at_maturity
        )

        return asset_entry, liability_entry

class EconomicSystem:
    def __init__(self):
        self.agents: Dict[str, Agent] = {}  # Current state
        self.asset_liability_pairs: List[AssetLiabilityPair] = []
        self.time_states: Dict[int, Dict[str, Agent]] = {}  # States at different time points
        self.current_time_state = 0  # Track current time state
        self.final_time_state = 999 
        self.simulation_finalized = False  # Track if simulation is finalized
        self.save_state(0) # Initialize state time 0
        self.scheduled_actions: Dict[int, List[Tuple[str, Dict]]] = {} #[time, [(action, args for function)]]

    def schedule_action(self, time_point: int, action_name: str, params: Dict):
        if time_point not in self.scheduled_actions:
            self.scheduled_actions[time_point] = []
        self.scheduled_actions[time_point].append((action_name, params))
        
    def validate_time_point(self, time_point: int, allow_t0: bool = True) -> None: 
        """Validate a time point string"""
        valid_points = (range(0,self.final_time_state) if allow_t0 
                        else range(1,self.final_time_state))
        if time_point not in valid_points:
            raise ValueError(f"Time point must in range 0 - {self.final_time_state}")

    def add_agent(self, agent: Agent):
        self.agents[agent.name] = agent

        # Auto-save state time 0 when agent is added
        if self.current_time_state == 0:
            self.save_state(0)

    def create_asset_liability_pair(self, pair: AssetLiabilityPair):
        self.asset_liability_pairs.append(pair)
        asset_entry, liability_entry = pair.create_entries()
        pair.asset_holder.add_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.add_liability(liability_entry)

        # Auto-save state based on current time point
        self.save_state(self.current_time_state)

    def get_time_points(self) -> List[str]:
        """Get all time points in order and converts int i to str ti"""
        times = range(0,self.final_time_state)
        return [f"t{time}" for time in times]

    def save_state(self, time_point: int):
        """Save current state for a given time point"""
        self.validate_time_point(time_point)
        self.time_states[time_point] = {}
        for name, agent in self.agents.items():
            agent_copy = deepcopy(agent)
            self.time_states[time_point][name] = agent_copy

        self.current_time_state = time_point

    def settle_entries(self, time_point: int):
        """Settle entries at a specific time point"""
        self.validate_time_point(time_point, allow_t0=False)

        # First save the current state at the previous time point
        prev_time = time_point - 1
        if prev_time not in self.time_states:
            self.save_state(prev_time)

        # Process all entries that are due at this time point
        for pair in self.asset_liability_pairs[:]:  # Create a copy to iterate over
            if (pair.maturity_type == MaturityType.FIXED_DATE and
                pair.maturity_date):
                # Check if the entry's maturity date matches our time state
                entry_time = pair.maturity_date

                if entry_time == time_point:
                    # Remove the original pair
                    self.asset_liability_pairs.remove(pair)
                    asset_entry, liability_entry = pair.create_entries()
                    pair.asset_holder.remove_asset(asset_entry)
                    if liability_entry:
                        pair.liability_holder.remove_liability(liability_entry)

                    # Handle settlement based on type
                    if pair.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
                        # Find the bank deposit that will be used for settlement
                        debtor_deposit = next(
                            (asset for asset in pair.liability_holder.assets
                             if asset.type == EntryType.DEPOSIT
                             and asset.initial_book_value >= pair.amount
                             and asset.denomination == pair.denomination),
                            None
                        )

                        if not debtor_deposit:
                            raise ValueError(f"No suitable deposit found for settlement")

                        # Get the bank that holds the deposit liability
                        bank = next(a for a in self.agents.values() if a.name == debtor_deposit.counterparty)

                        # Remove the original deposit from the debtor
                        pair.liability_holder.remove_asset(debtor_deposit)

                        # Remove the corresponding liability from the bank
                        bank_liability = next(
                            (l for l in bank.liabilities
                             if l.type == EntryType.DEPOSIT
                             and l.counterparty == pair.liability_holder.name
                             and l.initial_book_value == debtor_deposit.initial_book_value),
                            None
                        )
                        if bank_liability:
                            bank.remove_liability(bank_liability)

                        # Create new deposit entry for the creditor
                        settlement_pair = AssetLiabilityPair(
                            time=datetime.now(),
                            type=EntryType.DEPOSIT.value,
                            amount=pair.amount,
                            denomination=pair.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=pair.denomination,
                            asset_holder=pair.asset_holder,  # Creditor gets the deposit
                            liability_holder=bank,  # Bank keeps liability
                            asset_name=None
                        )

                        # Create entries with current time point as issuance time
                        new_asset_entry, new_liability_entry = settlement_pair.create_entries()
                        new_asset_entry.issuance_time = time_point
                        if new_liability_entry:
                            new_liability_entry.issuance_time = time_point

                        # Record settlement history
                        pair.asset_holder.record_settlement(
                            time_point=time_point,
                            original_entry=asset_entry,
                            settlement_result=new_asset_entry,
                            counterparty=pair.liability_holder.name,
                            as_asset_holder=True
                        )
                        pair.liability_holder.record_settlement(
                            time_point=time_point,
                            original_entry=liability_entry,
                            settlement_result=debtor_deposit,  # The deposit that was used
                            counterparty=pair.asset_holder.name,
                            as_asset_holder=False
                        )

                        # Add entries
                        settlement_pair.asset_holder.add_asset(new_asset_entry)
                        if new_liability_entry:
                            settlement_pair.liability_holder.add_liability(new_liability_entry)
                        self.asset_liability_pairs.append(settlement_pair)

                        # If there was remaining deposit amount, create a new deposit for the remainder
                        if debtor_deposit.initial_book_value > pair.amount:
                            remainder_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.DEPOSIT.value,
                                amount=debtor_deposit.initial_book_value - pair.amount,
                                denomination=pair.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=pair.denomination,
                                asset_holder=pair.liability_holder,  # Debtor keeps remainder
                                liability_holder=bank,  # Bank keeps liability
                                asset_name=None
                            )

                            # Create entries with current time point as issuance time
                            remainder_asset, remainder_liability = remainder_pair.create_entries()
                            remainder_asset.issuance_time = time_point
                            if remainder_liability:
                                remainder_liability.issuance_time = time_point

                            # Add entries
                            remainder_pair.asset_holder.add_asset(remainder_asset)
                            if remainder_liability:
                                remainder_pair.liability_holder.add_liability(remainder_liability)
                            self.asset_liability_pairs.append(remainder_pair)

                    elif pair.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
                        # Find and remove the non-financial asset from the liability holder
                        non_financial_asset = next(
                            (asset for asset in pair.liability_holder.assets
                             if asset.type == EntryType.NON_FINANCIAL
                             and asset.name == pair.asset_name
                             and asset.initial_book_value >= pair.amount),
                            None
                        )

                        if not non_financial_asset:
                            raise ValueError(f"Non-financial asset {pair.asset_name} not found for settlement")

                        # Remove the asset from the liability holder
                        pair.liability_holder.remove_asset(non_financial_asset)

                        # Create non-financial asset entry for the asset holder
                        settlement_pair = AssetLiabilityPair(
                            time=datetime.now(),
                            type=EntryType.NON_FINANCIAL.value,
                            amount=pair.amount,
                            denomination=pair.settlement_details.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=pair.settlement_details.denomination,
                            asset_holder=pair.asset_holder,  # Original creditor gets the goods
                            liability_holder=None,  # Non-financial assets have no liability holder
                            asset_name=pair.asset_name  # Use the asset_name directly
                        )
                        # Create entry with current time point as issuance time
                        new_asset_entry, _ = settlement_pair.create_entries()
                        new_asset_entry.issuance_time = time_point

                        # Record settlement history
                        pair.asset_holder.record_settlement(
                            time_point=time_point,
                            original_entry=asset_entry,
                            settlement_result=new_asset_entry,
                            counterparty=pair.liability_holder.name,
                            as_asset_holder=True
                        )
                        pair.liability_holder.record_settlement(
                            time_point=time_point,
                            original_entry=liability_entry,
                            settlement_result=non_financial_asset,  # The non-financial asset that was delivered
                            counterparty=pair.asset_holder.name,
                            as_asset_holder=False
                        )

                        # Add entry directly to avoid default t0 issuance time
                        settlement_pair.asset_holder.add_asset(new_asset_entry)
                        self.asset_liability_pairs.append(settlement_pair)

                        # If there was remaining amount in the non-financial asset, create a new entry for it
                        if non_financial_asset.initial_book_value > pair.amount:
                            remainder_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.NON_FINANCIAL.value,
                                amount=non_financial_asset.initial_book_value - pair.amount,
                                denomination=non_financial_asset.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=non_financial_asset.denomination,
                                asset_holder=pair.liability_holder,  # Original holder keeps remainder
                                liability_holder=None,
                                asset_name=non_financial_asset.name
                            )

                            # Create entry with current time point as issuance time
                            remainder_asset, _ = remainder_pair.create_entries()
                            remainder_asset.issuance_time = time_point

                            # Add entry
                            remainder_pair.asset_holder.add_asset(remainder_asset)
                            self.asset_liability_pairs.append(remainder_pair)

        # Auto-save state after settlements
        self.save_state(time_point)
        self.current_time_state = time_point

    def adj_book_values(self):
        """Adjusts current book value, needs input t(i)"""

        def adj_value(bal_entry: BalanceSheetEntry):
            t = self.current_time_state - bal_entry.issuance_time
            m = bal_entry.maturity_date - bal_entry.issuance_time # what happens if m = 0, can't divide by zero
            rate_of_adjustment = (bal_entry.cash_flow_at_maturity / bal_entry.initial_book_value)**(1/m) - 1
            bal_entry.current_book_value = bal_entry.initial_book_value * (1 + rate_of_adjustment)**t

        for agent in self.agents.values():
                for liability in agent.liabilities:
                    if liability.maturity_type == MaturityType.FIXED_DATE:
                        adj_value(liability)

                for asset in agent.assets:
                    if asset.maturity_type == MaturityType.FIXED_DATE:
                        adj_value(asset)

    def get_agents_at_time(self, time_point: int) -> Dict[str, Agent]:
        """Get agents state at a specific time point"""
        self.validate_time_point(time_point)

        # For t0, always show current state
        if time_point == 0:
            return {name: agent for name, agent in self.agents.items()}

        # For other time points, use saved state if available
        if time_point in self.time_states:
            return self.time_states[time_point]

        # If state not saved yet and we're looking at a future point,
        # we need to process settlements up to that point
        if time_point > self.current_time_state:
            # Save current state
            current_state = deepcopy(self.agents)
            current_time = self.current_time_state

            # Process settlements for each time point up to the requested one
            try:
                for t in range(self.current_time_state + 1, time_point + 1):
                    self.settle_entries(t)

                # Get the state after settlements
                result = {name: agent for name, agent in self.agents.items()}

                # Restore original state
                self.agents = current_state
                self.current_time_state = current_time

                return result
            except Exception as e:
                # If settlement fails, restore original state and return it
                self.agents = current_state
                self.current_time_state = current_time
                print(f"\nWarning: Could not process settlements ({str(e)})")
                return current_state

        # If none of the above, return current state
        return {name: agent for name, agent in self.agents.items()}

    def compute_changes(self, from_time: int, to_time: int) -> Dict[int, Dict[str, List]]: 
        """Compute changes between two time points"""
        if from_time not in self.time_states or to_time not in self.time_states:
            raise ValueError(f"Missing state for time point {from_time} or {to_time}")

        changes = {}
        for name, to_agent in self.time_states[to_time].items():
            from_agent = self.time_states[from_time][name]

            # Find new and removed assets
            new_assets = [a for a in to_agent.assets if not any(a.matches(from_a) for from_a in from_agent.assets)]
            removed_assets = [a for a in from_agent.assets if not any(a.matches(to_a) for to_a in to_agent.assets)]

            # Find new and removed liabilities
            new_liabilities = [l for l in to_agent.liabilities if not any(l.matches(from_l) for from_l in from_agent.liabilities)]
            removed_liabilities = [l for l in from_agent.liabilities if not any(l.matches(to_l) for to_l in to_agent.liabilities)]

            changes[name] = {
                'new_assets': new_assets,
                'removed_assets': removed_assets,
                'new_liabilities': new_liabilities,
                'removed_liabilities': removed_liabilities
            }

        return changes

    def can_settle_entry(self, agent: Agent, entry: BalanceSheetEntry) -> Tuple[bool, str]:
        """Check if an agent can settle a liability"""
        if entry.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
            # Check for sufficient deposits
            deposits = sum(asset.initial_book_value for asset in agent.assets
                          if asset.type == EntryType.DEPOSIT
                          and asset.is_asset
                          and asset.denomination == entry.denomination)
            if deposits < entry.initial_book_value:
                return False, f"Insufficient deposits: has {deposits} {entry.denomination}, needs {entry.initial_book_value}"

        elif entry.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            # Check for required non-financial asset
            has_asset = any(asset.type == EntryType.NON_FINANCIAL
                           and asset.name == entry.name
                           and asset.initial_book_value >= entry.initial_book_value
                           for asset in agent.assets)
            if not has_asset:
                return False, f"Missing required non-financial asset: {entry.name}"

        return True, "OK"

    def create_default_entries(self, failed_entry: BalanceSheetEntry) -> Tuple[BalanceSheetEntry, BalanceSheetEntry]: 
        """Create default claim and liability entries when settlement fails"""
        # Create default claim for the creditor
        default_claim = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.initial_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        # Create default liability for the debtor
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.initial_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default on {failed_entry.type.value}",
            issuance_time=self.current_time_state
        )

        return default_claim, default_liability
    
    def run_user_scheduled_actions(self):
        '''run actions explicitly scheduled by user'''
        actions = self.scheduled_actions.get(self.current_time_state, [])
        if not actions:
            return
         # skips if the time point has no scheduled actions
        
        action_map = {
            'Create Asset–Liability Pair': self.create_asset_liability_pair,
            'Issue Loan': self.disburse_loans,
            'Request Loan': self.submit_loan_requests,
            'Quote Loan Terms': self.return_loan_quotes,
            'Commit to Loan': self.commit_to_terms,
            'Repay Loan': self.execute_repayments,
            'Default on Loan': self.resolve_defaults,
            'Issue Bond': self.issue_securities,
            'Request Bond Quote': self.submit_trade_quotes,
            'Dealer Bond Quote': self.return_trade_quotes,
            'Commit to Bond Purchase': self.commit_to_terms,
            'Execute Bond Trade': self.dealer_execute_trades,
            'Repay Bond': self.execute_repayments,
            'Default on Bond': self.resolve_defaults,
            'Issue Share (One-Time)': self.issue_securities,
            'Request Share Quote': self.submit_trade_quotes,
            'Dealer Quote': self.return_trade_quotes,
            'Commit to Purchase': self.commit_to_terms,
            'Execute Trade': self.dealer_execute_trades,
            'Transfer Deposit': self.reallocate_deposits,
            'Net Interbank IOUs': self.update_interbank_ious,
            'Rollover IOU Request': self.evaluate_iou_rollovers,
            'Respond to Rollover': self.evaluate_iou_rollovers,
            'Request Overnight Loan': self.request_cb_loans,
            'Issue Overnight Loan': self.disburse_cb_loans,
            'Settle with Reserves': self.settle_final_reserve_obligations,
            'Regulatory Check (RRR, CAR)': self.check_regulatory_compliance,
            'Request Regulatory Support': self.request_regulatory_support,
            'Issue Regulatory Support': self.provide_regulatory_support,
            'Compute Metrics': self.compute_system_metrics,
            'Update Internal Info': self.update_agent_metrics,
            'Dealer Inventory Valuation': self.update_dealer_quotes
        }

        for action_name, params in actions:
            method = action_map.get(action_name)
            if not method:
                # skip unimplemented or unknown actions
                continue
            try:
                if params:
                    method(**params)
                else:
                    method()
            except TypeError:
                # fallback if method signature doesn't accept params
                method()
        # Clear executed actions for this time point
        self.scheduled_actions.pop(self.current_time_state, None)
    
    def update_agent_metrics(self):
        '''agents recalculate internal balance sheet metrics'''
        pass

    def nonbank_actions(self):
        '''general transactions for non-banks'''
        pass

    def issue_securities(self):
        '''companies and banks can issue new bonds and first time shares'''
        pass

    def update_dealer_quotes(self): ## think these three functions can be combined
        '''dealers compute their prices'''
        pass

    def submit_trade_quotes(self): ##
        '''agents send quote request to dealers'''
        pass

    def return_trade_quotes(self): ## 
        '''dealers return their quotes to agents'''
        pass 

    def submit_loan_requests(self): ### these four as well
        '''agents request banks for loans'''
        pass

    def return_loan_quotes(self): ###
        '''banks evaluate and quote terms'''
        pass

    def commit_to_terms(self): ###
        '''borrowers accept loan offers'''
        pass

    def disburse_loans(self): ###
        '''loan is created'''
        pass

    def finalize_trade_commitments(self):
        '''agents confirm the trades they plan to execute'''
        pass

    def dealer_execute_trades(self):
        '''dealers trade and update portfolios'''
        pass

    def reallocate_deposits(self):
        '''deposits are transferred and interbank IOUs are updated'''
        pass

    def verify_settlement_needs(self):
        '''system checks agents due obligations and means of payment'''
        maturing = [
            (agent, l) 
            for agent in self.agents.values()
            for l in agent.liabilities
            if l.maturity_type == MaturityType.FIXED_DATE 
               and l.maturity_date == self.current_time_state
        ]
        for agent, liability in maturing:
            can, reason = self.can_settle_entry(agent, liability)
            if not can:
                agent.status = "defaulting"

    def execute_repayments(self):
        '''scheduled payments are executed, updating balance sheets and flagging defaulting agents'''
        maturing = [
            (agent, l) 
            for agent in self.agents.values()
            for l in agent.liabilities
            if l.maturity_type == MaturityType.FIXED_DATE 
               and l.maturity_date == self.current_time_state
        ]
        for agent, liability in maturing:
            can_settle, reason = self.can_settle_entry(agent, liability)

            if not can_settle:
                print(f"\nDEFAULT DETECTED: {agent.name} cannot settle {liability.type.value} ({reason})")
                agent.status = 'defaulting'

            if self.current_time_state > 0:
                self.settle_entries(self.current_time_state)

    def resolve_nonbank_defaults(self):
        '''Process all non-bank agents marked as 'defaulting': recover assets and allocate to creditors.'''
        
        # find all non-bank agents in defaulting state
        defaulting_agents = [agent for agent in self.agents.values()
                              if agent.status == 'defaulting' and not agent.is_bank]

        for defaulting_agent in defaulting_agents:
            # aggregate transferable assets
            total_recovery = 0.0
            recovery_assets: List[BalanceSheetEntry] = []
            for asset in list(defaulting_agent.assets):
                if asset.settlement_details.type in (
                    SettlementType.MEANS_OF_PAYMENT,
                    SettlementType.CRYPTO,
                ):
                    total_recovery += asset.current_book_value
                    recovery_assets.append(asset)

            # identify creditors
            creditors: List[Tuple[Agent, BalanceSheetEntry]] = []
            total_claims = 0.0
            for other in self.agents.values():
                for claim in list(other.assets):
                    if claim.counterparty == defaulting_agent.name and claim.type != EntryType.DEFAULT:
                        creditors.append((other, claim))
                        total_claims += claim.current_book_value

            # if no resources or creditors, clear liabilities and continue
            if total_claims <= 0 or total_recovery <= 0 or not creditors:
                defaulting_agent.liabilities.clear()
                continue

            # remove assets from debtor
            for asset in recovery_assets:
                defaulting_agent.remove_asset(asset)

            # distribute recovery pro-rata
            for creditor, claim in creditors:
                share = claim.current_book_value / total_claims * total_recovery
                creditor.remove_asset(claim)
                default_entry = BalanceSheetEntry(
                    type=EntryType.DEFAULT,
                    is_asset=True,
                    counterparty=defaulting_agent.name,
                    initial_book_value=share,
                    denomination=claim.denomination,
                    maturity_type=MaturityType.ON_DEMAND,
                    maturity_date=None,
                    settlement_details=claim.settlement_details,
                    issuance_time=self.current_time_state,
                    name=f"Recovery from {defaulting_agent.name} default",
                )
                creditor.add_asset(default_entry)

            # clear debtor liabilities
            defaulting_agent.liabilities.clear()

    def update_interbank_ious(self):
        '''all intraday IOUs are netted'''
        pass

    def evaluate_iou_rollovers(self): 
        '''banks in defecit may request IOU rollovers, to which surplus banks respond'''
        pass

    def request_cb_loans(self): 
        '''remaining deficit banks request overnight funding from central bank'''
        pass

    def disburse_cb_loans(self): 
        '''central bank disburses to eligible deficit banks'''
        pass

    def settle_final_reserve_obligations(self):
        '''remaining obligations are settled, unresolved payment now leads to default'''
        pass

    def check_regulatory_compliance(self):
        '''system evaluates each banks RRR and CAR'''
        pass

    def request_regulatory_support(self): 
        '''banks failing RRR and CAR checks request assistance from central bank'''
        pass

    def provide_regulatory_support(self): 
        '''central bank provides liquid/capital to restore minimum compliance'''
        pass

    def compute_system_metrics(self):
        '''final state metrics are calculated'''
        pass

    def run_actions(self):
        self.update_agent_metrics()
        self.nonbank_actions()
        self.issue_securities()
        self.update_dealer_quotes()
        self.submit_trade_quotes()
        self.return_trade_quotes()
        self.submit_loan_requests()
        self.return_loan_quotes()
        self.commit_to_terms()
        self.disburse_loans()
        self.finalize_trade_commitments()
        self.dealer_execute_trades()
        self.reallocate_deposits()
        self.verify_settlement_needs()  
        self.execute_repayments()       
        self.resolve_nonbank_defaults() 
        self.update_interbank_ious()
        self.evaluate_iou_rollovers()
        self.request_cb_loans()
        self.disburse_cb_loans()
        self.settle_final_reserve_obligations()
        self.check_regulatory_compliance()
        self.request_regulatory_support()
        self.provide_regulatory_support()
        self.compute_system_metrics()

    def run_simulation(self, time_points) -> bool: 
        """Run the full simulation from t0 to ti, handling settlements and defaults"""
        print("\nStarting simulation from t0...")

        for time in range(time_points):
            print(f"\nProcessing time {time}...")
            self.run_actions()
            
        print("\nSimulation completed successfully!")
        return True

    def display_settlement_history(self):
        """Display settlement history for all agents"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        print("\nSettlement History:")
        for agent_name, agent in self.agents.items():
            print(f"\n{agent_name}'s Settlement History:")

            # Display settlements where agent was asset holder
            if agent.settlement_history.get('as_asset_holder', []):
                print("\n  As Asset Holder:")
                for settlement in agent.settlement_history['as_asset_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Asset: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].initial_book_value} {settlement['original_entry'].denomination}")
                    print(f"    Settled For: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].initial_book_value} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            # Display settlements where agent was liability holder
            if agent.settlement_history.get('as_liability_holder', []):
                print("\n  As Liability Holder:")
                for settlement in agent.settlement_history['as_liability_holder']:
                    print(f"\n    Time: {settlement['time_point']}")
                    print(f"    Original Liability: {settlement['original_entry'].type.value} "
                          f"of {settlement['original_entry'].initial_book_value} {settlement['original_entry'].denomination}")
                    print(f"    Settled With: {settlement['settlement_result'].type.value} "
                          f"of {settlement['settlement_result'].initial_book_value} {settlement['settlement_result'].denomination}")
                    print(f"    Counterparty: {settlement['counterparty']}")

            if (not agent.settlement_history.get('as_asset_holder') and
                not agent.settlement_history.get('as_liability_holder')):
                print("  No settlements recorded")

    def display_balance_sheets(self, time_point: int):
        """Display balance sheets for all agents at a specific time point"""
        if not self.agents:
            print("\nNo agents in the system yet!")
            return

        current_agents = self.get_agents_at_time(time_point).values()
        print(f"\nBalance sheets at t{time_point}:")

        for agent in current_agents:
            print(f"\n{agent.name} ({agent.type.value}):")
            print("Assets:")
            for asset in agent.assets:
                maturity_info = ""
                if asset.maturity_type == MaturityType.FIXED_DATE:
                    maturity_info = f" (matures at t{asset.maturity_date})"

                # Show appropriate entry type
                if asset.type == EntryType.PAYABLE:
                    entry_type = "receivable"
                elif asset.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery claim for {asset.name}" if asset.name else "delivery claim"
                elif asset.type == EntryType.DEFAULT:
                    entry_type = f"default claim ({asset.name})"
                else:
                    entry_type = asset.type.value

                # Skip if entry has matured and been settled
                if asset.maturity_type == MaturityType.FIXED_DATE and time_point > asset.maturity_date:
                    continue

                print(f"  - {entry_type}: {asset.current_book_value} {asset.denomination} "
                      f"(from {asset.counterparty if asset.counterparty else 'N/A'})"
                      f"{maturity_info} [issued at t{asset.issuance_time}]")

            print("Liabilities:")
            for liability in agent.liabilities:
                maturity_info = ""
                if liability.maturity_type == MaturityType.FIXED_DATE:
                    maturity_info = f" (matures at t{liability.maturity_date})"

                # Show appropriate entry type
                if liability.type == EntryType.DELIVERY_CLAIM:
                    entry_type = f"delivery promise for {liability.name}" if liability.name else "delivery promise"
                elif liability.type == EntryType.DEFAULT:
                    entry_type = f"default liability ({liability.name})"
                else:
                    entry_type = liability.type.value

                # Skip if entry has matured and been settled
                if liability.maturity_type == MaturityType.FIXED_DATE and time_point > asset.maturity_date:
                    continue

                print(f"  - {entry_type}: {liability.current_book_value} {liability.denomination} "
                      f"(to {liability.counterparty}){maturity_info} "
                      f"[issued at t{liability.issuance_time}]")

class ExcelExporter:
    def __init__(self, system: EconomicSystem):
        self.system = system

    def create_t_table(self, sheet, row_start: int, col_start: int, agent: Agent, time_point: int):
        thick = Side(style='thick', color='000000')

        # Add time point header
        time_header = sheet.cell(row=row_start, column=1)
        time_header.value = f"Time: t{time_point}"
        time_header.alignment = Alignment(horizontal="center")
        time_header.font = openpyxl.styles.Font(bold=True)

        name_cell = sheet.cell(row=row_start, column=col_start)
        name_cell.value = f"{agent.name} ({agent.type.value})"
        name_cell.alignment = Alignment(horizontal="center")

        # Set up headers and borders
        for i in range(10):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.border = Border(top=thick)

        for row in range(row_start + 1, row_start + 20):
            cell = sheet.cell(row=row, column=col_start + 4)
            cell.border = Border(right=thick)
            if row == row_start + 1:
                cell.border = Border(right=thick, top=thick)

        headers = ['Type', 'CP', 'Amount', 'Maturity', 'Settlement']
        for i, header in enumerate(headers):
            cell = sheet.cell(row=row_start + 1, column=col_start + i)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")
            cell = sheet.cell(row=row_start + 1, column=col_start + i + 5)
            cell.value = header
            cell.alignment = Alignment(horizontal="center")

        # Display balance sheet entries
        current_row = row_start + 2
        for entry in agent.assets:
            # Skip entries that were issued after the current time point
            if entry.issuance_time > time_point:
                continue
            # Skip already matured entries
            if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date and time_point > entry.maturity_date:
                continue

            # Show entry details
            entry_type = "receivable" if entry.type == EntryType.PAYABLE else entry.type.value
            if entry.type == EntryType.NON_FINANCIAL and entry.name:
                entry_type = f"{entry.type.value} ({entry.name})"
            sheet.cell(row=current_row, column=col_start).value = entry_type
            sheet.cell(row=current_row, column=col_start + 1).value = entry.counterparty if entry.counterparty else "N/A"
            sheet.cell(row=current_row, column=col_start + 2).value = f"+{entry.initial_book_value} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date is not None:
                maturity = f"t{time_point}"
            sheet.cell(row=current_row, column=col_start + 3).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 4).value = settlement

            current_row += 1

        current_row = row_start + 2
        for entry in agent.liabilities:
            # Skip entries that were issued after the current time point
            if entry.issuance_time > time_point:
                continue
            # Skip already matured entries
            if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date and time_point > entry.maturity_date:
                continue

            # Show entry details
            entry_type = entry.type.value
            if entry.type == EntryType.DELIVERY_CLAIM:
                entry_type = f"delivery promise for {entry.name}" if entry.name else "delivery promise"
            sheet.cell(row=current_row, column=col_start + 5).value = entry_type
            sheet.cell(row=current_row, column=col_start + 6).value = entry.counterparty
            sheet.cell(row=current_row, column=col_start + 7).value = f"+{entry.initial_book_value} {entry.denomination}"

            maturity = entry.maturity_type.value
            if entry.maturity_type == MaturityType.FIXED_DATE:
                maturity = f"t{entry.maturity_date}"
            sheet.cell(row=current_row, column=col_start + 8).value = f"{maturity} (issued at {entry.issuance_time})"

            settlement = entry.settlement_details.type.value
            if settlement != "none":
                settlement += f" ({entry.settlement_details.denomination})"
            sheet.cell(row=current_row, column=col_start + 9).value = settlement

            current_row += 1

        # Add totals
        total_row = current_row + 2
        sheet.cell(row=total_row, column=col_start).value = "Total Assets:"
        sheet.cell(row=total_row, column=col_start + 2).value = agent.get_total_assets()
        sheet.cell(row=total_row + 1, column=col_start).value = "Total Liabilities:"
        sheet.cell(row=total_row + 1, column=col_start + 2).value = agent.get_total_liabilities()
        sheet.cell(row=total_row + 2, column=col_start).value = "Net Worth:"
        sheet.cell(row=total_row + 2, column=col_start + 2).value = agent.get_net_worth()

        return total_row + 4

    def export_balance_sheets(self, filename: str, time: int):
        """Export balance sheets for all time points vertically stacked"""
        # First, ensure all settlements are processed
        if self.system.simulation_finalized:
            # Save current state
            current_state = deepcopy(self.system.agents)
            current_time = self.system.current_time_state

            # Process settlements if not already done
            try:
                if time not in self.system.time_states:
                    self.system.settle_entries(time)
                if time not in self.system.time_states:
                    self.system.settle_entries(time)
            except Exception as e:
                print(f"\nWarning: Settlement processing failed ({str(e)})")

            # Restore original state
            self.system.agents = current_state
            self.system.current_time_state = current_time

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Balance Sheets Over Time"

        current_row = 1

        for time_point in range(time + 1):
            agents = (self.system.time_states[time_point].values()
                     if time_point in self.system.time_states
                     else self.system.agents.values())

            # Add time point separator
            sheet.cell(row=current_row, column=1).value = "=" * 50
            current_row += 1

            col_start = 2
            max_row_in_timepoint = current_row

            for agent in agents:
                max_row_in_timepoint = max(
                    max_row_in_timepoint,
                    self.create_t_table(sheet, current_row, col_start, agent, time_point)
                )
                col_start += 10

            # Add system totals for this time point
            system_total_row = max_row_in_timepoint
            sheet.cell(row=system_total_row, column=1).value = f"System Totals at t{time_point}:"
            sheet.cell(row=system_total_row + 1, column=1).value = "Total Assets:"
            sheet.cell(row=system_total_row + 1, column=2).value = sum(agent.get_total_assets() for agent in agents)
            sheet.cell(row=system_total_row + 2, column=1).value = "Total Liabilities:"
            sheet.cell(row=system_total_row + 2, column=2).value = sum(agent.get_total_liabilities() for agent in agents)
            sheet.cell(row=system_total_row + 3, column=1).value = "Total Net Worth:"
            sheet.cell(row=system_total_row + 3, column=2).value = sum(agent.get_net_worth() for agent in agents)

            current_row = system_total_row + 5  # Leave space between time points

        # Adjust column widths
        for i in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(i)].width = 15

        wb.save(filename)

        # NEED TO BE UNCOMMENTED IF RUNNING IN COLAB, OR MODIFIED IF RUNNING ELSEWHERE
        # # Add Colab download capability
        # try:
        #     from google.colab import files
        #     files.download(filename)
        #     print(f"\nExcel file has been created and downloaded: {filename}")
        # except ImportError:
        #     print(f"\nExcel file has been created: {filename}")

def get_user_time_input(prompt: str) -> Optional[datetime]:
    """Helper function to get time state input"""
    print(prompt)
    print("Enter a time state (0, 1, 2...)") 
    try:
        time_state = int(input("Time state: ").strip().lower())
        if time_state <= 0:
            print("Please enter a non-negative time state")
            return None
        return time_state

    except ValueError:
        print("That's not a valid number!")
        return None

def create_agent_interactive(system: EconomicSystem):
    """Interactive function to create an agent"""
    print("\nCreating new agent:")
    name = input("Enter agent name: ")

    if name in system.agents:
        print(f"Error: Agent '{name}' already exists!")
        return

    print("\nAvailable agent types:")
    for i, agent_type in enumerate(AgentType, 1):
        print(f"{i}. {agent_type.value}")

    try:
        type_idx = int(input("\nSelect agent type (enter number): ")) - 1
        agent_type = list(AgentType)[type_idx]

        agent = Agent(name, agent_type)
        system.add_agent(agent)
        print(f"\nAgent '{name}' of type '{agent_type.value}' created successfully!")
        return agent
    except (ValueError, IndexError):
        print("Error: Invalid agent type selection")
        return None

def create_asset_liability_pair_interactive(system: EconomicSystem, default_denomination: Optional[str] = None):
    """Interactive function to create an asset-liability pair"""
    if len(system.agents) < 2:
        print("\nError: Need at least 2 agents to create an asset-liability pair!")
        return

    print("\nCreating new asset-liability pair:")

    # List available agents
    print("\nAvailable agents:")
    agents = list(system.agents.values())
    for i, agent in enumerate(agents, 1):
        print(f"{i}. {agent.name} ({agent.type.value})")

    try:
        # Get asset holder
        asset_idx = int(input("\nSelect asset holder (enter number): ")) - 1
        asset_holder = agents[asset_idx]

        print("\nIs this a financial or non-financial entry?")
        print("1. Financial (requires liability holder)")
        print("2. Non-financial (no liability holder)")
        entry_type_choice = int(input("Enter choice (1 or 2): "))

        liability_holder = None
        asset_name = None

        if entry_type_choice == 1:
            liability_idx = int(input("Select liability holder (enter number): ")) - 1
            liability_holder = agents[liability_idx]
            if liability_holder == asset_holder:
                print("Error: Asset holder and liability holder cannot be the same!")
                return

            # Get entry type for financial entries
            print("\nAvailable entry types:")
            print("1. loan")
            print("2. deposit")
            print("3. receivable-payable (creates a receivable for the asset holder and a payable for the liability holder)")
            print("4. bond")
            print("5. delivery claim (creates a delivery claim for the asset holder and a delivery promise for the liability holder)")
            type_idx = int(input("Select entry type (enter number): ")) - 1
            if type_idx == 2:  # receivable-payable
                entry_type = EntryType.PAYABLE
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                print("\nNote: For receivable-payable pairs:")
                print("- The asset holder will get a receivable")
                print("- The liability holder will get a payable")
                print("- Settlement type is automatically set to means of payment")
            elif type_idx == 4:  # delivery claim
                entry_type = EntryType.DELIVERY_CLAIM
                settlement_type = SettlementType.NON_FINANCIAL_ASSET
                print("\nNote: For delivery claim pairs:")
                print("- The asset holder will get a delivery claim")
                print("- The liability holder will get a delivery promise")
                print("- At maturity, the specified asset must be transferred from the liability holder to the asset holder")
                print("- The liability holder must have the asset at maturity to fulfill the promise")

                # Get asset name for delivery claim
                while True:
                    asset_name = input("\nEnter the name of the non-financial asset to be delivered (e.g., 'machine', 'building'): ").strip()
                    if asset_name:
                        # Check if liability holder has any non-financial assets with this name
                        has_asset = any(asset.type == EntryType.NON_FINANCIAL and asset.name == asset_name
                                      for asset in liability_holder.assets)
                        if has_asset:
                            print(f"\nNote: {liability_holder.name} currently has a {asset_name} in their balance sheet.")
                        else:
                            print(f"\nWarning: {liability_holder.name} does not currently have a {asset_name} in their balance sheet.")
                            print("They will need to acquire it before the maturity date to fulfill the delivery promise.")
                        break
                    print("Error: Delivery claim requires specifying the asset to be delivered!")

                # Get amount and denomination
                amount = float(input("\nEnter amount (value of the asset): "))
                if default_denomination:
                    denomination_prompt = f"Enter denomination (press Enter for default: {default_denomination}): "
                    denomination = input(denomination_prompt).strip() or default_denomination
                else:
                    denomination = input("Enter denomination (e.g., USD): ")

                try:
                    cash_flow_input = input("Enter cash flow at maturity (press Enter to use amount): ").strip()
                    cash_flow_at_maturity = float(cash_flow_input) if cash_flow_input else None
                except ValueError:
                    print("Invalid number. Using default.")
                    cash_flow_at_maturity = 0

                # Get maturity information
                print("\nSelect maturity type:")
                for i, mt in enumerate(MaturityType, 1):
                    print(f"{i}. {mt.value}")
                maturity_idx = int(input("Enter choice: ")) - 1
                maturity_type = list(MaturityType)[maturity_idx]

                maturity_date = None
                if maturity_type == MaturityType.FIXED_DATE:
                    maturity_date = get_user_time_input("\nEnter maturity date (when the asset must be delivered):")
                    if not maturity_date:
                        print("Error: Invalid date input!")
                        return

                # Create the delivery claim pair
                pair = AssetLiabilityPair(
                    time=datetime.now(),
                    type=entry_type.value,
                    amount=amount,
                    denomination=denomination,
                    maturity_type=maturity_type,
                    maturity_date=maturity_date,
                    settlement_type=settlement_type,
                    settlement_denomination=denomination,
                    asset_holder=asset_holder,
                    liability_holder=liability_holder,
                    asset_name=asset_name,
                    cash_flow_at_maturity=cash_flow_at_maturity
                )

                system.create_asset_liability_pair(pair)
                print("\nDelivery claim pair created successfully!")
                print(f"- {asset_holder.name} will receive a delivery claim for {asset_name}")
                print(f"- {liability_holder.name} will receive a delivery promise to deliver {asset_name}")
                if maturity_type == MaturityType.FIXED_DATE:
                    print(f"- The delivery must occur at {maturity_date.year}")
                return
            else:
                entry_type = list(et for et in EntryType if et not in [EntryType.NON_FINANCIAL, EntryType.DELIVERY_CLAIM])[type_idx]
        else:
            entry_type = EntryType.NON_FINANCIAL
            # Get asset name for non-financial entries
            while True:
                asset_name = input("\nEnter the name of the non-financial asset (e.g., 'machine', 'building'): ").strip()
                if asset_name:
                    break
                print("Error: Non-financial asset requires a name!")

        # Get amount and denomination
        amount = float(input("\nEnter amount: "))
        if default_denomination:
            denomination_prompt = f"Enter denomination (press Enter for default: {default_denomination}): "
            denomination = input(denomination_prompt).strip() or default_denomination
        else:
            denomination = input("Enter denomination (e.g., USD): ")

        # Set maturity information based on entry type
        if entry_type in [EntryType.NON_FINANCIAL, EntryType.DEPOSIT]:
            maturity_type = MaturityType.ON_DEMAND
            maturity_date = None
        else:
            # Get maturity information
            print("\nSelect maturity type:")
            for i, mt in enumerate(MaturityType, 1):
                print(f"{i}. {mt.value}")
            maturity_idx = int(input("Enter choice: ")) - 1
            maturity_type = list(MaturityType)[maturity_idx]

            maturity_date = None
            if maturity_type == MaturityType.FIXED_DATE:
                maturity_date = get_user_time_input("\nEnter maturity date:")
                if not maturity_date:
                    print("Error: Invalid date input!")
                    return

        # Set settlement information based on entry type
        if entry_type == EntryType.NON_FINANCIAL or entry_type == EntryType.DEPOSIT:
            settlement_type = SettlementType.NONE
            settlement_denomination = denomination
        elif entry_type == EntryType.PAYABLE:
            # Already set above for receivable-payable
            settlement_denomination = denomination
        else:
            # For other types (LOAN, BOND), allow any settlement type except NONE
            print("\nSelect settlement type:")
            settlement_types = [st for st in SettlementType if st != SettlementType.NONE]
            for i, st in enumerate(settlement_types, 1):
                print(f"{i}. {st.value}")
            settlement_idx = int(input("Enter choice: ")) - 1
            settlement_type = settlement_types[settlement_idx]

            if default_denomination:
                denomination_prompt = f"Enter settlement denomination (press Enter for default: {default_denomination}): "
                settlement_denomination = input(denomination_prompt).strip() or default_denomination
            else:
                settlement_denomination = input("\nEnter settlement denomination (e.g., USD): ")

        # Create the asset-liability pair
        pair = AssetLiabilityPair(
            time=datetime.now(),
            type=entry_type.value,
            amount=amount,
            denomination=denomination,
            maturity_type=maturity_type,
            maturity_date=maturity_date,
            settlement_type=settlement_type,
            settlement_denomination=settlement_denomination,
            asset_holder=asset_holder,
            liability_holder=liability_holder,
            asset_name=asset_name
        )

        system.create_asset_liability_pair(pair)
        print("\nAsset-liability pair created successfully!")

    except (ValueError, IndexError) as e:
        print(f"Error: {str(e)}")
        print("Please try again with valid inputs.")

def main():
    print("Welcome to the Economic Balance Sheet Simulation!")
    print("==============================================")
    print("\nThis simulation automatically saves states at each time point.")
    print("- t0: Initial state (auto-saved when creating agents and entries)")
    print("- t1: Intermediate state (auto-saved when settling entries)")
    print("- t2: Final state (auto-saved when settling entries)")

    # Get default denomination
    print("\nWould you like to set a default denomination (e.g., USD) for all entries?")
    print("You can still override this for individual entries if needed.")
    default_denomination = input("Enter default denomination (press Enter to skip): ").strip() or None

    system = EconomicSystem()

    while True:
        print("\nEconomic Balance Sheet Simulation")
        print("1. Create agent")
        print("2. Create asset-liability pair")
        print("3. View balance sheets")
        print("4. View settlement history")
        print("5. Simulate (finalize agents and pairs)")
        print("6. Export to Excel")
        print("7. Exit")

        choice = input("\nEnter your choice (1-7): ")

        if choice == '1':
            create_agent_interactive(system)

        elif choice == '2':
            create_asset_liability_pair_interactive(system, default_denomination)

        elif choice == '3':
            if system.simulation_finalized:
                try:
                    time_point = int(input(f"\nEnter time point in range 0 to {system.final_time_state} to view: ").strip().lower())
                    if time_point >= 0:
                        system.display_balance_sheets(time_point)
                    else:
                        print("\nInvalid time point. Please enter a non-negative integer")
                except TypeError:
                    print("Please input an integer 0,1,2...")
            else:
                system.display_balance_sheets(0)

        elif choice == '4':
            system.display_settlement_history()

        elif choice == '5':
            final_time = int(input("\nSelect number of days to simulate (enter number): "))
            if not isinstance(final_time,int):
                print("Error: Please enter an integer")
            elif final_time < 0:
                print("Error: Your integer must be non-negative")
            else:
                system.final_time_state = final_time
                print("\nFinalizing simulation setup...")
                print("Processing settlements...")
                system.simulation_finalized = True
            # Run the simulation
                success = system.run_simulation(final_time)
                if success:
                    print("\nAll settlements processed successfully!")
                    print(f"You can now view balance sheets from time point t0 to t{system.final_time_state}")
                else:
                    print("\nSimulation stopped due to settlement failure.")
                    print("Check settlement history for details.")
                print("\nNote: No more agents or asset-liability pairs can be added.")

        elif choice == '6':
            if EXCEL_AVAILABLE:
                filename = "balance_sheets.xlsx"
                exporter = ExcelExporter(system)
                exporter.export_balance_sheets(filename, system.final_time_state)
            else:
                print("\nError: Excel export is not available. Please install openpyxl package.")
                print("Run: pip install openpyxl==3.1.2")

        elif choice == '7':
            print("\nExiting simulation. Goodbye!")
            break

        else:
            print("\nInvalid choice. Please try again.")

if __name__ == "__main__":
    main()