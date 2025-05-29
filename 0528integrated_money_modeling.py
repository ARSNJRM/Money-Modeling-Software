try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl package not found. Excel export functionality will be disabled.")
    print("To enable Excel export, please install openpyxl using: pip install openpyxl")
    EXCEL_AVAILABLE = False

from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union, Any
from enum import Enum
from datetime import datetime, timedelta
from copy import deepcopy
import sys

# Core Enums and Classes
class AgentType(Enum):
    BANK = "bank"
    COMPANY = "company"
    HOUSEHOLD = "household"
    TREASURY = "treasury"
    CENTRAL_BANK = "central_bank"
    OTHER = "other"

class EntryType(Enum):
    LOAN = "loan"
    DEPOSIT = "deposit"
    PAYABLE = "payable"
    # Integration of bond types
    BOND = "bond" 
    BOND_ZERO_COUPON = "bond_zero_coupon"
    BOND_COUPON = "bond_coupon"
    BOND_AMORTIZING = "bond_amortizing"
    # Integration of types related to interday liquidity of banks
    INTRADAY_IOU = "intraday_iou"
    OVERNIGHT_LOAN = "overnight_loan"
    # Integration of share types
    SHARE = "share"
    # Other types
    DELIVERY_CLAIM = "delivery_claim"
    NON_FINANCIAL = "non_financial"
    DEFAULT = "default"

class MaturityType(Enum):
    ON_DEMAND = "on_demand"
    FIXED_DATE = "fixed_date"
    PERPETUAL = "perpetual"

class SettlementType(Enum):
    MEANS_OF_PAYMENT = "means_of_payment"
    SECURITIES = "securities"
    NON_FINANCIAL_ASSET = "non_financial_asset"
    SERVICES = "services"
    CRYPTO = "crypto"
    NONE = "none"

class BondType(Enum):
    ZERO_COUPON = 0  # zero coupon bond
    COUPON = 1       # coupon bond
    AMORTIZING = 2   # Amortized bonds

@dataclass
class SettlementDetails:
    type: SettlementType
    denomination: str 
    is_intraday: bool = False 

@dataclass
class BalanceSheetEntry:
    type: EntryType
    is_asset: bool 
    counterparty: Optional[str] 
    initial_book_value: float 
    denomination: str 
    maturity_type: MaturityType  
    maturity_date: Optional[int]  
    settlement_details: SettlementDetails 
    cash_flow_at_maturity: float = 0 
    name: Optional[str] = None  
    issuance_time: int = 0 
    current_book_value: float = 0 
    rollover_count: int = 0 
    expected_cash_flow: Optional[float] = None 
    parent_bond: Optional[str] = None 

    def matches(self, other: 'BalanceSheetEntry') -> bool:
        return (
            self.type == other.type and
            self.is_asset == other.is_asset and
            self.counterparty == other.counterparty and
            self.initial_book_value == other.initial_book_value and
            self.denomination == other.denomination and
            self.maturity_type == other.maturity_type and
            self.maturity_date == other.maturity_date and
            self.settlement_details.type == other.settlement_details.type and
            self.settlement_details.denomination == other.settlement_details.denomination and
            self.name == other.name and
            self.issuance_time == other.issuance_time
        )

    def __post_init__(self):
        if self.initial_book_value <= 0:
            raise ValueError("The amount must be positive")


        if self.issuance_time < 0:
            raise ValueError("Time of issue must be non-negative")


        if self.type != EntryType.NON_FINANCIAL and not self.counterparty:
            raise ValueError("Financial entries must have a counterparty")
        if self.type == EntryType.NON_FINANCIAL and self.counterparty:
            raise ValueError("Non-financial entries cannot have counterparties")


        if self.type == EntryType.NON_FINANCIAL and not self.name:
            raise ValueError("Non-financial entries must have a name")


        if self.type == EntryType.PAYABLE and self.settlement_details.type != SettlementType.MEANS_OF_PAYMENT:
            raise ValueError("Payable entries must have a means of payment settlement type")
            

        if self.type == EntryType.SHARE:
            self.maturity_type = MaturityType.PERPETUAL
            self.maturity_date = None
            

        if self.current_book_value == 0:
            self.current_book_value = self.initial_book_value

class SettlementFailure(Exception):
    def __init__(self, agent_name: str, entry: BalanceSheetEntry, reason: str):
        self.agent_name = agent_name
        self.entry = entry
        self.reason = reason
        super().__init__(f"Settlement failure {agent_name}: {reason}")

class BankIntradayModule:
    def __init__(self, bank_agent):
        self.bank = bank_agent
        self.settlement_log = [] 
        self.system = None 

    def _find_iou_counterparty(self, iou_entry):
        counterparty = next(
            (agent for agent in self.bank.system.agents.values() if agent.name == iou_entry.counterparty),
            None
        )
        if counterparty is None:
            self.settlement_log.append(f"Counterparty not found for IOU settlement: {iou_entry.counterparty}")
            raise ValueError(f"Counterparty not found for IOU settlement: {iou_entry.counterparty}")
        return counterparty

    def _settle_partial(self, iou_entry, time_point):
        """
        Attempts to settle as many IOUs as possible using available deposits. 
        Supports partial settlement (if full amount is not available).
        Returns the settlement amount (or 0 if not available).
        """
        # Find all deposits in the bank that match the denomination of the IOU, sorted in descending order of amount
        deposits = sorted(
            (a for a in self.bank.assets if a.type == EntryType.DEPOSIT and a.denomination == iou_entry.denomination),
            key=lambda x: x.initial_book_value,
            reverse=True
        )

        amount_to_settle = iou_entry.initial_book_value
        total_settled = 0.0

        for deposit in deposits:
            if amount_to_settle <= 0:
                break
            settle_amount = min(deposit.initial_book_value, amount_to_settle)

            # Remove or reduce deposits accordingly
            if deposit.initial_book_value == settle_amount:
                self.bank.remove_asset(deposit)
            else:
                deposit.initial_book_value -= settle_amount
                deposit.current_book_value -= settle_amount

            # Find counterparty banks receiving deposits
            receiver = self._find_iou_counterparty(iou_entry)

            # Create a new deposit asset for the recipient bank
            new_deposit = BalanceSheetEntry(
                type=EntryType.DEPOSIT,
                is_asset=True,
                counterparty=self.bank.name,
                initial_book_value=settle_amount,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination
                ),
                issuance_time=time_point
            )
            receiver.add_asset(new_deposit)

            total_settled += settle_amount
            amount_to_settle -= settle_amount

        if total_settled > 0:
            # Reduction of IOUs on settled amounts
            iou_entry.initial_book_value -= total_settled
            iou_entry.current_book_value -= total_settled

            if iou_entry.initial_book_value <= 0:
                self.bank.remove_liability(iou_entry)
            return total_settled
        else:
            return 0.0

    def _evaluate_rollover_proposal(self, iou_entry, receiver_bank):
        """
        Evaluate whether to allow rollovers (rollovers of day-ahead IOUs) based on the recipient bank's reserve deposits 
        and total deposits, using the Reserve Requirement Ratio (RRR).

        The logic is as follows:
        1. calculate the total reserve deposits (RD) held by the recipient bank.
        2. calculate the total bank deposits (BD) held by the recipient bank (excluding reserves). 3. calculate the excess reserves (SR).
        3. calculate excess reserves (SR) as SR = RD - (RRR * BD).
           - If SR is negative, it indicates a reserve deficiency.
        4. If the reserve is insufficient (SR < 0):
           - If the IOU amount is less than or equal to the underfunded amount, deny the rollover.
           - Otherwise, rollover is authorized, but only up to the amount in excess of the underfunded amount.
        5. if no reserve is insufficient (SR >= 0), grant the rollover in full.

        Returns: 
        bool: True if rollover is approved, False otherwise.
        """
        RRR = 0.10  # Reserve requirement ratios

        reserve_deposits = sum(
            asset.initial_book_value for asset in receiver_bank.assets
            if asset.type == EntryType.DEPOSIT and asset.denomination == "reserves"
        )
        bank_deposits = sum(
            asset.initial_book_value for asset in receiver_bank.assets
            if asset.type == EntryType.DEPOSIT and asset.denomination != "reserves"
        )

        surplus_reserve = reserve_deposits - (RRR * bank_deposits)

        if surplus_reserve < 0:
            deficit_reserve = abs(surplus_reserve)
            if iou_entry.initial_book_value <= deficit_reserve:
                return False
            else:
                return True
        else:
            return True

    def _handle_rollover(self, iou_entry, time_point):
        """Process IOU extensions: if accepted, create new IOUs with extended expiration dates; if rejected, process defaults"""
        receiver = self._find_iou_counterparty(iou_entry)
        if receiver and self._evaluate_rollover_proposal(iou_entry, receiver):
            new_iou = BalanceSheetEntry(
                type=EntryType.INTRADAY_IOU,
                is_asset=False,
                counterparty=iou_entry.counterparty,
                initial_book_value=iou_entry.initial_book_value,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.FIXED_DATE,
                maturity_date=time_point + 1,  # extend to the next time point
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination,
                    is_intraday=True
                ),
                issuance_time=time_point,
                rollover_count=iou_entry.rollover_count + 1,
                current_book_value=iou_entry.initial_book_value
            )

            # create a new asset for the receiver
            new_asset_for_receiver = BalanceSheetEntry(
                type=EntryType.INTRADAY_IOU,
                is_asset=True,
                counterparty=self.bank.name,
                initial_book_value=iou_entry.initial_book_value,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.FIXED_DATE,
                maturity_date=time_point + 1,
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination,
                    is_intraday=True
                ),
                issuance_time=time_point,
                rollover_count=iou_entry.rollover_count + 1,
                current_book_value=iou_entry.initial_book_value
            )

            # delete the old IOU
            self.bank.remove_liability(iou_entry)
            receiver.remove_asset(next(
                asset for asset in receiver.assets
                if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
                and asset.initial_book_value == iou_entry.initial_book_value
            ))

            # add the new IOU and asset
            self.bank.add_liability(new_iou)
            receiver.add_asset(new_asset_for_receiver)

            self.settlement_log.append(f"Extend maturity IOU: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
            return True
        else:
            # handle default if rollover is not approved
            self._handle_default(iou_entry, time_point)
            return False

    def _convert_to_overnight(self, iou_entry, time_point):
        """Convert day IOUs to overnight loans."""
        receiver = self._find_iou_counterparty(iou_entry)
        
        # create an overnight loan asset (for the bank)
        overnight_loan = BalanceSheetEntry(
            type=EntryType.OVERNIGHT_LOAN,
            is_asset=True,
            counterparty=self.bank.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=time_point + 1,  # matures at the next time point
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        # create an overnight loan liability (for the receiver)
        overnight_liability = BalanceSheetEntry(
            type=EntryType.OVERNIGHT_LOAN,
            is_asset=False,
            counterparty=receiver.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=time_point + 1,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        # delete the old IOU
        self.bank.remove_liability(iou_entry)
        receiver.remove_asset(next(
            asset for asset in receiver.assets
            if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
            and asset.initial_book_value == iou_entry.initial_book_value
        ))
        
        # add the new overnight loan and liability
        receiver.add_asset(overnight_loan)
        self.bank.add_liability(overnight_liability)
        
        self.settlement_log.append(f"Converting IOUs to Overnight Loans: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
        return True

    def _handle_default(self, iou_entry, time_point):
        """handle IOU defaults: create default assets and liabilities, remove old IOUs."""
        receiver = self._find_iou_counterparty(iou_entry)
        
        # create default asset (for the bank)
        default_asset = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=self.bank.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            name=f"default {self.bank.name}",
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        # create default liability (for the receiver)
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=receiver.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            name=f"default {receiver.name}",
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        # delete the old IOU
        self.bank.remove_liability(iou_entry)
        receiver.remove_asset(next(
            asset for asset in receiver.assets
            if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
            and asset.initial_book_value == iou_entry.initial_book_value
        ))
        
        # add the new default asset and liability
        receiver.add_asset(default_asset)
        self.bank.add_liability(default_liability)
        
        self.settlement_log.append(f"IOU default: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
        return True

    def process_intraday_settlements(self, time_point):
        """Handles settlement, rollover or default of all daytime IOUs."""
        # search for all IOUs that are due at the current time point
        due_ious = [
            liability for liability in self.bank.liabilities
            if liability.type == EntryType.INTRADAY_IOU
            and liability.maturity_date == time_point
        ]
        
        for iou in due_ious:
            settled_amount = self._settle_partial(iou, time_point)
            
            # if not fully settled, check for rollover or default
            if iou in self.bank.liabilities: 
                if iou.rollover_count < 2: 
                    self._handle_rollover(iou, time_point)
                else:
                    # exceeds the maximum rollover limit, convert to overnight loan
                    self._convert_to_overnight(iou, time_point)

class Agent:
    def __init__(self, name: str, agent_type: AgentType):
        self.name = name
        self.type = agent_type
        self.assets: List[BalanceSheetEntry] = []
        self.liabilities: List[BalanceSheetEntry] = []
        self.status: str = "operating"  # operating or bankrupt
        self.creation_time: datetime = datetime.now()
        # add settlement history
        self.settlement_history = {
            'as_asset_holder': [],
            'as_liability_holder': [] 
        }
        # add bank-specific properties
        if self.type == AgentType.BANK:
            self.intraday_module = BankIntradayModule(self)
            self.system = None 

        self.inventory_bond = 0
        self.inventory_cash = 0
        self.trade_count = 0
        self.total_volume = 0
        self.trade_volumes = []
        self.trade_prices = []

        self.reserves: float = 0.0
        self.customer_deposits: Dict[str, float] = {} 

    def add_asset(self, entry: BalanceSheetEntry):
        self.assets.append(entry)

    def add_liability(self, entry: BalanceSheetEntry):
        self.liabilities.append(entry)

    def remove_asset(self, entry: BalanceSheetEntry):
        self.assets = [e for e in self.assets if not e.matches(entry)]

    def remove_liability(self, entry: BalanceSheetEntry):
        self.liabilities = [e for e in self.liabilities if not e.matches(entry)]

    def get_balance_sheet(self) -> Dict:
        return {
            "assets": self.assets,
            "liabilities": self.liabilities
        }

    def get_total_assets(self) -> float:
        return sum(entry.current_book_value for entry in self.assets)

    def get_total_liabilities(self) -> float:
        return sum(entry.current_book_value for entry in self.liabilities)

    def get_net_worth(self) -> float:
        total_assets = sum(entry.current_book_value for entry in self.assets 
                          if entry.type != EntryType.SHARE)
        total_liabilities = sum(entry.current_book_value for entry in self.liabilities 
                               if entry.type != EntryType.SHARE)
        return total_assets - total_liabilities

    def get_type_specific_metrics(self) -> Dict:
        metrics = {
            "name": self.name,
            "type": self.type.value,
            "creation_time": self.creation_time,
            "status": self.status,
            "total_assets": self.get_total_assets(),
            "total_liabilities": self.get_total_liabilities(),
            "net_worth": self.get_net_worth()
        }

        if self.type == AgentType.BANK:
            metrics["capital_ratio"] = self.get_total_assets() / self.get_total_liabilities() if self.get_total_liabilities() > 0 else float('inf')
        elif self.type == AgentType.COMPANY:
            metrics["leverage_ratio"] = self.get_total_liabilities() / self.get_total_assets() if self.get_total_assets() > 0 else float('inf')
        elif self.type == AgentType.HOUSEHOLD:
            metrics["savings_rate"] = (self.get_total_assets() - self.get_total_liabilities()) / self.get_total_assets() if self.get_total_assets() > 0 else 0

        return metrics

    def record_settlement(self,
                         time_point: int,
                         original_entry: BalanceSheetEntry,
                         settlement_result: BalanceSheetEntry,
                         counterparty: str,
                         as_asset_holder: bool):

        settlement_record = {
            'time_point': time_point,
            'original_entry': deepcopy(original_entry),
            'settlement_result': deepcopy(settlement_result),
            'counterparty': counterparty,
            'timestamp': datetime.now()
        }
        if as_asset_holder:
            self.settlement_history['as_asset_holder'].append(settlement_record)
        else:
            self.settlement_history['as_liability_holder'].append(settlement_record)

    def update_inventory(self, bond_delta, cash_delta):

        self.inventory_bond += bond_delta
        self.inventory_cash += cash_delta
        

        if bond_delta != 0: 
            self.trade_count += 1
            self.total_volume += abs(bond_delta)
            self.trade_volumes.append(abs(bond_delta))
            self.trade_prices.append(-cash_delta / bond_delta) 

    def transfer_customer_deposit(self, sender_bank, receiver_bank, customer_a, customer_b, amount):
        sender_bank.customer_deposits[customer_a] -= amount
        receiver_bank.customer_deposits[customer_b] += amount

        pair = AssetLiabilityPair(
            time=0, 
            type=EntryType.INTRADAY_IOU.value,
            amount=amount, 
            denomination="USD", 
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=1,
            settlement_type=SettlementType.MEANS_OF_PAYMENT,
            settlement_denomination="USD",
            asset_holder=receiver_bank,
            liability_holder=sender_bank
        )
        self.system.create_asset_liability_pair(pair)

class AssetLiabilityPair:
    def __init__(self,
                 time: int,
                 type: str,
                 amount: float,
                 denomination: str,
                 maturity_type: MaturityType,
                 maturity_date: Optional[int],
                 settlement_type: SettlementType,
                 settlement_denomination: str,
                 asset_holder: Agent,
                 liability_holder: Optional[Agent] = None,
                 cash_flow_at_maturity: Optional[float] = 0,
                 asset_name: Optional[str] = None,
                 bond_type: Optional[BondType] = None,
                 coupon_rate: Optional[float] = None):
        self.time = time
        self.type = type
        self.amount = amount
        self.denomination = denomination
        self.maturity_type = maturity_type
        self.maturity_date = maturity_date
        self.cash_flow_at_maturity = cash_flow_at_maturity
        self.settlement_details = SettlementDetails(
            type=settlement_type,
            denomination=settlement_denomination
        )
        self.asset_holder = asset_holder
        self.liability_holder = liability_holder
        self.asset_name = asset_name
        self.bond_type = bond_type
        self.coupon_rate = coupon_rate
        self.connected_claims = [] 
        self.current_time_state = 0 

        if type == EntryType.NON_FINANCIAL.value:
            if liability_holder is not None:
                raise ValueError("Non-financial entries cannot have liability holders")
            if not asset_name:
                raise ValueError("Non-financial entries must have the name of the asset")
        else:
            if liability_holder is None:
                raise ValueError("Financial entries must have liability holders")

    def create_entries(self) -> Tuple[BalanceSheetEntry, Optional[BalanceSheetEntry]]:
        if self.type == EntryType.LOAN.value:
            if self.asset_holder.type != AgentType.BANK:
                raise ValueError("Only banks can hold loans as assets")
                
        if self.type == EntryType.DELIVERY_CLAIM.value:
            if not self.asset_name:
                raise ValueError("Delivery claims must specify assets to be delivered")

            settlement_details = SettlementDetails(
                type=SettlementType.NON_FINANCIAL_ASSET,
                denomination=self.settlement_details.denomination
            )

            # create delivery claim (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # create delivery claim (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.DELIVERY_CLAIM,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            return asset_entry, liability_entry

        # deal with payable entries
        elif self.type == EntryType.PAYABLE.value:
            settlement_details = SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=self.settlement_details.denomination
            )

            # create payable (asset)
            asset_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE, 
                is_asset=True,
                counterparty=self.liability_holder.name if self.liability_holder else None,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # create payable (liability)
            liability_entry = BalanceSheetEntry(
                type=EntryType.PAYABLE,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=settlement_details,
                name=None,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            return asset_entry, liability_entry
            
        # deal with share entries
        elif self.type == EntryType.SHARE.value:
            asset_entry = BalanceSheetEntry(
                type=EntryType.SHARE,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=MaturityType.PERPETUAL,
                maturity_date=None,
                settlement_details=SettlementDetails(
                    type=SettlementType.NONE, 
                    denomination="shares"
                ),
                name=None,
                issuance_time=self.current_time_state,
                current_book_value=self.amount
            )

            # create share liability
            liability_entry = BalanceSheetEntry(
                type=EntryType.SHARE,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=MaturityType.PERPETUAL,
                maturity_date=None,
                settlement_details=SettlementDetails(
                    type=SettlementType.NONE, 
                    denomination="shares"
                ),
                name=None,
                issuance_time=self.current_time_state,
                current_book_value=self.amount
            )
            return asset_entry, liability_entry

        # deal with non-financial entries
        if self.type == EntryType.NON_FINANCIAL.value:
            asset_entry = BalanceSheetEntry(
                type=EntryType.NON_FINANCIAL,
                is_asset=True,
                counterparty=None,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity,
                current_book_value=self.amount
            )
            return asset_entry, None

        # deal with bond entries
        if self.type in [EntryType.BOND_ZERO_COUPON.value, EntryType.BOND_COUPON.value, EntryType.BOND_AMORTIZING.value]:

            asset_entry = BalanceSheetEntry(
                type=EntryType(self.type),
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity,
                current_book_value=self.amount,
                expected_cash_flow=self._calculate_expected_cash_flow()
            )

            # create liability entry for the bond
            liability_entry = BalanceSheetEntry(
                type=EntryType(self.type),
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=None,
                issuance_time=self.current_time_state,
                cash_flow_at_maturity=self.cash_flow_at_maturity,
                current_book_value=self.amount,
                expected_cash_flow=self._calculate_expected_cash_flow()
            )

            # Creating Connected Debentures for Coupon Bonds and Amortizing Bonds
            if self.type in [EntryType.BOND_COUPON.value, EntryType.BOND_AMORTIZING.value]:
                claims = self.create_bond_claims()
                self.connected_claims = claims
                
                for claim in claims:
                    self.asset_holder.add_asset(claim)
                    
                    # create corresponding liability for the asset holder
                    liability = BalanceSheetEntry(
                        type=claim.type,
                        is_asset=False,
                        counterparty=self.asset_holder.name,
                        initial_book_value=claim.initial_book_value,
                        denomination=claim.denomination,
                        maturity_type=claim.maturity_type,
                        maturity_date=claim.maturity_date,
                        settlement_details=claim.settlement_details,
                        current_book_value=claim.current_book_value,
                        expected_cash_flow=claim.expected_cash_flow
                    )
                    self.liability_holder.add_liability(liability)

            return asset_entry, liability_entry

        # deal with other financial entries
        asset_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=True,
            counterparty=self.liability_holder.name if self.liability_holder else None,
            initial_book_value=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=self.asset_name,
            issuance_time=self.current_time_state,
            cash_flow_at_maturity=self.cash_flow_at_maturity,
            current_book_value=self.amount
        )

        liability_entry = BalanceSheetEntry(
            type=EntryType(self.type),
            is_asset=False,
            counterparty=self.asset_holder.name,
            initial_book_value=self.amount,
            denomination=self.denomination,
            maturity_type=self.maturity_type,
            maturity_date=self.maturity_date,
            settlement_details=self.settlement_details,
            name=None, 
            issuance_time=self.current_time_state,
            cash_flow_at_maturity=self.cash_flow_at_maturity,
            current_book_value=self.amount
        )

        return asset_entry, liability_entry

    def _calculate_expected_cash_flow(self) -> float:
        if self.type not in [EntryType.BOND_ZERO_COUPON.value, 
                           EntryType.BOND_COUPON.value, 
                           EntryType.BOND_AMORTIZING.value]:
            return 1.0  # Default cash flows for non-bond entries

        if self.type == EntryType.BOND_ZERO_COUPON.value:
            return self.amount / self.amount 

        elif self.type == EntryType.BOND_COUPON.value:
            if not self.coupon_rate:
                raise ValueError("Coupon bonds require a coupon rate")
            
            if self.maturity_date == 1: 
                # Maturity at t1: one coupon payment + principal
                return (self.amount * self.coupon_rate + self.amount) / self.amount
            else:  
                # Maturity at t2: two coupon payments + principal
                return (2 * self.amount * self.coupon_rate + self.amount) / self.amount

        elif self.type == EntryType.BOND_AMORTIZING.value:
            if not self.coupon_rate:
                raise ValueError("Amortization of bonds requires interest rate")
            
            if self.maturity_date == 1: 
                # Maturity at t1: lump sum payment of principal + interest
                return (self.amount * (1 + self.coupon_rate)) / self.amount
            else: 
                # Maturity at t2: two payments of principal/2 + interest
                principal_t1 = self.amount / 2
                principal_t2 = self.amount / 2
                interest_t1 = self.amount * self.coupon_rate
                interest_t2 = principal_t2 * self.coupon_rate
                return (principal_t1 + interest_t1 + principal_t2 + interest_t2) / self.amount

    def create_bond_claims(self) -> List[BalanceSheetEntry]:
        claims = []
        if self.type in [EntryType.BOND_COUPON.value, EntryType.BOND_AMORTIZING.value]:
            schedule = self._create_bond_payment_schedule()
            
            # create a unique bond ID based on type and instance ID
            bond_id = f"bond_{self.type}_{id(self)}"
            
            for date, amount, payment_type in schedule:
                bv = amount
                cf = amount / self.amount
                
                # create claim for the bond payment
                claim = BalanceSheetEntry(
                    type=EntryType.PAYABLE,
                    is_asset=True,
                    counterparty=self.liability_holder.name,
                    initial_book_value=amount,
                    denomination=self.denomination,
                    maturity_type=MaturityType.FIXED_DATE,
                    maturity_date=date,
                    settlement_details=SettlementDetails(
                        type=SettlementType.MEANS_OF_PAYMENT,
                        denomination=self.denomination
                    ),
                    name=f"{payment_type} for {bond_id}",
                    issuance_time=self.current_time_state,
                    parent_bond=bond_id,
                    current_book_value=bv,
                    expected_cash_flow=cf
                )
                claims.append(claim)
                
        return claims

    def _create_bond_payment_schedule(self) -> List[Tuple[int, float, str]]:
        schedule = []
        
        if self.type == EntryType.BOND_COUPON.value:
            if not self.coupon_rate:
                raise ValueError("Coupon bonds require a coupon rate")
                
            coupon_amount = self.amount * self.coupon_rate
            

            if self.maturity_date > 1: 
                schedule.append((1, coupon_amount, "Coupon")) 
                
            final_payment = coupon_amount + self.amount
            schedule.append((self.maturity_date, final_payment, "Coupon+Principal"))
                
        elif self.type == EntryType.BOND_AMORTIZING.value:
            if not self.coupon_rate:
                raise ValueError("Amortization of bonds requires interest rate")
                
            if self.maturity_date == 1:  
                payment = self.amount * (1 + self.coupon_rate)
                schedule.append((1, payment, "Principal+Interest"))
            else: 
                principal_t1 = self.amount / 2
                interest_t1 = self.amount * self.coupon_rate
                payment_t1 = principal_t1 + interest_t1
                schedule.append((1, payment_t1, "Principal+Interest"))
                
                principal_t2 = self.amount / 2
                interest_t2 = principal_t2 * self.coupon_rate
                payment_t2 = principal_t2 + interest_t2
                schedule.append((self.maturity_date, payment_t2, "Principal+Interest"))
                
        return schedule

    def calculate_share_value(self):
        if self.type != EntryType.SHARE.value:
            return

        if not self.asset_holder or not self.liability_holder:
            raise ValueError("Shares lack asset/liability holders")

        issuer_net_worth = self.liability_holder.get_net_worth()
        if issuer_net_worth <= 0:
            raise ValueError("Issuer net worth must be positive")

        if self.amount <= 0:
            raise ValueError("Number of shares must be positive")

        return issuer_net_worth / self.amount

class EconomicSystem:
    def __init__(self):
        self.agents: Dict[str, Agent] = {} 
        self.asset_liability_pairs: List[AssetLiabilityPair] = []
        self.time_states: Dict[int, Dict[str, Agent]] = {} 
        self.current_time_state = 0  
        self.simulation_finalized = False 
        self.scheduled_actions = {} 
        self.save_state(0)

    def validate_time_point(self, time_point: int, allow_t0: bool = True) -> None:
        valid_points = [0, 1, 2] if allow_t0 else [1, 2]
        if time_point not in valid_points:
            valid_points_str = ", ".join(map(str, valid_points))
            raise ValueError(f"Timepoint must be {valid_points_str}")

    def add_agent(self, agent: Agent):
        self.agents[agent.name] = agent
        if agent.type == AgentType.BANK:
            agent.system = self
            agent.intraday_module.system = self

        if self.current_time_state == 0:
            self.save_state(0)

    def create_asset_liability_pair(self, pair: AssetLiabilityPair):

        pair.current_time_state = self.current_time_state
        self.asset_liability_pairs.append(pair)
        asset_entry, liability_entry = pair.create_entries()
        pair.asset_holder.add_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.add_liability(liability_entry)

        self.save_state(self.current_time_state)

    def get_time_points(self) -> List[int]:
        return [0, 1, 2]

    def save_state(self, time_point: int):
        self.validate_time_point(time_point)
        self.time_states[time_point] = {}
        for name, agent in self.agents.items():
            agent_copy = Agent(agent.name, agent.type)
            agent_copy.assets = deepcopy(agent.assets)
            agent_copy.liabilities = deepcopy(agent.liabilities)
            self.time_states[time_point][name] = agent_copy

        self.current_time_state = time_point

    def settle_entries(self, time_point: int):
        self.validate_time_point(time_point, allow_t0=False)

        prev_time = 0 if time_point == 1 else 1
        if prev_time not in self.time_states:
            self.save_state(prev_time)

        for pair in self.asset_liability_pairs[:]: 
            if (pair.maturity_type == MaturityType.FIXED_DATE and
                pair.maturity_date == time_point):
     
                self.asset_liability_pairs.remove(pair)
                asset_entry, liability_entry = pair.create_entries()
                pair.asset_holder.remove_asset(asset_entry)
                if liability_entry:
                    pair.liability_holder.remove_liability(liability_entry)

                if pair.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
                    debtor_deposit = next(
                        (asset for asset in pair.liability_holder.assets
                         if asset.type == EntryType.DEPOSIT
                         and asset.current_book_value >= pair.amount
                         and asset.denomination == pair.denomination),
                        None
                    )

                    if not debtor_deposit:
                        raise ValueError(f"Suitable deposits not found for settlement")

                    # Access to banks holding deposit liabilities
                    bank = next(a for a in self.agents.values() if a.name == debtor_deposit.counterparty)

                    # Removal of the original deposit from the debtor
                    pair.liability_holder.remove_asset(debtor_deposit)

                    # Remove the corresponding liability from the bank
                    bank_liability = next(
                        (l for l in bank.liabilities
                         if l.type == EntryType.DEPOSIT
                         and l.counterparty == pair.liability_holder.name
                         and l.current_book_value == debtor_deposit.current_book_value),
                        None
                    )
                    if bank_liability:
                        bank.remove_liability(bank_liability)

                    # create new settlement pair
                    settlement_pair = AssetLiabilityPair(
                        time=time_point,
                        type=EntryType.DEPOSIT.value,
                        amount=pair.amount,
                        denomination=pair.denomination,
                        maturity_type=MaturityType.ON_DEMAND,
                        maturity_date=None,
                        settlement_type=SettlementType.NONE,
                        settlement_denomination=pair.denomination,
                        asset_holder=pair.asset_holder, 
                        liability_holder=bank 
                    )

                    # use current time point as issuance time
                    settlement_pair.current_time_state = time_point
                    new_asset_entry, new_liability_entry = settlement_pair.create_entries()

                    # record settlement history
                    pair.asset_holder.record_settlement(
                        time_point=time_point,
                        original_entry=asset_entry,
                        settlement_result=new_asset_entry,
                        counterparty=pair.liability_holder.name,
                        as_asset_holder=True
                    )
                    pair.liability_holder.record_settlement(
                        time_point=time_point,
                        original_entry=liability_entry,
                        settlement_result=debtor_deposit, 
                        counterparty=pair.asset_holder.name,
                        as_asset_holder=False
                    )

                    settlement_pair.asset_holder.add_asset(new_asset_entry)
                    if new_liability_entry:
                        settlement_pair.liability_holder.add_liability(new_liability_entry)
                    self.asset_liability_pairs.append(settlement_pair)

                    # if debtor deposit has remaining balance, create a new entry
                    if debtor_deposit.current_book_value > pair.amount:
                        remainder_pair = AssetLiabilityPair(
                            time=time_point,
                            type=EntryType.DEPOSIT.value,
                            amount=debtor_deposit.current_book_value - pair.amount,
                            denomination=pair.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=pair.denomination,
                            asset_holder=pair.liability_holder, 
                            liability_holder=bank 
                        )

                        
                        remainder_pair.current_time_state = time_point
                        remainder_asset, remainder_liability = remainder_pair.create_entries()

                        
                        remainder_pair.asset_holder.add_asset(remainder_asset)
                        if remainder_liability:
                            remainder_pair.liability_holder.add_liability(remainder_liability)
                        self.asset_liability_pairs.append(remainder_pair)

                elif pair.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
                    # Locate and remove non-financial assets from liability holders
                    non_financial_asset = next(
                        (asset for asset in pair.liability_holder.assets
                         if asset.type == EntryType.NON_FINANCIAL
                         and asset.name == pair.asset_name
                         and asset.current_book_value >= pair.amount),
                        None
                    )

                    if not non_financial_asset:
                        raise ValueError(f"Non-financial assets not found for settlement {pair.asset_name}")

                    # delete the non-financial asset from the liability holder
                    pair.liability_holder.remove_asset(non_financial_asset)

                    # Create non-financial asset entries for asset holders
                    settlement_pair = AssetLiabilityPair(
                        time=time_point,
                        type=EntryType.NON_FINANCIAL.value,
                        amount=pair.amount,
                        denomination=pair.settlement_details.denomination,
                        maturity_type=MaturityType.ON_DEMAND,
                        maturity_date=None,
                        settlement_type=SettlementType.NONE,
                        settlement_denomination=pair.settlement_details.denomination,
                        asset_holder=pair.asset_holder,  
                        liability_holder=None, 
                        asset_name=pair.asset_name 
                    )

                    settlement_pair.current_time_state = time_point
                    new_asset_entry, _ = settlement_pair.create_entries()


                    pair.asset_holder.record_settlement(
                        time_point=time_point,
                        original_entry=asset_entry,
                        settlement_result=new_asset_entry,
                        counterparty=pair.liability_holder.name,
                        as_asset_holder=True
                    )
                    pair.liability_holder.record_settlement(
                        time_point=time_point,
                        original_entry=liability_entry,
                        settlement_result=non_financial_asset, 
                        counterparty=pair.asset_holder.name,
                        as_asset_holder=False
                    )

                    settlement_pair.asset_holder.add_asset(new_asset_entry)
                    self.asset_liability_pairs.append(settlement_pair)

                    # if non-financial asset has remaining balance, create a new entry
                    if non_financial_asset.current_book_value > pair.amount:
                        remainder_pair = AssetLiabilityPair(
                            time=time_point,
                            type=EntryType.NON_FINANCIAL.value,
                            amount=non_financial_asset.current_book_value - pair.amount,
                            denomination=non_financial_asset.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=non_financial_asset.denomination,
                            asset_holder=pair.liability_holder, 
                            liability_holder=None,
                            asset_name=non_financial_asset.name
                        )

                        remainder_pair.current_time_state = time_point
                        remainder_asset, _ = remainder_pair.create_entries()

                        remainder_pair.asset_holder.add_asset(remainder_asset)
                        self.asset_liability_pairs.append(remainder_pair)

        # Handling of interbank liquidity
        for agent in self.agents.values():
            if agent.type == AgentType.BANK:
                agent.intraday_module.process_intraday_settlements(time_point)


        self.save_state(time_point)

    def adj_book_values(self):

        for agent in self.agents.values():
            for liability in agent.liabilities:
                if liability.maturity_type == MaturityType.FIXED_DATE and liability.maturity_date is not None:
                    self._adjust_value(liability)

            for asset in agent.assets:
                if asset.maturity_type == MaturityType.FIXED_DATE and asset.maturity_date is not None:
                    self._adjust_value(asset)

    def _adjust_value(self, bal_entry: BalanceSheetEntry):
        t = self.current_time_state - bal_entry.issuance_time
        if bal_entry.maturity_date == bal_entry.issuance_time:  
            bal_entry.current_book_value = bal_entry.initial_book_value
            return
            
        m = bal_entry.maturity_date - bal_entry.issuance_time
        if m == 0: 
             bal_entry.current_book_value = bal_entry.initial_book_value
             return
        rate_of_adjustment = (bal_entry.cash_flow_at_maturity / bal_entry.initial_book_value)**(1/m) - 1
        bal_entry.current_book_value = bal_entry.initial_book_value * (1 + rate_of_adjustment)**t

    def get_agents_at_time(self, time_point: int) -> Dict[str, Agent]:
        self.validate_time_point(time_point)

        if time_point == 0:
            return {name: agent for name, agent in self.agents.items()}
        if time_point in self.time_states:
            return self.time_states[time_point]

        if time_point > self.current_time_state:
            current_state = deepcopy(self.agents)
            current_time = self.current_time_state

            try:
                for t in range(self.current_time_state + 1, time_point + 1):
                    self.settle_entries(t)

                # get the state after settlement
                result = {name: agent for name, agent in self.agents.items()}

                
                self.agents = current_state
                self.current_time_state = current_time

                return result
            except Exception as e:
                # if settlement fails, restore the previous state
                self.agents = current_state
                self.current_time_state = current_time
                print(f"\nWarning: Unable to settle ({str(e)})")
                return current_state

        return {name: agent for name, agent in self.agents.items()}

    def compute_changes(self, from_time: int, to_time: int) -> Dict[str, Dict[str, List]]:
        if from_time not in self.time_states or to_time not in self.time_states:
            raise ValueError(f"Lack status for state {from_time} or {to_time}")

        changes = {}
        for name, to_agent in self.time_states[to_time].items():
            from_agent = self.time_states[from_time][name]

            new_assets = [a for a in to_agent.assets if not any(a.matches(from_a) for from_a in from_agent.assets)]
            removed_assets = [a for a in from_agent.assets if not any(a.matches(to_a) for to_a in to_agent.assets)]

            new_liabilities = [l for l in to_agent.liabilities if not any(l.matches(from_l) for from_l in from_agent.liabilities)]
            removed_liabilities = [l for l in from_agent.liabilities if not any(l.matches(to_l) for to_l in to_agent.liabilities)]

            changes[name] = {
                'new_assets': new_assets,
                'removed_assets': removed_assets,
                'new_liabilities': new_liabilities,
                'removed_liabilities': removed_liabilities
            }

        return changes

    def can_settle_entry(self, agent: Agent, entry: BalanceSheetEntry) -> Tuple[bool, str]:
        if entry.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:

            deposits = sum(asset.current_book_value for asset in agent.assets
                          if asset.type == EntryType.DEPOSIT
                          and asset.is_asset
                          and asset.denomination == entry.denomination)
            if deposits < entry.current_book_value:
                return False, f"Insufficient deposit: has {deposits} {entry.denomination}, needs {entry.current_book_value}"

        elif entry.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            has_asset = any(asset.type == EntryType.NON_FINANCIAL
                           and asset.name == entry.name
                           and asset.current_book_value >= entry.current_book_value
                           for asset in agent.assets)
            if not has_asset:
                return False, f"Lack necessary non-financial asset: {entry.name}"

        return True, "OK"

    def create_default_entries(self, failed_entry: BalanceSheetEntry) -> Tuple[BalanceSheetEntry, BalanceSheetEntry]:
        # create default claim for the creditor
        default_claim = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.current_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default {failed_entry.type.value}",
            issuance_time=self.current_time_state,
            current_book_value=failed_entry.current_book_value
        )

        # create default liability for the debtor
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.current_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default {failed_entry.type.value}",
            issuance_time=self.current_time_state,
            current_book_value=failed_entry.current_book_value
        )

        return default_claim, default_liability

    def schedule_action(self, time_point: int, action_type: str, params: Dict[str, Any]):
        if time_point not in self.scheduled_actions:
            self.scheduled_actions[time_point] = []
        
        self.scheduled_actions[time_point].append({
            'type': action_type,
            'params': params
        })

    def run_user_scheduled_actions(self):
        actions = self.scheduled_actions.get(self.current_time_state, [])
        if not actions:
            return 
        
        action_map = {
            'Create Asset–Liability Pair': self.create_asset_liability_pair,
            'Issue Loan': self.disburse_loans,
            'Request Loan': self.submit_loan_requests,
            'Quote Loan Terms': self.return_loan_quotes,
            'Commit to Loan': self.commit_to_terms,
            'Repay Loan': self.execute_repayments,
            'Default on Loan': self.resolve_defaults,
            'Issue Bond': self.issue_securities,
            'Request Bond Quote': self.submit_trade_quotes,
            'Dealer Bond Quote': self.return_trade_quotes,
            'Commit to Bond Purchase': self.commit_to_terms,
            'Execute Bond Trade': self.dealer_execute_trades,
            'Repay Bond': self.execute_repayments,
            'Default on Bond': self.resolve_defaults,
            'Issue Share': self.issue_shares,
            'Request Share Quote': self.submit_trade_quotes,
            'Dealer Share Quote': self.return_trade_quotes,
            'Commit to Share Purchase': self.commit_to_terms,
            'Execute Share Trade': self.dealer_execute_trades,
            'Pay Dividend': self.pay_dividends,
            'Transfer Customer Deposit': self.transfer_customer_deposit,
            'Process Intraday Settlements': self.process_intraday_settlements
        }
        
        for action in actions:
            action_type = action['type']
            params = action['params']
            
            if action_type in action_map:
                try:
                    action_map[action_type](**params)
                    print(f"Opearation success: {action_type}, parameters: {params}")
                except Exception as e:
                    print(f"Operation failure {action_type}: {e}")
            else:
                print(f"Warning: undefined operation: '{action_type}'")

    
    def disburse_loans(self, **params):
        lender = self.agents[params['lender']]
        borrower = self.agents[params['borrower']]
        amount = params['amount']
        interest_rate = params.get('interest_rate', 0.05)
        maturity = params.get('maturity', 1) 
        
        # create loan entry
        loan_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.LOAN.value,
            amount=amount,
            denomination=params.get('denomination', 'USD'),
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=maturity,
            settlement_type=SettlementType.MEANS_OF_PAYMENT,
            settlement_denomination=params.get('denomination', 'USD'),
            asset_holder=lender,
            liability_holder=borrower,
            cash_flow_at_maturity=amount * (1 + interest_rate)
        )
        
        # create deposit entry for the borrower
        deposit_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.DEPOSIT.value,
            amount=amount,
            denomination=params.get('denomination', 'USD'),
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_type=SettlementType.NONE,
            settlement_denomination=params.get('denomination', 'USD'),
            asset_holder=borrower,
            liability_holder=lender
        )
        
        self.create_asset_liability_pair(loan_pair)
        self.create_asset_liability_pair(deposit_pair)

    def submit_loan_requests(self, **params):
        print(f"Loan request received: {params}")
        pass

    def return_loan_quotes(self, **params):
        print(f"Loan quotes returned: {params}")
        pass

    def commit_to_terms(self, **params):
        print(f"Committed terms: {params}")
        pass

    def execute_repayments(self, **params):
        print(f"Repayments executed: {params}")
        pass

    def resolve_defaults(self, **params):
        print(f"Default resolved: {params}")
        pass

    def issue_securities(self, **params):
        issuer = self.agents[params['issuer']]
        buyer = self.agents[params['buyer']]
        amount = params['amount']
        bond_type_val = params.get('bond_type', BondType.ZERO_COUPON.value) 
        coupon_rate = params.get('coupon_rate', 0.0)
        maturity = params.get('maturity', 1)
        price = params.get('price', amount)
        
        entry_type = EntryType.BOND.value
        bond_type_enum = BondType(bond_type_val)
        if bond_type_enum == BondType.ZERO_COUPON:
            entry_type = EntryType.BOND_ZERO_COUPON.value
        elif bond_type_enum == BondType.COUPON:
            entry_type = EntryType.BOND_COUPON.value
        elif bond_type_enum == BondType.AMORTIZING:
            entry_type = EntryType.BOND_AMORTIZING.value

        bond_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=entry_type,
            amount=amount,
            denomination=params.get('denomination', 'USD'),
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=maturity,
            settlement_type=SettlementType.MEANS_OF_PAYMENT,
            settlement_denomination=params.get('denomination', 'USD'),
            asset_holder=buyer,
            liability_holder=issuer,
            bond_type=bond_type_enum,
            coupon_rate=coupon_rate
        )
        
        deposit_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.DEPOSIT.value,
            amount=price,
            denomination=params.get('denomination', 'USD'),
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_type=SettlementType.NONE,
            settlement_denomination=params.get('denomination', 'USD'),
            asset_holder=issuer,
            liability_holder=buyer
        )
        
        self.create_asset_liability_pair(bond_pair)
        self.create_asset_liability_pair(deposit_pair)

    def submit_trade_quotes(self, **params):
        print(f"Trade quotes submitted: {params}")
        pass

    def return_trade_quotes(self, **params):
        print(f"Trade quotes returned: {params}")
        pass

    def dealer_execute_trades(self, **params):
        print(f"Trade executed: {params}")
        pass

    def issue_shares(self, **params):
        issuer = self.agents[params['issuer']]
        buyer = self.agents[params['buyer']]
        amount = params['amount']
        price = params.get('price', amount)
        
        share_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.SHARE.value,
            amount=amount,
            denomination=params.get('denomination', 'shares'),
            maturity_type=MaturityType.PERPETUAL,
            maturity_date=None,
            settlement_type=SettlementType.NONE,
            settlement_denomination='shares',
            asset_holder=buyer,
            liability_holder=issuer
        )
        
        deposit_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.DEPOSIT.value,
            amount=price,
            denomination=params.get('price_denomination', 'USD'),
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_type=SettlementType.NONE,
            settlement_denomination=params.get('price_denomination', 'USD'),
            asset_holder=issuer,
            liability_holder=buyer
        )
        
        self.create_asset_liability_pair(share_pair)
        self.create_asset_liability_pair(deposit_pair)

    def pay_dividends(self, **params):
        issuer = self.agents[params['issuer']]
        dividend_per_share = params['dividend_per_share']
        
        shareholders = {}
        for agent_name, agent in self.agents.items():
            for asset in agent.assets:
                if asset.type == EntryType.SHARE and asset.counterparty == issuer.name:
                    shareholders[agent_name] = asset.current_book_value 
        
        for shareholder_name, shares in shareholders.items():
            shareholder = self.agents[shareholder_name]
            dividend_amount = shares * dividend_per_share
            
     
            dividend_pair = AssetLiabilityPair(
                time=self.current_time_state,
                type=EntryType.PAYABLE.value,
                amount=dividend_amount,
                denomination=params.get('denomination', 'USD'),
                maturity_type=MaturityType.FIXED_DATE,
                maturity_date=self.current_time_state, 
                settlement_type=SettlementType.MEANS_OF_PAYMENT,
                settlement_denomination=params.get('denomination', 'USD'),
                asset_holder=shareholder,
                liability_holder=issuer
            )
            
            self.create_asset_liability_pair(dividend_pair)

    def transfer_customer_deposit(self, **params):
        sender_bank = self.agents[params['sender_bank']]
        receiver_bank = self.agents[params['receiver_bank']]
        customer_a = params['customer_a']
        customer_b = params['customer_b']
        amount = params['amount']
        
        # update customer deposits
        if customer_a not in sender_bank.customer_deposits:
            sender_bank.customer_deposits[customer_a] = 0
        if customer_b not in receiver_bank.customer_deposits:
            receiver_bank.customer_deposits[customer_b] = 0
            
        sender_bank.customer_deposits[customer_a] -= amount
        receiver_bank.customer_deposits[customer_b] += amount
        
        # create daytime IOU entry
        iou_pair = AssetLiabilityPair(
            time=self.current_time_state,
            type=EntryType.INTRADAY_IOU.value,
            amount=amount,
            denomination=params.get('denomination', 'USD'),
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=self.current_time_state + 1, 
            settlement_type=SettlementType.MEANS_OF_PAYMENT,
            settlement_denomination=params.get('denomination', 'USD'),
            asset_holder=receiver_bank,
            liability_holder=sender_bank
        )
        
        self.create_asset_liability_pair(iou_pair)

    def process_intraday_settlements(self, **params):
        for agent in self.agents.values():
            if agent.type == AgentType.BANK:
                agent.intraday_module.process_intraday_settlements(self.current_time_state)


def export_to_excel(system: EconomicSystem, filename: str = "economic_system.xlsx"):
    if not EXCEL_AVAILABLE:
        print("Unable to export to Excel: openpyxl module is not available.")
        return

    wb = openpyxl.Workbook()
    
    agents_sheet = wb.active
    agents_sheet.title = "Agents"
    agents_sheet.append(["Name", "Type", "Total Assets", "Total Liabilities", "Net Worth", "Status"])
    
    for agent in system.agents.values():
        metrics = agent.get_type_specific_metrics()
        agents_sheet.append([
            agent.name,
            agent.type.value,
            metrics["total_assets"],
            metrics["total_liabilities"],
            metrics["net_worth"],
            agent.status
        ])
    
    for time_point in system.get_time_points():
        if time_point in system.time_states:
            agents = system.time_states[time_point]
            
            time_sheet = wb.create_sheet(f"Time {time_point}")
            time_sheet.append(["Agent", "Type", "Total Assets", "Total Liabilities", "Net Worth"])
            
            for name, agent in agents.items():
                time_sheet.append([
                    agent.name,
                    agent.type.value,
                    sum(asset.current_book_value for asset in agent.assets),
                    sum(liability.current_book_value for liability in agent.liabilities),
                    sum(asset.current_book_value for asset in agent.assets) - 
                    sum(liability.current_book_value for liability in agent.liabilities)
                ])
            
            for name, agent in agents.items():
                detail_sheet = wb.create_sheet(f"T{time_point} - {name}")
                
                detail_sheet.append(["ASSETS"])
                detail_sheet.append(["Type", "Counterparty", "Amount", "Denomination", 
                                    "Maturity", "Maturity Date", "Settlement", "Name"])
                
                for asset in agent.assets:
                    maturity_date = asset.maturity_date if asset.maturity_date is not None else "N/A"
                    detail_sheet.append([
                        asset.type.value,
                        asset.counterparty if asset.counterparty else "N/A",
                        asset.current_book_value,
                        asset.denomination,
                        asset.maturity_type.value,
                        maturity_date,
                        asset.settlement_details.type.value,
                        asset.name if asset.name else "N/A"
                    ])

                detail_sheet.append([])
                
                detail_sheet.append(["LIABILITIES"])
                detail_sheet.append(["Type", "Counterparty", "Amount", "Denomination", 
                                    "Maturity", "Maturity Date", "Settlement", "Name"])
                
                for liability in agent.liabilities:
                    maturity_date = liability.maturity_date if liability.maturity_date is not None else "N/A"
                    detail_sheet.append([
                        liability.type.value,
                        liability.counterparty,
                        liability.current_book_value,
                        liability.denomination,
                        liability.maturity_type.value,
                        maturity_date,
                        liability.settlement_details.type.value,
                        liability.name if liability.name else "N/A"
                    ])
    
    try:
        wb.save(filename)
        print(f"Exported to {filename}")
    except Exception as e:
        print(f"Export failure: {e}")


def get_input(prompt: str, input_type=str, choices=None):

    while True:
        try:
            value = input(prompt).strip()
            if not value and input_type != str:
                print("Input cannot be empty")
                continue
            
            if input_type == bool:
                if value.lower() in ['y', 'yes', 'true', '1']:
                    return True
                elif value.lower() in ['n', 'no', 'false', '0']:
                    return False
                else:
                    print("Please enter y/n or true/false")
            else:
                converted_value = input_type(value)
                if choices and converted_value not in choices:
                    print(f"Invalid choice. Please choose within {choices}.")
                else:
                    return converted_value
        except ValueError:
            print(f"Invalid input. Please enter {input_type.__name__} type.")
        except Exception as e:
            print(f"Error: {e}")

def select_agent(system: EconomicSystem, prompt: str) -> Optional[Agent]:
    agents = list(system.agents.keys())
    if not agents:
        print("No agents available.")
        return None
    print("Available agents:")
    for i, name in enumerate(agents):
        print(f"  {i + 1}. {name}")
    while True:
        choice = get_input(f"{prompt} (Enter number): ", int)
        if 1 <= choice <= len(agents):
            return system.agents[agents[choice - 1]]
        else:
            print("Invalid number")

def select_enum(enum_class, prompt: str):
    print(f"Available: ({enum_class.__name__}):")
    members = list(enum_class)
    for i, member in enumerate(members):
        print(f"  {i + 1}. {member.value}")
    while True:
        choice = get_input(f"{prompt} (Enter number): ", int)
        if 1 <= choice <= len(members):
            return members[choice - 1]
        else:
            print("Invalid number")


def main_interactive():
    system = EconomicSystem()
    print("Welcome to the Economic Systems Simulator!")

    while True:
        print("\n--- Menu ---")
        print("1. Create agent")
        print("2. Create asset-liability pair")
        print("3. Operations")
        print("4. Settle due entries")
        print("5. Run scheduled actions")
        print("6. View balance sheets")
        print("7. View status")
        print("8. Export to Excel")
        print("9. Exit")

        choice = get_input("Please select: ", int, choices=list(range(1, 10)))

        try:
            if choice == 1:
                name = get_input("Enter agent name: ")
                agent_type = select_enum(AgentType, "Enter agent type")
                agent = Agent(name, agent_type)
                system.add_agent(agent)
                print(f"Agent '{name}' ({agent_type.value}) added.")

            elif choice == 2:
                print("\n--- Create asset-liability pair ---")
                asset_holder = select_agent(system, "select asset holder")
                if not asset_holder:
                    continue
                
                entry_type = select_enum(EntryType, "select entry type")
                
                liability_holder = None
                asset_name = None
                bond_type = None
                coupon_rate = None
                
                if entry_type != EntryType.NON_FINANCIAL:
                    liability_holder = select_agent(system, "select liability holder")
                    if not liability_holder:
                        continue
                else:
                    asset_name = get_input("Enter name for non-financial asset: ")

                amount = get_input("Enter amount: ", float)
                denomination = get_input("Enter denomination (USD, shares, reserves, etc.): ")
                maturity_type = select_enum(MaturityType, "select maturity type")
                maturity_date = None
                if maturity_type == MaturityType.FIXED_DATE:
                    maturity_date = get_input("Enter maturity date: ", int, choices=[1, 2])
                
                settlement_type = select_enum(SettlementType, "select settlement type")
                settlement_denomination = denomination
                if settlement_type != SettlementType.NONE:
                     settlement_denomination = get_input("Enter denomination: ")
                
                cash_flow_at_maturity = 0
                if maturity_type == MaturityType.FIXED_DATE:
                     cash_flow_at_maturity = get_input("Enter cash flow at maturity (if applicable): ", float)
                     
                if entry_type in [EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
                    bond_type = select_enum(BondType, "Enter bond type")
                    coupon_rate = get_input("Enter coupon rate: ", float)
                elif entry_type == EntryType.BOND_ZERO_COUPON:
                     bond_type = BondType.ZERO_COUPON

                pair = AssetLiabilityPair(
                    time=system.current_time_state,
                    type=entry_type.value,
                    amount=amount,
                    denomination=denomination,
                    maturity_type=maturity_type,
                    maturity_date=maturity_date,
                    settlement_type=settlement_type,
                    settlement_denomination=settlement_denomination,
                    asset_holder=asset_holder,
                    liability_holder=liability_holder,
                    cash_flow_at_maturity=cash_flow_at_maturity,
                    asset_name=asset_name,
                    bond_type=bond_type,
                    coupon_rate=coupon_rate
                )
                system.create_asset_liability_pair(pair)
                print("asset-liability pair created successfully.")

            elif choice == 3:
                print("\n--- Operations ---")
                time_point = get_input("Enter executing time: ", int, choices=[1, 2])
                
                action_types = [
                    'Issue Loan', 'Repay Loan', 'Issue Bond', 'Repay Bond', 
                    'Issue Share', 'Pay Dividend', 'Transfer Customer Deposit', 
                    'Process Intraday Settlements'
                ]
                print("Applicable types:")
                for i, atype in enumerate(action_types):
                    print(f"  {i + 1}. {atype}")
                action_choice = get_input("Select action type: ", int, choices=list(range(1, len(action_types) + 1)))
                action_type = action_types[action_choice - 1]
                
                params = {}
                if action_type == 'Issue Loan':
                    lender = select_agent(system, "Select lender")
                    borrower = select_agent(system, "Select borrower")
                    if lender and borrower:
                        params['lender'] = lender.name
                        params['borrower'] = borrower.name
                        params['amount'] = get_input("Enter amount: ", float)
                        params['interest_rate'] = get_input("Enter interest rate: ", float)
                        params['maturity'] = get_input("Enter maturity date: ", int, choices=[1, 2])
                        params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Repay Loan':
                     borrower = select_agent(system, "Select borrower")
                     lender = select_agent(system, "Select lender")
                     if borrower and lender:
                         params['borrower'] = borrower.name
                         params['lender'] = lender.name
                         params['amount'] = get_input("Enter amount: ", float)
                         params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Issue Bond':
                    issuer = select_agent(system, "Select issuer")
                    buyer = select_agent(system, "Select buyer")
                    if issuer and buyer:
                        params['issuer'] = issuer.name
                        params['buyer'] = buyer.name
                        params['amount'] = get_input("Enter amount: ", float)
                        bond_type_enum = select_enum(BondType, "Select bond type")
                        params['bond_type'] = bond_type_enum.value
                        if bond_type_enum != BondType.ZERO_COUPON:
                             params['coupon_rate'] = get_input("Enter coupon rate: ", float)
                        params['maturity'] = get_input("Enter maturity date: ", int, choices=[1, 2])
                        params['price'] = get_input("Enter price: ", float)
                        params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Repay Bond':
                     issuer = select_agent(system, "Select issuer")
                     holder = select_agent(system, "Select holder")
                     if issuer and holder:
                         params['issuer'] = issuer.name
                         params['holder'] = holder.name
                         params['amount'] = get_input("Enter amount: ", float)
                         params['interest'] = get_input("Enter interest: ", float)
                         params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Issue Share':
                    issuer = select_agent(system, "Select issuer")
                    buyer = select_agent(system, "Select buyer")
                    if issuer and buyer:
                        params['issuer'] = issuer.name
                        params['buyer'] = buyer.name
                        params['amount'] = get_input("Enter amount: ", float)
                        params['price'] = get_input("Enter price: ", float)
                        params['price_denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Pay Dividend':
                     issuer = select_agent(system, "Select issuer")
                     if issuer:
                         params['issuer'] = issuer.name
                         params['dividend_per_share'] = get_input("Enter dividend per share: ", float)
                         params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Transfer Customer Deposit':
                    sender_bank = select_agent(system, "Select sender bank")
                    receiver_bank = select_agent(system, "Select receiver bank")
                    if sender_bank and receiver_bank:
                        params['sender_bank'] = sender_bank.name
                        params['receiver_bank'] = receiver_bank.name
                        params['customer_a'] = get_input("Enter payer: ")
                        params['customer_b'] = get_input("Enter receiver: ")
                        params['amount'] = get_input("Enter amount: ", float)
                        params['denomination'] = get_input("Enter denomination: ", str)
                elif action_type == 'Process Intraday Settlements':
                    pass 
                
                if params or action_type == 'Process Intraday Settlements': 
                    system.schedule_action(time_point, action_type, params)
                    print(f"Operation '{action_type}' at timepoint {time_point}。")
                else:
                     print("Operation cancelled")

            elif choice == 4:
                time_point = get_input("Enter settle timepoint: ", int, choices=[1, 2])
                system.settle_entries(time_point)
                print(f"Timepoint {time_point} settled")

            elif choice == 5:
                time_point = get_input("Enter operation timepoint: ", int, choices=[1, 2])
                if system.current_time_state != time_point:
                     print(f"Warning: Current timepoint {system.current_time_state}, but you choose to operate {time_point}.")
                print(f"--- Operating {time_point} ---")
                system.run_user_scheduled_actions()
                print(f"Timepoint {time_point} operated")

            elif choice == 6:
                agent = select_agent(system, "Select agent to view balance sheet")
                if agent:
                    print(f"\n--- {agent.name} balance sheet (Timepoint {system.current_time_state}) ---")
                    bs = agent.get_balance_sheet()
                    print("Asset:")
                    if not bs['assets']:
                        print("  (None)")
                    for entry in bs['assets']:
                        print(f"  - {entry.type.value}: {entry.current_book_value} {entry.denomination} (Counterparty: {entry.counterparty or 'N/A'}, maturity date: {entry.maturity_date or entry.maturity_type.value})")
                    print("Liabilities:")
                    if not bs['liabilities']:
                        print("  (None)")
                    for entry in bs['liabilities']:
                         print(f"  - {entry.type.value}: {entry.current_book_value} {entry.denomination} (Counterparty: {entry.counterparty}, maturity date: {entry.maturity_date or entry.maturity_type.value})")
                    metrics = agent.get_type_specific_metrics()
                    print(f"Total asset: {metrics['total_assets']:.2f}")
                    print(f"Total liability: {metrics['total_liabilities']:.2f}")
                    print(f"Net worth: {metrics['net_worth']:.2f}")

            elif choice == 7:
                time_point = get_input("Enter timepoint to be checked: ", int, choices=[0, 1, 2])
                agents_state = system.get_agents_at_time(time_point)
                print(f"\n--- System status (Timepoint {time_point}) ---")
                if not agents_state:
                    print("No status at given timepoint.")
                else:
                    for name, agent in agents_state.items():
                        metrics = agent.get_type_specific_metrics()
                        print(f"  {name} ({agent.type.value}): Asset={metrics['total_assets']:.2f}, Liability={metrics['total_liabilities']:.2f}, Net worth={metrics['net_worth']:.2f}")

            elif choice == 8:
                filename = get_input("Enter export filename: ", str)
                if not filename.endswith(".xlsx"):
                    filename += ".xlsx"
                export_to_excel(system, filename)

            elif choice == 9:
                print("Exiting the simulator.")
                sys.exit()
                
        except ValueError as e:
            print(f"\nError: {e}")
        except KeyError as e:
             print(f"\nError: couldn't find agent '{e}'")
        except Exception as e:
            print(f"\nUnexpected error: {e}")

if __name__ == "__main__":
    main_interactive()

    # # --- 原来的演示代码 (注释掉) ---
    # # 创建经济系统
    # system = EconomicSystem()
    # 
    # # 创建代理
    # central_bank = Agent("Central Bank", AgentType.CENTRAL_BANK)
    # bank_a = Agent("Bank A", AgentType.BANK)
    # bank_b = Agent("Bank B", AgentType.BANK)
    # company = Agent("Company", AgentType.COMPANY)
    # household = Agent("Household", AgentType.HOUSEHOLD)
    # 
    # # 添加代理到系统
    # system.add_agent(central_bank)
    # system.add_agent(bank_a)
    # system.add_agent(bank_b)
    # system.add_agent(company)
    # system.add_agent(household)
    # 
    # # 创建初始资产-负债对
    # 
    # # 1. 中央银行向银行A发行准备金
    # reserves_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.DEPOSIT.value,
    #     amount=100.0,
    #     denomination="reserves",
    #     maturity_type=MaturityType.ON_DEMAND,
    #     maturity_date=None,
    #     settlement_type=SettlementType.NONE,
    #     settlement_denomination="reserves",
    #     asset_holder=bank_a,
    #     liability_holder=central_bank
    # )
    # system.create_asset_liability_pair(reserves_pair)
    # 
    # # 2. 银行A向公司发放贷款
    # loan_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.LOAN.value,
    #     amount=50.0,
    #     denomination="USD",
    #     maturity_type=MaturityType.FIXED_DATE,
    #     maturity_date=1,  # t1到期
    #     settlement_type=SettlementType.MEANS_OF_PAYMENT,
    #     settlement_denomination="USD",
    #     asset_holder=bank_a,
    #     liability_holder=company,
    #     cash_flow_at_maturity=55.0  # 包含利息
    # )
    # system.create_asset_liability_pair(loan_pair)
    # 
    # # 3. 银行A为公司创建存款（贷款资金）
    # deposit_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.DEPOSIT.value,
    #     amount=50.0,
    #     denomination="USD",
    #     maturity_type=MaturityType.ON_DEMAND,
    #     maturity_date=None,
    #     settlement_type=SettlementType.NONE,
    #     settlement_denomination="USD",
    #     asset_holder=company,
    #     liability_holder=bank_a
    # )
    # system.create_asset_liability_pair(deposit_pair)
    # 
    # # 4. 公司向家庭发行股份
    # share_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.SHARE.value,
    #     amount=100.0,  # 股份数量
    #     denomination="shares",
    #     maturity_type=MaturityType.PERPETUAL,
    #     maturity_date=None,
    #     settlement_type=SettlementType.NONE,
    #     settlement_denomination="shares",
    #     asset_holder=household,
    #     liability_holder=company
    # )
    # system.create_asset_liability_pair(share_pair)
    # 
    # # 5. 家庭向公司支付股份购买资金
    # share_payment_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.DEPOSIT.value,
    #     amount=20.0,  # 股份价格
    #     denomination="USD",
    #     maturity_type=MaturityType.ON_DEMAND,
    #     maturity_date=None,
    #     settlement_type=SettlementType.NONE,
    #     settlement_denomination="USD",
    #     asset_holder=company,
    #     liability_holder=bank_b
    # )
    # system.create_asset_liability_pair(share_payment_pair)
    # 
    # # 6. 银行B为家庭创建存款
    # household_deposit_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.DEPOSIT.value,
    #     amount=30.0,
    #     denomination="USD",
    #     maturity_type=MaturityType.ON_DEMAND,
    #     maturity_date=None,
    #     settlement_type=SettlementType.NONE,
    #     settlement_denomination="USD",
    #     asset_holder=household,
    #     liability_holder=bank_b
    # )
    # system.create_asset_liability_pair(household_deposit_pair)
    # 
    # # 7. 公司向家庭发行债券
    # bond_pair = AssetLiabilityPair(
    #     time=0,
    #     type=EntryType.BOND_COUPON.value,
    #     amount=10.0,
    #     denomination="USD",
    #     maturity_type=MaturityType.FIXED_DATE,
    #     maturity_date=2,  # t2到期
    #     settlement_type=SettlementType.MEANS_OF_PAYMENT,
    #     settlement_denomination="USD",
    #     asset_holder=household,
    #     liability_holder=company,
    #     bond_type=BondType.COUPON,
    #     coupon_rate=0.05  # 5%息票率
    # )
    # system.create_asset_liability_pair(bond_pair)
    # 
    # # 调度操作
    # 
    # # 在t1时，公司向银行A还款
    # system.schedule_action(1, 'Repay Loan', {
    #     'borrower': 'Company',
    #     'lender': 'Bank A',
    #     'amount': 55.0,
    #     'denomination': 'USD'
    # })
    # 
    # # 在t1时，公司支付股息
    # system.schedule_action(1, 'Pay Dividend', {
    #     'issuer': 'Company',
    #     'dividend_per_share': 0.1,  # 每股0.1美元
    #     'denomination': 'USD'
    # })
    # 
    # # 在t2时，公司偿还债券
    # system.schedule_action(2, 'Repay Bond', {
    #     'issuer': 'Company',
    #     'holder': 'Household',
    #     'amount': 10.0,
    #     'interest': 0.5,  # 最后一次息票
    #     'denomination': 'USD'
    # })
    # 
    # # 处理t1结算
    # system.settle_entries(1)
    # 
    # # 运行t1的调度操作
    # system.run_user_scheduled_actions()
    # 
    # # 处理t2结算
    # system.settle_entries(2)
    # 
    # # 运行t2的调度操作
    # system.run_user_scheduled_actions()
    # 
    # # 导出到Excel
    # export_to_excel(system, "economic_simulation_results.xlsx")
    # 
    # print("模拟完成！结果已导出到economic_simulation_results.xlsx")
    pass # 占位符，因为之前的main被注释了
