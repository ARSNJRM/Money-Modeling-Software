# economic_sim_web_integrated.py
import os
from io import BytesIO
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple, Union, Any
from enum import Enum
from datetime import datetime, timedelta
from copy import deepcopy
from flask import send_file, flash, url_for, Flask, request, render_template_string, redirect, session, jsonify
import json
import uuid
from pathlib import Path
import copy
import sys

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    print("Warning: openpyxl package not found. Excel export functionality will be disabled.")
    print("To enable Excel export, please install openpyxl using: pip install openpyxl")
    EXCEL_AVAILABLE = False

# ======== 核心枚举和类定义（整合版本） ========

class AgentType(Enum):
    BANK = "bank"
    COMPANY = "company"
    HOUSEHOLD = "household"
    TREASURY = "treasury"
    CENTRAL_BANK = "central_bank"
    OTHER = "other"

class EntryType(Enum):
    LOAN = "bank_client_loan"  # 更新命名
    DEPOSIT = "bank_deposit"   # 更新命名
    RESERVE_DEPOSIT = "reserve_deposit"  # 新增
    PAYABLE = "payable"
    # Integration of bond types
    BOND = "bond" 
    BOND_ZERO_COUPON = "bond_zero_coupon"
    BOND_COUPON = "bond_coupon"
    BOND_AMORTIZING = "bond_amortizing"
    # 新增政府债券类型
    GOVERNMENT_BOND_ZERO_COUPON = "government_bond_zero_coupon"
    GOVERNMENT_BOND_COUPON = "government_bond_coupon"
    GOVERNMENT_BOND_AMORTIZING = "government_bond_amortizing"
    # 新增央行债券类型
    CENTRAL_BANK_BOND_ZERO_COUPON = "central_bank_bond_zero_coupon"
    CENTRAL_BANK_BOND_COUPON = "central_bank_bond_coupon"
    CENTRAL_BANK_BOND_AMORTIZING = "central_bank_bond_amortizing"
    # Integration of types related to interday liquidity of banks
    INTRADAY_IOU = "interbank_intraday_iou"  # 更新命名
    OVERNIGHT_LOAN = "interbank_overnight_loan"  # 更新命名
    # Integration of share types
    SHARE = "share"
    # Other types
    DELIVERY_CLAIM = "delivery_claim"
    NON_FINANCIAL = "non_financial"
    DEFAULT = "default"

class MaturityType(Enum):
    ON_DEMAND = "on_demand"
    FIXED_DATE = "fixed_date"
    PERPETUAL = "perpetual"

class SettlementType(Enum):
    MEANS_OF_PAYMENT = "means_of_payment"
    SECURITIES = "securities"
    NON_FINANCIAL_ASSET = "non_financial_asset"
    SERVICES = "services"
    CRYPTO = "crypto"
    NONE = "none"

class BondType(Enum):
    ZERO_COUPON = 0  # zero coupon bond
    COUPON = 1       # coupon bond
    AMORTIZING = 2   # Amortized bonds

@dataclass
class SettlementDetails:
    type: SettlementType
    denomination: str 
    is_intraday: bool = False 

@dataclass
class BalanceSheetEntry:
    type: EntryType
    is_asset: bool 
    counterparty: Optional[str] 
    initial_book_value: float 
    denomination: str 
    maturity_type: MaturityType  
    maturity_date: Optional[int]  
    settlement_details: SettlementDetails 
    cash_flow_at_maturity: float = 0 
    name: Optional[str] = None  
    issuance_time: int = 0 
    current_book_value: float = 0 
    rollover_count: int = 0 
    expected_cash_flow: Optional[float] = None 
    parent_bond: Optional[str] = None 

    def matches(self, other: 'BalanceSheetEntry') -> bool:
        return (
            self.type == other.type and
            self.is_asset == other.is_asset and
            self.counterparty == other.counterparty and
            self.initial_book_value == other.initial_book_value and
            self.denomination == other.denomination and
            self.maturity_type == other.maturity_type and
            self.maturity_date == other.maturity_date and
            self.settlement_details.type == other.settlement_details.type and
            self.settlement_details.denomination == other.settlement_details.denomination and
            self.name == other.name and
            self.issuance_time == other.issuance_time
        )

    def __post_init__(self):
        if self.initial_book_value <= 0:
            raise ValueError("The amount must be positive")


        if self.issuance_time < 0:
            raise ValueError("Time of issue must be non-negative")

        if self.type != EntryType.NON_FINANCIAL and not self.counterparty:
            raise ValueError("Financial entries must have a counterparty")
        if self.type == EntryType.NON_FINANCIAL and self.counterparty:
            raise ValueError("Non-financial entries cannot have counterparties")

        if self.type == EntryType.NON_FINANCIAL and not self.name:
            raise ValueError("Non-financial entries must have a name")

        if self.type == EntryType.PAYABLE and self.settlement_details.type != SettlementType.MEANS_OF_PAYMENT:
            raise ValueError("Payable entries must have a means of payment settlement type")
            
        if self.type == EntryType.SHARE:
            self.maturity_type = MaturityType.PERPETUAL
            self.maturity_date = None
            
        if self.current_book_value == 0:
            self.current_book_value = self.initial_book_value
            
        # 处理Web UI的兼容性
        if isinstance(self.maturity_type, str):
            self.maturity_type = MaturityType(self.maturity_type)
        if isinstance(self.settlement_details, dict):
            self.settlement_details = SettlementDetails(**self.settlement_details)

    def to_dict(self):
        """用于Web UI的序列化 - 统一使用数字"""
        return {
            "type": self.type.value,
            "is_asset": self.is_asset,
            "counterparty": self.counterparty,
            "amount": self.initial_book_value,  # 为了兼容性保留amount字段
            "initial_book_value": self.initial_book_value,
            "denomination": self.denomination,
            "maturity_type": self.maturity_type.value,
            "maturity_date": self.maturity_date,
            "settlement_details": {
                "type": self.settlement_details.type.value,
                "denomination": self.settlement_details.denomination,
                "is_intraday": getattr(self.settlement_details, 'is_intraday', False)
            },
            "name": self.name,
            "issuance_time": self.issuance_time,  # 直接使用数字，不转换
            "current_book_value": self.current_book_value,
            "cash_flow_at_maturity": self.cash_flow_at_maturity,
            "rollover_count": self.rollover_count,
            "expected_cash_flow": self.expected_cash_flow,
            "parent_bond": self.parent_bond
        }

class AssetLiabilityConstraints:
    """Defines and validates constraints for asset-liability pairs"""
    
    @staticmethod
    def validate_asset_holder(entry_type: EntryType, agent_type: AgentType) -> Tuple[bool, str]:
        """
        Validates if an agent type can hold a specific asset type.
        Returns (is_valid, error_message)
        """
        constraints = {
            EntryType.LOAN: [AgentType.BANK],
            EntryType.DEPOSIT: [AgentType.COMPANY, AgentType.HOUSEHOLD, AgentType.OTHER],
            EntryType.RESERVE_DEPOSIT: [AgentType.BANK, AgentType.TREASURY],
            EntryType.INTRADAY_IOU: [AgentType.BANK],
            EntryType.OVERNIGHT_LOAN: [AgentType.BANK],
            # Other asset types remain unconstrained for now
        }
        
        if entry_type in constraints:
            allowed_types = constraints[entry_type]
            if agent_type not in allowed_types:
                allowed_names = [t.value for t in allowed_types]
                return False, f"{entry_type.value} assets can only be held by: {', '.join(allowed_names)}"
        
        return True, ""
    
    @staticmethod
    def validate_liability_holder(entry_type: EntryType, agent_type: AgentType) -> Tuple[bool, str]:
        """
        Validates if an agent type can hold a specific liability type.
        Returns (is_valid, error_message)
        """
        constraints = {
            EntryType.BOND_ZERO_COUPON: [AgentType.COMPANY, AgentType.BANK, AgentType.CENTRAL_BANK, AgentType.TREASURY, AgentType.OTHER],
            EntryType.BOND_COUPON: [AgentType.COMPANY, AgentType.BANK, AgentType.CENTRAL_BANK, AgentType.TREASURY, AgentType.OTHER],
            EntryType.BOND_AMORTIZING: [AgentType.COMPANY, AgentType.BANK, AgentType.CENTRAL_BANK, AgentType.TREASURY, AgentType.OTHER],
            EntryType.GOVERNMENT_BOND_ZERO_COUPON: [AgentType.TREASURY],
            EntryType.GOVERNMENT_BOND_COUPON: [AgentType.TREASURY],
            EntryType.GOVERNMENT_BOND_AMORTIZING: [AgentType.TREASURY],
            EntryType.CENTRAL_BANK_BOND_ZERO_COUPON: [AgentType.CENTRAL_BANK],
            EntryType.CENTRAL_BANK_BOND_COUPON: [AgentType.CENTRAL_BANK],
            EntryType.CENTRAL_BANK_BOND_AMORTIZING: [AgentType.CENTRAL_BANK],
            EntryType.DEPOSIT: [AgentType.BANK],
            EntryType.RESERVE_DEPOSIT: [AgentType.CENTRAL_BANK],
            EntryType.INTRADAY_IOU: [AgentType.BANK],
            EntryType.OVERNIGHT_LOAN: [AgentType.BANK],
            EntryType.SHARE: [AgentType.BANK, AgentType.COMPANY, AgentType.OTHER],
        }
        
        if entry_type in constraints:
            allowed_types = constraints[entry_type]
            if agent_type not in allowed_types:
                allowed_names = [t.value for t in allowed_types]
                return False, f"{entry_type.value} liabilities can only be held by: {', '.join(allowed_names)}"
        
        return True, ""
    
    @staticmethod
    def validate_asset_liability_pair(asset_holder, liability_holder, entry_type: EntryType) -> Tuple[bool, str]:
        """
        Validates an entire asset-liability pair against all constraints.
        Returns (is_valid, error_message)
        """
        # Check that asset holder and liability holder are different agents
        if liability_holder is not None and asset_holder.name == liability_holder.name:
            return False, "An agent cannot be both the asset holder and liability holder of the same entry"
        
        # Validate asset holder constraints
        is_valid, error_msg = AssetLiabilityConstraints.validate_asset_holder(entry_type, asset_holder.type)
        if not is_valid:
            return False, f"Asset holder constraint violation: {error_msg}"
        
        # Validate liability holder constraints (if liability holder exists)
        if liability_holder is not None:
            is_valid, error_msg = AssetLiabilityConstraints.validate_liability_holder(entry_type, liability_holder.type)
            if not is_valid:
                return False, f"Liability holder constraint violation: {error_msg}"
        
        return True, ""

class SettlementFailure(Exception):
    def __init__(self, agent_name: str, entry: BalanceSheetEntry, reason: str):
        self.agent_name = agent_name
        self.entry = entry
        self.reason = reason
        super().__init__(f"Settlement failure {agent_name}: {reason}")


class BankIntradayModule:
    def __init__(self, bank_agent):
        self.bank = bank_agent
        self.settlement_log = [] 
        self.system = None 

    def _find_iou_counterparty(self, iou_entry):
        counterparty = next(
            (agent for agent in self.bank.system.agents.values() if agent.name == iou_entry.counterparty),
            None
        )
        if counterparty is None:
            self.settlement_log.append(f"Counterparty not found for IOU settlement: {iou_entry.counterparty}")
            raise ValueError(f"Counterparty not found for IOU settlement: {iou_entry.counterparty}")
        return counterparty

    def _settle_partial(self, iou_entry, time_point):
        """
        Attempts to settle as many IOUs as possible using available deposits. 
        Supports partial settlement (if full amount is not available).
        Returns the settlement amount (or 0 if not available).
        """
        # Find all deposits in the bank that match the denomination of the IOU, sorted in descending order of amount
        deposits = sorted(
            (a for a in self.bank.assets if a.type == EntryType.DEPOSIT and a.denomination == iou_entry.denomination),
            key=lambda x: x.initial_book_value,
            reverse=True
        )

        amount_to_settle = iou_entry.initial_book_value
        total_settled = 0.0

        for deposit in deposits:
            if amount_to_settle <= 0:
                break
            settle_amount = min(deposit.initial_book_value, amount_to_settle)

            # Remove or reduce deposits accordingly
            if deposit.initial_book_value == settle_amount:
                self.bank.remove_asset(deposit)
            else:
                deposit.initial_book_value -= settle_amount
                deposit.current_book_value -= settle_amount

            receiver = self._find_iou_counterparty(iou_entry)

            new_deposit = BalanceSheetEntry(
                type=EntryType.DEPOSIT,
                is_asset=True,
                counterparty=self.bank.name,
                initial_book_value=settle_amount,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.ON_DEMAND,
                maturity_date=None,
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination
                ),
                issuance_time=time_point
            )
            receiver.add_asset(new_deposit)

            total_settled += settle_amount
            amount_to_settle -= settle_amount

        if total_settled > 0:
            iou_entry.initial_book_value -= total_settled
            iou_entry.current_book_value -= total_settled

            if iou_entry.initial_book_value <= 0:
                self.bank.remove_liability(iou_entry)
            return total_settled
        else:
            return 0.0

    def _evaluate_rollover_proposal(self, iou_entry, receiver_bank):
        """
        Evaluate whether to allow rollovers (rollovers of day-ahead IOUs) based on the recipient bank's reserve deposits 
        and total deposits, using the Reserve Requirement Ratio (RRR).

        The logic is as follows:
        1. calculate the total reserve deposits (RD) held by the recipient bank.
        2. calculate the total bank deposits (BD) held by the recipient bank (excluding reserves). 3. calculate the excess reserves (SR).
        3. calculate excess reserves (SR) as SR = RD - (RRR * BD).
           - If SR is negative, it indicates a reserve deficiency.
        4. If the reserve is insufficient (SR < 0):
           - If the IOU amount is less than or equal to the underfunded amount, deny the rollover.
           - Otherwise, rollover is authorized, but only up to the amount in excess of the underfunded amount.
        5. if no reserve is insufficient (SR >= 0), grant the rollover in full.

        Returns: 
        bool: True if rollover is approved, False otherwise.
        """
        RRR = 0.10  

        reserve_deposits = sum(
            asset.initial_book_value for asset in receiver_bank.assets
            if asset.type == EntryType.DEPOSIT and asset.denomination == "reserves"
        )
        bank_deposits = sum(
            asset.initial_book_value for asset in receiver_bank.assets
            if asset.type == EntryType.DEPOSIT and asset.denomination != "reserves"
        )

        surplus_reserve = reserve_deposits - (RRR * bank_deposits)

        if surplus_reserve < 0:
            deficit_reserve = abs(surplus_reserve)
            if iou_entry.initial_book_value <= deficit_reserve:
                return False
            else:
                return True
        else:
            return True

    def _handle_rollover(self, iou_entry, time_point):
        """Process IOU extensions: if accepted, create new IOUs with extended expiration dates; if rejected, process defaults"""
        receiver = self._find_iou_counterparty(iou_entry)
        if receiver and self._evaluate_rollover_proposal(iou_entry, receiver):
            new_iou = BalanceSheetEntry(
                type=EntryType.INTRADAY_IOU,
                is_asset=False,
                counterparty=iou_entry.counterparty,
                initial_book_value=iou_entry.initial_book_value,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.FIXED_DATE,
                maturity_date=time_point + 1,  # extend to the next time point
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination,
                    is_intraday=True
                ),
                issuance_time=time_point,
                rollover_count=iou_entry.rollover_count + 1,
                current_book_value=iou_entry.initial_book_value
            )

            new_asset_for_receiver = BalanceSheetEntry(
                type=EntryType.INTRADAY_IOU,
                is_asset=True,
                counterparty=self.bank.name,
                initial_book_value=iou_entry.initial_book_value,
                denomination=iou_entry.denomination,
                maturity_type=MaturityType.FIXED_DATE,
                maturity_date=time_point + 1,
                settlement_details=SettlementDetails(
                    type=SettlementType.MEANS_OF_PAYMENT,
                    denomination=iou_entry.denomination,
                    is_intraday=True
                ),
                issuance_time=time_point,
                rollover_count=iou_entry.rollover_count + 1,
                current_book_value=iou_entry.initial_book_value
            )

            self.bank.remove_liability(iou_entry)
            receiver.remove_asset(next(
                asset for asset in receiver.assets
                if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
                and asset.initial_book_value == iou_entry.initial_book_value
            ))

            self.bank.add_liability(new_iou)
            receiver.add_asset(new_asset_for_receiver)

            self.settlement_log.append(f"Extend maturity IOU: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
            return True
        else:
            self._handle_default(iou_entry, time_point)
            return False

    def _convert_to_overnight(self, iou_entry, time_point):
        """Convert day IOUs to overnight loans."""
        receiver = self._find_iou_counterparty(iou_entry)
        
        overnight_loan = BalanceSheetEntry(
            type=EntryType.OVERNIGHT_LOAN,
            is_asset=True,
            counterparty=self.bank.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=time_point + 1,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        overnight_liability = BalanceSheetEntry(
            type=EntryType.OVERNIGHT_LOAN,
            is_asset=False,
            counterparty=receiver.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.FIXED_DATE,
            maturity_date=time_point + 1,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        self.bank.remove_liability(iou_entry)
        receiver.remove_asset(next(
            asset for asset in receiver.assets
            if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
            and asset.initial_book_value == iou_entry.initial_book_value
        ))
        
        receiver.add_asset(overnight_loan)
        self.bank.add_liability(overnight_liability)
        
        self.settlement_log.append(f"Converting IOUs to Overnight Loans: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
        return True

    def _handle_default(self, iou_entry, time_point):
        """处理IOU违约"""
        receiver = self._find_iou_counterparty(iou_entry)
        
        default_asset = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=self.bank.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            name=f"default {self.bank.name}",
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=receiver.name,
            initial_book_value=iou_entry.initial_book_value,
            denomination=iou_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=iou_entry.denomination
            ),
            name=f"default {receiver.name}",
            issuance_time=time_point,
            current_book_value=iou_entry.initial_book_value
        )
        
        self.bank.remove_liability(iou_entry)
        receiver.remove_asset(next(
            asset for asset in receiver.assets
            if asset.counterparty == self.bank.name and asset.type == EntryType.INTRADAY_IOU
            and asset.initial_book_value == iou_entry.initial_book_value
        ))
        
        receiver.add_asset(default_asset)
        self.bank.add_liability(default_liability)
        
        self.settlement_log.append(f"IOU default: {iou_entry.initial_book_value} {iou_entry.denomination} from {self.bank.name} to {receiver.name}")
        return True

    def process_intraday_settlements(self, time_point):
        """Handles settlement, rollover or default of all daytime IOUs."""
        # search for all IOUs that are due at the current time point
        due_ious = [
            liability for liability in self.bank.liabilities
            if liability.type == EntryType.INTRADAY_IOU
            and liability.maturity_date == time_point
        ]
        
        for iou in due_ious:
            settled_amount = self._settle_partial(iou, time_point)
            
            if iou in self.bank.liabilities: 
                if iou.rollover_count < 2: 
                    self._handle_rollover(iou, time_point)
                else:
                    self._convert_to_overnight(iou, time_point)

class Agent:
    def __init__(self, name: str, agent_type: AgentType):
        self.name = name
        self.type = agent_type
        self.assets: List[BalanceSheetEntry] = []
        self.liabilities: List[BalanceSheetEntry] = []
        self.status: str = "operating"
        self.creation_time: datetime = datetime.now()
        self.settlement_history = {
            'as_asset_holder': [],
            'as_liability_holder': [] 
        }
        # 银行特定属性
        if self.type == AgentType.BANK:
            self.intraday_module = BankIntradayModule(self)
            self.system = None 

        # 新增的交易相关属性
        self.inventory_bond = 0
        self.inventory_cash = 0
        self.trade_count = 0
        self.total_volume = 0
        self.trade_volumes = []
        self.trade_prices = []
        self.reserves: float = 0.0
        self.customer_deposits: Dict[str, float] = {}

    def add_asset(self, entry: BalanceSheetEntry):
        self.assets.append(entry)

    def add_liability(self, entry: BalanceSheetEntry):
        self.liabilities.append(entry)

    def remove_asset(self, entry: BalanceSheetEntry):
        self.assets = [e for e in self.assets if not e.matches(entry)]

    def remove_liability(self, entry: BalanceSheetEntry):
        self.liabilities = [e for e in self.liabilities if not e.matches(entry)]

    def get_balance_sheet(self) -> Dict:
        return {
            "assets": self.assets,
            "liabilities": self.liabilities
        }

    def get_total_assets(self) -> float:
        return sum(entry.current_book_value for entry in self.assets)

    def get_total_liabilities(self) -> float:
        return sum(entry.current_book_value for entry in self.liabilities)

    def get_net_worth(self) -> float:
        total_assets = sum(entry.current_book_value for entry in self.assets 
                          if entry.type != EntryType.SHARE)
        total_liabilities = sum(entry.current_book_value for entry in self.liabilities 
                               if entry.type != EntryType.SHARE)
        return total_assets - total_liabilities

    def get_type_specific_metrics(self) -> Dict:
        metrics = {
            "name": self.name,
            "type": self.type.value,
            "creation_time": self.creation_time,
            "status": self.status,
            "total_assets": self.get_total_assets(),
            "total_liabilities": self.get_total_liabilities(),
            "net_worth": self.get_net_worth()
        }

        if self.type == AgentType.BANK:
            metrics["capital_ratio"] = self.get_total_assets() / self.get_total_liabilities() if self.get_total_liabilities() > 0 else float('inf')
        elif self.type == AgentType.COMPANY:
            metrics["leverage_ratio"] = self.get_total_liabilities() / self.get_total_assets() if self.get_total_assets() > 0 else float('inf')
        elif self.type == AgentType.HOUSEHOLD:
            metrics["savings_rate"] = (self.get_total_assets() - self.get_total_liabilities()) / self.get_total_assets() if self.get_total_assets() > 0 else 0

        return metrics

    def record_settlement(self,
                         time_point: int,
                         original_entry: BalanceSheetEntry,
                         settlement_result: BalanceSheetEntry,
                         counterparty: str,
                         as_asset_holder: bool):
        settlement_record = {
            'time_point': time_point,  # 加t前缀显示
            'original_entry': deepcopy(original_entry),
            'settlement_result': deepcopy(settlement_result),
            'counterparty': counterparty,
            'timestamp': datetime.now()
        }
        if as_asset_holder:
            self.settlement_history['as_asset_holder'].append(settlement_record)
        else:
            self.settlement_history['as_liability_holder'].append(settlement_record)

    def update_inventory(self, bond_delta, cash_delta):
        self.inventory_bond += bond_delta
        self.inventory_cash += cash_delta
        
        if bond_delta != 0: 
            self.trade_count += 1
            self.total_volume += abs(bond_delta)
            self.trade_volumes.append(abs(bond_delta))
            self.trade_prices.append(-cash_delta / bond_delta)

    def to_dict(self):
        """用于Web UI的序列化 - 统一使用数字"""
        return {
            "name": self.name,
            "type": self.type.value,
            "assets": [a.to_dict() for a in self.assets],
            "liabilities": [l.to_dict() for l in self.liabilities],
            "settlement_history": {
                'as_asset_holder': [
                    {
                        'time_point': r['time_point'],  # 保持数字格式
                        'original_entry': r['original_entry'].to_dict(),
                        'settlement_result': r['settlement_result'].to_dict(),
                        'counterparty': r['counterparty'],
                        'timestamp': r['timestamp'].isoformat()
                    } for r in self.settlement_history['as_asset_holder']
                ],
                'as_liability_holder': [
                    {
                        'time_point': r['time_point'],  # 保持数字格式
                        'original_entry': r['original_entry'].to_dict(),
                        'settlement_result': r['settlement_result'].to_dict(),
                        'counterparty': r['counterparty'],
                        'timestamp': r['timestamp'].isoformat()
                    } for r in self.settlement_history['as_liability_holder']
                ]
            },
            "creation_time": self.creation_time.isoformat()
        }

from typing import Tuple, Optional, List

# Assuming the Enums (AgentType, EntryType, MaturityType, SettlementType, BondType)
# and dataclasses (SettlementDetails, BalanceSheetEntry)
# and Agent class are defined as in your provided 'economic_sim_web_integrated.py'

class AssetLiabilityPair:
    def __init__(self,
                 time: int,
                 type: str,
                 amount: float,
                 denomination: str,
                 maturity_type: MaturityType,
                 maturity_date: Optional[int],
                 settlement_type: SettlementType,
                 settlement_denomination: str,
                 asset_holder: Agent,
                 liability_holder: Optional[Agent] = None,
                 cash_flow_at_maturity: Optional[float] = 0,
                 asset_name: Optional[str] = None,
                 bond_type: Optional[BondType] = None,
                 coupon_rate: Optional[float] = None,
                 scheduled_time_point: Optional[int] = None):  # NEW: For scheduled actions
        self.time = time
        self.type = type
        self.amount = amount
        self.denomination = denomination
        self.maturity_type = MaturityType(maturity_type) if isinstance(maturity_type, str) else maturity_type
        self.maturity_date = maturity_date
        
        # Ensure cash_flow_at_maturity is at least the amount if not specified otherwise for non-interest/non-discount items
        self.cash_flow_at_maturity = cash_flow_at_maturity if cash_flow_at_maturity is not None else amount

        self.settlement_details = SettlementDetails(
            type=SettlementType(settlement_type) if isinstance(settlement_type, str) else settlement_type,
            denomination=settlement_denomination
        )
        self.asset_holder = asset_holder
        self.liability_holder = liability_holder
        self.asset_name = asset_name
        self.bond_type = BondType(bond_type) if isinstance(bond_type, (str, int)) and bond_type is not None else bond_type # Handle int from HTML form
        self.coupon_rate = coupon_rate
        self.connected_claims: List[BalanceSheetEntry] = []
        self.current_time_state = 0 # This should be set by EconomicSystem
        self.scheduled_time_point = scheduled_time_point  # NEW: For scheduled actions

        # Validations from original __init__
        entry_type_enum = EntryType(self.type)
        if entry_type_enum == EntryType.NON_FINANCIAL:
            if liability_holder is not None:
                raise ValueError("Non-financial entries cannot have liability holders")
            if not asset_name:
                raise ValueError("Non-financial entries must have the name of the asset")
        else: # Financial entries
            if liability_holder is None:
                raise ValueError(f"Financial entry type '{self.type}' must have a liability holder.")
            if asset_holder.name == liability_holder.name:
                raise ValueError("Asset holder and liability holder cannot be the same for financial entries.")

        if entry_type_enum == EntryType.SHARE:
            self.maturity_type = MaturityType.PERPETUAL
            self.maturity_date = None
            self.settlement_details = SettlementDetails(type=SettlementType.NONE, denomination="shares")


    def _calculate_expected_cash_flow(self) -> float:
        entry_type_enum = EntryType(self.type)
        if entry_type_enum not in [EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
            return 1.0 # Or perhaps self.amount if it's meant to be absolute

        if self.amount == 0: return 1.0 # Avoid division by zero

        if entry_type_enum == EntryType.BOND_ZERO_COUPON:
            # Assuming cash_flow_at_maturity holds the face value for a ZCB bought at 'amount'
            # If 'amount' is face value, and it's issued at discount, this is different.
            # Based on _adjust_value, cash_flow_at_maturity is likely the redemption value.
            # If expected_cash_flow is a ratio to initial_book_value for valuation:
            return self.cash_flow_at_maturity / self.amount if self.cash_flow_at_maturity is not None else 1.0


        elif entry_type_enum == EntryType.BOND_COUPON:
            if not self.coupon_rate:
                raise ValueError("Coupon bonds require a coupon rate")
            
            total_coupons = 0
            # Simplified: assumes 1 or 2 periods. A real calculation needs all coupon dates.
            if self.maturity_date == 1: # Matures in 1 period
                total_coupons = self.amount * self.coupon_rate
            elif self.maturity_date == 2: # Matures in 2 periods
                total_coupons = 2 * (self.amount * self.coupon_rate) # Coupon each period
            else: # For t0 issuance and other maturity dates, needs more robust period calculation
                # Assuming maturity_date is number of periods for simplicity here.
                if self.maturity_date and self.maturity_date > 0:
                    total_coupons = self.maturity_date * (self.amount * self.coupon_rate)


            return (self.amount + total_coupons) / self.amount # Total received / initial investment

        elif entry_type_enum == EntryType.BOND_AMORTIZING:
            if not self.coupon_rate:
                raise ValueError("Amortizing bonds require an interest rate")
            
            # This is a simplified placeholder. Real amortization sum of payments is complex.
            # Let's assume total cash received, similar to coupon bond for this ratio.
            # For a simple 2-period amortization:
            if self.maturity_date == 1:
                return (self.amount * (1 + self.coupon_rate)) / self.amount # All paid in one go
            elif self.maturity_date == 2:
                # Simplified: P1=Amt/2 + Int_Full, P2=Amt/2 + Int_Half
                interest_t1 = self.amount * self.coupon_rate
                principal_t1 = self.amount / 2 # Simplified
                payment_t1 = principal_t1 + interest_t1

                remaining_principal_t2 = self.amount - principal_t1
                interest_t2 = remaining_principal_t2 * self.coupon_rate
                principal_t2 = remaining_principal_t2
                payment_t2 = principal_t2 + interest_t2
                return (payment_t1 + payment_t2) / self.amount
            else: # Default for 0 or other periods
                return (self.amount * (1 + self.coupon_rate * (self.maturity_date if self.maturity_date else 1))) / self.amount


        return 1.0 # Default


    def _create_bond_payment_schedule(self) -> List[Tuple[int, float, str]]:
        schedule = []
        entry_type_enum = EntryType(self.type)

        if entry_type_enum == EntryType.BOND_COUPON:
            if not self.coupon_rate:
                raise ValueError("Coupon bonds require a coupon rate")
            if not self.maturity_date or self.maturity_date <= 0:
                raise ValueError("Coupon bonds require a positive maturity date (number of periods)")

            coupon_amount = self.amount * self.coupon_rate
            for period in range(1, self.maturity_date): # Coupons before maturity
                schedule.append((self.issuance_time + period, coupon_amount, "Coupon"))
            
            final_payment = coupon_amount + self.amount # Final coupon + principal
            schedule.append((self.issuance_time + self.maturity_date, final_payment, "Coupon+Principal"))

        elif entry_type_enum == EntryType.BOND_AMORTIZING:
            if not self.coupon_rate:
                raise ValueError("Amortizing bonds require an interest rate")
            if not self.maturity_date or self.maturity_date <= 0:
                raise ValueError("Amortizing bonds require a positive maturity date (number of periods)")

            # Simplified equal principal payments + interest on outstanding
            # More complex amortization (e.g., equal total payments) would require a different formula
            principal_payment_part = self.amount / self.maturity_date
            outstanding_principal = self.amount
            for period in range(1, self.maturity_date + 1):
                interest_payment = outstanding_principal * self.coupon_rate
                total_payment = principal_payment_part + interest_payment
                schedule.append((self.issuance_time + period, total_payment, "Principal+Interest"))
                outstanding_principal -= principal_payment_part
        return schedule

    def create_bond_claims(self) -> List[BalanceSheetEntry]:
        claims = []
        entry_type_enum = EntryType(self.type)
        if entry_type_enum in [EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING]:
            schedule = self._create_bond_payment_schedule()
            # Use a more stable bond identifier if possible, e.g., from a unique ID service or hash
            bond_id = f"bond_{self.asset_holder.name}_{self.liability_holder.name}_{self.amount}_{self.current_time_state}_{id(self)}"

            for date, payment_amount, payment_type_desc in schedule:
                claim = BalanceSheetEntry(
                    type=EntryType.PAYABLE, # From bond issuer's view it's a payable, for holder it's receivable
                    is_asset=True, # This is the asset for the bond_holder
                    counterparty=self.liability_holder.name, # Owed by the bond issuer
                    initial_book_value=payment_amount,
                    denomination=self.denomination,
                    maturity_type=MaturityType.FIXED_DATE,
                    maturity_date=date,
                    settlement_details=SettlementDetails(
                        type=SettlementType.MEANS_OF_PAYMENT, # Typically settled in cash
                        denomination=self.denomination
                    ),
                    name=f"{payment_type_desc} for {bond_id}",
                    issuance_time=self.current_time_state, # Claims are established when bond is issued
                    parent_bond=bond_id,
                    current_book_value=payment_amount, # Initial value of this specific claim
                    cash_flow_at_maturity=payment_amount # This claim pays its face value
                )
                claims.append(claim)
        return claims

    def create_entries(self) -> Tuple[Optional[BalanceSheetEntry], Optional[BalanceSheetEntry]]:
        def find_existing_entry(entries: List[BalanceSheetEntry], is_asset: bool) -> Optional[BalanceSheetEntry]:
            for entry in entries:
                if (
                    entry.type == EntryType(self.type) and
                    entry.is_asset == is_asset and
                    entry.counterparty == (self.liability_holder.name if is_asset else self.asset_holder.name) and
                    entry.denomination == self.denomination and
                    entry.maturity_type == self.maturity_type and
                    entry.maturity_date == self.maturity_date
                ):
                    return entry
            return None
        
        asset_entry: Optional[BalanceSheetEntry] = None
        liability_entry: Optional[BalanceSheetEntry] = None
        
        entry_type_enum = EntryType(self.type)
        current_time = self.current_time_state

        # --- Validation specific to certain types before creation ---
        if entry_type_enum == EntryType.LOAN:
            if self.asset_holder.type != AgentType.BANK:
                raise ValueError("Only banks (typically) hold loans as assets in this model.")

        # --- Non-Financial Entry ---
        if entry_type_enum == EntryType.NON_FINANCIAL:
            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=None,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name,
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.amount
            )
            liability_entry = None

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            return asset_entry, liability_entry

        # --- Delivery Claim ---
        elif entry_type_enum == EntryType.DELIVERY_CLAIM:
            if not self.asset_name:
                raise ValueError("Delivery claims must specify the asset to be delivered.")
            if not self.liability_holder:
                raise ValueError("Delivery claims must have a liability holder (promisor).")

            actual_settlement_details = SettlementDetails(
                type=SettlementType.NON_FINANCIAL_ASSET,
                denomination=self.settlement_details.denomination
            )
            
            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=actual_settlement_details,
                name=self.asset_name,
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.amount
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=actual_settlement_details,
                name=self.asset_name,
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.amount
            )

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

        # --- Share ---
        elif entry_type_enum == EntryType.SHARE:
            if not self.liability_holder:
                raise ValueError("Shares must have a liability holder (the issuer).")
                
            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=MaturityType.PERPETUAL,
                maturity_date=None,
                settlement_details=SettlementDetails(type=SettlementType.NONE, denomination="shares"),
                name=f"Shares in {self.liability_holder.name}",
                issuance_time=current_time,
                current_book_value=self.amount
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=MaturityType.PERPETUAL,
                maturity_date=None,
                settlement_details=SettlementDetails(type=SettlementType.NONE, denomination="shares"),
                name="Share Capital",
                issuance_time=current_time,
                current_book_value=self.amount
            )

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

        # --- All Bond Types (including new government and central bank bonds) ---
        elif entry_type_enum in [
            EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING,
            EntryType.GOVERNMENT_BOND_ZERO_COUPON, EntryType.GOVERNMENT_BOND_COUPON, EntryType.GOVERNMENT_BOND_AMORTIZING,
            EntryType.CENTRAL_BANK_BOND_ZERO_COUPON, EntryType.CENTRAL_BANK_BOND_COUPON, EntryType.CENTRAL_BANK_BOND_AMORTIZING
        ]:
            if not self.liability_holder:
                raise ValueError("Bonds must have a liability holder (the issuer).")

            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name if self.asset_name else f"{entry_type_enum.value} Bond",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity,
                expected_cash_flow=self._calculate_expected_cash_flow()
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=f"{entry_type_enum.value} Bond Issued",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity,
                expected_cash_flow=self._calculate_expected_cash_flow()
            )
            
            # Create and link subsidiary claims for coupon and amortizing bonds
            if entry_type_enum in [
                EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING,
                EntryType.GOVERNMENT_BOND_COUPON, EntryType.GOVERNMENT_BOND_AMORTIZING,
                EntryType.CENTRAL_BANK_BOND_COUPON, EntryType.CENTRAL_BANK_BOND_AMORTIZING
            ]:
                self.connected_claims = self.create_bond_claims()
                
                for asset_claim_for_bond_holder in self.connected_claims:
                    self.asset_holder.add_asset(asset_claim_for_bond_holder)
                    
                    liability_for_issuer_for_claim = BalanceSheetEntry(
                        type=EntryType.PAYABLE,
                        is_asset=False,
                        counterparty=self.asset_holder.name,
                        initial_book_value=asset_claim_for_bond_holder.initial_book_value,
                        denomination=asset_claim_for_bond_holder.denomination,
                        maturity_type=asset_claim_for_bond_holder.maturity_type,
                        maturity_date=asset_claim_for_bond_holder.maturity_date,
                        settlement_details=asset_claim_for_bond_holder.settlement_details,
                        name=asset_claim_for_bond_holder.name,
                        issuance_time=asset_claim_for_bond_holder.issuance_time,
                        parent_bond=asset_claim_for_bond_holder.parent_bond,
                        current_book_value=asset_claim_for_bond_holder.current_book_value,
                        cash_flow_at_maturity=asset_claim_for_bond_holder.cash_flow_at_maturity
                    )
                    self.liability_holder.add_liability(liability_for_issuer_for_claim)

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

        # --- Payable ---
        elif entry_type_enum == EntryType.PAYABLE:
            if not self.liability_holder:
                raise ValueError("Payables must have a liability holder (the debtor).")
            
            actual_settlement_details = SettlementDetails(
                type=SettlementType.MEANS_OF_PAYMENT,
                denomination=self.settlement_details.denomination
            )
            
            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=actual_settlement_details,
                name=self.asset_name if self.asset_name else "Account Receivable",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=actual_settlement_details,
                name=self.asset_name if self.asset_name else "Account Payable",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

        # --- Reserve Deposit (NEW) ---
        elif entry_type_enum == EntryType.RESERVE_DEPOSIT:
            if not self.liability_holder:
                raise ValueError("Reserve deposits must have a liability holder (central bank).")
            
            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name="Reserve Deposit",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name="Reserve Deposit Liability",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

        # --- Other Financial Entries (LOAN, DEPOSIT, INTRADAY_IOU, OVERNIGHT_LOAN, etc.) ---
        else:
            if not self.liability_holder:
                raise ValueError(f"Financial entry type '{self.type}' must have a liability holder.")

            asset_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=True,
                counterparty=self.liability_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name if self.asset_name else f"{entry_type_enum.value} held",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )
            
            liability_entry = BalanceSheetEntry(
                type=entry_type_enum,
                is_asset=False,
                counterparty=self.asset_holder.name,
                initial_book_value=self.amount,
                denomination=self.denomination,
                maturity_type=self.maturity_type,
                maturity_date=self.maturity_date,
                settlement_details=self.settlement_details,
                name=self.asset_name if self.asset_name else f"{entry_type_enum.value} owed",
                issuance_time=current_time,
                current_book_value=self.amount,
                cash_flow_at_maturity=self.cash_flow_at_maturity
            )

            # Check for existing entries to merge
            existing_asset_entry = find_existing_entry(self.asset_holder.assets, is_asset=True)
            if existing_asset_entry:
                existing_asset_entry.initial_book_value += self.amount
                existing_asset_entry.current_book_value += self.amount
                asset_entry = None

            existing_liability_entry = find_existing_entry(self.liability_holder.liabilities, is_asset=False)
            if existing_liability_entry:
                existing_liability_entry.initial_book_value += self.amount
                existing_liability_entry.current_book_value += self.amount
                liability_entry = None

            return asset_entry, liability_entry

    def to_dict(self):
        """For Web UI serialization, if needed."""
        return {
            "time": self.time,
            "type": self.type, # Keep as string value
            "amount": self.amount,
            "denomination": self.denomination,
            "maturity_type": self.maturity_type.value,
            "maturity_date": self.maturity_date,
            "settlement_type": self.settlement_details.type.value,
            "settlement_denomination": self.settlement_details.denomination,
            "asset_holder": self.asset_holder.name,
            "liability_holder": self.liability_holder.name if self.liability_holder else None,
            "asset_name": self.asset_name,
            "bond_type": self.bond_type.value if self.bond_type else None,
            "coupon_rate": self.coupon_rate,
            "cash_flow_at_maturity": self.cash_flow_at_maturity,
            "current_time_state": self.current_time_state,
            "scheduled_time_point": self.scheduled_time_point  # NEW: Include scheduled time
        }

class EconomicSystem:
    def __init__(self):
        self.agents: Dict[str, Agent] = {} 
        self.asset_liability_pairs: List[AssetLiabilityPair] = []
        self.time_states: Dict[int, Dict[str, Agent]] = {} 
        self.current_time_state = 0  
        self.simulation_finalized = False 
        self.scheduled_actions = {} 
        self.money_system={}
        self.time_series_tracker={}
        self.flux_and_reflux={}
        self.means_of_payment={}
        #debt system
        self.debt_map={}
        self.debt_stock={}
        self.debt_flows={}
        self.maturity_profile={'t+1':{},'t+n':{}}
        self.maturity={}
        self.debt_position={}
        self.debt_ratio=[{},{},{}]
        self.default_tracker={}
        self.save_state(0)

    def validate_time_point(self, time_point: int, allow_t0: bool = True) -> None:
        valid_points = [0, 1, 2] if allow_t0 else [1, 2]
        if time_point not in valid_points:
            valid_points_str = ", ".join(map(str, valid_points))
            raise ValueError(f"Timepoint must be {valid_points_str}")

    def add_agent(self, agent: Agent):
        self.agents[agent.name] = agent
        if agent.type == AgentType.BANK:
            agent.system = self
            agent.intraday_module.system = self

        if self.current_time_state == 0:
            self.save_state(0)


    def create_asset_liability_pair(self, pair: AssetLiabilityPair):
        # Validate constraints before creating the pair
        entry_type = EntryType(pair.type)
        is_valid, error_msg = AssetLiabilityConstraints.validate_asset_liability_pair(
            pair.asset_holder, 
            pair.liability_holder, 
            entry_type
        )
        
        if not is_valid:
            raise ValueError(f"Constraint violation: {error_msg}")

        pair.current_time_state = self.current_time_state
        self.asset_liability_pairs.append(pair)
        asset_entry, liability_entry = pair.create_entries()
        if asset_entry:  # 检查是否为None（合并情况）
            pair.asset_holder.add_asset(asset_entry)
        if liability_entry:
            pair.liability_holder.add_liability(liability_entry)

        self.save_state(self.current_time_state)

    def get_time_points(self) -> List[int]:
        return [0, 1, 2]

    def save_state(self, time_point: int):
        self.validate_time_point(time_point)
        self.time_states[time_point] = {}
        for name, agent in self.agents.items():
            agent_copy = Agent(agent.name, agent.type)
            agent_copy.assets = deepcopy(agent.assets)
            agent_copy.liabilities = deepcopy(agent.liabilities)
            agent_copy.settlement_history = deepcopy(agent.settlement_history)
            agent_copy.status = agent.status
            agent_copy.creation_time = agent.creation_time
            self.time_states[time_point][name] = agent_copy

        self.current_time_state = time_point

    # NEW: Enhanced scheduled actions management
    def schedule_asset_liability_creation(self, time_point: int, pair_data: Dict[str, Any]):
        """Schedule an asset-liability pair creation for a future time point"""
        if time_point not in self.scheduled_actions:
            self.scheduled_actions[time_point] = []
        
        self.scheduled_actions[time_point].append({
            'type': 'create_asset_liability_pair',
            'data': pair_data
        })

    def execute_scheduled_actions(self, time_point: int):
        """Execute all scheduled actions for a given time point"""
        if time_point not in self.scheduled_actions:
            return
        
        actions = self.scheduled_actions[time_point]
        for action in actions:
            if action['type'] == 'create_asset_liability_pair':
                try:
                    # Recreate the asset-liability pair from stored data
                    data = action['data']
                    
                    # Validate that agents still exist
                    asset_holder = self.agents.get(data['asset_holder_name'])
                    liability_holder = self.agents.get(data['liability_holder_name']) if data.get('liability_holder_name') else None
                    
                    if not asset_holder:
                        print(f"Warning: Asset holder {data['asset_holder_name']} no longer exists, skipping scheduled action")
                        continue
                    
                    if data.get('liability_holder_name') and not liability_holder:
                        print(f"Warning: Liability holder {data['liability_holder_name']} no longer exists, skipping scheduled action")
                        continue
                    
                    # Create the pair
                    pair = AssetLiabilityPair(
                        time=time_point,
                        type=data['type'],
                        amount=data['amount'],
                        denomination=data['denomination'],
                        maturity_type=data['maturity_type'],
                        maturity_date=data['maturity_date'],
                        settlement_type=data['settlement_type'],
                        settlement_denomination=data['settlement_denomination'],
                        asset_holder=asset_holder,
                        liability_holder=liability_holder,
                        cash_flow_at_maturity=data.get('cash_flow_at_maturity', 0),
                        asset_name=data.get('asset_name'),
                        bond_type=data.get('bond_type'),
                        coupon_rate=data.get('coupon_rate')
                    )
                    
                    self.create_asset_liability_pair(pair)
                    print(f"Executed scheduled action: Created {data['type']} pair at t{time_point}")
                    
                except Exception as e:
                    print(f"Error executing scheduled action: {e}")
        
        # Clear executed actions
        del self.scheduled_actions[time_point]

    # CRITICAL FIX: Enhanced settle_entries method with proper interest handling
    def settle_entries(self, time_point: str):
        """Settle entries at a specific time point"""
        self.validate_time_point(time_point, allow_t0=False)

        # First save the current state at the previous time point
        prev_time = 't0' if time_point == 't1' else 't1'
        if prev_time not in self.time_states:
            self.save_state(prev_time)

        # Process all entries that are due at this time point
        for pair in self.asset_liability_pairs[:]:  # Create a copy to iterate over
            if (pair.maturity_type == MaturityType.FIXED_DATE and
                pair.maturity_date):
                # Check if the entry's maturity date matches our time state
                entry_time = 't1'
                if pair.maturity_date.year == 2100:  # t2
                    entry_time = 't2'

                if entry_time == time_point:
                    # Remove the original pair
                    self.asset_liability_pairs.remove(pair)
                    asset_entry, liability_entry = pair.create_entries()
                    pair.asset_holder.remove_asset(asset_entry)
                    if liability_entry:
                        pair.liability_holder.remove_liability(liability_entry)

                    # Handle settlement based on type
                    if pair.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
                        # Find the bank deposit that will be used for settlement
                        debtor_deposit = next(
                            (asset for asset in pair.liability_holder.assets
                             if asset.type == EntryType.DEPOSIT
                             and asset.current_book_value >= asset_entry.current_book_value  # CRITICAL: Use current_book_value
                             and asset.denomination == pair.denomination),
                            None
                        )
                        self.money_system[pair.asset_holder]=self.money_system.get(pair.asset_holder,[])+[f'{time_point}','asset',pair.amount,'+']

                        if not debtor_deposit:
                            # CRITICAL: Check using current_book_value instead of amount
                            total_available = sum(asset.current_book_value for asset in pair.liability_holder.assets
                                                if asset.type == EntryType.DEPOSIT and asset.denomination == pair.denomination)
                            raise ValueError(f"Insufficient deposits for settlement. Need {asset_entry.current_book_value}, have {total_available}")

                        # Get the bank that holds the deposit liability
                        bank = next(a for a in self.agents.values() if a.name == debtor_deposit.counterparty)
                        self.money_system[bank]=self.money_system.get(bank,[])+[f'{time_point}','liability',pair.amount,f'change of liability holder to {pair.asset_holder}']
                        # CRITICAL: Use current_book_value or cash_flow_at_maturity for settlement amount
                        settlement_amount = asset_entry.current_book_value
                        # Remove the original deposit from the debtor
                        pair.liability_holder.remove_asset(debtor_deposit)

                        # Remove the corresponding liability from the bank
                        bank_liability = next(
                            (l for l in bank.liabilities
                             if l.type == EntryType.DEPOSIT
                             and l.counterparty == pair.liability_holder.name
                             and l.current_book_value == debtor_deposit.current_book_value),
                            None
                        )
                        if bank_liability:
                            bank.remove_liability(bank_liability)
                        # CRITICAL: Check if receiver is the bank itself (special case)
                        if pair.asset_holder.name == bank.name:
                            # Special case: receiver is the bank, this is internal transfer
                            # Bank receives payment, reducing customer liability and increasing net assets
                            
                            # Record settlement history
                            pair.asset_holder.record_settlement(
                                time_point=time_point,
                                original_entry=asset_entry,
                                settlement_result=debtor_deposit,  # Bank received customer deposit
                                counterparty=pair.liability_holder.name,
                                as_asset_holder=True
                            )
                            pair.liability_holder.record_settlement(
                                time_point=time_point,
                                original_entry=liability_entry,
                                settlement_result=debtor_deposit,
                                counterparty=pair.asset_holder.name,
                                as_asset_holder=False
                            )
                            
                            # Handle remaining deposit if any
                            if debtor_deposit.current_book_value > settlement_amount:
                                remainder_amount = debtor_deposit.current_book_value - settlement_amount
                                
                                remainder_pair = AssetLiabilityPair(
                                    time=time_point,
                                    type=EntryType.DEPOSIT.value,
                                    amount=remainder_amount,
                                    denomination=pair.denomination,
                                    maturity_type=MaturityType.ON_DEMAND,
                                    maturity_date=None,
                                    settlement_type=SettlementType.NONE,
                                    settlement_denomination=pair.denomination,
                                    asset_holder=pair.liability_holder,  # Customer
                                    liability_holder=bank  # Bank
                                )
                                
                                remainder_pair.current_time_state = time_point
                                remainder_asset, remainder_liability = remainder_pair.create_entries()
                                
                                remainder_pair.asset_holder.add_asset(remainder_asset)
                                if remainder_liability:
                                    remainder_pair.liability_holder.add_liability(remainder_liability)
                                self.asset_liability_pairs.append(remainder_pair)
                                
                        else:

                            # Create new deposit entry for the creditor
                            settlement_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.DEPOSIT.value,
                                amount=pair.amount,
                                denomination=pair.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=pair.denomination,
                                asset_holder=pair.asset_holder,  # Creditor gets the deposit
                                liability_holder=bank,  # Bank keeps liability
                                asset_name=None
                            )

                            # Create entries with current time point as issuance time
                            settlement_pair.current_time_state = time_point
                            new_asset_entry, new_liability_entry = settlement_pair.create_entries()
                            new_asset_entry.issuance_time = time_point
                            if new_liability_entry:
                                new_liability_entry.issuance_time = time_point

                            # Record settlement history
                            pair.asset_holder.record_settlement(
                                time_point=time_point,
                                original_entry=asset_entry,
                                settlement_result=new_asset_entry,
                                counterparty=pair.liability_holder.name,
                                as_asset_holder=True
                            )
                            pair.liability_holder.record_settlement(
                                time_point=time_point,
                                original_entry=liability_entry,
                                settlement_result=debtor_deposit,  # The deposit that was used
                                counterparty=pair.asset_holder.name,
                                as_asset_holder=False
                            )

                            # Add entries
                            settlement_pair.asset_holder.add_asset(new_asset_entry)
                            if new_liability_entry:
                                settlement_pair.liability_holder.add_liability(new_liability_entry)
                            self.asset_liability_pairs.append(settlement_pair)
                            # Handle remaining deposit
                            if debtor_deposit.current_book_value > settlement_amount:
                                remainder_amount = debtor_deposit.current_book_value - settlement_amount
                                remainder_pair = AssetLiabilityPair(
                                    time=datetime.now(),
                                    type=EntryType.DEPOSIT.value,
                                    amount=remainder_amount,
                                    denomination=pair.denomination,
                                    maturity_type=MaturityType.ON_DEMAND,
                                    maturity_date=None,
                                    settlement_type=SettlementType.NONE,
                                    settlement_denomination=pair.denomination,
                                    asset_holder=pair.liability_holder,  # Debtor keeps remainder
                                    liability_holder=bank,  # Bank keeps liability
                                    asset_name=None
                                )

                                # Create entries with current time point as issuance time
                                remainder_pair.current_time_state = time_point
                                remainder_asset, remainder_liability = remainder_pair.create_entries()
                                remainder_asset.issuance_time = time_point
                                if remainder_liability:
                                    remainder_liability.issuance_time = time_point

                                # Add entries
                                remainder_pair.asset_holder.add_asset(remainder_asset)
                                if remainder_liability:
                                    remainder_pair.liability_holder.add_liability(remainder_liability)
                                self.asset_liability_pairs.append(remainder_pair)

                    elif pair.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
                        # Find and remove the non-financial asset from the liability holder
                        non_financial_asset = next(
                            (asset for asset in pair.liability_holder.assets
                             if asset.type == EntryType.NON_FINANCIAL
                             and asset.name == pair.asset_name
                             and asset.current_book_value >= pair.amount),
                            None
                        )

                        if not non_financial_asset:
                            raise ValueError(f"Non-financial asset {pair.asset_name} not found for settlement")

                        # Remove the asset from the liability holder
                        pair.liability_holder.remove_asset(non_financial_asset)

                        # Create non-financial asset entry for the asset holder
                        settlement_pair = AssetLiabilityPair(
                            time=datetime.now(),
                            type=EntryType.NON_FINANCIAL.value,
                            amount=pair.amount,
                            denomination=pair.settlement_details.denomination,
                            maturity_type=MaturityType.ON_DEMAND,
                            maturity_date=None,
                            settlement_type=SettlementType.NONE,
                            settlement_denomination=pair.settlement_details.denomination,
                            asset_holder=pair.asset_holder,  # Original creditor gets the goods
                            liability_holder=None,  # Non-financial assets have no liability holder
                            asset_name=pair.asset_name  # Use the asset_name directly
                        )
                        # Create entry with current time point as issuance time
                        settlement_pair.current_time_state = time_point
                        new_asset_entry, _ = settlement_pair.create_entries()
                        new_asset_entry.issuance_time = time_point

                        # Record settlement history
                        pair.asset_holder.record_settlement(
                            time_point=time_point,
                            original_entry=asset_entry,
                            settlement_result=new_asset_entry,
                            counterparty=pair.liability_holder.name,
                            as_asset_holder=True
                        )
                        pair.liability_holder.record_settlement(
                            time_point=time_point,
                            original_entry=liability_entry,
                            settlement_result=non_financial_asset,  # The non-financial asset that was delivered
                            counterparty=pair.asset_holder.name,
                            as_asset_holder=False
                        )

                        # Add entry directly to avoid default t0 issuance time
                        settlement_pair.asset_holder.add_asset(new_asset_entry)
                        self.asset_liability_pairs.append(settlement_pair)

                        # If there was remaining amount in the non-financial asset, create a new entry for it
                        if non_financial_asset.current_book_value > pair.amount:
                            remainder_pair = AssetLiabilityPair(
                                time=datetime.now(),
                                type=EntryType.NON_FINANCIAL.value,
                                amount=non_financial_asset.amount - pair.amount,
                                denomination=non_financial_asset.denomination,
                                maturity_type=MaturityType.ON_DEMAND,
                                maturity_date=None,
                                settlement_type=SettlementType.NONE,
                                settlement_denomination=non_financial_asset.denomination,
                                asset_holder=pair.liability_holder,  # Original holder keeps remainder
                                liability_holder=None,
                                asset_name=non_financial_asset.name
                            )

                            # Create entry with current time point as issuance time
                            remainder_pair.current_time_state = time_point
                            remainder_asset, _ = remainder_pair.create_entries()
                            remainder_asset.issuance_time = time_point

                            # Add entry
                            remainder_pair.asset_holder.add_asset(remainder_asset)
                            self.asset_liability_pairs.append(remainder_pair)
        # Handle interbank intraday liquidity
        for agent in self.agents.values():
            if agent.type == AgentType.BANK:
                agent.intraday_module.process_intraday_settlements(time_point)
                            
        # Auto-save state after settlements
        self.save_state(time_point)
        self.current_time_state = time_point

    # CRITICAL FIX: Implement adj_book_values method
    def adj_book_values(self):
        """Adjust book values for all entries based on time progression and cash flows"""
        for agent in self.agents.values():
            for liability in agent.liabilities:
                if liability.maturity_type == MaturityType.FIXED_DATE and liability.maturity_date is not None:
                    self._adjust_value(liability)

            for asset in agent.assets:
                if asset.maturity_type == MaturityType.FIXED_DATE and asset.maturity_date is not None:
                    self._adjust_value(asset)

    def _adjust_value(self, bal_entry: BalanceSheetEntry):
        """Adjust individual balance sheet entry value based on time and cash flows"""
        t = self.current_time_state - bal_entry.issuance_time
        if bal_entry.maturity_date == bal_entry.issuance_time:  
            bal_entry.current_book_value = bal_entry.initial_book_value
            return
            
        m = bal_entry.maturity_date - bal_entry.issuance_time
        if m == 0: 
            bal_entry.current_book_value = bal_entry.initial_book_value
            return
        
        # Calculate rate of adjustment based on cash flow at maturity
        if bal_entry.cash_flow_at_maturity and bal_entry.initial_book_value > 0:
            rate_of_adjustment = (bal_entry.cash_flow_at_maturity / bal_entry.initial_book_value)**(1/m) - 1
            bal_entry.current_book_value = bal_entry.initial_book_value * (1 + rate_of_adjustment)**t
        else:
            # If no cash flow at maturity specified, no adjustment
            bal_entry.current_book_value = bal_entry.initial_book_value

    def get_agents_at_time(self, time_point: int) -> Dict[str, Agent]:
        self.validate_time_point(time_point)

        if time_point in self.time_states:
            return self.time_states[time_point]

        if time_point > self.current_time_state:
            current_state = deepcopy(self.agents)
            current_time = self.current_time_state

            try:
                for t in range(self.current_time_state + 1, time_point + 1):
                    self.settle_entries(t)

                result = {name: agent for name, agent in self.agents.items()}
                    
                self.agents = current_state
                self.current_time_state = current_time

                return result
            except Exception as e:
                self.agents = current_state
                self.current_time_state = current_time
                print(f"\nWarning: Unable to settle ({str(e)})")
                return current_state

        return {name: agent for name, agent in self.agents.items()}

    def compute_changes(self, from_time: str, to_time: str) -> Dict[str, Dict[str, List]]:
        """Compute changes between two time points"""
        if from_time not in self.time_states or to_time not in self.time_states:
            raise ValueError(f"Missing state for time point {from_time} or {to_time}")
        changes = {}
        for name, to_agent in self.time_states[to_time].items():
            from_agent = self.time_states[from_time][name]

            # Find new and removed assets
            new_assets = [a for a in to_agent.assets if not any(a.matches(from_a) for from_a in from_agent.assets)]
            removed_assets = [a for a in from_agent.assets if not any(a.matches(to_a) for to_a in to_agent.assets)]

            # Find new and removed liabilities
            new_liabilities = [l for l in to_agent.liabilities if not any(l.matches(from_l) for from_l in from_agent.liabilities)]
            for liability in new_liabilities:
                    if liability.maturity_type == MaturityType.FIXED_DATE:
                        maturity_time='t'+str(int((liability.maturity_date.year-2000)/50))
                        asset_holder = next(a for a in self.agents.values()
                                        if a.name == liability.counterparty)

                        asset_entry = next(a for a in asset_holder.assets
                                        if a.matches(liability))
                        liability_holder = next(a for a in self.agents.values()
                                        if a.name == asset_entry.counterparty)
                        
                        if int(maturity_time[-1])-int(liability.issuance_time[-1])<=1:
                            self.maturity_profile['t+1'][liability.type]=self.maturity_profile['t+1'][liability.type]+liability.denomination if self.maturity_profile['t+1'][liability.type] else liability.denomination
                            self.maturity[liability_holder][to_time]['short-term']=self.maturity[liability_holder][to_time].get('short-term',0)+1
                        else:
                            self.maturity_profile['t+n'][liability.type]=self.maturity_profile['t+n'][liability.type]+liability.denomination if self.maturity_profile['t+n'][liability.type] else liability.denomination
                            self.maturity[liability_holder][to_time]['long-term']=self.maturity[liability_holder][to_time].get('long-term',0)+1
            removed_liabilities = [l for l in from_agent.liabilities if not any(l.matches(to_l) for to_l in to_agent.liabilities)]

            changes[name] = {
                'new_assets': new_assets,
                'removed_assets': removed_assets,
                'new_liabilities': new_liabilities,
                'removed_liabilities': removed_liabilities
            }

        return changes

    def can_settle_entry(self, agent: Agent, entry: BalanceSheetEntry) -> Tuple[bool, str]:
        if entry.settlement_details.type == SettlementType.MEANS_OF_PAYMENT:
            deposits = sum(asset.current_book_value for asset in agent.assets
                          if asset.type == EntryType.DEPOSIT
                          and asset.is_asset
                          and asset.denomination == entry.denomination)
            if deposits < entry.current_book_value:
                return False, f"Insufficient deposit: has {deposits} {entry.denomination}, needs {entry.current_book_value}"

        elif entry.settlement_details.type == SettlementType.NON_FINANCIAL_ASSET:
            has_asset = any(asset.type == EntryType.NON_FINANCIAL
                           and asset.name == entry.name
                           and asset.current_book_value >= entry.current_book_value
                           for asset in agent.assets)
            if not has_asset:
                return False, f"Lack necessary non-financial asset: {entry.name}"

        return True, "OK"

    def create_default_entries(self, failed_entry: BalanceSheetEntry) -> Tuple[BalanceSheetEntry, BalanceSheetEntry]:
        default_claim = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=True,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.current_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default {failed_entry.type.value}",
            issuance_time=self.current_time_state,
            current_book_value=failed_entry.current_book_value
        )

        default_liability = BalanceSheetEntry(
            type=EntryType.DEFAULT,
            is_asset=False,
            counterparty=failed_entry.counterparty,
            initial_book_value=failed_entry.current_book_value,
            denomination=failed_entry.denomination,
            maturity_type=MaturityType.ON_DEMAND,
            maturity_date=None,
            settlement_details=failed_entry.settlement_details,
            name=f"Default {failed_entry.type.value}",
            issuance_time=self.current_time_state,
            current_book_value=failed_entry.current_book_value
        )

        return default_claim, default_liability

    # ENHANCED: Improved run_simulation with proper book value adjustments
    def run_simulation(self) -> bool:
        """Run the full simulation from t0 through t2, handling settlements and defaults"""
        print("\nStarting simulation from t0...")
        for agent in self.agents.values():
            self.maturity[agent.name]['t0']={'short-term':0,'long-term':0,'most':None}
            count,default_key,default_value={},None,0
            for liability in agent.liabilities:
                if liability.issuance_time=='t0' and liability.type in (EntryType.LOAN, EntryType.BOND, EntryType.DELIVERY_CLAIM,EntryType.PAYABLE):
                    self.debt_position[agent.name]['t0']['gross_debt_owed']=self.debt_position[agent.name]['t0'].get('gross_debt_owed',0)+liability.denomination
                    asset_holder = next(a for a in self.agents.values()
                                        if a.name == liability.counterparty)
                    self.debt_position[asset_holder.name]['t0']['gross_claims']=self.debt_position[asset_holder.name]['t0'].get('gross_claims',0)+liability.denomination
                    self.debt_map[agent.name]['t0'][liability.type]=liability.denomination
                    if liability.maturity_type == MaturityType.FIXED_DATE:
                        maturity_time='t'+str(int((liability.maturity_date.year-2000)/50))
                        count[maturity_time]=count.get(maturity_time,0)+1
                        if int(maturity_time[-1])-int(liability.issuance_time[-1])<=1:
                            self.maturity_profile['t+1'][liability.type]=self.maturity_profile['t+1'][liability.type]+liability.denomination if self.maturity_profile['t+1'][liability.type] else liability.denomination
                            self.maturity[agent.name]['t0']['short-term']+=1
                        else:
                            self.maturity_profile['t+n'][liability.type]=self.maturity_profile['t+n'][liability.type]+liability.denomination if self.maturity_profile['t+n'][liability.type] else liability.denomination
                            self.maturity[agent.name]['t0']['long-term']+=1
                    if self.debt_stock['t0'][liability.type][agent.type]:
                        self.debt_stock['t0'][liability.type][agent.type]=self.debt_stock['t0'][liability.type][agent.type]+liability.denomination
                    else:
                        self.debt_stock['t0'][liability.type][agent.type]=0
                        self.debt_stock['t0'][liability.type][agent.type]=self.debt_stock['t0'][liability.type][agent.type]+liability.denomination
            for key,value in count.items():
                if value>default_value:
                    default_value=value
                    default_key=key
            self.maturity[agent.name]['t0']['most']=default_key
        for agent in self.agents.values():
            self.debt_position[agent.name]['t0']['net_debt']=self.debt_position[agent.name]['t0'].get('gross_debt_owed',0)-self.debt_position[agent.name]['t0'].get('gross_claims',0)
            for i in range(3):
                self.debt_ratio[i][agent.name]['total_debt']=self.debt_position[agent.name]['t0'].get('gross_debt_owed',0)
            self.debt_ratio[2][agent.name]['net_worth']=agent.get_net_worth()
            self.debt_ratio[2][agent.name]['debt_to_equity']=self.debt_ratio[i][agent.name]['total_debt']/self.debt_ratio[2][agent.name]['net_worth']
            self.debt_ratio[0]['total_assets']=agent.get_total_assets()
            self.debt_ratio[0]['debt_to_asset']=self.debt_ratio[i][agent.name]['total_debt']/self.debt_ratio[0]['total_assets']
            # System-wide Debt-to-GDP Proxy(work out what is nominal output)
            # single time point or changing in time?
        for agent in self.agents.values():
            for asset in agent.assets:
                if asset.issuance_time=='t0' and asset.type == EntryType.DEPOSIT:
                    self.money_system[agent]=self.money_system.get(agent,[])+[('t0','asset',asset.amount)]
                    if self.time_series_tracker.get('t0',{})=={}:
                        self.time_series_tracker['t0']={}
                    changes=self.time_series_tracker['t0']
                    changes[asset.type]=changes.get(asset.type,0)+asset.amount
        for time_point in ['t1', 't2']:
            print(f"\nProcessing {time_point}...")

            # Get all entries that mature at this time point
            maturing_entries = []
            for agent in self.agents.values():
                for liability in agent.liabilities:
                    if (liability.maturity_type == MaturityType.FIXED_DATE and
                        ((time_point == 't1' and liability.maturity_date.year == 2050) or
                         (time_point == 't2' and liability.maturity_date.year == 2100))):
                        maturing_entries.append((agent, liability))

            # Try to settle each entry
            for agent, liability in maturing_entries:
                can_settle, reason = self.can_settle_entry(agent, liability)

                if not can_settle:
                    print(f"\nDEFAULT DETECTED: {agent.name} cannot settle {liability.type.value}")
                    temp=f"\nDEFAULT DETECTED: {agent.name} cannot settle {liability.type.value}/n Reason: {reason}"
                    value=liability.denomination
                    print(f"Reason: {reason}")

                    # Find the corresponding asset holder
                    asset_holder = next(a for a in self.agents.values()
                                      if a.name == liability.counterparty)

                    # Remove the original asset-liability pair
                    asset_entry = next(a for a in asset_holder.assets
                                     if a.matches(liability))
                    asset_holder.remove_asset(asset_entry)
                    agent.remove_liability(liability)

                    # Create and add default entries
                    default_claim, default_liability = self.create_default_entries(liability)
                    asset_holder.add_asset(default_claim)
                    agent.add_liability(default_liability)

                    # Save state after default
                    self.save_state(time_point)
                    class Node:
                        def __init__(self, element = None,length=None, prev = [], next = []):
                            self._element = element
                            self._length=length
                            self._prev = prev
                            self._next = next
                    if self.default_tracker[liability.maturity_date.year][asset_holder]==None:
                        self.default_tracker[liability.maturity_date.year][asset_holder]=[]
                    a=Node(value,0)
                    self.default_tracker[liability.maturity_date.year][asset_holder].append((temp,a))
                    if self.default_tracker[liability.maturity_date.year-50][agent]:
                        count=0
                        test=True
                        for i in self.default_tracker[liability.maturity_date.year-50][agent]:
                            if i[-1]._element>a._element:
                                i[-1]._length+=1
                                i[-1]._next.append(a)
                                a._prev.append(i[-1])
                                test=False
                                break
                        if test:
                            for i in self.default_tracker[liability.maturity_date.year-50][agent]:
                                i[-1]._length+=1
                                i[-1]._next.append(a)
                                a._prev.append(i[-1])
                    def recur(q,count):
                        if len(a._prev)==0:
                            return
                        if count!=0:
                            for n in q._prev:
                                    n._length+=1
                                    recur(n,count+1)
                        else:
                            for n in q._prev:
                                if len(n._prev)!=0:
                                    for g in n._prev:
                                        g._length+=1
                                        recur(g,count+1)
                    recur(a,0)

            # If we get here, try to settle all entries for this time point
            self.settle_entries(time_point)
            last_time='t'+str(int(time_point[-1])-1)
            self.compute_changes(last_time,time_point)
            self.debt_flows[f'{last_time} - {time_point}']={'issued': 0, 'repaid': 0, 'net_change': 0}
            for agent in self.agents.values():
                for liability in agent.liabilities:
                    if liability.type in (EntryType.LOAN, EntryType.BOND, EntryType.DELIVERY_CLAIM,EntryType.PAYABLE):
                        self.debt_map[agent.name][time_point][liability.type]=liability.denomination
                        self.debt_position[agent.name][time_point]['gross_debt_owed']=self.debt_position[agent.name][time_point].get('gross_debt_owed',0)+liability.denomination
                        asset_holder = next(a for a in self.agents.values()
                                        if a.name == liability.counterparty)
                        self.debt_position[asset_holder.name][time_point]['gross_claims']=self.debt_position[asset_holder.name][time_point].get('gross_claims',0)+liability.denomination
                        if self.debt_stock[time_point][liability.type][agent.type]:
                            self.debt_stock[time_point][liability.type][agent.type]=self.debt_stock[time_point][liability.type][agent.type]+liability.denomination
                        else:
                            self.debt_stock[time_point][liability.type][agent.type]=0
                            self.debt_stock[time_point][liability.type][agent.type]=self.debt_stock[time_point][liability.type][agent.type]+liability.denomination
            for agent in self.agents.values():
                self.debt_position[asset_holder.name][time_point]['net_debt']=self.debt_position[asset_holder.name][time_point].get('gross_debt_owed',0)-self.debt_position[asset_holder.name][time_point].get('gross_claims',0)
            for liability_type in self.debt_stock[time_point].keys():
                if self.debt_stock[last_time].get(liability_type,None)==None:
                    for agent_value in self.debt_stock[time_point][liability_type].values():
                        self.debt_flows[f'{last_time} - {time_point}']['issued']+=agent_value
                else:
                    for agent_type,agent_value in self.debt_stock[time_point][liability_type].items():
                        if self.debt_stock[last_time][liability_type].get(agent_type,None)==None:
                            self.debt_flows[f'{last_time} - {time_point}']['issued']+=agent_value
            
            for liability_type in self.debt_stock[last_time].keys():
                if self.debt_stock[time_point].get(liability_type,None)==None:
                    for agent_value in self.debt_stock[last_time][liability_type].values():
                        self.debt_flows[f'{last_time} - {time_point}']['repaid']+=agent_value
                else:
                    for agent_type,agent_value in self.debt_stock[last_time][liability_type].items():
                        if self.debt_stock[time_point][liability_type].get(agent_type,None)==None:
                            self.debt_flows[f'{last_time} - {time_point}']['repaid']+=agent_value
            self.debt_flows[f'{last_time} - {time_point}']['net_change']=self.debt_flows[f'{last_time} - {time_point}']['issued']-self.debt_flows[f'{last_time} - {time_point}']['repaid']
            
            for agent in self.agents.values():
                for asset in agent.assets:
                    if asset.type == EntryType.DEPOSIT:
                        if self.time_series_tracker.get(time_point,{})=={}:
                            self.time_series_tracker[time_point]={}
                        changes=self.time_series_tracker[time_point]
                        changes[asset.type]=changes.get(asset.type,0)+asset.amount          
            for time_point,assets in self.time_series_tracker.items():
                for asset in assets.keys():
                    if time_point!='t0':
                        last='t'+str(int(time_point[-1])-1)
                        change=self.time_series_tracker[last].get(asset)-assets.get(asset)
                        self.means_of_payment[asset]=self.means_of_payment[asset]+[{"time": time_point, "total": assets.get(asset,0), "change": change}]
                        if change!=0.0:
                            if change<0:
                                self.flux_and_reflux[time_point]=self.flux_and_reflux.get(time_point,[])+[(time_point,'+'+str(change),'flux')]
                            else:
                                self.flux_and_reflux[time_point]=self.flux_and_reflux.get(time_point,[])+[(time_point,'-'+str(change),'reflux')]
                    else:
                        self.means_of_payment[asset]=[{"time": time_point, "total": assets.get(asset,0), "change": None}]
            max_length,max_default=0,None
            for date,agents in self.default_tracker.items():
                for agent,defaults in agents.items():
                    for default in defaults:
                        if default[-1]._length>max_length:
                            max_length=default[-1]._length
                            max_default=default
            self.default_tracker['max']=max_default
                    
                    
                               
        print("\nSimulation completed successfully!")
        print(self.time_series_tracker)
        print(self.flux_and_reflux)
        print(self.means_of_payment)
        return True

    
    def to_dict(self):
        """Serialize for Web UI with enhanced scheduled actions"""
        return {
            "agents": {name: agent.to_dict() for name, agent in self.agents.items()},
            "asset_liability_pairs": [pair.to_dict() for pair in self.asset_liability_pairs],
            "time_states": {
                tp: {
                    agent_name: {
                        "type": agent_state.type.value,
                        "assets": [a.to_dict() for a in agent_state.assets],
                        "liabilities": [l.to_dict() for l in agent_state.liabilities],
                        "creation_time": agent_state.creation_time.isoformat()
                    } for agent_name, agent_state in agents.items()
                } for tp, agents in self.time_states.items()
            },
            "current_time_state": self.current_time_state,
            "simulation_finalized": self.simulation_finalized,
            "scheduled_actions": self.scheduled_actions  # NEW: Include scheduled actions
        }

    @classmethod
    def from_dict(cls, data):
        """Enhanced deserialization with new EntryType support and backward compatibility"""
        system = cls()
        
        # Helper function to convert old entry type names to new ones
        def convert_entry_type(old_type_str):
            """Convert old entry type names to new EntryType enum values"""
            # Mapping for backward compatibility
            type_mapping = {
                # Old format -> New format
                "loan": EntryType.LOAN,  # now "bank_client_loan"
                "deposit": EntryType.DEPOSIT,  # now "bank_deposit"
                "intraday_iou": EntryType.INTRADAY_IOU,  # now "interbank_intraday_iou"
                "overnight_loan": EntryType.OVERNIGHT_LOAN,  # now "interbank_overnight_loan"
                
                # Direct mappings for existing types
                "bank_client_loan": EntryType.LOAN,
                "bank_deposit": EntryType.DEPOSIT,
                "reserve_deposit": EntryType.RESERVE_DEPOSIT,
                "interbank_intraday_iou": EntryType.INTRADAY_IOU,
                "interbank_overnight_loan": EntryType.OVERNIGHT_LOAN,
                "payable": EntryType.PAYABLE,
                "bond": EntryType.BOND,
                "bond_zero_coupon": EntryType.BOND_ZERO_COUPON,
                "bond_coupon": EntryType.BOND_COUPON,
                "bond_amortizing": EntryType.BOND_AMORTIZING,
                
                # New government bond types
                "government_bond_zero_coupon": EntryType.GOVERNMENT_BOND_ZERO_COUPON,
                "government_bond_coupon": EntryType.GOVERNMENT_BOND_COUPON,
                "government_bond_amortizing": EntryType.GOVERNMENT_BOND_AMORTIZING,
                
                # New central bank bond types
                "central_bank_bond_zero_coupon": EntryType.CENTRAL_BANK_BOND_ZERO_COUPON,
                "central_bank_bond_coupon": EntryType.CENTRAL_BANK_BOND_COUPON,
                "central_bank_bond_amortizing": EntryType.CENTRAL_BANK_BOND_AMORTIZING,
                
                "share": EntryType.SHARE,
                "delivery_claim": EntryType.DELIVERY_CLAIM,
                "non_financial": EntryType.NON_FINANCIAL,
                "default": EntryType.DEFAULT
            }
            
            if old_type_str in type_mapping:
                return type_mapping[old_type_str]
            else:
                # Try direct conversion first
                try:
                    return EntryType(old_type_str)
                except ValueError:
                    raise ValueError(f"Unknown entry type: {old_type_str}")

        # Helper function to convert time representations
        def convert_time_value(time_value):
            """Convert various time representations to integer"""
            if time_value is None:
                return None
            if isinstance(time_value, int):
                return time_value
            if isinstance(time_value, str):
                if time_value.startswith('t'):
                    return int(time_value[1:])
                elif 'T' in time_value:  # ISO datetime format
                    dt = datetime.fromisoformat(time_value)
                    if dt.year == 2050:
                        return 1
                    elif dt.year == 2100:
                        return 2
                    else:
                        return 0
                else:
                    return int(time_value)
            return 0

        # Helper function to create BalanceSheetEntry from data
        def create_balance_sheet_entry(entry_data):
            """Create a BalanceSheetEntry from serialized data"""
            # Convert entry type
            entry_type = convert_entry_type(entry_data["type"])
            
            # Convert time values
            issuance_time = convert_time_value(entry_data.get("issuance_time", 0))
            maturity_date = convert_time_value(entry_data.get("maturity_date"))
            
            # Handle settlement details
            settlement_details_data = entry_data.get("settlement_details", {})
            if isinstance(settlement_details_data, dict):
                settlement_details = SettlementDetails(
                    type=SettlementType(settlement_details_data.get("type", "none")),
                    denomination=settlement_details_data.get("denomination", "USD"),
                    is_intraday=settlement_details_data.get("is_intraday", False)
                )
            else:
                # Fallback for older formats
                settlement_details = SettlementDetails(
                    type=SettlementType.NONE,
                    denomination="USD"
                )
            
            return BalanceSheetEntry(
                type=entry_type,
                is_asset=entry_data["is_asset"],
                counterparty=entry_data.get("counterparty"),
                initial_book_value=entry_data.get("initial_book_value", entry_data.get("amount", 0)),
                denomination=entry_data.get("denomination", "USD"),
                maturity_type=MaturityType(entry_data.get("maturity_type", "on_demand")),
                maturity_date=maturity_date,
                settlement_details=settlement_details,
                name=entry_data.get("name"),
                issuance_time=issuance_time,
                current_book_value=entry_data.get("current_book_value", entry_data.get("initial_book_value", entry_data.get("amount", 0))),
                cash_flow_at_maturity=entry_data.get("cash_flow_at_maturity", 0),
                rollover_count=entry_data.get("rollover_count", 0),
                expected_cash_flow=entry_data.get("expected_cash_flow"),
                parent_bond=entry_data.get("parent_bond")
            )

        # Deserialize Agents
        agents_data = data.get("agents", {})
        for agent_name, agent_data in agents_data.items():
            if "type" not in agent_data:
                raise ValueError(f"Agent {agent_name} is missing 'type' field in configuration data.")
            
            try:
                agent_type = AgentType(agent_data["type"])
            except ValueError:
                raise ValueError(f"Invalid agent type '{agent_data['type']}' for agent {agent_name}")
            
            agent = Agent(name=agent_data.get("name", agent_name), agent_type=agent_type)
            
            # Set agent metadata
            if "creation_time" in agent_data:
                agent.creation_time = datetime.fromisoformat(agent_data["creation_time"])
            agent.status = agent_data.get("status", "operating")
            
            # Deserialize assets with error handling
            agent.assets = []
            for asset_data in agent_data.get("assets", []):
                try:
                    entry = create_balance_sheet_entry(asset_data)
                    agent.assets.append(entry)
                except Exception as e:
                    print(f"Warning: Could not deserialize asset for {agent_name}: {e}")
                    continue
            
            # Deserialize liabilities with error handling
            agent.liabilities = []
            for liability_data in agent_data.get("liabilities", []):
                try:
                    entry = create_balance_sheet_entry(liability_data)
                    agent.liabilities.append(entry)
                except Exception as e:
                    print(f"Warning: Could not deserialize liability for {agent_name}: {e}")
                    continue
            
            # Deserialize settlement history with backward compatibility
            if "settlement_history" in agent_data:
                settlement_history = agent_data["settlement_history"]
                
                # Process as_asset_holder records
                for record in settlement_history.get('as_asset_holder', []):
                    try:
                        # Convert time_point if needed
                        time_point = convert_time_value(record['time_point'])
                        record['time_point'] = time_point
                        
                        # Convert entry objects if they exist
                        if 'original_entry' in record and isinstance(record['original_entry'], dict):
                            record['original_entry'] = create_balance_sheet_entry(record['original_entry'])
                        if 'settlement_result' in record and isinstance(record['settlement_result'], dict):
                            record['settlement_result'] = create_balance_sheet_entry(record['settlement_result'])
                            
                    except Exception as e:
                        print(f"Warning: Could not process settlement history for {agent_name}: {e}")
                        continue
                
                # Process as_liability_holder records
                for record in settlement_history.get('as_liability_holder', []):
                    try:
                        # Convert time_point if needed
                        time_point = convert_time_value(record['time_point'])
                        record['time_point'] = time_point
                        
                        # Convert entry objects if they exist
                        if 'original_entry' in record and isinstance(record['original_entry'], dict):
                            record['original_entry'] = create_balance_sheet_entry(record['original_entry'])
                        if 'settlement_result' in record and isinstance(record['settlement_result'], dict):
                            record['settlement_result'] = create_balance_sheet_entry(record['settlement_result'])
                            
                    except Exception as e:
                        print(f"Warning: Could not process settlement history for {agent_name}: {e}")
                        continue
                
                agent.settlement_history = settlement_history
            
            # Add agent to system
            system.agents[agent.name] = agent
            
            # Set bank-specific attributes
            if agent.type == AgentType.BANK:
                agent.system = system
                agent.intraday_module.system = system

        # Deserialize AssetLiabilityPairs with enhanced error handling
        system.asset_liability_pairs = []
        for pair_data in data.get("asset_liability_pairs", []):
            try:
                # Convert time values
                time = convert_time_value(pair_data.get("time", 0))
                maturity_date = convert_time_value(pair_data.get("maturity_date"))
                
                # Get agents
                asset_holder_name = pair_data.get("asset_holder")
                liability_holder_name = pair_data.get("liability_holder")
                
                asset_holder = system.agents.get(asset_holder_name)
                liability_holder = system.agents.get(liability_holder_name) if liability_holder_name else None
                
                if not asset_holder:
                    print(f"Warning: Asset holder '{asset_holder_name}' not found for asset-liability pair")
                    continue
                
                if liability_holder_name and not liability_holder:
                    print(f"Warning: Liability holder '{liability_holder_name}' not found for asset-liability pair")
                    continue
                
                # Convert bond type if present
                bond_type = None
                if pair_data.get("bond_type") is not None:
                    try:
                        bond_type = BondType(pair_data["bond_type"])
                    except (ValueError, TypeError):
                        bond_type = None
                
                # Convert entry type
                entry_type_str = pair_data.get("type", "")
                entry_type = convert_entry_type(entry_type_str)
                
                # Create AssetLiabilityPair
                pair = AssetLiabilityPair(
                    time=time,
                    type=entry_type.value,  # Use the converted entry type value
                    amount=pair_data.get("amount", 0),
                    denomination=pair_data.get("denomination", "USD"),
                    maturity_type=MaturityType(pair_data.get("maturity_type", "on_demand")),
                    maturity_date=maturity_date,
                    settlement_type=SettlementType(pair_data.get("settlement_type", "none")),
                    settlement_denomination=pair_data.get("settlement_denomination", "USD"),
                    asset_holder=asset_holder,
                    liability_holder=liability_holder,
                    asset_name=pair_data.get("asset_name"),
                    bond_type=bond_type,
                    coupon_rate=pair_data.get("coupon_rate"),
                    cash_flow_at_maturity=pair_data.get("cash_flow_at_maturity", 0),
                    scheduled_time_point=pair_data.get("scheduled_time_point")
                )
                
                system.asset_liability_pairs.append(pair)
                
            except Exception as e:
                print(f"Warning: Could not deserialize asset-liability pair: {e}")
                continue

        # Deserialize Time States with enhanced error handling
        system.time_states = {}
        time_states_data = data.get("time_states", {})
        
        for tp_str, agents_data in time_states_data.items():
            try:
                time_point = int(tp_str) if isinstance(tp_str, str) and tp_str.isdigit() else tp_str
                if isinstance(time_point, str):
                    time_point = convert_time_value(time_point)
                
                system.time_states[time_point] = {}
                
                for agent_name, agent_data in agents_data.items():
                    try:
                        agent_type = AgentType(agent_data["type"])
                        agent_copy = Agent(agent_name, agent_type)
                        
                        if "creation_time" in agent_data:
                            agent_copy.creation_time = datetime.fromisoformat(agent_data["creation_time"])
                        
                        # Copy assets
                        agent_copy.assets = []
                        for asset_data in agent_data.get("assets", []):
                            try:
                                entry = create_balance_sheet_entry(asset_data)
                                agent_copy.assets.append(entry)
                            except Exception as e:
                                print(f"Warning: Could not deserialize time state asset for {agent_name} at t{time_point}: {e}")
                                continue
                        
                        # Copy liabilities
                        agent_copy.liabilities = []
                        for liability_data in agent_data.get("liabilities", []):
                            try:
                                entry = create_balance_sheet_entry(liability_data)
                                agent_copy.liabilities.append(entry)
                            except Exception as e:
                                print(f"Warning: Could not deserialize time state liability for {agent_name} at t{time_point}: {e}")
                                continue
                        
                        system.time_states[time_point][agent_name] = agent_copy
                        
                    except Exception as e:
                        print(f"Warning: Could not deserialize time state for agent {agent_name} at t{time_point}: {e}")
                        continue
                        
            except Exception as e:
                print(f"Warning: Could not deserialize time state t{tp_str}: {e}")
                continue

        # Restore system state with validation
        system.current_time_state = data.get("current_time_state", 0)
        system.simulation_finalized = data.get("simulation_finalized", False)
        
        # NEW: Restore scheduled actions with enhanced support
        scheduled_actions_data = data.get("scheduled_actions", {})
        system.scheduled_actions = {}
        
        for time_point_str, actions in scheduled_actions_data.items():
            try:
                time_point = int(time_point_str)
                system.scheduled_actions[time_point] = []
                
                for action in actions:
                    try:
                        # Validate action structure
                        if isinstance(action, dict) and 'type' in action:
                            # Convert any entry types in action data
                            if action['type'] == 'create_asset_liability_pair' and 'data' in action:
                                action_data = action['data']
                                if 'type' in action_data:
                                    try:
                                        # Convert old entry type to new format
                                        old_type = action_data['type']
                                        new_entry_type = convert_entry_type(old_type)
                                        action_data['type'] = new_entry_type.value
                                    except ValueError as e:
                                        print(f"Warning: Could not convert scheduled action entry type '{old_type}': {e}")
                                        continue
                            
                            system.scheduled_actions[time_point].append(action)
                        else:
                            print(f"Warning: Invalid scheduled action format at t{time_point}: {action}")
                            
                    except Exception as e:
                        print(f"Warning: Could not process scheduled action at t{time_point}: {e}")
                        continue
                        
            except (ValueError, TypeError) as e:
                print(f"Warning: Invalid time point in scheduled actions: {time_point_str}")
                continue

        # Validate system integrity
        try:
            # Check that all agents exist
            if not system.agents:
                print("Warning: No agents found in deserialized system")
            
            # Check that current_time_state is valid
            if system.current_time_state not in [0, 1, 2]:
                print(f"Warning: Invalid current_time_state {system.current_time_state}, resetting to 0")
                system.current_time_state = 0
            
            # Ensure t0 state exists
            if 0 not in system.time_states and system.agents:
                system.save_state(0)
                
        except Exception as e:
            print(f"Warning: System integrity check failed: {e}")

        print(f"Successfully deserialized system with {len(system.agents)} agents, "
            f"{len(system.asset_liability_pairs)} pairs, "
            f"{len(system.time_states)} time states, "
            f"and {sum(len(actions) for actions in system.scheduled_actions.values())} scheduled actions")
        
        return system
# ======== Web应用扩展 ========
app = Flask(__name__)
app.secret_key = os.urandom(24).hex()
system = EconomicSystem()

# ======== Web界面HTML模板 ========
HTML_BASE = '''
<!DOCTYPE html>
<html>
<head>
    <title>Economic Sim Web</title>
    <link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css">
    <style>
        .container { 
            max-width: 800px;  
            margin: 20px auto; 
        }
        .form-section { 
            margin: 20px 0; 
            padding: 20px; 
            border: 1px solid #ddd;  
        }
        .visualization { 
            height: 400px; 
            background: #f5f5f5;  
            margin: 20px 0; 
        }
        .list-group-item form {
            margin-left: 10px;
        }
        .btn-sm {
            padding: 0.15rem 0.5rem;
            font-size: 0.875rem;
        }
        .scheduled-actions {
            background-color: #f8f9fa;
            border: 1px solid #dee2e6;
            border-radius: 0.375rem;
            padding: 1rem;
            margin: 1rem 0;
        }
    </style>
</head>
<body>
    <div class="container">
        {% with messages = get_flashed_messages(with_categories=true) %}  
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }} alert-dismissible fade show">
                        {{ message }}
                        <button type="button" class="btn-close" data-bs-dismiss="alert"></button>  
                    </div>
                {% endfor %}
            {% endif %}
        {% endwith %}
        
        <div class="d-flex justify-content-between mb-3">  
            <h2>Economic Balance Sheet Simulator</h2>
            <div>
                <a href="/output" class="btn btn-primary me-2">Output</a>
                <form action="/simulate" method="post" class="d-inline"> 
                    <button class="btn btn-warning">Run Simulation</button> 
                </form>
            </div>
        </div>
        
        <div class="row">
            <div class="col-md-4">
                <div class="list-group">
                    <a href="/" class="list-group-item list-group-item-action">Home</a>
                    <a href="/agents" class="list-group-item list-group-item-action">View Agents</a>
                    <a href="/settlement_history" class="list-group-item list-group-item-action">View Settlement History</a>
                    <a href="/scheduled_actions" class="list-group-item list-group-item-action">Scheduled Actions</a>
                    <a href="/config" class="list-group-item list-group-item-action">Configurations</a>
                </div>
            </div>
            <div class="col-md-8">
                {% block content %}{% endblock %}
            </div>
        </div>
    </div>
</body>
</html>
'''

# ENHANCED HOME_CONTENT with Interest Rate field and Operation Time Point
HOME_CONTENT = '''
<div class="form-section">
    <h4>Create Agent</h4>
    <form action="/create_agent" method="post">
        <div class="mb-3">
            <input type="text" name="name" placeholder="Agent Name" class="form-control" required>
        </div>
        <select name="type" class="form-select mb-3">
            {% for type in agent_types %}
            <option value="{{ type.value }}">{{ type.value.title() }}</option>
            {% endfor %}
        </select>
        <button type="submit" class="btn btn-primary">Create</button>
    </form>
</div>

<!-- NEW: Show pending scheduled actions -->
{% if scheduled_actions %}
<div class="scheduled-actions">
    <h5>Pending Scheduled Actions</h5>
    {% for time_point, actions in scheduled_actions.items() %}
        <div class="mb-2">
            <strong>t{{ time_point }}:</strong>
            {% for action in actions %}
                <span class="badge bg-info me-1">{{ action.type }}</span>
            {% endfor %}
        </div>
    {% endfor %}
</div>
{% endif %}

<div class="card">
    <div class="card-body">
        <h4>Create Asset-Liability Pair</h4>
        
        <form action="/create_pair" method="post" onsubmit="return validateForm()">
            <!-- NEW: Operation Time Point Selection -->
            <div class="mb-3">
                <label class="form-label">Operation Time Point</label>
                <select name="operation_time_point" class="form-select" id="operationTimePoint">
                    <option value="0">t0 (Immediate Creation)</option>
                    <option value="1">t1 (Schedule for t1)</option>
                    <option value="2">t2 (Schedule for t2)</option>
                </select>
                <div class="form-text">Select when this asset-liability pair should be created</div>
            </div>

            <!-- Asset Holder - MODIFIED -->
            <div class="mb-3">
                <label class="form-label">Asset Holder</label>
                <select name="asset_holder" class="form-select" required id="assetHolder">
                    {% for agent in agents %}
                    <option value="{{ agent.name }}" data-agent-type="{{ agent.type.value }}">{{ agent.name }} ({{ agent.type.value }})</option>
                    {% endfor %}
                </select>
            </div>

            <!-- Entry Type -->
            <div class="mb-3">
                <label class="form-label">Entry Type</label>
                <select name="entry_type" class="form-select" id="entryType" required>
                    {% for et in entry_types %}
                    <option value="{{ et.value }}">
                        {% if et.value == 'government_bond_zero_coupon' %}
                            Government Bond
                        {% elif et.value == 'central_bank_bond_zero_coupon' %}
                            Central Bank Bond
                        {% else %}
                            {{ et.value|replace('_', ' ')|title }}
                        {% endif %}
                    </option>
                    {% endfor %}
                </select>
            </div>

            <!-- Liability Holder - MODIFIED -->
            <div class="mb-3" id="liabilityHolderGroup">
                <label class="form-label">Liability Holder</label>
                <select name="liability_holder" class="form-select" id="liabilityHolder">
                    <option value="" data-agent-type="">None</option>
                    {% for agent in agents %}
                    <option value="{{ agent.name }}" data-agent-type="{{ agent.type.value }}">{{ agent.name }} ({{ agent.type.value }})</option>
                    {% endfor %}
                </select>
            </div>

            <!-- Asset Name (conditional) -->
            <div class="mb-3" id="assetNameGroup" style="display:none;">
                <label class="form-label">Asset Name</label>
                <input type="text" name="asset_name" class="form-control" id="assetName">
            </div>

            <!-- Bond Category - NEW -->
            <div class="mb-3" id="bondCategoryGroup" style="display:none;">
                <label class="form-label">Bond Category</label>
                <select name="bond_category" class="form-select" id="bondCategory">
                    <option value="corporate">Corporate/Private Bond</option>
                    <option value="government">Government Bond</option>
                    <option value="central_bank">Central Bank Bond</option>
                </select>
            </div>

            <!-- Bond Type (conditional) -->
            <div class="mb-3" id="bondTypeGroup" style="display:none;">
                <label class="form-label">Bond Type</label>
                <select name="bond_type" class="form-select" id="bondType">
                    <option value="0">Zero Coupon</option>
                    <option value="1">Coupon Bond</option>
                    <option value="2">Amortizing Bond</option>
                </select>
            </div>

            <!-- Coupon Rate (conditional) -->
            <div class="mb-3" id="couponRateGroup" style="display:none;">
                <label class="form-label">Coupon/Interest Rate</label>
                <input type="number" step="0.001" name="coupon_rate" class="form-control" id="couponRate" placeholder="e.g., 0.05 for 5%">
            </div>

            <!-- NEW: Interest Rate for Loans -->
            <div class="mb-3" id="interestRateGroup" style="display:none;">
                <label class="form-label">Interest Rate</label>
                <input type="number" step="0.001" name="interest_rate" class="form-control" id="interestRate" value="0" placeholder="e.g., 0.05 for 5%">
                <div class="form-text">Interest rate for loans and overnight loans</div>
            </div>

            <!-- Amount & Denomination -->
            <div class="row mb-3">
                <div class="col">
                    <label class="form-label">Amount</label>
                    <input type="number" step="0.01" name="amount" class="form-control" required>
                </div>
                <div class="col">
                    <label class="form-label">Denomination</label>
                    <input type="text" name="denomination" class="form-control" value="USD" required>
                </div>
            </div>

            <!-- Maturity -->
            <div class="row mb-3" id="maturityGroup">
                <div class="col">
                    <label class="form-label">Maturity Type</label>
                    <select name="maturity_type" class="form-select" id="maturityType">
                        {% for mt in MaturityType %}
                        <option value="{{ mt.value }}" 
                                {% if mt == MaturityType.ON_DEMAND %}selected{% endif %}>
                            {{ mt.value|replace('_', ' ')|title }}
                        </option>
                        {% endfor %}
                    </select>
                </div>
                <div class="col" id="maturityDateGroup" style="display:none;">
                    <label class="form-label">Maturity Time Point</label>
                    <select name="maturity_time_point" class="form-select" id="maturityDate">
                        <option value="1">t1</option>
                        <option value="2">t2</option>
                    </select>
                </div>
            </div>

            <!-- Settlement Type -->
            <div class="mb-3" id="settlementGroup">
                <label class="form-label">Settlement Type</label>
                <select name="settlement_type" class="form-select" id="settlementType">
                    {% for st in SettlementType if st != SettlementType.NONE %}
                        <option value="{{ st.value }}" 
                                {% if st == SettlementType.MEANS_OF_PAYMENT %}selected{% endif %}>
                            {{ st.value|replace('_', ' ')|title }}
                        </option>
                    {% endfor %}
                </select>
            </div>

            <button type="submit" class="btn btn-primary">Create</button>
        </form>
    </div>
</div>

<script>
// 约束配置
const CONSTRAINTS = {
    'bank_client_loan': {
        assetHolders: ['bank'],
        liabilityHolders: ['company', 'household', 'other']
    },
    'loan': {  // 向后兼容
        assetHolders: ['bank'],
        liabilityHolders: ['company', 'household', 'other']
    },
    'bank_deposit': {
        assetHolders: ['company', 'household', 'other'],
        liabilityHolders: ['bank']
    },
    'deposit': {  // 向后兼容
        assetHolders: ['company', 'household', 'other'],
        liabilityHolders: ['bank']
    },
    'reserve_deposit': {
        assetHolders: ['bank', 'treasury'],
        liabilityHolders: ['central_bank']
    },
    'interbank_intraday_iou': {
        assetHolders: ['bank'],
        liabilityHolders: ['bank']
    },
    'intraday_iou': {  // 向后兼容
        assetHolders: ['bank'],
        liabilityHolders: ['bank']
    },
    'interbank_overnight_loan': {
        assetHolders: ['bank'],
        liabilityHolders: ['bank']
    },
    'overnight_loan': {  // 向后兼容
        assetHolders: ['bank'],
        liabilityHolders: ['bank']
    },
    'bond_zero_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['bank', 'company', 'central_bank', 'treasury', 'other']
    },
    'bond_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['bank', 'company', 'central_bank', 'treasury', 'other']
    },
    'bond_amortizing': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['bank', 'company', 'central_bank', 'treasury', 'other']
    },
    'government_bond_zero_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['treasury']
    },
    'government_bond_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['treasury']
    },
    'government_bond_amortizing': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['treasury']
    },
    'central_bank_bond_zero_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['central_bank']
    },
    'central_bank_bond_coupon': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['central_bank']
    },
    'central_bank_bond_amortizing': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['central_bank']
    },
    'share': {
        assetHolders: ['bank', 'company', 'household', 'other'],
        liabilityHolders: ['bank', 'company', 'other']
    }
};

// 债券类别配置
const BOND_CATEGORIES = {
    'corporate': [
        { value: '0', name: 'Zero Coupon', entry_type: 'bond_zero_coupon' },
        { value: '1', name: 'Coupon Bond', entry_type: 'bond_coupon' },
        { value: '2', name: 'Amortizing Bond', entry_type: 'bond_amortizing' }
    ],
    'government': [
        { value: '0', name: 'Zero Coupon', entry_type: 'government_bond_zero_coupon' },
        { value: '1', name: 'Coupon Bond', entry_type: 'government_bond_coupon' },
        { value: '2', name: 'Amortizing Bond', entry_type: 'government_bond_amortizing' }
    ],
    'central_bank': [
        { value: '0', name: 'Zero Coupon', entry_type: 'central_bank_bond_zero_coupon' },
        { value: '1', name: 'Coupon Bond', entry_type: 'central_bank_bond_coupon' },
        { value: '2', name: 'Amortizing Bond', entry_type: 'central_bank_bond_amortizing' }
    ]
};

// 添加约束验证函数
function validateConstraints() {
    const entryType = document.getElementById('entryType').value;
    const assetHolderSelect = document.getElementById('assetHolder');
    const liabilityHolderSelect = document.getElementById('liabilityHolder');
    
    // 获取选中的agent类型
    const assetHolderType = assetHolderSelect.selectedOptions[0]?.dataset.agentType;
    const liabilityHolderType = liabilityHolderSelect.selectedOptions[0]?.dataset.agentType;
    
    // 对于债券类型，需要根据类别确定具体的entry type
    let actualEntryType = entryType;
    if (entryType === 'bond') {
        const bondCategory = document.getElementById('bondCategory')?.value || 'corporate';
        const bondTypeValue = document.getElementById('bondType').value;
        const bondConfig = BOND_CATEGORIES[bondCategory]?.find(b => b.value === bondTypeValue);
        if (bondConfig) {
            actualEntryType = bondConfig.entry_type;
        }
    }
    
    // 检查约束
    if (CONSTRAINTS[actualEntryType]) {
        const constraint = CONSTRAINTS[actualEntryType];
        
        if (!constraint.assetHolders.includes(assetHolderType)) {
            alert(`${actualEntryType.replace(/_/g, ' ')} assets can only be held by: ${constraint.assetHolders.join(', ')}`);
            return false;
        }
        
        if (liabilityHolderType && !constraint.liabilityHolders.includes(liabilityHolderType)) {
            alert(`${actualEntryType.replace(/_/g, ' ')} liabilities can only be held by: ${constraint.liabilityHolders.join(', ')}`);
            return false;
        }
    }
    
    return true;
}

// 更新债券类型选项
function updateBondTypeOptions() {
    const bondCategory = document.getElementById('bondCategory')?.value || 'corporate';
    const bondTypeSelect = document.getElementById('bondType');
    
    if (!bondTypeSelect) return;
    
    // 清空现有选项
    bondTypeSelect.innerHTML = '';
    
    // 添加新选项
    if (BOND_CATEGORIES[bondCategory]) {
        BOND_CATEGORIES[bondCategory].forEach(bond => {
            const option = document.createElement('option');
            option.value = bond.value;
            option.textContent = bond.name;
            bondTypeSelect.appendChild(option);
        });
    }
    
    // 触发债券字段更新
    updateBondFields();
}

// Form validation - 修改后的版本
function validateForm() {
    const entryType = document.getElementById('entryType').value;
    const assetHolder = document.getElementById('assetHolder').value;
    const liabilityHolder = document.getElementById('liabilityHolder').value;
    const maturityType = document.getElementById('maturityType');

    // Validate non-financial entries
    if (entryType === 'non_financial') {
        if (liabilityHolder) {
            alert("Non-financial entries cannot have a liability holder!");
            return false;
        }
        if (document.getElementById('assetName').value.trim() === "") {
            alert("Non-financial entries must specify an asset name!");
            return false;
        }
    }

    // Check asset holder and liability holder are different
    if (liabilityHolder && assetHolder === liabilityHolder) {
        alert("Asset holder and liability holder cannot be the same!");
        return false;
    }

    // Validate bond entries
    if (entryType.includes('bond') || entryType === 'bond') {
        const bondType = document.getElementById('bondType').value;
        const couponRate = document.getElementById('couponRate').value;
        
        if ((bondType === '1' || bondType === '2') && !couponRate) {
            alert("Coupon and amortizing bonds require an interest rate!");
            return false;
        }
    }

    // 添加约束验证
    if (!validateConstraints()) {
        return false;
    }

    return true;
}

// Filter liability holders
function filterLiabilityHolders() {
    const assetHolderSelect = document.getElementById('assetHolder');
    const liabilityHolderSelect = document.getElementById('liabilityHolder');
    const selectedAssetHolder = assetHolderSelect.value;

    Array.from(liabilityHolderSelect.options).forEach(option => {
        option.disabled = (option.value === selectedAssetHolder);
        if (option.value === "") {
            option.disabled = false;
        }
    });
}

// Limit maturity_type options
function limitMaturityTypeOptions(allowedOptions) {
    const maturityTypeSelect = document.getElementById('maturityType');
    const allOptions = maturityTypeSelect.querySelectorAll('option');
    
    allOptions.forEach(option => {
        if (allowedOptions.includes(option.value)) {
            option.style.display = 'block';
            option.disabled = false;
        } else {
            option.style.display = 'none';
            option.disabled = true;
        }
    });
}

// Reset maturity_type options
function resetMaturityTypeOptions() {
    const maturityTypeSelect = document.getElementById('maturityType');
    const allOptions = maturityTypeSelect.querySelectorAll('option');
    
    allOptions.forEach(option => {
        option.style.display = 'block';
        option.disabled = false;
    });
}

// NEW: Update form fields based on entry type including interest rate - 修改后的版本
function updateFormFields() {
    const entryType = document.getElementById('entryType').value;
    const formElements = {
        liability: {
            group: document.getElementById('liabilityHolderGroup'),
            field: document.getElementById('liabilityHolder')
        },
        assetName: {
            group: document.getElementById('assetNameGroup'),
            field: document.getElementById('assetName')
        },
        bondCategory: {  // 新增
            group: document.getElementById('bondCategoryGroup'),
            field: document.getElementById('bondCategory')
        },
        bondType: {
            group: document.getElementById('bondTypeGroup'),
            field: document.getElementById('bondType')
        },
        couponRate: {
            group: document.getElementById('couponRateGroup'),
            field: document.getElementById('couponRate')
        },
        interestRate: {
            group: document.getElementById('interestRateGroup'),
            field: document.getElementById('interestRate')
        },
        maturity: {
            group: document.getElementById('maturityGroup'),
            type: document.getElementById('maturityType'),
            dateGroup: document.getElementById('maturityDateGroup'),
            date: document.getElementById('maturityDate')
        },
        settlement: {
            group: document.getElementById('settlementGroup'),
            field: document.getElementById('settlementType')
        }
    };

    // Reset all fields
    resetMaturityTypeOptions();
    formElements.liability.group.style.display = 'block';
    formElements.liability.field.disabled = false;
    formElements.liability.field.required = true;
    formElements.assetName.group.style.display = 'none';
    formElements.assetName.field.required = false;
    
    // 新增：重置债券相关字段
    if (formElements.bondCategory.group) {
        formElements.bondCategory.group.style.display = 'none';
    }
    formElements.bondType.group.style.display = 'none';
    formElements.couponRate.group.style.display = 'none';
    formElements.interestRate.group.style.display = 'none';
    
    formElements.maturity.group.style.display = 'block';
    formElements.maturity.type.disabled = false;
    formElements.maturity.date.disabled = false;
    formElements.settlement.group.style.display = 'block';
    formElements.settlement.field.disabled = false;

    // Handle specific entry types
    if (entryType === 'non_financial') {
        resetMaturityTypeOptions();
        
        formElements.settlement.field.value = 'none';
        formElements.maturity.type.value = 'on_demand';
        formElements.liability.field.value = '';
        
        formElements.liability.group.style.display = 'none';
        formElements.maturity.group.style.display = 'block';
        formElements.settlement.group.style.display = 'none';
        formElements.assetName.group.style.display = 'block';
        formElements.assetName.field.required = true;
        
        formElements.liability.field.disabled = true;
        formElements.liability.field.required = false;
        formElements.maturity.type.disabled = false;
        formElements.maturity.date.disabled = false;
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'delivery_claim') {
        resetMaturityTypeOptions();
        formElements.assetName.group.style.display = 'block';
        formElements.assetName.field.required = true;
        formElements.settlement.field.value = 'non_financial_asset';
        formElements.settlement.field.disabled = true;

    } else if (entryType === 'bond' || entryType === 'government_bond_zero_coupon' || entryType === 'central_bank_bond_zero_coupon') {
        resetMaturityTypeOptions();
        
        // 显示债券类别选择
        if (formElements.bondCategory.group) {
            formElements.bondCategory.group.style.display = 'block';
            
            // 根据entry type预设债券类别
            if (entryType === 'government_bond_zero_coupon') {
                formElements.bondCategory.field.value = 'government';
                formElements.bondCategory.field.disabled = true; // 锁定选择
            } else if (entryType === 'central_bank_bond_zero_coupon') {
                formElements.bondCategory.field.value = 'central_bank';
                formElements.bondCategory.field.disabled = true; // 锁定选择
            } else {
                formElements.bondCategory.field.value = 'corporate';
                formElements.bondCategory.field.disabled = false; // 允许选择
            }
        }
        
        formElements.bondType.group.style.display = 'block';
        updateBondTypeOptions();
        updateBondFields();
        
    } else if (entryType === 'share') {
        limitMaturityTypeOptions(['perpetual']);
        formElements.maturity.type.value = 'perpetual';
        formElements.maturity.type.disabled = true;
        formElements.maturity.dateGroup.style.display = 'none';
        formElements.settlement.field.value = 'none';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'payable') {
        resetMaturityTypeOptions();
        formElements.settlement.field.value = 'means_of_payment';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'deposit' || entryType === 'bank_deposit') {
        limitMaturityTypeOptions(['on_demand', 'fixed_date']);
        
        if (!formElements.maturity.type.value || formElements.maturity.type.value === 'perpetual') {
            formElements.maturity.type.value = 'on_demand';
        }
        formElements.settlement.field.value = 'none';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'reserve_deposit') {  // 新增
        limitMaturityTypeOptions(['on_demand', 'fixed_date']);
        
        if (!formElements.maturity.type.value || formElements.maturity.type.value === 'perpetual') {
            formElements.maturity.type.value = 'on_demand';
        }
        formElements.settlement.field.value = 'none';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'intraday_iou' || entryType === 'interbank_intraday_iou') {
        limitMaturityTypeOptions(['fixed_date']);
        formElements.maturity.type.value = 'fixed_date';
        formElements.maturity.type.disabled = true;
        formElements.maturity.dateGroup.style.display = 'block';
        formElements.settlement.field.value = 'means_of_payment';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType === 'loan' || entryType === 'overnight_loan' || 
               entryType === 'bank_client_loan' || entryType === 'interbank_overnight_loan') {
        // NEW: Show interest rate field for loans and overnight loans
        resetMaturityTypeOptions();
        formElements.interestRate.group.style.display = 'block';
        formElements.settlement.field.value = 'means_of_payment';
        formElements.settlement.field.disabled = true;
        
    } else if (entryType.includes('bond')) {  // 处理具体的债券类型
        resetMaturityTypeOptions();
        formElements.bondType.group.style.display = 'block';
        
        // 根据具体的债券类型设置债券类别
        if (entryType.includes('government')) {
            if (formElements.bondCategory.field) {
                formElements.bondCategory.field.value = 'government';
            }
        } else if (entryType.includes('central_bank')) {
            if (formElements.bondCategory.field) {
                formElements.bondCategory.field.value = 'central_bank';
            }
        } else {
            if (formElements.bondCategory.field) {
                formElements.bondCategory.field.value = 'corporate';
            }
        }
        
        if (formElements.bondCategory.group) {
            formElements.bondCategory.group.style.display = 'block';
        }
        
        updateBondFields();
        
    } else {
        resetMaturityTypeOptions();
    }

    updateMaturityDateVisibility();
    updateDepositSettlementType();
}

function updateDepositSettlementType() {
    const entryType = document.getElementById('entryType').value;
    const maturityType = document.getElementById('maturityType').value;
    const settlementField = document.getElementById('settlementType');
    
    if ((entryType === 'deposit' || entryType === 'bank_deposit' || entryType === 'reserve_deposit') && maturityType === 'fixed_date') {
        settlementField.value = 'means_of_payment';
        settlementField.disabled = true;
    } else if (entryType === 'deposit' || entryType === 'bank_deposit' || entryType === 'reserve_deposit') {
        settlementField.value = 'none';
        settlementField.disabled = true;
    }
}

// Update bond-specific fields
function updateBondFields() {
    const bondType = document.getElementById('bondType').value;
    const couponRateGroup = document.getElementById('couponRateGroup');
    
    if (bondType === '1' || bondType === '2') {
        couponRateGroup.style.display = 'block';
        document.getElementById('couponRate').required = true;
    } else {
        couponRateGroup.style.display = 'none';
        document.getElementById('couponRate').required = false;
    }
}

// Update maturity date visibility
function updateMaturityDateVisibility() {
    const maturityType = document.getElementById('maturityType').value;
    const maturityDateGroup = document.getElementById('maturityDateGroup');
    
    if (maturityType === 'fixed_date') {
        maturityDateGroup.style.display = 'block';
    } else {
        maturityDateGroup.style.display = 'none';
    }
    
    updateDepositSettlementType();
}

// Initialize form on page load - 修改后的版本
document.addEventListener('DOMContentLoaded', () => {
    filterLiabilityHolders();
    updateFormFields();
    
    document.getElementById('assetHolder').addEventListener('change', filterLiabilityHolders);
    document.getElementById('entryType').addEventListener('change', updateFormFields);
    document.getElementById('maturityType').addEventListener('change', updateMaturityDateVisibility);
    document.getElementById('bondType').addEventListener('change', updateBondFields);
    
    // 新增：添加债券类别变化监听器
    const bondCategoryElement = document.getElementById('bondCategory');
    if (bondCategoryElement) {
        bondCategoryElement.addEventListener('change', updateBondTypeOptions);
    }
    
    // Validate agent name uniqueness
    const agentForm = document.querySelector('form[action="/create_agent"]');
    if (agentForm) {
        agentForm.addEventListener('submit', function(e) {
            const existingAgents = Array.from(document.querySelectorAll('#assetHolder option'))
                                       .map(opt => opt.value);
            const newName = document.querySelector('input[name="name"]').value;
            
            if (existingAgents.includes(newName)) {
                e.preventDefault();
                alert('Agent name already exists!');
            }
        });
    }
});
</script>
'''

# ======== Web路由处理 ========
@app.route('/')
def home():
    # 只保留高层次的债券类型
    available_entry_types = [
        EntryType.LOAN, 
        EntryType.DEPOSIT,
        EntryType.RESERVE_DEPOSIT,
        EntryType.PAYABLE,
        EntryType.BOND,  # 企业债券
        # 添加高层次的政府和央行债券类型
        EntryType.GOVERNMENT_BOND_ZERO_COUPON,  # 重用作为"政府债券"的代表
        EntryType.CENTRAL_BANK_BOND_ZERO_COUPON,  # 重用作为"央行债券"的代表
        EntryType.INTRADAY_IOU,
        EntryType.OVERNIGHT_LOAN,
        EntryType.SHARE,
        EntryType.DELIVERY_CLAIM, 
        EntryType.NON_FINANCIAL
    ]
    
    return render_template_string(
        HTML_BASE + HOME_CONTENT,
        agent_types=AgentType,
        agents=system.agents.values(),
        entry_types=available_entry_types,
        MaturityType=MaturityType,
        SettlementType=SettlementType,
        EntryType=EntryType,
        scheduled_actions=system.scheduled_actions
    )

@app.route('/create_agent', methods=['POST'])
def create_agent():
    name = request.form['name']
    agent_type = AgentType(request.form['type'])
    
    if name in system.agents:
        flash(f"Agent '{name}' already exists!", "danger")
        return redirect(url_for('home'))
    
    agent = Agent(name, agent_type)
    system.add_agent(agent)
    flash(f"Agent '{name}' created successfully!", "success")
    return redirect(url_for('home'))

# ENHANCED: create_pair route with interest rate and scheduling support
@app.route('/create_pair', methods=['POST'])
def create_pair():
    try:
        # Basic field validation
        asset_holder_name = request.form['asset_holder']
        entry_type_str = request.form['entry_type']
        amount = float(request.form['amount'])
        denomination = request.form.get('denomination', 'USD')
        
        # NEW: Get operation time point
        operation_time_point = int(request.form.get('operation_time_point', '0'))

        # Convert string to EntryType enum
        try:
            entry_type = EntryType(entry_type_str)
        except ValueError:
            raise ValueError(f"Invalid entry type: {entry_type_str}")

        # Get Agent instances
        asset_holder = system.agents.get(asset_holder_name)
        if not asset_holder:
            raise ValueError(f"Asset holder {asset_holder_name} not found")

        # Initialize parameters
        liability_holder = None
        asset_name = None
        settlement_type = SettlementType.NONE
        maturity_type = MaturityType.ON_DEMAND
        maturity_date = None
        bond_type = None
        coupon_rate = None
        cash_flow_at_maturity = 0

        # Handle non-financial assets
        if entry_type == EntryType.NON_FINANCIAL:
            asset_name = request.form.get('asset_name')
            settlement_type = SettlementType.NONE
            maturity_type = MaturityType.ON_DEMAND
            liability_holder = None
            if not asset_name:
                raise ValueError("Non-financial entries must have an asset name")
                
        else:
            # Financial entries must have liability holder
            liability_holder_name = request.form.get('liability_holder')
            if not liability_holder_name:
                raise ValueError("Liability holder is required for financial entries")
            
            liability_holder = system.agents.get(liability_holder_name)
            if not liability_holder:
                raise ValueError(f"Liability holder {liability_holder_name} not found")
            
            if liability_holder == asset_holder:
                raise ValueError("Asset holder and liability holder cannot be the same")

            # Handle different types of financial entries
            if entry_type == EntryType.DELIVERY_CLAIM:
                asset_name = request.form.get('asset_name')
                if not asset_name:
                    raise ValueError("Delivery claim requires asset name")
                settlement_type = SettlementType.NON_FINANCIAL_ASSET
                
            elif entry_type == EntryType.PAYABLE:
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                
            elif entry_type == EntryType.DEPOSIT:
                settlement_type = SettlementType.NONE
                maturity_type = MaturityType.ON_DEMAND
                
            elif entry_type == EntryType.RESERVE_DEPOSIT:
                settlement_type = SettlementType.NONE
                maturity_type = MaturityType.ON_DEMAND
                
            elif entry_type == EntryType.SHARE:
                settlement_type = SettlementType.NONE
                maturity_type = MaturityType.PERPETUAL
                
            elif entry_type == EntryType.INTRADAY_IOU:
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                maturity_type = MaturityType.FIXED_DATE
                
            elif entry_type in [EntryType.LOAN, EntryType.OVERNIGHT_LOAN]:
                settlement_type = SettlementType(request.form.get('settlement_type', SettlementType.MEANS_OF_PAYMENT.value))
                
                # Handle interest rate for loans
                interest_rate = float(request.form.get('interest_rate', 0))
                if interest_rate > 0:
                    cash_flow_at_maturity = amount * (1 + interest_rate)
                else:
                    cash_flow_at_maturity = amount
                
            # NEW: Handle bond types - simplified logic
            elif entry_type in [EntryType.BOND, EntryType.GOVERNMENT_BOND_ZERO_COUPON, EntryType.CENTRAL_BANK_BOND_ZERO_COUPON]:
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                
                # Determine bond category based on entry type
                if entry_type == EntryType.GOVERNMENT_BOND_ZERO_COUPON:
                    bond_category = 'government'
                elif entry_type == EntryType.CENTRAL_BANK_BOND_ZERO_COUPON:
                    bond_category = 'central_bank'
                else:  # EntryType.BOND
                    bond_category = request.form.get('bond_category', 'corporate')
                
                # Get bond type from form
                bond_type_value = int(request.form.get('bond_type', '0'))
                bond_type = BondType(bond_type_value)
                
                # Map to specific EntryType based on category and bond type
                if bond_category == 'government':
                    if bond_type == BondType.ZERO_COUPON:
                        entry_type = EntryType.GOVERNMENT_BOND_ZERO_COUPON
                    elif bond_type == BondType.COUPON:
                        entry_type = EntryType.GOVERNMENT_BOND_COUPON
                    elif bond_type == BondType.AMORTIZING:
                        entry_type = EntryType.GOVERNMENT_BOND_AMORTIZING
                elif bond_category == 'central_bank':
                    if bond_type == BondType.ZERO_COUPON:
                        entry_type = EntryType.CENTRAL_BANK_BOND_ZERO_COUPON
                    elif bond_type == BondType.COUPON:
                        entry_type = EntryType.CENTRAL_BANK_BOND_COUPON
                    elif bond_type == BondType.AMORTIZING:
                        entry_type = EntryType.CENTRAL_BANK_BOND_AMORTIZING
                else:  # corporate
                    if bond_type == BondType.ZERO_COUPON:
                        entry_type = EntryType.BOND_ZERO_COUPON
                    elif bond_type == BondType.COUPON:
                        entry_type = EntryType.BOND_COUPON
                    elif bond_type == BondType.AMORTIZING:
                        entry_type = EntryType.BOND_AMORTIZING
                
                # Get coupon rate for coupon and amortizing bonds
                if bond_type in [BondType.COUPON, BondType.AMORTIZING]:
                    coupon_rate = float(request.form.get('coupon_rate', 0))
                    if coupon_rate <= 0:
                        raise ValueError("Coupon and amortizing bonds require a positive interest rate")

            # Handle all other specific bond types (if any remain from old data)
            elif entry_type in [
                EntryType.BOND_ZERO_COUPON, EntryType.BOND_COUPON, EntryType.BOND_AMORTIZING,
                EntryType.GOVERNMENT_BOND_COUPON, EntryType.GOVERNMENT_BOND_AMORTIZING,
                EntryType.CENTRAL_BANK_BOND_COUPON, EntryType.CENTRAL_BANK_BOND_AMORTIZING
            ]:
                settlement_type = SettlementType.MEANS_OF_PAYMENT
                
                # Determine bond type from entry type
                if 'zero_coupon' in entry_type.value:
                    bond_type = BondType.ZERO_COUPON
                elif 'coupon' in entry_type.value and 'zero_coupon' not in entry_type.value:
                    bond_type = BondType.COUPON
                elif 'amortizing' in entry_type.value:
                    bond_type = BondType.AMORTIZING
                
                # Get coupon rate for coupon and amortizing bonds
                if bond_type in [BondType.COUPON, BondType.AMORTIZING]:
                    coupon_rate = float(request.form.get('coupon_rate', 0))
                    if coupon_rate <= 0:
                        raise ValueError("Coupon and amortizing bonds require a positive interest rate")

            # Handle maturity date for non-perpetual entries
            if entry_type != EntryType.SHARE:
                maturity_type = MaturityType(request.form.get('maturity_type', MaturityType.ON_DEMAND.value))
                if maturity_type == MaturityType.FIXED_DATE:
                    maturity_time_point = int(request.form.get('maturity_time_point', '1'))
                    maturity_date = maturity_time_point

            # Handle settlement type (if not already set)
            if settlement_type == SettlementType.NONE and entry_type not in [
                EntryType.DEPOSIT, EntryType.SHARE, EntryType.NON_FINANCIAL, EntryType.RESERVE_DEPOSIT
            ]:
                settlement_type = SettlementType(request.form.get('settlement_type', SettlementType.MEANS_OF_PAYMENT.value))

            # Calculate cash flow at maturity for bond types if not already set
            if cash_flow_at_maturity == 0:
                if bond_type:
                    if bond_type == BondType.ZERO_COUPON:
                        cash_flow_at_maturity = amount
                    elif bond_type == BondType.COUPON:
                        if coupon_rate:
                            cash_flow_at_maturity = amount * (1 + coupon_rate)
                        else:
                            cash_flow_at_maturity = amount
                    elif bond_type == BondType.AMORTIZING:
                        if coupon_rate and maturity_date:
                            if maturity_date == 1:
                                cash_flow_at_maturity = amount * (1 + coupon_rate)
                            else:
                                cash_flow_at_maturity = amount * (1 + coupon_rate * 1.5)
                        else:
                            cash_flow_at_maturity = amount
                else:
                    cash_flow_at_maturity = amount

        # NEW: Validate constraints before creating pair
        if liability_holder:
            is_valid, error_msg = AssetLiabilityConstraints.validate_asset_liability_pair(
                asset_holder, liability_holder, entry_type
            )
            if not is_valid:
                raise ValueError(f"Constraint violation: {error_msg}")

        # NEW: Handle scheduling vs immediate creation
        if operation_time_point == 0:
            # Immediate creation
            pair = AssetLiabilityPair(
                time=system.current_time_state,
                type=entry_type.value,
                amount=amount,
                denomination=denomination,
                maturity_type=maturity_type,
                maturity_date=maturity_date,
                settlement_type=settlement_type,
                settlement_denomination=denomination,
                asset_holder=asset_holder,
                liability_holder=liability_holder,
                asset_name=asset_name,
                bond_type=bond_type,
                coupon_rate=coupon_rate,
                cash_flow_at_maturity=cash_flow_at_maturity
            )

            system.create_asset_liability_pair(pair)
            
            # Create user-friendly success message
            if entry_type == EntryType.NON_FINANCIAL:
                flash(f"✅ Non-financial asset '{asset_name}' created for {asset_holder.name}!", "success")
            elif entry_type == EntryType.RESERVE_DEPOSIT:
                flash(f"✅ Reserve deposit of {amount} {denomination} created between {asset_holder.name} and {liability_holder.name}!", "success")
            elif 'government' in entry_type.value:
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                flash(f"✅ Government {bond_type_name} bond created successfully!", "success")
            elif 'central_bank' in entry_type.value:
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                flash(f"✅ Central bank {bond_type_name} bond created successfully!", "success")
            elif entry_type.value.startswith('bond'):
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                flash(f"✅ Corporate {bond_type_name} bond created successfully!", "success")
            else:
                entry_display = entry_type.value.replace('_', ' ').title()
                flash(f"✅ {entry_display} created successfully!", "success")
                
        else:
            # Schedule for future creation
            pair_data = {
                'type': entry_type.value,
                'amount': amount,
                'denomination': denomination,
                'maturity_type': maturity_type.value,
                'maturity_date': maturity_date,
                'settlement_type': settlement_type.value,
                'settlement_denomination': denomination,
                'asset_holder_name': asset_holder.name,
                'liability_holder_name': liability_holder.name if liability_holder else None,
                'asset_name': asset_name,
                'bond_type': bond_type.value if bond_type else None,
                'coupon_rate': coupon_rate,
                'cash_flow_at_maturity': cash_flow_at_maturity
            }
            
            system.schedule_asset_liability_creation(operation_time_point, pair_data)
            
            # Create user-friendly success message for scheduled actions
            entry_description = entry_type.value.replace('_', ' ')
            if entry_type == EntryType.NON_FINANCIAL:
                entry_description = f"non-financial asset '{asset_name}'"
            elif 'government' in entry_type.value:
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                entry_description = f"government {bond_type_name} bond"
            elif 'central_bank' in entry_type.value:
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                entry_description = f"central bank {bond_type_name} bond"
            elif entry_type.value.startswith('bond'):
                bond_type_name = bond_type.name.lower().replace('_', ' ') if bond_type else 'bond'
                entry_description = f"corporate {bond_type_name} bond"
            
            flash(f"📅 {entry_description.title()} scheduled for creation at t{operation_time_point}!", "success")
        
        return redirect(url_for('home'))

    except (KeyError, ValueError) as e:
        error_msg = str(e)
        if isinstance(e, KeyError):
            error_msg = f"Missing required field: {e}"
        
        # Provide specific error messages for common constraint violations
        if "Constraint violation" in error_msg:
            flash(f"❌ {error_msg}", "danger")
        elif "interest rate" in error_msg.lower():
            flash(f"⚠️ Interest Rate Error: {error_msg}", "warning")
        elif "asset name" in error_msg.lower():
            flash(f"⚠️ Asset Name Required: {error_msg}", "warning")
        elif "liability holder" in error_msg.lower():
            flash(f"⚠️ Liability Holder Error: {error_msg}", "warning")
        else:
            flash(f"❌ Validation Error: {error_msg}", "danger")
            
        app.logger.warning(f"Validation error in create_pair: {error_msg}")
        return redirect(url_for('home'))

    except Exception as e:
        app.logger.exception("Unexpected error in create_pair")
        flash("❌ Internal server error. Please contact support.", "danger")
        return redirect(url_for('home'))

# NEW: Route to view and manage scheduled actions
@app.route('/scheduled_actions')
def scheduled_actions():
    actions_by_time = {}
    for time_point, actions in system.scheduled_actions.items():
        actions_by_time[time_point] = []
        for action in actions:
            if action['type'] == 'create_asset_liability_pair':
                data = action['data']
                action_desc = {
                    'type': 'Asset-Liability Pair Creation',
                    'details': f"{data['type']} - {data['amount']} {data['denomination']} from {data['asset_holder_name']} to {data.get('liability_holder_name', 'None')}",
                    'action_data': data
                }
                actions_by_time[time_point].append(action_desc)
            else:
                actions_by_time[time_point].append({
                    'type': action['type'],
                    'details': str(action.get('params', {})),
                    'action_data': action
                })
    
    return render_template_string(
        HTML_BASE + '''
        <div class="card">
            <div class="card-header">
                <h5>Scheduled Actions</h5>
            </div>
            <div class="card-body">
                {% if actions_by_time %}
                    {% for time_point, actions in actions_by_time.items() %}
                        <h6>Time Point t{{ time_point }}:</h6>
                        <div class="list-group mb-3">
                            {% for action in actions %}
                                <div class="list-group-item d-flex justify-content-between align-items-center">
                                    <div>
                                        <strong>{{ action.type }}</strong><br>
                                        <small class="text-muted">{{ action.details }}</small>
                                    </div>
                                    <form action="/cancel_scheduled_action" method="post" class="d-inline">
                                        <input type="hidden" name="time_point" value="{{ time_point }}">
                                        <input type="hidden" name="action_index" value="{{ loop.index0 }}">
                                        <button type="submit" class="btn btn-danger btn-sm" 
                                                onclick="return confirm('Cancel this scheduled action?')">
                                            Cancel
                                        </button>
                                    </form>
                                </div>
                            {% endfor %}
                        </div>
                    {% endfor %}
                {% else %}
                    <p class="text-muted">No scheduled actions.</p>
                {% endif %}
            </div>
        </div>
        ''',
        actions_by_time=actions_by_time
    )

@app.route('/cancel_scheduled_action', methods=['POST'])
def cancel_scheduled_action():
    try:
        time_point = int(request.form['time_point'])
        action_index = int(request.form['action_index'])
        
        if time_point in system.scheduled_actions:
            if 0 <= action_index < len(system.scheduled_actions[time_point]):
                removed_action = system.scheduled_actions[time_point].pop(action_index)
                if not system.scheduled_actions[time_point]:  # Remove empty time point
                    del system.scheduled_actions[time_point]
                flash("Scheduled action cancelled successfully!", "success")
            else:
                flash("Invalid action index!", "danger")
        else:
            flash("Time point not found!", "danger")
            
    except (ValueError, KeyError) as e:
        flash(f"Error cancelling action: {e}", "danger")
    
    return redirect(url_for('scheduled_actions'))

@app.route('/agents')
def list_agents():
    agents_html = "<ul class='list-group'>"
    for agent in system.agents.values():
        agents_html += f'''
        <li class="list-group-item">
            <h5><a href="/agent/{agent.name}" class="text-decoration-none">{agent.name}</a> ({agent.type.value})</h5>
            <div class="row">
                <div class="col">
                    <h6>Assets</h6>
                    <ul class="list-group">'''
        
        for a in agent.assets:
            entry_name = a.name + ": " if a.type == EntryType.NON_FINANCIAL else ""
            entry_type = a.type.value.replace('_', ' ').title()
            # Show both initial and current book value if different
            value_display = f"{a.initial_book_value}"
            if a.current_book_value != a.initial_book_value:
                value_display = f"{a.initial_book_value} (current: {a.current_book_value:.2f})"
            
            agents_html += f'''<li class="list-group-item">
                {entry_name}{value_display} {a.denomination} ({entry_type}) 
                [issued at t{a.issuance_time}]</li>'''
        
        agents_html += '''</ul>
                </div>
                <div class="col">
                    <h6>Liabilities</h6>
                    <ul class="list-group">'''
        
        for l in agent.liabilities:
            entry_type = l.type.value.replace('_', ' ').title()
            # Show both initial and current book value if different
            value_display = f"{l.initial_book_value}"
            if l.current_book_value != l.initial_book_value:
                value_display = f"{l.initial_book_value} (current: {l.current_book_value:.2f})"
                
            agents_html += f'''<li class="list-group-item">
                {value_display} {l.denomination} ({entry_type}) 
                [issued at t{l.issuance_time}]</li>'''
        
        agents_html += '''</ul>
                </div>
            </div>
        </li>
        '''
    agents_html += "</ul>"
    
    return render_template_string(HTML_BASE + '''
        <div class="form-section">
            <h4>All Agents</h4>
            ''' + agents_html + '''
        </div>
    ''')

@app.route('/agent/<name>')
def agent_detail(name):
    agent = system.agents.get(name)
    if not agent:
        return "Agent not found", 404
    
    session_key = f'edit_mode_{name}'
    edit_mode = session.get(session_key, False)
    
    return render_template_string(
        AGENT_TEMPLATE,
        agent=agent,
        edit_mode=edit_mode,
        EntryType=EntryType
    )

AGENT_TEMPLATE = HTML_BASE + '''
<div class="card mb-3">
    <div class="card-header d-flex justify-content-between align-items-center"> 
        <h5>
        <a href="{{ url_for('agent_detail', name=agent.name) }}" class="text-decoration-none">
            {{ agent.name }}
        </a> 
        ({{ agent.type.value }})
        </h5>
        <div>
            <a href="{{ url_for('toggle_edit_mode', name=agent.name) }}" 
               class="btn btn-sm {{ 'btn-warning' if edit_mode else 'btn-outline-secondary' }}">
                {{ 'Exit Edit' if edit_mode else 'Edit' }}
            </a>
            {% if edit_mode and agent.assets|length == 0 and agent.liabilities|length == 0 %}
            <form action="{{ url_for('delete_agent', name=agent.name) }}" 
                  method="post" 
                  class="d-inline ms-2">
                <button type="submit" 
                        class="btn btn-sm btn-danger"
                        onclick="return confirm('Are you sure you want to delete this agent?')"> 
                    Delete Agent
                </button>
            </form>
            {% endif %}
        </div>
    </div>

    <div class="card-body">
        <div class="row">
            <div class="col-md-6">
                <h6 class="mb-3">Assets</h6>
                <ul class="list-group">
                    {% for asset in agent.assets %}
                    <li class="list-group-item d-flex justify-content-between align-items-center">
                        <div class="me-3">
                            <span class="badge bg-primary me-2">{{ asset.type.value|upper|replace('_', ' ') }}</span>
                            {{ asset.initial_book_value }} {{ asset.denomination }}
                            {% if asset.current_book_value != asset.initial_book_value %}
                            <small class="text-muted">(Current: {{ asset.current_book_value|round(2) }})</small>
                            {% endif %}
                            {% if asset.cash_flow_at_maturity and asset.cash_flow_at_maturity != asset.initial_book_value %}
                            <small class="text-info">(CF@Mat: {{ asset.cash_flow_at_maturity|round(2) }})</small>
                            {% endif %}
                            {% if asset.counterparty %}
                            <div class="text-muted small mt-1">From {{ asset.counterparty }}</div>
                            {% endif %}
                            {% if asset.name %}
                            <div class="text-muted small">{{ asset.name }}</div>
                            {% endif %}
                            {% if asset.maturity_type.value == 'fixed_date' %}
                            <div class="text-muted small">Matures at t{{ asset.maturity_date }}</div>
                            {% endif %}
                        </div>
                        {% if edit_mode %}
                        <form action="{{ url_for('delete_entry') }}" 
                              method="post"
                              onsubmit="return confirm('Are you sure you want to delete this asset?')">
                            <input type="hidden" name="entry_type" value="asset">
                            <input type="hidden" name="agent_name" value="{{ agent.name }}">
                            <input type="hidden" name="amount" value="{{ asset.initial_book_value }}">
                            <input type="hidden" name="denomination" value="{{ asset.denomination }}">
                            <input type="hidden" name="counterparty" value="{{ asset.counterparty or '' }}">
                            <input type="hidden" name="entry_type_value" value="{{ asset.type.value }}">
                            <button type="submit" 
                                    class="btn btn-danger btn-sm"
                                    title="Delete Asset">
                                &times;
                            </button>
                        </form>
                        {% endif %}
                    </li>
                    {% endfor %}
                </ul>
            </div>

            <div class="col-md-6">
                <h6 class="mb-3">Liabilities</h6>
                <ul class="list-group">
                    {% for liability in agent.liabilities %}
                    <li class="list-group-item d-flex justify-content-between align-items-center">
                        <div class="me-3">
                            <span class="badge bg-danger me-2">{{ liability.type.value|upper|replace('_', ' ') }}</span>
                            {{ liability.initial_book_value }} {{ liability.denomination }}
                            {% if liability.current_book_value != liability.initial_book_value %}
                            <small class="text-muted">(Current: {{ liability.current_book_value|round(2) }})</small>
                            {% endif %}
                            {% if liability.cash_flow_at_maturity and liability.cash_flow_at_maturity != liability.initial_book_value %}
                            <small class="text-info">(CF@Mat: {{ liability.cash_flow_at_maturity|round(2) }})</small>
                            {% endif %}
                            <div class="text-muted small mt-1">To {{ liability.counterparty }}</div>
                            {% if liability.name %}
                            <div class="text-muted small">{{ liability.name }}</div>
                            {% endif %}
                            {% if liability.maturity_type.value == 'fixed_date' %}
                            <div class="text-muted small">Matures at t{{ liability.maturity_date }}</div>
                            {% endif %}
                        </div>
                        {% if edit_mode %}
                        <form action="{{ url_for('delete_entry') }}" 
                              method="post"
                              onsubmit="return confirm('Are you sure you want to delete this liability?')">
                            <input type="hidden" name="entry_type" value="liability">
                            <input type="hidden" name="agent_name" value="{{ agent.name }}">
                            <input type="hidden" name="amount" value="{{ liability.initial_book_value }}">
                            <input type="hidden" name="denomination" value="{{ liability.denomination }}">
                            <input type="hidden" name="counterparty" value="{{ liability.counterparty or '' }}">
                            <input type="hidden" name="entry_type_value" value="{{ liability.type.value }}">
                            <button type="submit" 
                                    class="btn btn-danger btn-sm"
                                    title="Delete Liability">
                                &times;
                            </button>
                        </form>
                        {% endif %}
                    </li>
                    {% endfor %}
                </ul>
            </div>
        </div>
        
        <div class="mt-3">
            <div class="row">
                <div class="col-md-4">
                    <strong>Total Assets:</strong> {{ agent.get_total_assets()|round(2) }}
                </div>
                <div class="col-md-4">
                    <strong>Total Liabilities:</strong> {{ agent.get_total_liabilities()|round(2) }}
                </div>
                <div class="col-md-4">
                    <strong>Net Worth:</strong> {{ agent.get_net_worth()|round(2) }}
                </div>
            </div>
        </div>
    </div>
</div>
'''

@app.route('/settlement_history')
def settlement_history():
    agents = system.agents.values()
    return render_template_string(
        HTML_BASE + '''
        <div class="card">
            <div class="card-header">
                <h5>All Agents' Settlement History</h5>
            </div>
            <div class="card-body">
                {% for agent in agents %}
                <div class="mb-4">
                    <h6>{{ agent.name }} ({{ agent.type.value }})</h6>
                    <div class="ms-4">
                        <h7>As Asset Holder:</h7>
                        {% for record in agent.settlement_history['as_asset_holder'] %}
                        <div class="alert alert-secondary mt-2">
                            <p class="mb-1"><strong>Time:</strong> t{{ record.time_point }}</p>
                            <p class="mb-1"><strong>Type:</strong> {{ record.original_entry.type.value|replace('_', ' ')|title }}</p>
                            <p class="mb-1"><strong>Amount:</strong> {{ record.original_entry.initial_book_value }} {{ record.original_entry.denomination }}</p>
                            <p class="mb-0"><strong>Settled:</strong> {{ record.settlement_result.type.value|replace('_', ' ')|title }} ({{ record.settlement_result.initial_book_value }} {{ record.settlement_result.denomination }})</p>
                        </div>
                        {% else %}
                        <div class="alert alert-light">No asset holder settlements</div>
                        {% endfor %}

                        <h7 class="mt-3">As Liability Holder:</h7>
                        {% for record in agent.settlement_history['as_liability_holder'] %}
                        <div class="alert alert-warning mt-2">
                            <p class="mb-1"><strong>Time:</strong> t{{ record.time_point }}</p>
                            <p class="mb-1"><strong>Type:</strong> {{ record.original_entry.type.value|replace('_', ' ')|title }}</p>
                            <p class="mb-1"><strong>Amount:</strong> {{ record.original_entry.initial_book_value }} {{ record.original_entry.denomination }}</p>
                            <p class="mb-0"><strong>Settled:</strong> {{ record.settlement_result.type.value|replace('_', ' ')|title }} ({{ record.settlement_result.initial_book_value }} {{ record.settlement_result.denomination }})</p>
                        </div>
                        {% else %}
                        <div class="alert alert-light">No liability holder settlements</div>
                        {% endfor %}
                    </div>
                </div>
                <hr>
                {% endfor %}
            </div>
        </div>
        ''',
        agents=agents
    )

# ENHANCED: run_simulation with improved error handling
@app.route('/simulate', methods=['POST'])
def run_simulation():
    try:
        result = system.run_simulation()

        if isinstance(result, dict):
            system.agents = result
            success = True
        else:
            success = result

        if success:
            flash("Simulation completed successfully!", "success")
        else:
            flash("Simulation failed due to payment defaults", "danger")
            # Reset to t0 state after failure
            if 0 in system.time_states:
                for agent_name, agent_state in system.time_states[0].items():
                    if agent_name in system.agents:
                        system.agents[agent_name].assets = deepcopy(agent_state.assets)
                        system.agents[agent_name].liabilities = deepcopy(agent_state.liabilities)
                        system.agents[agent_name].settlement_history = deepcopy(agent_state.settlement_history)
            
            system.current_time_state = 0
            flash("System automatically reset to t0 after failure", "info")
            
        return redirect('/')
    except Exception as e:
        flash(f"Simulation error: {str(e)}", "danger")
        # Reset to t0 state on exception
        if 0 in system.time_states:
            for agent_name, agent_state in system.time_states[0].items():
                if agent_name in system.agents:
                    system.agents[agent_name].assets = deepcopy(agent_state.assets)
                    system.agents[agent_name].liabilities = deepcopy(agent_state.liabilities)
                    system.agents[agent_name].settlement_history = deepcopy(agent_state.settlement_history)
        
        system.current_time_state = 0
        flash("System reset to t0 due to error", "info")
        return redirect('/')

@app.route('/export')
def export_data():
    if not EXCEL_AVAILABLE:
        return "Excel support not enabled. Install openpyxl first.", 503
    
    try:
        output = BytesIO()
        exporter = ExcelExporter(system)
        exporter.export_balance_sheets(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name="economic_simulation.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        app.logger.error(f"Export failed: {str(e)}")
        return f"Export failed: {str(e)}", 500

@app.context_processor
def inject_enums():
    return {
        'agent_types': AgentType,
        'entry_types': [et for et in EntryType if et != EntryType.DEFAULT],
        'settlement_types': SettlementType,
        'maturity_types': MaturityType
    }

@app.route('/toggle_edit_mode/<name>')
def toggle_edit_mode(name):
    session_key = f'edit_mode_{name}'
    session[session_key] = not session.get(session_key, False) 
    return redirect(url_for('agent_detail', name=name))

@app.route('/delete_entry', methods=['POST'])
def delete_entry():
    try:
        agent_name = request.form['agent_name']
        entry_type = request.form['entry_type']  # asset/liability
        amount = float(request.form['amount'])
        denomination = request.form['denomination']
        counterparty = request.form.get('counterparty', '').strip() or None
        entry_type_value = EntryType(request.form['entry_type_value'])
        
        agent = system.agents.get(agent_name)
        if not agent:
            flash("Agent not found", "danger")
            return redirect(url_for('home'))

        target_entry = None
        if entry_type == 'asset':
            target_entry = next((a for a in agent.assets 
                               if a.initial_book_value == amount
                               and a.denomination == denomination
                               and a.counterparty == counterparty
                               and a.type == entry_type_value), None)
            if target_entry:
                if entry_type_value != EntryType.NON_FINANCIAL and counterparty:
                    counterpart = system.agents.get(counterparty)
                    if counterpart:
                        counterpart_liability = next((l for l in counterpart.liabilities 
                                                     if l.initial_book_value == amount
                                                     and l.denomination == denomination
                                                     and l.counterparty == agent_name
                                                     and l.type == entry_type_value), None)
                        if counterpart_liability:
                            counterpart.remove_liability(counterpart_liability)
                agent.remove_asset(target_entry)
                flash("Asset deleted successfully", "success")
        elif entry_type == 'liability':
            target_entry = next((l for l in agent.liabilities 
                               if l.initial_book_value == amount
                               and l.denomination == denomination
                               and l.counterparty == counterparty
                               and l.type == entry_type_value), None)
            if target_entry:
                if entry_type_value != EntryType.NON_FINANCIAL and counterparty:
                    counterpart = system.agents.get(counterparty)
                    if counterpart:
                        counterpart_asset = next((a for a in counterpart.assets 
                                                if a.initial_book_value == amount
                                                and a.denomination == denomination
                                                and a.counterparty == agent_name
                                                and a.type == entry_type_value), None)
                        if counterpart_asset:
                            counterpart.remove_asset(counterpart_asset)
                agent.remove_liability(target_entry)
                flash("Liability deleted successfully", "success")
        
        if target_entry:
            system.asset_liability_pairs = [
                pair for pair in system.asset_liability_pairs
                if not (
                    (pair.asset_holder.name == agent_name and
                     any(a.matches(target_entry) for a in pair.asset_holder.assets)) or
                    (pair.liability_holder and 
                     pair.liability_holder.name == agent_name and
                     any(l.matches(target_entry) for l in pair.liability_holder.liabilities))
                )
            ]
        else:
            flash("Entry not found", "danger")
        
        return redirect(url_for('agent_detail', name=agent_name))
    
    except Exception as e:
        flash(f"Error deleting entry: {str(e)}", "danger")
        return redirect(url_for('home'))

@app.route('/delete_agent/<name>', methods=['POST'])
def delete_agent(name):
    agent = system.agents.get(name)
    if not agent:
        flash("Agent not found", "danger")
        return redirect(url_for('home'))
    
    if len(agent.assets) > 0 or len(agent.liabilities) > 0:
        flash("Cannot delete agent with existing assets/liabilities", "danger")
        return redirect(url_for('agent_detail', name=name))
    
    for time_point in system.time_states:
        if name in system.time_states[time_point]:
            del system.time_states[time_point][name]
    
    del system.agents[name]
    flash("Agent deleted successfully", "success")
    return redirect(url_for('home'))

# Configuration management routes remain unchanged
CONFIG_DIR = Path("configs")
CONFIG_DIR.mkdir(exist_ok=True)

@app.route('/config')
def config_manager():
    configs = []
    for config_file in CONFIG_DIR.glob("*.json"):
        try:
            with open(config_file) as f:
                config_data = json.load(f)
                configs.append({
                    "id": config_file.stem,
                    "name": config_data["name"],
                    "timestamp": config_data["timestamp"],
                    "agent_count": len(config_data["system"]["agents"]),
                    "pair_count": len(config_data["system"]["asset_liability_pairs"])
                })
        except (json.JSONDecodeError, KeyError) as e:
            app.logger.error(f"Error loading config {config_file}: {e}")
            continue
    return render_template_string(HTML_BASE + '''
    <div class="card">
        <div class="card-header d-flex justify-content-between">
            <h4>Saved Configurations</h4>
            <a href="/save_config" class="btn btn-primary">Save Current Config</a>
        </div>
        <div class="card-body">
            <table class="table">
                <thead>
                    <tr>
                        <th>Name</th>
                        <th>Saved At</th>
                        <th>Agents</th>
                        <th>Pairs</th>
                        <th>Actions</th>
                    </tr>
                </thead>
                <tbody>
                    {% for config in configs %}
                    <tr>
                        <td>{{ config.name }}</td>
                        <td>{{ config.timestamp }}</td>
                        <td>{{ config.agent_count }}</td>
                        <td>{{ config.pair_count }}</td>
                        <td>
                            <form action="/load_config/{{ config.id }}" method="post" class="d-inline">
                                <button type="submit" class="btn btn-sm btn-success">Load</button>
                            </form>
                            <form action="/delete_config/{{ config.id }}" method="post" class="d-inline">
                                <button type="submit" class="btn btn-sm btn-danger">Delete</button>
                            </form>
                        </td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
    </div>
    ''', configs=configs)

@app.route('/save_config', methods=['GET', 'POST'])
def save_config():
    if request.method == 'GET':
        return render_template_string(HTML_BASE + '''
        <div class="card">
            <div class="card-header">
                <h4>Save Current Configuration</h4>
            </div>
            <div class="card-body">
                <form method="POST">
                    <div class="mb-3">
                        <label class="form-label">Configuration Name</label>
                        <input type="text" name="config_name" class="form-control" required>
                    </div>
                    <button type="submit" class="btn btn-primary">Save</button>
                </form>
            </div>
        </div>
        ''')
    
    config_name = request.form["config_name"]
    config_id = str(uuid.uuid4())
    
    config_data = {
        "name": config_name,
        "timestamp": datetime.now().isoformat(),
        "system": system.to_dict()
    }
    
    with open(CONFIG_DIR / f"{config_id}.json", "w") as f:
        json.dump(config_data, f, indent=2)
    
    return redirect("/config")

@app.route('/load_config/<config_id>', methods=['POST'])
def load_config(config_id):
    config_file = CONFIG_DIR / f"{config_id}.json"
    with open(config_file) as f:
        config_data = json.load(f)
    
    # Clear current system
    system.agents.clear()
    system.asset_liability_pairs.clear()
    system.time_states.clear()
    system.scheduled_actions.clear()  # NEW: Clear scheduled actions
    
    # Rebuild from config
    new_system = EconomicSystem.from_dict(config_data["system"])
    system.agents = new_system.agents
    system.asset_liability_pairs = new_system.asset_liability_pairs
    system.time_states = new_system.time_states
    system.current_time_state = new_system.current_time_state
    system.simulation_finalized = new_system.simulation_finalized
    system.scheduled_actions = new_system.scheduled_actions  # NEW: Load scheduled actions
    system.save_state(0)
    
    return redirect("/")

@app.route('/delete_config/<config_id>', methods=['POST'])
def delete_config(config_id):
    config_file = CONFIG_DIR / f"{config_id}.json"
    config_file.unlink()
    return redirect("/config")

@app.route('/output')
def output_interface():
    # Render HTML template with React component (unchanged from original)
    output_template = '''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Balance Sheet Output</title>
        <link href="https://cdn.jsdelivr.net/npm/tailwindcss@2.2.19/dist/tailwind.min.css" rel="stylesheet">
        <script src="https://unpkg.com/react@17/umd/react.development.js"></script>
        <script src="https://unpkg.com/react-dom@17/umd/react-dom.development.js"></script>
        <script src="https://unpkg.com/@babel/standalone/babel.min.js"></script>
    </head>
    <body>
        <div id="balance-sheet-output-app"></div>
        
        <script type="text/babel">
            // HoverDetails component for micro interactions with fixed positioning
            const HoverDetails = ({ entry, targetRect, onClose }) => {
              if (!entry) return null;
              
              const tooltipRef = React.useRef(null);
              
              React.useEffect(() => {
                if (tooltipRef.current && targetRect) {
                  const tooltip = tooltipRef.current;
                  const tooltipRect = tooltip.getBoundingClientRect();
                  
                  let top = targetRect.top - tooltipRect.height - 10;
                  let left = targetRect.left + (targetRect.width / 2) - (tooltipRect.width / 2);
                  
                  if (top < 10) top = targetRect.bottom + 10;
                  if (left < 10) left = 10;
                  if (left + tooltipRect.width > window.innerWidth - 10) {
                    left = window.innerWidth - tooltipRect.width - 10;
                  }
                  
                  tooltip.style.top = `${top}px`;
                  tooltip.style.left = `${left}px`;
                }
              }, [targetRect]);
              
              return (
                <div 
                  ref={tooltipRef}
                  className="fixed z-10 bg-white p-4 shadow-lg rounded-md border border-gray-300 w-80"
                  style={{ top: 0, left: 0 }}
                >
                  <button 
                    onClick={onClose}
                    className="absolute top-2 right-2 text-gray-400 hover:text-gray-600 text-lg font-bold"
                  >
                    ×
                  </button>
                  
                  <h4 className="font-bold mb-3 border-b pb-2 text-lg">Entry Details</h4>
                  
                  <div className="text-sm space-y-2">
                    <div className="grid grid-cols-3 gap-1">
                      <span className="font-medium">Type:</span>
                      <span className="col-span-2">{entry.type}</span>
                    </div>
                    
                    <div className="grid grid-cols-3 gap-1">
                      <span className="font-medium">Amount:</span>
                      <span className="col-span-2">{entry.amount} {entry.denomination}</span>
                    </div>
                    
                    {entry.current_book_value && entry.current_book_value !== entry.amount && (
                      <div className="grid grid-cols-3 gap-1">
                        <span className="font-medium">Current Value:</span>
                        <span className="col-span-2">{entry.current_book_value.toFixed(2)} {entry.denomination}</span>
                      </div>
                    )}
                    
                    {entry.cash_flow_at_maturity && entry.cash_flow_at_maturity !== entry.amount && (
                      <div className="grid grid-cols-3 gap-1">
                        <span className="font-medium">Cash Flow @ Mat:</span>
                        <span className="col-span-2">{entry.cash_flow_at_maturity.toFixed(2)} {entry.denomination}</span>
                      </div>
                    )}
                    
                    {entry.counterparty && (
                      <div className="grid grid-cols-3 gap-1">
                        <span className="font-medium">Counterparty:</span>
                        <span className="col-span-2">{entry.counterparty}</span>
                      </div>
                    )}
                    
                    <div className="grid grid-cols-3 gap-1">
                      <span className="font-medium">Maturity:</span>
                      <span className="col-span-2">{entry.maturity_type}</span>
                    </div>
                    
                    {entry.maturity_date && (
                      <div className="grid grid-cols-3 gap-1">
                        <span className="font-medium">Maturity Date:</span>
                        <span className="col-span-2">t{entry.maturity_date}</span>
                      </div>
                    )}
                    
                    <div className="grid grid-cols-3 gap-1">
                      <span className="font-medium">Issuance:</span>
                      <span className="col-span-2">t{entry.issuance_time}</span>
                    </div>
                    
                    {entry.name && (
                      <div className="grid grid-cols-3 gap-1">
                        <span className="font-medium">Name:</span>
                        <span className="col-span-2">{entry.name}</span>
                      </div>
                    )}
                  </div>
                </div>
              );
            };

            const BalanceSheetOutputInterface = () => {
              // State management
              const [timePoints, setTimePoints] = React.useState(['t0', 't2']); 
              const [scopeOption, setScopeOption] = React.useState('all-agents-single-time');
              const [mappingType, setMappingType] = React.useState('full');
              const [showLabels, setShowLabels] = React.useState(true);
              const [highlightCounterparty, setHighlightCounterparty] = React.useState(false);
              
              const [selectedAgentName, setSelectedAgentName] = React.useState('');
              const [selectedSingleTimePoint, setSelectedSingleTimePoint] = React.useState('t0');
              
              const [hoverEntry, setHoverEntry] = React.useState(null);
              const [targetRect, setTargetRect] = React.useState(null);
              const [linkedEntries, setLinkedEntries] = React.useState([]);
              
              const [loading, setLoading] = React.useState(true);
              const [agentData, setAgentData] = React.useState({});
              const [availableTimePoints, setAvailableTimePoints] = React.useState([]);
              const [availableAgents, setAvailableAgents] = React.useState([]);
              
              const isSingleTimeMode = scopeOption.endsWith('single-time');
              
              // Load data on component mount
              React.useEffect(() => {
                setLoading(true);
                fetch('/api/balance-sheet-data')
                  .then(response => response.json())
                  .then(data => {
                    setAgentData(data.agents_by_time);
                    setAvailableTimePoints(data.time_points);
                    
                    const firstTimePoint = data.time_points[0];
                    const agentsAtFirstTime = data.agents_by_time[firstTimePoint] || {};
                    const agentNames = Object.keys(agentsAtFirstTime);
                    setAvailableAgents(agentNames);
                    
                    if (agentNames.length > 0) {
                      setSelectedAgentName(agentNames[0]);
                    }
                    
                    setTimePoints([data.time_points[0], data.time_points[data.time_points.length - 1]]);
                    setSelectedSingleTimePoint(data.time_points[0]);
                    
                    setLoading(false);
                  })
                  .catch(error => {
                    console.error('Error loading balance sheet data:', error);
                    setLoading(false);
                  });
              }, []);
              
              React.useEffect(() => {
                if (isSingleTimeMode && mappingType === 'change-only') {
                  setMappingType('full');
                }
              }, [scopeOption, isSingleTimeMode]);
              
              React.useEffect(() => {
                if (scopeOption.endsWith('single-time')) {
                  setTimePoints([selectedSingleTimePoint]);
                } else if (timePoints.length === 1) {
                  setTimePoints([timePoints[0], 't2']);
                }
              }, [scopeOption, selectedSingleTimePoint]);
              
              const handleScopeChange = (e) => {
                const newScope = e.target.value;
                setScopeOption(newScope);
                
                if (newScope.endsWith('single-time')) {
                  setTimePoints([selectedSingleTimePoint]);
                } else {
                  setTimePoints(['t0', 't2']);
                }
              };
              
              const addTimePoint = (point) => {
                if (!timePoints.includes(point)) {
                  const newTimePoints = [...timePoints];
                  if (point === 't1' && timePoints.includes('t0') && timePoints.includes('t2')) {
                    newTimePoints.splice(1, 0, 't1');
                  } else {
                    newTimePoints.push(point);
                    newTimePoints.sort();
                  }
                  setTimePoints(newTimePoints);
                }
              };
              
              const removeTimePoint = (point) => {
                if (scopeOption.endsWith('single-time')) {
                  return;
                }
                
                if (timePoints.length > 2) {
                  setTimePoints(timePoints.filter(p => p !== point));
                }
              };
              
              const expandToFullTimeline = () => {
                if (scopeOption.endsWith('single-time')) {
                  return;
                }
                setTimePoints(['t0', 't1', 't2']);
              };
              
              const handleSingleTimePointChange = (e) => {
                const newTimePoint = e.target.value;
                setSelectedSingleTimePoint(newTimePoint);
                
                if (scopeOption.endsWith('single-time')) {
                  setTimePoints([newTimePoint]);
                }
              };
              
              const getTotal = (entries) => {
                return entries.reduce((sum, entry) => sum + (entry.current_book_value || entry.amount), 0);
              };
              
              const computeChanges = (agent, fromTime, toTime) => {
                const getEntriesAtTime = (entries, time) => {
                  const timeInt = parseInt(time.slice(1));
                  return entries.filter(entry => 
                    entry.issuance_time <= timeInt && 
                    (entry.maturity_type !== "fixed_date" || 
                      (entry.maturity_date === null || timeInt < entry.maturity_date))
                  );
                };
                
                const fromTimeInt = parseInt(fromTime.slice(1));
                const toTimeInt = parseInt(toTime.slice(1));
                
                const fromAssets = getEntriesAtTime(agent.assets, fromTime);
                const toAssets = getEntriesAtTime(agent.assets, toTime);
                const fromLiabilities = getEntriesAtTime(agent.liabilities, fromTime);
                const toLiabilities = getEntriesAtTime(agent.liabilities, toTime);
                
                const newAssets = toAssets.filter(toEntry => 
                  !fromAssets.some(fromEntry => 
                    fromEntry.type === toEntry.type && 
                    fromEntry.amount === toEntry.amount && 
                    fromEntry.counterparty === toEntry.counterparty
                  )
                );
                
                const removedAssets = fromAssets.filter(fromEntry => 
                  !toAssets.some(toEntry => 
                    toEntry.type === fromEntry.type && 
                    toEntry.amount === fromEntry.amount && 
                    toEntry.counterparty === fromEntry.counterparty
                  )
                );
                
                const newLiabilities = toLiabilities.filter(toEntry => 
                  !fromLiabilities.some(fromEntry => 
                    fromEntry.type === toEntry.type && 
                    toEntry.amount === fromEntry.amount && 
                    toEntry.counterparty === fromEntry.counterparty
                  )
                );
                
                const removedLiabilities = fromLiabilities.filter(fromEntry => 
                  !toLiabilities.some(toEntry => 
                    toEntry.type === fromEntry.type && 
                    toEntry.amount === fromEntry.amount && 
                    toEntry.counterparty === fromEntry.counterparty
                  )
                );
                
                return {
                  newAssets,
                  removedAssets,
                  newLiabilities,
                  removedLiabilities
                };
              };
              
              const formatEntryType = (type, isAsset) => {
                let formatted = type.replace(/_/g, ' ');
                
                if (type === 'payable' && isAsset) {
                  return 'receivable';
                } else if (type === 'delivery_claim' && !isAsset) {
                  return 'delivery promise';
                }
                return formatted;
              };
              
              const renderTChart = (agent, timePoint) => {
                const timeIdx = parseInt(timePoint.slice(1));

                const filteredAssets = agent.assets.filter(asset =>
                  asset.issuance_time <= timeIdx && 
                  (asset.maturity_type !== "fixed_date" || 
                    (asset.maturity_date === null || timeIdx < asset.maturity_date))
                );

                const filteredLiabilities = agent.liabilities.filter(liability =>
                  liability.issuance_time <= timeIdx && 
                  (liability.maturity_type !== "fixed_date" || 
                    (liability.maturity_date === null || timeIdx < liability.maturity_date))
                );
                
                const totalAssets = getTotal(filteredAssets);
                const totalLiabilities = getTotal(filteredLiabilities);
                const netWorth = totalAssets - totalLiabilities;
                
                return (
                  <div className="flex flex-col p-4 bg-white rounded-lg shadow-md mb-6 border border-gray-300">
                    {showLabels && (
                      <>
                        <h3 className="text-lg font-bold text-center mb-1">{agent.name} ({agent.type})</h3>
                        <p className="text-sm text-center mb-3">Time: {timePoint}</p>
                      </>
                    )}
                    
                    <div className="flex">
                      {/* Assets Side */}
                      <div className="w-1/2 border-r border-gray-300">
                        <div className="p-2 font-bold text-center bg-blue-100">ASSETS</div>
                        <div className="grid grid-cols-4 gap-1 p-1 text-xs font-bold bg-gray-100">
                          <div className="px-1">Type</div>
                          <div className="px-1">CP</div>
                          <div className="px-1 col-span-2">Amount/Info</div>
                        </div>
                        
                        {filteredAssets.map((asset, idx) => (
                          <div 
                            key={idx} 
                            className={`grid grid-cols-4 gap-1 p-1 text-xs border-b border-gray-100 hover:bg-blue-50 cursor-pointer
                              ${highlightCounterparty && linkedEntries.includes(`${asset.counterparty}-${asset.type}-${asset.amount}`) ? 'bg-yellow-100' : ''}
                            `}
                            onClick={(e) => {
                              e.stopPropagation();
                              setTargetRect(e.currentTarget.getBoundingClientRect());
                              setHoverEntry({...asset, agentName: agent.name, isAsset: true});
                              
                              if (highlightCounterparty && asset.counterparty) {
                                const entrySignature = `${agent.name}-${asset.type}-${asset.amount}`;
                                setLinkedEntries([entrySignature]);
                              }
                            }}
                          >
                            <div className="truncate px-1">{formatEntryType(asset.type, true)}</div>
                            <div className="truncate px-1">{asset.counterparty || 'N/A'}</div>
                            <div className="col-span-2 px-1">
                              <div>{asset.amount} {asset.denomination}</div>
                              {asset.current_book_value && asset.current_book_value !== asset.amount && (
                                <div className="text-gray-500 text-xs">Current: {asset.current_book_value.toFixed(2)}</div>
                              )}
                              {asset.cash_flow_at_maturity && asset.cash_flow_at_maturity !== asset.amount && (
                                <div className="text-blue-600 text-xs">CF@Mat: {asset.cash_flow_at_maturity.toFixed(2)}</div>
                              )}
                              <div className="text-gray-500 text-xs">{asset.maturity_type}, t{asset.issuance_time}</div>
                            </div>
                          </div>
                        ))}
                        
                        <div className="grid grid-cols-4 gap-1 p-1 text-xs font-bold mt-2 border-t border-gray-300">
                          <div className="px-1">Total:</div>
                          <div></div>
                          <div className="col-span-2 px-1">{totalAssets.toFixed(2)}</div>
                        </div>
                        
                      </div>
                      
                      {/* Liabilities Side */}
                      <div className="w-1/2">
                        <div className="p-2 font-bold text-center bg-red-100">LIABILITIES</div>
                        <div className="grid grid-cols-4 gap-1 p-1 text-xs font-bold bg-gray-100">
                          <div className="px-1">Type</div>
                          <div className="px-1">CP</div>
                          <div className="px-1 col-span-2">Amount/Info</div>
                        </div>
                        
                        {filteredLiabilities.map((liability, idx) => (
                          <div 
                            key={idx} 
                            className={`grid grid-cols-4 gap-1 p-1 text-xs border-b border-gray-100 hover:bg-red-50 cursor-pointer
                              ${highlightCounterparty && linkedEntries.includes(`${liability.counterparty}-${liability.type}-${liability.amount}`) ? 'bg-yellow-100' : ''}
                            `}
                            onClick={(e) => {
                              e.stopPropagation();
                              setTargetRect(e.currentTarget.getBoundingClientRect());
                              setHoverEntry({...liability, agentName: agent.name, isAsset: false});
                              
                              if (highlightCounterparty && liability.counterparty) {
                                const entrySignature = `${agent.name}-${liability.type}-${liability.amount}`;
                                setLinkedEntries([entrySignature]);
                              }
                            }}
                          >
                            <div className="truncate px-1">{formatEntryType(liability.type, false)}</div>
                            <div className="truncate px-1">{liability.counterparty}</div>
                            <div className="col-span-2 px-1">
                              <div>{liability.amount} {liability.denomination}</div>
                              {liability.current_book_value && liability.current_book_value !== liability.amount && (
                                <div className="text-gray-500 text-xs">Current: {liability.current_book_value.toFixed(2)}</div>
                              )}
                              {liability.cash_flow_at_maturity && liability.cash_flow_at_maturity !== liability.amount && (
                                <div className="text-blue-600 text-xs">CF@Mat: {liability.cash_flow_at_maturity.toFixed(2)}</div>
                              )}
                              <div className="text-gray-500 text-xs">{liability.maturity_type}, t{liability.issuance_time}</div>
                            </div>
                          </div>
                        ))}
                        
                        <div className="grid grid-cols-4 gap-1 p-1 text-xs font-bold mt-2 border-t border-gray-300">
                          <div className="px-1">Total:</div>
                          <div></div>
                          <div className="col-span-2 px-1">{totalLiabilities.toFixed(2)}</div>
                        </div>
                        
                        <div className="grid grid-cols-4 gap-1 p-1 text-xs font-bold bg-gray-200">
                          <div className="px-1">Net Worth:</div>
                          <div></div>
                          <div className="col-span-2 px-1">{netWorth.toFixed(2)}</div>
                        </div>
                      </div>
                    </div>
                  </div>
                );
              };
              
              const renderBalanceSheets = () => {
                if (loading) {
                  return <div className="text-center py-8">Loading balance sheet data...</div>;
                }
                
                let agentsToRender = [];
                
                if (scopeOption.startsWith('single-agent')) {
                  const timePointToUse = scopeOption.endsWith('single-time') ? 
                    selectedSingleTimePoint : timePoints[0];
                  const agentsAtTime = agentData[timePointToUse] || {};
                  
                  if (selectedAgentName && agentsAtTime[selectedAgentName]) {
                    agentsToRender = [agentsAtTime[selectedAgentName]];
                  } else if (Object.keys(agentsAtTime).length > 0) {
                    const firstAgentKey = Object.keys(agentsAtTime)[0];
                    agentsToRender = [agentsAtTime[firstAgentKey]];
                  }
                } else if (scopeOption.startsWith('all-agents')) {
                  const relevantTimePoints = scopeOption.endsWith('single-time')
                    ? [selectedSingleTimePoint]
                    : timePoints;
                
                  const allAgentNames = new Set();
                  relevantTimePoints.forEach(tp => {
                    if (agentData[tp]) {
                      Object.keys(agentData[tp]).forEach(name => allAgentNames.add(name));
                    }
                  });
                
                  agentsToRender = Array.from(allAgentNames);
                }    

                if (scopeOption.endsWith('single-time')) {
                  const singleTimePoint = selectedSingleTimePoint;
                  
                  return (
                    <div className="flex flex-row flex-nowrap overflow-x-auto -mx-2">
                      {agentsToRender.map(agent => {
                        if (typeof agent === 'string') {
                          const actualAgent = agentData[singleTimePoint]?.[agent];
                          if (!actualAgent) return null;
                          return (
                            <div
                              key={`${agent}-${singleTimePoint}`}
                              className="flex-shrink-0 p-2"
                              style={{ minWidth: '320px' }}
                            >
                              {renderTChart(actualAgent, singleTimePoint)}
                            </div>
                          );
                        }
                        
                        return (
                          <div
                            key={`${agent.name}-${singleTimePoint}`}
                            className="flex-shrink-0 p-2"
                            style={{ minWidth: '320px' }}
                          >
                            {renderTChart(agent, singleTimePoint)}
                          </div>
                        );
                      })}
                    </div>
                  );
                }
                
                if (mappingType === 'change-only' && timePoints.length >= 2) {
                  const changes = [];
                  
                  for (let i = 1; i < timePoints.length; i++) {
                    const fromTime = timePoints[i-1];
                    const toTime = timePoints[i];
                    
                    agentsToRender.forEach(agentNameOrObject => {
                      const agentName = typeof agentNameOrObject === 'string' ? 
                        agentNameOrObject : agentNameOrObject.name;
                      
                      const fromAgent = agentData[fromTime]?.[agentName];
                      const toAgent = agentData[toTime]?.[agentName];
                      
                      if (fromAgent && toAgent) {
                        changes.push({
                          fromTime,
                          toTime,
                          agent: toAgent,
                          changes: computeChanges(toAgent, fromTime, toTime)
                        });
                      }
                    });
                  }
                  
                  return (
                    <div className="flex flex-wrap">
                      {changes.map((change, idx) => (
                        <div key={idx} className="w-full mb-8">
                          <h2 className="text-xl font-bold mb-4 text-center bg-gray-100 py-2">
                            Changes: {change.fromTime} → {change.toTime} for {change.agent.name}
                          </h2>
                          
                          <div className="grid grid-cols-1 md:grid-cols-2 gap-4 p-4 bg-white rounded-lg shadow-md border border-gray-300">
                            <div>
                              <h3 className="font-bold mb-2">New Assets</h3>
                              <ul className="list-disc pl-5">
                                {change.changes.newAssets.map((asset, i) => (
                                  <li key={i} className="mb-1 text-sm">
                                    {formatEntryType(asset.type, true)}: {asset.amount} {asset.denomination} (from {asset.counterparty || 'N/A'})
                                  </li>
                                ))}
                                {change.changes.newAssets.length === 0 && <li className="text-gray-500">None</li>}
                              </ul>
                              
                              <h3 className="font-bold mt-4 mb-2">Removed Assets</h3>
                              <ul className="list-disc pl-5">
                                {change.changes.removedAssets.map((asset, i) => (
                                  <li key={i} className="mb-1 text-sm">
                                    {formatEntryType(asset.type, true)}: {asset.amount} {asset.denomination} (from {asset.counterparty || 'N/A'})
                                  </li>
                                ))}
                                {change.changes.removedAssets.length === 0 && <li className="text-gray-500">None</li>}
                              </ul>
                            </div>
                            
                            <div>
                              <h3 className="font-bold mb-2">New Liabilities</h3>
                              <h3 className="font-bold mb-2">New Assets</h3>
                              <ul className="list-disc pl-5">
                                {change.changes.newLiabilities.map((liability, i) => (
                                  <li key={i} className="mb-1 text-sm">
                                    {formatEntryType(liability.type, false)}: {liability.amount} {liability.denomination} (to {liability.counterparty})
                                  </li>
                                ))}
                                {change.changes.newLiabilities.length === 0 && <li className="text-gray-500">None</li>}
                              </ul>
                              
                              <h3 className="font-bold mt-4 mb-2">Removed Liabilities</h3>
                              <ul className="list-disc pl-5">
                                {change.changes.removedLiabilities.map((liability, i) => (
                                  <li key={i} className="mb-1 text-sm">
                                    {formatEntryType(liability.type, false)}: {liability.amount} {liability.denomination} (to {liability.counterparty})
                                  </li>
                                ))}
                                {change.changes.removedLiabilities.length === 0 && <li className="text-gray-500">None</li>}
                              </ul>
                            </div>
                          </div>
                        </div>
                      ))}
                    </div>
                  );
                }

                return (
                  <div className="flex flex-wrap">
                    {timePoints.map(timePoint => (
                      <div key={timePoint} className="w-full mb-8">
                        <h2 className="text-xl font-bold mb-4 text-center bg-gray-100 py-2">
                          Time Point: {timePoint}
                        </h2>
                        <div className="flex flex-row flex-nowrap overflow-x-auto -mx-2">
                          {agentsToRender.map(agentName => {
                            const agent = agentData[timePoint]?.[agentName];
                            if (!agent) return null;
                            return (
                              <div
                                key={`${agentName}-${timePoint}`}
                                className="flex-shrink-0 p-2"
                                style={{ minWidth: '320px' }}
                              >
                                {renderTChart(agent, timePoint)}
                              </div>
                            );
                          })}
                        </div>
                      </div>
                    ))}
                  </div>
                );
              };
              
              const handleClickAway = () => {
                setHoverEntry(null);
                setLinkedEntries([]);
                setTargetRect(null);
              };

              const showAgentSelection = scopeOption.startsWith('single-agent');
              const showTimePointSelection = scopeOption.endsWith('single-time');

              return (
                <div 
                  className="px-4 py-2 bg-gray-50 min-h-screen"
                  onClick={handleClickAway}
                >
                  <div className="max-w-full mx-auto px-4">
                    {hoverEntry && targetRect && (
                      <HoverDetails
                        entry={hoverEntry}
                        targetRect={targetRect}
                        onClose={handleClickAway}
                      />
                    )}
                    
                    <h1 className="text-2xl font-bold mb-6">Balance Sheet Output Interface</h1>
                    
                    <div className="bg-white p-4 rounded-lg shadow-md mb-8">
                      <div className="flex flex-wrap justify-between items-center mb-4">
                        <h2 className="text-lg font-bold">View Configuration</h2>
                        
                        <a 
                          href="/export" 
                          className="bg-green-600 hover:bg-green-700 text-white font-bold py-2 px-4 rounded inline-flex items-center"
                        >
                          Export to Excel
                        </a>
                      </div>
                      
                      <div className="grid grid-cols-1 md:grid-cols-2 gap-4">
                        <div>
                          <label className="block text-sm font-medium text-gray-700 mb-1">Scope:</label>
                          <select 
                            value={scopeOption} 
                            onChange={handleScopeChange}
                            className="block w-full py-2 px-3 border border-gray-300 bg-white rounded-md shadow-sm focus:outline-none focus:ring-blue-500 focus:border-blue-500 sm:text-sm mb-2"
                          >
                            <option value="single-agent-single-time">Single agent at single time point</option>
                            <option value="single-agent-across-time">Single agent across time</option>
                            <option value="all-agents-single-time">All agents at single time point</option>
                            <option value="all-agents-selected-times">All agents across selected times</option>
                          </select>
                          
                          {showAgentSelection && (
                            <div className="mt-2">
                              <label className="block text-sm font-medium text-gray-700 mb-1">Select Agent:</label>
                              <select 
                                value={selectedAgentName} 
                                onChange={(e) => setSelectedAgentName(e.target.value)}
                                className="block w-full py-2 px-3 border border-gray-300 bg-white rounded-md shadow-sm focus:outline-none focus:ring-blue-500 focus:border-blue-500 sm:text-sm"
                              >
                                {availableAgents.map(agentName => (
                                  <option key={agentName} value={agentName}>{agentName}</option>
                                ))}
                              </select>
                            </div>
                          )}
                          
                          {showTimePointSelection && (
                            <div className="mt-2">
                              <label className="block text-sm font-medium text-gray-700 mb-1">Select Time Point:</label>
                              <select 
                                value={selectedSingleTimePoint} 
                                onChange={handleSingleTimePointChange}
                                className="block w-full py-2 px-3 border border-gray-300 bg-white rounded-md shadow-sm focus:outline-none focus:ring-blue-500 focus:border-blue-500 sm:text-sm"
                              >
                                {availableTimePoints.map(timePoint => (
                                  <option key={timePoint} value={timePoint}>{timePoint}</option>
                                ))}
                              </select>
                            </div>
                          )}
                        </div>
                        
                        <div>
                          <label className="block text-sm font-medium text-gray-700 mb-1">Mapping Type:</label>
                          <div className="flex">
                            <button 
                              onClick={() => setMappingType('full')} 
                              className={`px-3 py-2 text-sm rounded-l ${mappingType === 'full' ? 'bg-blue-600 text-white' : 'bg-gray-200 text-gray-700'}`}
                            >
                              Full Balance Sheet
                            </button>
                            <button 
                              onClick={() => !isSingleTimeMode && setMappingType('change-only')}
                              disabled={isSingleTimeMode} 
                              className={`px-3 py-2 text-sm rounded-r ${
                                mappingType === 'change-only' ? 'bg-blue-600 text-white' : 
                                isSingleTimeMode ? 'bg-gray-100 text-gray-400 cursor-not-allowed' : 'bg-gray-200 text-gray-700'
                              }`}
                              title={isSingleTimeMode ? "Change Only view requires multiple time points" : ""}
                            >
                              Change Only
                            </button>
                          </div>
                          
                          {isSingleTimeMode && (
                            <p className="text-xs text-gray-500 mt-1">
                              "Change Only" view is not available in single time point mode
                            </p>
                          )}
                        </div>
                      </div>
                      
                      {!scopeOption.endsWith('single-time') && (
                        <div className="mt-4 border-t pt-4">
                          <label className="block text-sm font-medium text-gray-700 mb-2">Time Navigation:</label>
                          <div className="flex items-center gap-2">
                            <div className="flex">
                              {['t0', 't1', 't2'].map(point => {
                                const isSelected = timePoints.includes(point);
                                return (
                                  <button
                                    key={point}
                                    onClick={() => isSelected ? removeTimePoint(point) : addTimePoint(point)}
                                    className={`px-4 py-2 ${isSelected ? 'bg-blue-600 text-white' : 'bg-gray-200 text-gray-700'} ${point === 't0' ? 'rounded-l' : ''} ${point === 't2' ? 'rounded-r' : ''}`}
                                  >
                                    {point}
                                  </button>
                                );
                              })}
                            </div>
                            <button
                              onClick={expandToFullTimeline}
                              className="ml-4 px-3 py-2 bg-blue-100 text-blue-800 rounded hover:bg-blue-200"
                            >
                              Expand to Full Timeline
                            </button>
                          </div>
                        </div>
                      )}
                      
                      <div className="mt-4 border-t pt-4">
                        <label className="block text-sm font-medium text-gray-700 mb-2">Display Options:</label>
                        <div className="flex gap-4">
                          <label className="inline-flex items-center">
                            <input
                              type="checkbox"
                              checked={showLabels}
                              onChange={() => setShowLabels(!showLabels)}
                              className="h-4 w-4 text-blue-600 focus:ring-blue-500 border-gray-300 rounded"
                            />
                            <span className="ml-2 text-sm">Labels</span>
                          </label>
                          
                          <label className="inline-flex items-center">
                            <input
                              type="checkbox"
                              checked={highlightCounterparty}
                              onChange={() => setHighlightCounterparty(!highlightCounterparty)}
                              className="h-4 w-4 text-blue-600 focus:ring-blue-500 border-gray-300 rounded"
                            />
                            <span className="ml-2 text-sm">Highlight Counterparty</span>
                          </label>
                        </div>
                      </div>
                    </div>
                    
                    <div className="bg-white p-4 rounded-lg shadow-md border border-gray-300">
                      {renderBalanceSheets()}
                    </div>
                  </div>
                </div>
              );
            };

            ReactDOM.render(
              <BalanceSheetOutputInterface />,
              document.getElementById('balance-sheet-output-app')
            );
        </script>
    </body>
    </html>
    '''  
    return output_template

@app.route('/api/balance-sheet-data')
def get_balance_sheet_data():
    """Return JSON API for balance sheet data with enhanced current book values"""
    
    time_points = ['t0', 't1', 't2']
    
    agents_by_time = {}
    for i, time_point in enumerate(time_points):
        agents_original = system.get_agents_at_time(i)
        agents = {name: copy.deepcopy(agent) for name, agent in agents_original.items()}
        
        agents_data = {}
        for name, agent in agents.items():
            agents_data[name] = {
                'name': agent.name,
                'type': agent.type.value,
                'assets': [
                    {
                        'type': asset.type.value,
                        'is_asset': asset.is_asset,
                        'counterparty': asset.counterparty,
                        'amount': asset.initial_book_value,
                        'initial_book_value': asset.initial_book_value,
                        'current_book_value': asset.current_book_value,
                        'denomination': asset.denomination,
                        'maturity_type': asset.maturity_type.value,
                        'maturity_date': asset.maturity_date,
                        'settlement_type': asset.settlement_details.type.value,
                        'settlement_denomination': asset.settlement_details.denomination,
                        'name': asset.name,
                        'issuance_time': asset.issuance_time,
                        'cash_flow_at_maturity': asset.cash_flow_at_maturity,
                        'expected_cash_flow': asset.expected_cash_flow,
                        'parent_bond': asset.parent_bond,
                        'rollover_count': asset.rollover_count
                    }
                    for asset in agent.assets
                ],
                'liabilities': [
                    {
                        'type': liability.type.value,
                        'is_asset': liability.is_asset,
                        'counterparty': liability.counterparty,
                        'amount': liability.initial_book_value,
                        'initial_book_value': liability.initial_book_value,
                        'current_book_value': liability.current_book_value,
                        'denomination': liability.denomination,
                        'maturity_type': liability.maturity_type.value,
                        'maturity_date': liability.maturity_date,
                        'settlement_type': liability.settlement_details.type.value,
                        'settlement_denomination': liability.settlement_details.denomination,
                        'name': liability.name,
                        'issuance_time': liability.issuance_time,
                        'cash_flow_at_maturity': liability.cash_flow_at_maturity,
                        'expected_cash_flow': liability.expected_cash_flow,
                        'parent_bond': liability.parent_bond,
                        'rollover_count': liability.rollover_count
                    }
                    for liability in agent.liabilities
                ]
            }
        
        agents_by_time[time_point] = agents_data
    
    return jsonify({
        'time_points': time_points,
        'agents_by_time': agents_by_time
    })

class ExcelExporter:
    def __init__(self, system: EconomicSystem):
        self.system = system

    def create_t_table(self, sheet, row_start: int, col_start: int, agent: Agent, time_point: str):
        """Create a T-account format balance sheet for a single agent at a specific time point."""
        thick = Side(style='thick', color='000000')
        thin = Side(style='thin', color='000000')

        # Agent header with name and type
        name_cell = sheet.cell(row=row_start, column=col_start)
        name_cell.value = f"{agent.name} ({agent.type.value})"
        name_cell.alignment = Alignment(horizontal="center")
        name_cell.font = openpyxl.styles.Font(bold=True)
        sheet.merge_cells(start_row=row_start, start_column=col_start,
                          end_row=row_start, end_column=col_start + 9)

        # Time point header
        time_header = sheet.cell(row=row_start + 1, column=col_start)
        time_header.value = f"Time: {time_point}"
        time_header.alignment = Alignment(horizontal="center")
        time_header.font = openpyxl.styles.Font(bold=True)
        sheet.merge_cells(start_row=row_start + 1, start_column=col_start,
                          end_row=row_start + 1, end_column=col_start + 9)

        # Add borders to the T-chart
        for i in range(10): # Covers 5 columns for assets, 5 for liabilities
            cell = sheet.cell(row=row_start + 2, column=col_start + i)
            cell.border = Border(top=thick)

        # Add vertical border between assets and liabilities
        for r in range(row_start + 2, row_start + 20):
            cell = sheet.cell(row=r, column=col_start + 4)
            if r == row_start + 2:
                cell.border = Border(right=thick, top=thick)
            else:
                cell.border = Border(right=thick)
        
        # Add headers for assets and liabilities
        asset_header = sheet.cell(row=row_start + 2, column=col_start)
        asset_header.value = "ASSETS"
        asset_header.font = openpyxl.styles.Font(bold=True)
        asset_header.alignment = Alignment(horizontal="center")
        sheet.merge_cells(start_row=row_start + 2, start_column=col_start,
                          end_row=row_start + 2, end_column=col_start + 3)

        liability_header = sheet.cell(row=row_start + 2, column=col_start + 5)
        liability_header.value = "LIABILITIES"
        liability_header.font = openpyxl.styles.Font(bold=True)
        liability_header.alignment = Alignment(horizontal="center")
        sheet.merge_cells(start_row=row_start + 2, start_column=col_start + 5,
                          end_row=row_start + 2, end_column=col_start + 9)

        # Headers for columns
        headers = ['Type', 'CP', 'Amount', 'Current', 'CF@Mat']
        for i, header in enumerate(headers):
            row = row_start + 3
            # Asset headers
            cell_asset = sheet.cell(row=row, column=col_start + i)
            cell_asset.value = header
            cell_asset.alignment = Alignment(horizontal="center")
            cell_asset.font = openpyxl.styles.Font(bold=True)
            cell_asset.border = Border(bottom=thin)
            if i == 4:
                cell_asset.border = Border(bottom=thin, right=thick)

            # Liability headers
            cell_liability = sheet.cell(row=row, column=col_start + 5 + i)
            cell_liability.value = header
            cell_liability.alignment = Alignment(horizontal="center")
            cell_liability.font = openpyxl.styles.Font(bold=True)
            cell_liability.border = Border(bottom=thin)

        current_row_asset = row_start + 4
        current_row_liability = row_start + 4
        
        time_point_int = int(time_point[1:]) if time_point.startswith('t') else 0
        
        # Filter assets for the current time point
        for entry in agent.assets:
            if entry.issuance_time > time_point_int:
                continue
            if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date is not None and time_point_int > entry.maturity_date:
                continue

            entry_type_display = entry.type.value
            if entry.type == EntryType.PAYABLE:
                entry_type_display = "receivable"
            elif entry.type == EntryType.NON_FINANCIAL and entry.name:
                entry_type_display = f"{entry.type.value} ({entry.name})"
            elif entry.type == EntryType.DELIVERY_CLAIM and entry.name:
                entry_type_display = f"delivery claim for {entry.name}"
            elif entry.type == EntryType.DEFAULT:
                entry_type_display = f"default claim ({entry.name})"
            
            sheet.cell(row=current_row_asset, column=col_start).value = entry_type_display
            sheet.cell(row=current_row_asset, column=col_start + 1).value = entry.counterparty if entry.counterparty else "N/A"
            sheet.cell(row=current_row_asset, column=col_start + 2).value = f"{entry.initial_book_value} {entry.denomination}"
            sheet.cell(row=current_row_asset, column=col_start + 3).value = f"{entry.current_book_value:.2f}"
            sheet.cell(row=current_row_asset, column=col_start + 4).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
            
            # Add right border for last asset column
            cell = sheet.cell(row=current_row_asset, column=col_start + 4)
            cell.border = Border(right=thick)
            
            current_row_asset += 1

        # Filter liabilities for the current time point
        for entry in agent.liabilities:
            if entry.issuance_time > time_point_int:
                continue
            if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date is not None and time_point_int > entry.maturity_date:
                continue

            entry_type_display = entry.type.value
            if entry.type == EntryType.DELIVERY_CLAIM:
                entry_type_display = f"delivery promise for {entry.name}" if entry.name else "delivery promise"
            elif entry.type == EntryType.DEFAULT:
                entry_type_display = f"default liability ({entry.name})"

            sheet.cell(row=current_row_liability, column=col_start + 5).value = entry_type_display
            sheet.cell(row=current_row_liability, column=col_start + 6).value = entry.counterparty
            sheet.cell(row=current_row_liability, column=col_start + 7).value = f"{entry.initial_book_value} {entry.denomination}"
            sheet.cell(row=current_row_liability, column=col_start + 8).value = f"{entry.current_book_value:.2f}"
            sheet.cell(row=current_row_liability, column=col_start + 9).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
            
            current_row_liability += 1

        max_entries_row = max(current_row_asset, current_row_liability)
        total_row = max(max_entries_row, row_start + 12) 

        for i in range(5): # Asset columns
            sheet.cell(row=total_row, column=col_start + i).border = Border(top=thin)
        sheet.cell(row=total_row, column=col_start + 4).border = Border(top=thin, right=thick)

        for i in range(5): # Liability columns
            sheet.cell(row=total_row, column=col_start + 5 + i).border = Border(top=thin)
        
        sheet.cell(row=total_row, column=col_start).value = "Total Assets:"
        sheet.cell(row=total_row, column=col_start).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=total_row, column=col_start + 2).value = agent.get_total_assets()
        sheet.cell(row=total_row, column=col_start + 2).font = openpyxl.styles.Font(bold=True)
        
        sheet.cell(row=total_row, column=col_start + 5).value = "Total Liabilities:"
        sheet.cell(row=total_row, column=col_start + 5).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=total_row, column=col_start + 7).value = agent.get_total_liabilities()
        sheet.cell(row=total_row, column=col_start + 7).font = openpyxl.styles.Font(bold=True)
        
        sheet.cell(row=total_row + 1, column=col_start).value = "Net Worth:"
        sheet.cell(row=total_row + 1, column=col_start).font = openpyxl.styles.Font(bold=True)
        sheet.cell(row=total_row + 1, column=col_start + 2).value = agent.get_net_worth()
        sheet.cell(row=total_row + 1, column=col_start + 2).font = openpyxl.styles.Font(bold=True)
        
        sheet.cell(row=total_row + 1, column=col_start + 4).border = Border(right=thick)
        
        return total_row + 3

    def create_changes_sheet(self, wb):
        """Create a sheet showing changes between time points."""
        sheet = wb.create_sheet(title="Changes Between Time Points")
        
        # Column headers
        headers = ["Agent", "Time Change", "Changes", "Type", "Amount", "Denomination", 
                   "Counterparty", "Entry Type", "Maturity", "Settlement", "Current Value", "CF@Mat"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style='thick'))
            
        current_row = 2
        
        # Get time transitions - use system's time points
        time_points = self.system.get_time_points()
        transitions = []
        for i in range(len(time_points) - 1):
            transitions.append((time_points[i], time_points[i + 1]))
        
        # Process each transition
        for from_time, to_time in transitions:
            # Skip if we don't have both time states
            if from_time not in self.system.time_states or to_time not in self.system.time_states:
                continue
                
            # Get changes for this transition
            changes = self.system.compute_changes(from_time, to_time)
            
            # Add changes for each agent
            for agent_name, agent_changes in changes.items():
                # New assets
                for entry in agent_changes['new_assets']:
                    sheet.cell(row=current_row, column=1).value = agent_name
                    sheet.cell(row=current_row, column=2).value = f"t{from_time} → t{to_time}"
                    sheet.cell(row=current_row, column=3).value = "New Asset"
                    sheet.cell(row=current_row, column=4).value = "Asset"
                    sheet.cell(row=current_row, column=5).value = entry.initial_book_value
                    sheet.cell(row=current_row, column=6).value = entry.denomination
                    sheet.cell(row=current_row, column=7).value = entry.counterparty
                    sheet.cell(row=current_row, column=8).value = entry.type.value
                    
                    maturity = entry.maturity_type.value
                    if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date:
                        maturity = f"t{entry.maturity_date}"
                    sheet.cell(row=current_row, column=9).value = maturity
                    
                    settlement = entry.settlement_details.type.value
                    if settlement != "none":
                        settlement += f" ({entry.settlement_details.denomination})"
                    sheet.cell(row=current_row, column=10).value = settlement
                    sheet.cell(row=current_row, column=11).value = f"{entry.current_book_value:.2f}"
                    sheet.cell(row=current_row, column=12).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
                    
                    current_row += 1
                
                # Removed assets
                for entry in agent_changes['removed_assets']:
                    sheet.cell(row=current_row, column=1).value = agent_name
                    sheet.cell(row=current_row, column=2).value = f"t{from_time} → t{to_time}"
                    sheet.cell(row=current_row, column=3).value = "Removed Asset"
                    sheet.cell(row=current_row, column=4).value = "Asset"
                    sheet.cell(row=current_row, column=5).value = entry.initial_book_value
                    sheet.cell(row=current_row, column=6).value = entry.denomination
                    sheet.cell(row=current_row, column=7).value = entry.counterparty
                    sheet.cell(row=current_row, column=8).value = entry.type.value
                    
                    maturity = entry.maturity_type.value
                    if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date:
                        maturity = f"t{entry.maturity_date}"
                    sheet.cell(row=current_row, column=9).value = maturity
                    
                    settlement = entry.settlement_details.type.value
                    if settlement != "none":
                        settlement += f" ({entry.settlement_details.denomination})"
                    sheet.cell(row=current_row, column=10).value = settlement
                    sheet.cell(row=current_row, column=11).value = f"{entry.current_book_value:.2f}"
                    sheet.cell(row=current_row, column=12).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
                    
                    current_row += 1
                
                # New liabilities
                for entry in agent_changes['new_liabilities']:
                    sheet.cell(row=current_row, column=1).value = agent_name
                    sheet.cell(row=current_row, column=2).value = f"t{from_time} → t{to_time}"
                    sheet.cell(row=current_row, column=3).value = "New Liability"
                    sheet.cell(row=current_row, column=4).value = "Liability"
                    sheet.cell(row=current_row, column=5).value = entry.initial_book_value
                    sheet.cell(row=current_row, column=6).value = entry.denomination
                    sheet.cell(row=current_row, column=7).value = entry.counterparty
                    sheet.cell(row=current_row, column=8).value = entry.type.value
                    
                    maturity = entry.maturity_type.value
                    if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date:
                        maturity = f"t{entry.maturity_date}"
                    sheet.cell(row=current_row, column=9).value = maturity
                    
                    settlement = entry.settlement_details.type.value
                    if settlement != "none":
                        settlement += f" ({entry.settlement_details.denomination})"
                    sheet.cell(row=current_row, column=10).value = settlement
                    sheet.cell(row=current_row, column=11).value = f"{entry.current_book_value:.2f}"
                    sheet.cell(row=current_row, column=12).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
                    
                    current_row += 1
                
                # Removed liabilities
                for entry in agent_changes['removed_liabilities']:
                    sheet.cell(row=current_row, column=1).value = agent_name
                    sheet.cell(row=current_row, column=2).value = f"t{from_time} → t{to_time}"
                    sheet.cell(row=current_row, column=3).value = "Removed Liability"
                    sheet.cell(row=current_row, column=4).value = "Liability"
                    sheet.cell(row=current_row, column=5).value = entry.initial_book_value
                    sheet.cell(row=current_row, column=6).value = entry.denomination
                    sheet.cell(row=current_row, column=7).value = entry.counterparty
                    sheet.cell(row=current_row, column=8).value = entry.type.value
                    
                    maturity = entry.maturity_type.value
                    if entry.maturity_type == MaturityType.FIXED_DATE and entry.maturity_date:
                        maturity = f"t{entry.maturity_date}"
                    sheet.cell(row=current_row, column=9).value = maturity
                    
                    settlement = entry.settlement_details.type.value
                    if settlement != "none":
                        settlement += f" ({entry.settlement_details.denomination})"
                    sheet.cell(row=current_row, column=10).value = settlement
                    sheet.cell(row=current_row, column=11).value = f"{entry.current_book_value:.2f}"
                    sheet.cell(row=current_row, column=12).value = f"{entry.cash_flow_at_maturity:.2f}" if entry.cash_flow_at_maturity else "N/A"
                    
                    current_row += 1
        
        # Format the sheet
        for col in range(1, 13):
            sheet.column_dimensions[get_column_letter(col)].width = 18
            
        return sheet

    def create_settlement_history_sheet(self, wb):
        """Create a sheet showing settlement history."""
        sheet = wb.create_sheet(title="Settlement History")
        
        # Column headers
        headers = ["Agent", "Role", "Time Point", "Original Entry Type", "Original Amount", 
                  "Original Denomination", "Original Current Value", "Original CF@Mat",
                  "Settled With/For Type", "Settled Amount", "Settled Denomination", 
                  "Settled Current Value", "Settled CF@Mat", "Counterparty"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style='thick'))
            
        current_row = 2
        
        # Process settlement history for each agent
        for agent_name, agent in self.system.agents.items():
            # Process asset holder settlements
            for settlement in agent.settlement_history.get('as_asset_holder', []):
                sheet.cell(row=current_row, column=1).value = agent_name
                sheet.cell(row=current_row, column=2).value = "Asset Holder"
                sheet.cell(row=current_row, column=3).value = f"t{settlement['time_point']}"
                sheet.cell(row=current_row, column=4).value = settlement['original_entry'].type.value
                sheet.cell(row=current_row, column=5).value = settlement['original_entry'].initial_book_value
                sheet.cell(row=current_row, column=6).value = settlement['original_entry'].denomination
                sheet.cell(row=current_row, column=7).value = f"{settlement['original_entry'].current_book_value:.2f}"
                sheet.cell(row=current_row, column=8).value = f"{settlement['original_entry'].cash_flow_at_maturity:.2f}" if settlement['original_entry'].cash_flow_at_maturity else "N/A"
                sheet.cell(row=current_row, column=9).value = settlement['settlement_result'].type.value
                sheet.cell(row=current_row, column=10).value = settlement['settlement_result'].initial_book_value
                sheet.cell(row=current_row, column=11).value = settlement['settlement_result'].denomination
                sheet.cell(row=current_row, column=12).value = f"{settlement['settlement_result'].current_book_value:.2f}"
                sheet.cell(row=current_row, column=13).value = f"{settlement['settlement_result'].cash_flow_at_maturity:.2f}" if settlement['settlement_result'].cash_flow_at_maturity else "N/A"
                sheet.cell(row=current_row, column=14).value = settlement['counterparty']
                current_row += 1
                
            # Process liability holder settlements
            for settlement in agent.settlement_history.get('as_liability_holder', []):
                sheet.cell(row=current_row, column=1).value = agent_name
                sheet.cell(row=current_row, column=2).value = "Liability Holder"
                sheet.cell(row=current_row, column=3).value = f"t{settlement['time_point']}"
                sheet.cell(row=current_row, column=4).value = settlement['original_entry'].type.value
                sheet.cell(row=current_row, column=5).value = settlement['original_entry'].initial_book_value
                sheet.cell(row=current_row, column=6).value = settlement['original_entry'].denomination
                sheet.cell(row=current_row, column=7).value = f"{settlement['original_entry'].current_book_value:.2f}"
                sheet.cell(row=current_row, column=8).value = f"{settlement['original_entry'].cash_flow_at_maturity:.2f}" if settlement['original_entry'].cash_flow_at_maturity else "N/A"
                sheet.cell(row=current_row, column=9).value = settlement['settlement_result'].type.value
                sheet.cell(row=current_row, column=10).value = settlement['settlement_result'].initial_book_value
                sheet.cell(row=current_row, column=11).value = settlement['settlement_result'].denomination
                sheet.cell(row=current_row, column=12).value = f"{settlement['settlement_result'].current_book_value:.2f}"
                sheet.cell(row=current_row, column=13).value = f"{settlement['settlement_result'].cash_flow_at_maturity:.2f}" if settlement['settlement_result'].cash_flow_at_maturity else "N/A"
                sheet.cell(row=current_row, column=14).value = settlement['counterparty']
                current_row += 1
        
        # Format the sheet
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 18
            
        return sheet

    def create_system_metrics_sheet(self, wb):
        """Create a sheet with system-wide metrics across time points."""
        sheet = wb.create_sheet(title="System Metrics")
        
        # Column headers for time points
        sheet.cell(row=1, column=1).value = "Metric"
        sheet.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True)
        
        time_points = self.system.get_time_points()
        for i, tp in enumerate(time_points, 1):
            cell = sheet.cell(row=1, column=i+1)
            cell.value = f"t{tp}"
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Add system-wide metrics
        metrics = [
            "Total Assets", 
            "Total Liabilities",
            "Total Net Worth",
            "Number of Agents",
            "Number of Asset-Liability Pairs",
            "Number of Scheduled Actions"
        ]
        
        for i, metric in enumerate(metrics, 1):
            sheet.cell(row=i+1, column=1).value = metric
            sheet.cell(row=i+1, column=1).font = openpyxl.styles.Font(bold=True)
            
        # Populate metrics for each time point
        for j, tp in enumerate(time_points, 1):
            agents = (self.system.time_states[tp].values() 
                    if tp in self.system.time_states 
                    else self.system.agents.values())
            
            # Total assets
            total_assets = sum(agent.get_total_assets() for agent in agents)
            sheet.cell(row=2, column=j+1).value = total_assets
            
            # Total liabilities
            total_liabilities = sum(agent.get_total_liabilities() for agent in agents)
            sheet.cell(row=3, column=j+1).value = total_liabilities
            
            # Total net worth
            total_net_worth = sum(agent.get_net_worth() for agent in agents)
            sheet.cell(row=4, column=j+1).value = total_net_worth
            
            # Number of agents
            sheet.cell(row=5, column=j+1).value = len(agents)
            
            # Number of asset-liability pairs (only for t0)
            if j == 1:  # Only for t0
                sheet.cell(row=6, column=j+1).value = len(self.system.asset_liability_pairs)
            
            # Number of scheduled actions
            if j == 1:  # Only for t0
                scheduled_count = sum(len(actions) for actions in self.system.scheduled_actions.values())
                sheet.cell(row=7, column=j+1).value = scheduled_count
        
        # Format the sheet
        for col in range(1, len(time_points) + 2):
            sheet.column_dimensions[get_column_letter(col)].width = 18
            
        return sheet

    def create_agent_summaries_sheet(self, wb):
        """Create a sheet with agent summaries across all time points."""
        sheet = wb.create_sheet(title="Agent Summaries")
        
        # Headers
        headers = ["Agent", "Type", "Time Point", "Total Assets", "Total Liabilities", "Net Worth", 
                   "Asset Count", "Liability Count", "Status"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style='thick'))
            
        current_row = 2
        
        # Process each agent at each time point
        time_points = self.system.get_time_points()
        for tp in time_points:
            agents = (self.system.time_states[tp].values() 
                    if tp in self.system.time_states 
                    else self.system.agents.values())
                    
            for agent in agents:
                sheet.cell(row=current_row, column=1).value = agent.name
                sheet.cell(row=current_row, column=2).value = agent.type.value
                sheet.cell(row=current_row, column=3).value = f"t{tp}"
                sheet.cell(row=current_row, column=4).value = agent.get_total_assets()
                sheet.cell(row=current_row, column=5).value = agent.get_total_liabilities()
                sheet.cell(row=current_row, column=6).value = agent.get_net_worth()
                sheet.cell(row=current_row, column=7).value = len(agent.assets)
                sheet.cell(row=current_row, column=8).value = len(agent.liabilities)
                sheet.cell(row=current_row, column=9).value = getattr(agent, 'status', 'operating')
                current_row += 1
        
        # Format the sheet
        for col in range(1, 10):
            sheet.column_dimensions[get_column_letter(col)].width = 18
            
        return sheet

    def create_scheduled_actions_sheet(self, wb):
        """Create a sheet showing scheduled actions."""
        sheet = wb.create_sheet(title="Scheduled Actions")
        
        # Column headers
        headers = ["Time Point", "Action Type", "Asset Holder", "Liability Holder", "Entry Type", 
                   "Amount", "Denomination", "Maturity Type", "Maturity Date", "Settlement Type", 
                   "Asset Name", "Bond Type", "Coupon Rate", "Cash Flow @ Maturity"]
        
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style='thick'))
            
        current_row = 2
        
        # Process scheduled actions
        for time_point, actions in self.system.scheduled_actions.items():
            for action in actions:
                if action['type'] == 'create_asset_liability_pair':
                    data = action['data']
                    sheet.cell(row=current_row, column=1).value = f"t{time_point}"
                    sheet.cell(row=current_row, column=2).value = action['type']
                    sheet.cell(row=current_row, column=3).value = data.get('asset_holder_name', '')
                    sheet.cell(row=current_row, column=4).value = data.get('liability_holder_name', '')
                    sheet.cell(row=current_row, column=5).value = data.get('type', '')
                    sheet.cell(row=current_row, column=6).value = data.get('amount', 0)
                    sheet.cell(row=current_row, column=7).value = data.get('denomination', '')
                    sheet.cell(row=current_row, column=8).value = data.get('maturity_type', '')
                    sheet.cell(row=current_row, column=9).value = data.get('maturity_date', '')
                    sheet.cell(row=current_row, column=10).value = data.get('settlement_type', '')
                    sheet.cell(row=current_row, column=11).value = data.get('asset_name', '')
                    sheet.cell(row=current_row, column=12).value = data.get('bond_type', '')
                    sheet.cell(row=current_row, column=13).value = data.get('coupon_rate', '')
                    sheet.cell(row=current_row, column=14).value = data.get('cash_flow_at_maturity', '')
                    current_row += 1
        
        # Format the sheet
        for col in range(1, 15):
            sheet.column_dimensions[get_column_letter(col)].width = 18
            
        return sheet

    def export_balance_sheets(self, output_stream: BytesIO):
        """Export all data to an Excel workbook."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Balance Sheets"
        
        current_row_global = 1
        time_points_str_display = ['t0', 't1', 't2']
        time_points_int_fetch = [0, 1, 2]

        for idx, time_point_int in enumerate(time_points_int_fetch):
            time_point_str = time_points_str_display[idx]
            
            # Get agents at this time point, handle potential empty states
            try:
                agents = list(self.system.get_agents_at_time(time_point_int).values())
            except Exception as e:
                print(f"Warning: Could not get agents for time point {time_point_int}: {e}")
                continue

            if idx > 0:
                current_row_global +=1
                sheet.cell(row=current_row_global, column=1).value = "'" + "=" * 100
                current_row_global += 1
            
            time_header_row = current_row_global
            time_header = sheet.cell(row=time_header_row, column=1)
            time_header.value = f"Time Point: {time_point_str}"
            time_header.font = openpyxl.styles.Font(bold=True, size=16)
            sheet.merge_cells(start_row=time_header_row, start_column=1, end_row=time_header_row, end_column=10)
            current_row_global += 2

            col_start_agent = 1
            row_for_this_timepoint_start = current_row_global
            max_row_for_agents_in_timepoint = row_for_this_timepoint_start

            for agent_idx, agent in enumerate(agents):
                if agent_idx > 0:
                    col_start_agent += 11
                
                # Move to next row if too many agents horizontally
                if col_start_agent > 3 * 11:
                    current_row_global = max_row_for_agents_in_timepoint + 2
                    row_for_this_timepoint_start = current_row_global
                    max_row_for_agents_in_timepoint = row_for_this_timepoint_start
                    col_start_agent = 1

                agent_end_row = self.create_t_table(sheet, row_for_this_timepoint_start, col_start_agent, agent, time_point_str)
                max_row_for_agents_in_timepoint = max(max_row_for_agents_in_timepoint, agent_end_row)
            
            current_row_global = max_row_for_agents_in_timepoint + 1

            # System Totals for this time_point
            sheet.cell(row=current_row_global, column=1).value = f"System Totals at {time_point_str}:"
            sheet.cell(row=current_row_global, column=1).font = openpyxl.styles.Font(bold=True, size=12)
            current_row_global += 1
            
            sheet.cell(row=current_row_global, column=1).value = "Total Assets:"
            sheet.cell(row=current_row_global, column=2).value = sum(a.get_total_assets() for a in agents)
            current_row_global += 1
            
            sheet.cell(row=current_row_global, column=1).value = "Total Liabilities:"
            sheet.cell(row=current_row_global, column=2).value = sum(a.get_total_liabilities() for a in agents)
            current_row_global += 1
            
            sheet.cell(row=current_row_global, column=1).value = "Total Net Worth:"
            sheet.cell(row=current_row_global, column=2).value = sum(a.get_net_worth() for a in agents)
            current_row_global += 2

        # Add additional sheets with error handling
        try:
            self.create_changes_sheet(wb)
        except Exception as e:
            print(f"Warning: Could not create changes sheet: {e}")
            
        try:
            self.create_settlement_history_sheet(wb)
        except Exception as e:
            print(f"Warning: Could not create settlement history sheet: {e}")
            
        try:
            self.create_system_metrics_sheet(wb)
        except Exception as e:
            print(f"Warning: Could not create system metrics sheet: {e}")
            
        try:
            self.create_agent_summaries_sheet(wb)
        except Exception as e:
            print(f"Warning: Could not create agent summaries sheet: {e}")
            
        try:
            self.create_scheduled_actions_sheet(wb)
        except Exception as e:
            print(f"Warning: Could not create scheduled actions sheet: {e}")

        # Auto-size columns for better readability
        try:
            for column_cells in sheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)
        except Exception as e:
            print(f"Warning: Could not auto-size columns: {e}")

        wb.save(output_stream)

# ======== 运行应用 ========
if __name__ == '__main__':
    app.run(debug=True, port=5000)
